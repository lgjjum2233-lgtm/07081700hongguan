from __future__ import annotations

import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "（6月11日V11）TREE宏观分析_百分号格式修正.xlsx"
OUTPUT = ROOT / "（6月11日V12）TREE宏观分析_新增重点跟踪项.xlsx"
AUDIT = ROOT / "20260611_TREE重点跟踪项复核.xlsx"
SHEET = "重点策略跟踪情况(V3.0)"

HEADER_ROW = 5
INSERT_COL = 17  # after P: 环比; before Q: 数据折线图

COL_BIG = 1
COL_DIM = 2
COL_SUB = 3
COL_NAME = 4
COL_FREQ = 9
COL_CODE = 10
COL_CALIBER = 11
COL_QOQ = 16
COL_FOCUS = 17


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def compact(value: Any) -> str:
    return text(value).replace("\n", "").replace(" ", "")


def has_any(haystack: str, words: list[str]) -> bool:
    return any(word in haystack for word in words)


def valid_code(code: str) -> bool:
    return bool(code) and code not in {"—", "-", "无", "None"}


def is_macro_row(big: str, dim: str) -> bool:
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    # 延续前面用户确认过的处理边界：中国产业和资本市场不进入宏观数据着色/跟踪项判断。
    if big == "中国基本面" and dim == "产业":
        return False
    return True


def decide_focus(
    big: str,
    dim: str,
    sub: str,
    name: str,
    freq: str,
    code: str,
    caliber: str,
) -> tuple[str, str]:
    if not is_macro_row(big, dim):
        return "-", "非中国/美国宏观基本面，或属于中国产业/资本市场，暂不纳入本列判断。"
    if not valid_code(code):
        return "-", "定性指标或暂无有效底层数据代码。"

    t = compact(f"{big} {dim} {sub} {name} {freq} {code} {caliber}")

    if "环比" in t:
        return "环比", "名称或官方口径已经是环比，优先沿用该指标自身的主读数。"

    rate_words = [
        "利率",
        "收益率",
        "利差",
        "SHIBOR",
        "DR001",
        "DR007",
        "R001",
        "R007",
        "SOFR",
        "IORB",
        "LPR",
        "存款准备金率",
    ]
    if has_any(t, rate_words):
        return "环比", "利率、收益率、利差类指标，市场通常关注当期水平相对上期的边际变化。"

    liquidity_balance_words = ["美联储资产负债表", "TGA余额", "隔夜逆回购规模"]
    if has_any(t, liquidity_balance_words):
        return "环比", "美国流动性余额类高频指标，投研中更常观察周度/日度边际抽水或放水。"

    omo_words = ["公开市场操作", "净投放", "总投放", "常备借贷便利", "中期借贷便利", "抵押补充贷款", "其他政策工具"]
    if has_any(t, omo_words):
        return "环比", "央行操作和净投放类指标本身就是边际流量，更适合看上期到本期的变化。"

    credit_flow_words = ["社会融资", "新增人民币贷款", "政府债券", "票据融资", "企业债券融资"]
    if has_any(t, credit_flow_words):
        return "同比增量", "信用流量类指标市场常用同比多增/少增判断信用扩张强弱。"

    balance_diff_words = ["银行结售汇差额", "贸易差额", "差额/净额", "净流入"]
    if has_any(t, balance_diff_words):
        return "同比增量", "差额和净流量容易受季节性影响，绝对同比增减更利于判断真实边际方向。"

    diffusion_words = ["PMI", "ISM", "消费者信心", "乐观指数", "扩散指数", "新订单"]
    if has_any(t, diffusion_words):
        return "环比", "景气指数和扩散指数没有稳定同比含义，通常看最新读数较上期改善或走弱。"

    high_frequency_level_words = ["猪肉", "布伦特原油", "农产品批发价格指数"]
    if has_any(t, high_frequency_level_words):
        return "环比", "高频价格水平更适合观察短期边际变化，便于捕捉通胀压力的最新方向。"

    point_change_words = ["杠杆率", "财政赤字脉冲", "赤字率", "贡献率", "利润率"]
    if has_any(t, point_change_words):
        return "同比增量", "比例、脉冲和贡献率类指标用百分点变化更直观，避免同比率被低基数扭曲。"

    labor_rate_words = ["失业率", "非农数据"]
    if has_any(t, labor_rate_words):
        return "环比", "就业类月度数据市场更关注本期相对上期的边际变化和劳动力市场拐点。"

    published_yoy_words = ["同比", "累计同比", "当月同比", "增速"]
    if has_any(t, published_yoy_words):
        return "同比", "该指标的官方或主流投研口径本身就是同比增速，优先看同比读数。"

    amount_growth_words = [
        "累计值",
        "当月值",
        "规模现值",
        "名义现值",
        "资产规模",
        "余额/存量",
        "水平值/现值",
        "财政收入",
        "财政支出",
        "财政存款",
        "GDP总量",
        "社会消费品零售总额",
        "固定资产投资累计值",
        "银行信贷",
        "M1货币供应量",
        "M2货币供应量",
        "企业部门工商业贷款",
        "居民消费贷款",
        "居民房地产贷款",
    ]
    if has_any(t, amount_growth_words):
        return "同比", "金额、余额和规模类指标通常要剔除季节性，市场更常用同比增速判断趋势。"

    return "同比", "默认按宏观增长类指标处理，优先观察同比趋势。"


def copy_style(source, target) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy.copy(source.alignment)
    if source.font:
        target.font = copy.copy(source.font)
    if source.fill:
        target.fill = copy.copy(source.fill)
    if source.border:
        target.border = copy.copy(source.border)
    if source.protection:
        target.protection = copy.copy(source.protection)


def move_charts_right(ws, insert_col: int) -> None:
    zero_col = insert_col - 1
    for chart in ws._charts:
        anchor = chart.anchor
        if not hasattr(anchor, "_from"):
            continue
        if anchor._from.col >= zero_col:
            anchor._from.col += 1
        if hasattr(anchor, "to") and anchor.to.col >= zero_col:
            anchor.to.col += 1


def collect_rows(ws):
    rows = []
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, COL_BIG).value):
            big = text(ws.cell(row, COL_BIG).value)
        if text(ws.cell(row, COL_DIM).value):
            dim = text(ws.cell(row, COL_DIM).value)
        if text(ws.cell(row, COL_SUB).value):
            sub = text(ws.cell(row, COL_SUB).value)

        name = text(ws.cell(row, COL_NAME).value)
        if not name:
            continue
        rows.append(
            {
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "freq": text(ws.cell(row, COL_FREQ).value),
                "code": text(ws.cell(row, COL_CODE).value),
                "caliber": text(ws.cell(row, COL_CALIBER).value),
            }
        )
    return rows


def build_audit(rows: list[dict[str, str]], decisions: dict[int, tuple[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重点跟踪项判断"
    headers = ["TREE行号", "一级分类", "维度", "子维度", "指标名称", "发布频率", "指标代码", "数据口径", "重点跟踪项", "判断依据"]
    ws.append(headers)
    for row in rows:
        focus, reason = decisions[row["row"]]
        ws.append(
            [
                row["row"],
                row["big"],
                row["dim"],
                row["sub"],
                row["name"],
                row["freq"],
                row["code"],
                row["caliber"],
                focus,
                reason,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [10, 14, 14, 20, 34, 10, 24, 16, 12, 58]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(AUDIT)


def main() -> None:
    wb = openpyxl.load_workbook(INPUT)
    ws = wb[SHEET]
    rows = collect_rows(ws)
    decisions: dict[int, tuple[str, str]] = {}
    for row in rows:
        decisions[row["row"]] = decide_focus(
            row["big"],
            row["dim"],
            row["sub"],
            row["name"],
            row["freq"],
            row["code"],
            row["caliber"],
        )

    ws.insert_cols(INSERT_COL)
    move_charts_right(ws, INSERT_COL)

    # Copy the visual style of the adjacent 环比 column into the new text column.
    ws.column_dimensions[get_column_letter(COL_FOCUS)].width = 13
    for row in range(HEADER_ROW, ws.max_row + 1):
        copy_style(ws.cell(row, COL_QOQ), ws.cell(row, COL_FOCUS))
        ws.cell(row, COL_FOCUS).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if row > HEADER_ROW:
            ws.cell(row, COL_FOCUS).fill = PatternFill(fill_type=None)

    ws.cell(HEADER_ROW, COL_FOCUS).value = "重点跟踪项"
    ws.cell(HEADER_ROW, COL_FOCUS).font = copy.copy(ws.cell(HEADER_ROW, COL_QOQ).font)

    for row in rows:
        focus, _reason = decisions[row["row"]]
        cell = ws.cell(row["row"], COL_FOCUS)
        cell.value = focus
        cell.number_format = "@"

    wb.save(OUTPUT)
    build_audit(rows, decisions)
    print(f"saved: {OUTPUT}")
    print(f"audit: {AUDIT}")


if __name__ == "__main__":
    main()
