from __future__ import annotations

import argparse
import copy
import json
import math
import sys
import xml.etree.ElementTree as ET
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils import get_column_letter

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import add_tree_yoy_column_20260611 as yoy_base  # noqa: E402
import apply_macro_change_history_colors as color_base  # noqa: E402


TREE_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5
CODE_COL = 10
CURRENT_COL = 11
DATE_COL = 12
YOY_COL = 14
CALIBER_COL = 16
LOOKBACK_DAYS = 365
MIN_HISTORY_VALUES = 4
MIN_PERCENTILE_TO_FILL = 0.55


@dataclass(frozen=True)
class YoyPoint:
    when: date
    value: float
    source_value: float
    prior_when: date | None = None
    prior_value: float | None = None


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    return yoy_base.to_number(value)


def parse_date(value: Any) -> date | None:
    return yoy_base.parse_date(value)


def percentile(sorted_values: list[float], value: float) -> float:
    if not sorted_values:
        return 0.0
    return bisect_right(sorted_values, value) / len(sorted_values)


def scale_for_yoy_series(current_display: Any, current_raw: float | None) -> float:
    current_number = to_number(current_display)
    if current_number is None or current_raw is None or abs(current_raw) <= 1e-12:
        return 1.0
    ratio = current_number / current_raw
    if 0.0001 <= abs(ratio) <= 10000:
        return ratio
    return 1.0


def build_yoy_history(
    series: yoy_base.MacroSeries,
    current_date: date,
    current_display: Any,
    caliber: str,
) -> tuple[list[YoyPoint], str]:
    points = sorted((d, v) for d, v in series.points if d <= current_date)
    if not points:
        return [], "无底层历史数据"

    start_date = current_date - timedelta(days=LOOKBACK_DAYS)
    is_yoy = yoy_base.is_yoy_caliber(caliber, series.name)
    history: list[YoyPoint] = []

    if is_yoy:
        current_point = yoy_base.closest_current_period(points, current_date)
        scale = scale_for_yoy_series(current_display, current_point[1] if current_point else None)
        for d, value in points:
            if start_date <= d <= current_date:
                history.append(YoyPoint(d, value * scale, value))
        return history, "原指标已是同比口径，历史序列按日报原值同比序列缩放后比较"

    for d, value in points:
        if not (start_date <= d <= current_date):
            continue
        prior = yoy_base.closest_prior_period(points, d)
        if prior is None:
            continue
        prior_dt, prior_value = prior
        if abs(prior_value) <= 1e-12:
            continue
        rate = (value - prior_value) / abs(prior_value)
        if math.isfinite(rate):
            history.append(YoyPoint(d, rate, value, prior_dt, prior_value))
    return history, "非同比原始数据按(当期值-去年同期值)/ABS(去年同期值)逐期计算"


def color_for_history(current_value: Any, history: list[YoyPoint]) -> tuple[str | None, str, dict[str, Any]]:
    latest = to_number(current_value)
    values = [pt.value for pt in history if math.isfinite(pt.value)]
    detail = {
        "latest_yoy_change_rate": latest,
        "history_count": len(values),
        "history_start": history[0].when.isoformat() if history else "",
        "history_end": history[-1].when.isoformat() if history else "",
        "percentile": None,
        "intensity": None,
    }
    if latest is None or not math.isfinite(latest):
        return None, color_base.BLACK, detail
    if abs(latest) <= 1e-12:
        detail["percentile"] = 0.0
        detail["intensity"] = 0.0
        return None, color_base.BLACK, detail
    if len(values) < MIN_HISTORY_VALUES:
        return None, color_base.BLACK, detail

    history_abs = sorted(abs(value) for value in values)
    pct = percentile(history_abs, abs(latest))
    detail["percentile"] = pct
    if pct < MIN_PERCENTILE_TO_FILL:
        detail["intensity"] = 0.0
        return None, color_base.BLACK, detail

    intensity = (pct - MIN_PERCENTILE_TO_FILL) / (1.0 - MIN_PERCENTILE_TO_FILL)
    intensity = max(0.15, min(1.0, intensity))
    detail["intensity"] = intensity
    if latest > 0:
        fill = color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
    else:
        fill = color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
    return fill, color_base.text_color_for_fill(fill), detail


def iter_tree_rows(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    big = ""
    dim = ""
    sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        big_value = as_text(ws.cell(row, 1).value)
        dim_value = as_text(ws.cell(row, 2).value)
        sub_value = as_text(ws.cell(row, 3).value)
        if big_value:
            big = big_value
            dim = ""
            sub = ""
        if dim_value:
            dim = dim_value
            sub = ""
        if sub_value:
            sub = sub_value
        name = as_text(ws.cell(row, 4).value)
        code = yoy_base.norm_code(ws.cell(row, CODE_COL).value)
        if not name and not code:
            continue
        rows.append(
            {
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "code": code,
                "current": ws.cell(row, CURRENT_COL).value,
                "current_date": ws.cell(row, DATE_COL).value,
                "yoy": ws.cell(row, YOY_COL).value,
                "caliber": as_text(ws.cell(row, CALIBER_COL).value),
            }
        )
    return rows


def is_macro_row(row: dict[str, Any]) -> bool:
    return yoy_base.is_macro_tree_row(row["big"], row["dim"])


def read_xml(zf: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(zf.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def write_audit(path: Path, audit_rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "同比变化率颜色复核"
    headers = [
        "TREE行号",
        "大分类",
        "细分维度",
        "子维度",
        "指标名称",
        "指标代码",
        "数据日期",
        "同比变化率",
        "是否宏观计算区",
        "计算方法",
        "历史样本数",
        "历史起点",
        "历史终点",
        "历史分位",
        "颜色强度",
        "填充色",
        "字体色",
        "备注",
    ]
    ws.append(headers)
    for item in audit_rows:
        ws.append([item.get(header, "") for header in headers])
    for idx, width in enumerate([10, 18, 18, 18, 34, 22, 14, 14, 14, 34, 12, 14, 14, 12, 12, 14, 12, 48], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def apply_colors(input_tree: Path, daily_path: Path, output_tree: Path, audit_output: Path) -> dict[str, Any]:
    macro_series = yoy_base.read_macro_series(daily_path)
    wb = openpyxl.load_workbook(input_tree, data_only=True, read_only=False)
    ws = wb[TREE_SHEET]
    rows = iter_tree_rows(ws)
    wb.close()

    summaries = {
        "rows": len(rows),
        "macro_rows": 0,
        "styled_rows": 0,
        "positive_styled_rows": 0,
        "negative_styled_rows": 0,
        "dash_rows": 0,
        "no_history_rows": 0,
        "small_rows": 0,
        "unmatched_rows": 0,
    }
    audit_rows: list[dict[str, Any]] = []
    style_cache: dict[tuple[int, str, str], int] = {}
    replacements: dict[str, bytes] = {}

    with ZipFile(input_tree, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        sheet_path = color_base.workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)

        for row in rows:
            cell = color_base.find_cell(sheet_root, row["row"], YOY_COL)
            base_style_id = int(cell.attrib.get("s", "0"))
            fill_rgb: str | None = None
            font_rgb = color_base.BLACK
            method = ""
            note = ""
            detail: dict[str, Any] = {
                "history_count": 0,
                "history_start": "",
                "history_end": "",
                "percentile": None,
                "intensity": None,
            }

            if row["yoy"] == "-":
                summaries["dash_rows"] += 1
                note = "同比变化率为-，不填色"
            elif not is_macro_row(row):
                note = "非宏观计算区，不填色"
            else:
                summaries["macro_rows"] += 1
                series = macro_series.get(row["code"])
                current_dt = parse_date(row["current_date"])
                if series is None or current_dt is None:
                    summaries["unmatched_rows"] += 1
                    note = "未匹配到底层宏观数据或日期为空"
                else:
                    history, method = build_yoy_history(series, current_dt, row["yoy"], row["caliber"])
                    fill_rgb, font_rgb, detail = color_for_history(row["yoy"], history)
                    if fill_rgb:
                        summaries["styled_rows"] += 1
                        if to_number(row["yoy"]) and to_number(row["yoy"]) > 0:
                            summaries["positive_styled_rows"] += 1
                        else:
                            summaries["negative_styled_rows"] += 1
                    elif detail.get("history_count", 0) < MIN_HISTORY_VALUES and to_number(row["yoy"]) not in (None, 0.0):
                        summaries["no_history_rows"] += 1
                        note = "近一年同比变化率历史样本不足"
                    else:
                        summaries["small_rows"] += 1
                        note = "历史分位未达到填色阈值或数值接近0"

            cell.attrib["s"] = str(
                color_base.ensure_cell_style(styles_root, base_style_id, fill_rgb, font_rgb, style_cache)
            )
            audit_rows.append(
                {
                    "TREE行号": row["row"],
                    "大分类": row["big"],
                    "细分维度": row["dim"],
                    "子维度": row["sub"],
                    "指标名称": row["name"],
                    "指标代码": row["code"],
                    "数据日期": row["current_date"],
                    "同比变化率": row["yoy"],
                    "是否宏观计算区": "是" if is_macro_row(row) else "否",
                    "计算方法": method,
                    "历史样本数": detail.get("history_count", 0),
                    "历史起点": detail.get("history_start", ""),
                    "历史终点": detail.get("history_end", ""),
                    "历史分位": detail.get("percentile"),
                    "颜色强度": detail.get("intensity"),
                    "填充色": fill_rgb or "",
                    "字体色": font_rgb,
                    "备注": note,
                }
            )

        replacements[sheet_path] = write_xml(sheet_root)
        replacements["xl/styles.xml"] = write_xml(styles_root)

        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, replacements[name] if name in replacements else zin.read(name))

    write_audit(audit_output, audit_rows)
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        "lookback_days": LOOKBACK_DAYS,
        "min_history_values": MIN_HISTORY_VALUES,
        "min_percentile_to_fill": MIN_PERCENTILE_TO_FILL,
        **summaries,
    }


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()
    result = apply_colors(
        Path(args.tree),
        Path(args.daily),
        Path(args.output_tree),
        Path(args.audit_output),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
