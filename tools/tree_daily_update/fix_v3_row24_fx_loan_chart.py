from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from sync_tree_new_indicators_from_daily import (
    chart_rows,
    find_cell,
    read_xml,
    set_date,
    set_number,
    sheet_drawing_path,
    update_chart_series,
    workbook_sheet_path,
    write_xml,
)
from sync_tree_v30_full import build_sources, scale_for_cell


WORKSPACE = Path(__file__).resolve().parents[2]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

import sync_tree_20260604 as base  # noqa: E402


SHEET = "重点策略跟踪情况(V3)"
ROW = 24
CODE = "M5206732"
CURRENT_COL = 15
DATE_COL = 16
CHANGE_COL = 17


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill the V3 display-sheet FX loan row and its trend chart from the daily source.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default=date.today().isoformat())
    args = parser.parse_args()

    if Path(args.input).resolve() == Path(args.output).resolve():
        raise ValueError("Input and output workbook paths must be different.")

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
    by_code, _by_name, best_trend = build_sources(Path(args.daily), as_of)
    source = by_code[base.code_key(CODE)]
    source = base.with_best_trend(source, best_trend)

    wb = openpyxl.load_workbook(args.input, data_only=False, read_only=True)
    ws = wb[SHEET]

    with ZipFile(args.input, "r") as zin:
        sheet_path = workbook_sheet_path(zin, SHEET)
        sheet_root = read_xml(zin, sheet_path)
        drawing_path = sheet_drawing_path(zin, sheet_path)
        chart_map = chart_rows(zin, drawing_path) if drawing_path else {}
        replacements: dict[str, bytes] = {}

        set_number(
            find_cell(sheet_root, ROW, CURRENT_COL),
            scale_for_cell(source.current, ws.cell(ROW, CURRENT_COL).number_format),
        )
        set_date(find_cell(sheet_root, ROW, DATE_COL), source.data_date)
        set_number(
            find_cell(sheet_root, ROW, CHANGE_COL),
            scale_for_cell(source.change, ws.cell(ROW, CHANGE_COL).number_format),
        )

        chart_info = chart_map.get(ROW)
        chart_points = 0
        if chart_info:
            values = base.trim_trend_values(source, base.frequency_take_count(ws, ROW, source))
            if values:
                replacements[chart_info["chart_path"]] = update_chart_series(zin.read(chart_info["chart_path"]), values)
                chart_points = len(values)

        replacements[sheet_path] = write_xml(sheet_root)
        with ZipFile(args.output, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    print(
        json.dumps(
            {
                "output": args.output,
                "row": ROW,
                "code": CODE,
                "date": source.data_date,
                "current": source.current,
                "change": source.change,
                "chart_points": chart_points,
            },
            ensure_ascii=False,
            indent=2,
            default=json_default,
        )
    )


if __name__ == "__main__":
    main()
