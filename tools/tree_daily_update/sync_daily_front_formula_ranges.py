from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def load_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def quote_sheet_name(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def last_header_col(ws) -> int:
    last = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value not in (None, ""):
            last = col
    return last


def update_formula_ranges(formula: str, sheet_last_cols: dict[str, str]) -> tuple[str, int]:
    updated = formula
    changes = 0
    for sheet_name, last_col_letter in sheet_last_cols.items():
        refs = [sheet_name + "!", quote_sheet_name(sheet_name) + "!"]
        for ref in refs:
            prefix = re.escape(ref)
            pattern = re.compile(rf"({prefix}\$F\$(\d+):\$)([A-Z]{{1,4}})(\$(\d+))")

            def repl(match: re.Match[str]) -> str:
                nonlocal changes
                start_row = match.group(2)
                end_row = match.group(5)
                if start_row != end_row:
                    return match.group(0)
                old_col = match.group(3)
                if old_col == last_col_letter:
                    return match.group(0)
                changes += 1
                return match.group(1) + last_col_letter + match.group(4)

            updated = pattern.sub(repl, updated)
    return updated, changes


def sync_formula_ranges(input_path: Path, output_path: Path, config_path: Path) -> dict[str, Any]:
    config = load_config(config_path)
    wb = openpyxl.load_workbook(input_path)

    sheet_last_cols: dict[str, str] = {}
    for sheet_name in config["base_sheets"]:
        if sheet_name not in wb.sheetnames:
            continue
        sheet_last_cols[sheet_name] = get_column_letter(last_header_col(wb[sheet_name]))

    changed_cells: list[dict[str, Any]] = []
    replacements = 0
    for sheet_name in config["front_sheets"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                new_value, count = update_formula_ranges(value, sheet_last_cols)
                if count:
                    cell.value = new_value
                    replacements += count
                    changed_cells.append({"sheet": sheet_name, "cell": cell.coordinate, "replacements": count})

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)
    return {
        "input": str(input_path),
        "output": str(output_path),
        "base_last_cols": sheet_last_cols,
        "changed_cells": len(changed_cells),
        "range_replacements": replacements,
        "sample": changed_cells[:20],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Expand daily front-sheet formula ranges to current base-sheet date columns.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--config", default=str(Path(__file__).with_name("config.json")))
    args = parser.parse_args()

    result = sync_formula_ranges(Path(args.input), Path(args.output), Path(args.config))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
