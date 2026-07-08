from __future__ import annotations

import argparse
import json
import math
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils.datetime import to_excel


TREE_SHEET = "重点策略跟踪情况(V3.0)"
DISPLAY_SHEET = "重点策略跟踪情况(V3)"
DAILY_INDEX_SHEET = "指数走势"

TARGET_ROWS = {
    81: {"tree_name": "上证指数：PE", "daily_row": 33, "code": "000001.SH|pe_ttm"},
    82: {"tree_name": "深成指数：PE", "daily_row": 34, "code": "399001.SZ|pe_ttm"},
    83: {"tree_name": "沪深300：PE", "daily_row": 35, "code": "000300.SH|pe_ttm"},
    84: {"tree_name": "创业板指：PE", "daily_row": 36, "code": "399006.SZ|pe_ttm"},
    85: {"tree_name": "恒生指数：PE", "daily_row": 37, "code": "HSI.HI|pe_ttm"},
    86: {"tree_name": "恒生科技：PE", "daily_row": 38, "code": "HSTECH.HI|pe_ttm"},
}

TEMPLATE_CHART_ROW = 80

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}

for prefix, uri in [
    ("", NS["main"]),
    ("r", NS["officeRel"]),
    ("rel", NS["rel"]),
    ("xdr", NS["xdr"]),
    ("a", NS["a"]),
    ("c", NS["c"]),
]:
    ET.register_namespace(prefix, uri)


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_number(value: float) -> str:
    return format(float(value), ".12g")


def parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = as_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return datetime(parsed.year, parsed.month, parsed.day)
        except ValueError:
            pass
    return None


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def sheet_drawing_path(z: ZipFile, sheet_path: str) -> str | None:
    sheet = read_xml(z, sheet_path)
    drawing = sheet.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib[f"{{{NS['officeRel']}}}id"]
    rels_name = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
    rels = read_xml(z, rels_name)
    for rel in rels:
        if rel.attrib["Id"] == rid:
            return rel_target(sheet_path, rel.attrib["Target"])
    return None


def drawing_rels_path(drawing_path: str) -> str:
    p = PurePosixPath(drawing_path)
    return str(p.parent / "_rels" / (p.name + ".rels"))


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    rows = sheet_data.findall("main:row", NS)
    insert_at = len(rows)
    for idx, row in enumerate(rows):
        if int(row.attrib["r"]) > row_num:
            insert_at = idx
            break
    sheet_data.insert(insert_at, new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def find_existing_cell(sheet_root: ET.Element, ref: str) -> ET.Element | None:
    row_num = int("".join(ch for ch in ref if ch.isdigit()))
    row = find_row(sheet_root, row_num)
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    return None


def cell_style(sheet_root: ET.Element, ref: str, fallback: int | None = None) -> int | None:
    cell = find_existing_cell(sheet_root, ref)
    if cell is not None and "s" in cell.attrib:
        return int(cell.attrib["s"])
    return fallback


def remove_children(cell: ET.Element) -> None:
    for child in list(cell):
        cell.remove(child)


def apply_style(cell: ET.Element, style_id: int | None) -> None:
    if style_id is not None:
        cell.attrib["s"] = str(style_id)


def set_inline_string(cell: ET.Element, value: Any, style_id: int | None = None) -> None:
    remove_children(cell)
    cell.attrib["t"] = "inlineStr"
    apply_style(cell, style_id)
    is_el = ET.SubElement(cell, q("is"))
    t = ET.SubElement(is_el, q("t"))
    t.text = as_text(value)


def set_number(cell: ET.Element, value: Any, style_id: int | None = None) -> None:
    number = to_number(value)
    if number is None:
        remove_children(cell)
        cell.attrib.pop("t", None)
        apply_style(cell, style_id)
        return
    remove_children(cell)
    cell.attrib.pop("t", None)
    apply_style(cell, style_id)
    v = ET.SubElement(cell, q("v"))
    v.text = format_number(number)


def set_date(cell: ET.Element, value: Any, style_id: int | None = None) -> str:
    parsed = parse_date(value)
    if parsed is None:
        remove_children(cell)
        cell.attrib.pop("t", None)
        apply_style(cell, style_id)
        return ""
    remove_children(cell)
    cell.attrib.pop("t", None)
    apply_style(cell, style_id)
    v = ET.SubElement(cell, q("v"))
    v.text = format_number(to_excel(parsed))
    return parsed.strftime("%Y-%m-%d")


def max_chart_number(names: list[str]) -> int:
    max_num = 0
    for name in names:
        if name.startswith("xl/charts/chart") and name.endswith(".xml"):
            raw = name.removeprefix("xl/charts/chart").removesuffix(".xml")
            if raw.isdigit():
                max_num = max(max_num, int(raw))
    return max_num


def max_rel_id(rels_root: ET.Element) -> int:
    max_num = 0
    for rel in rels_root:
        rid = rel.attrib.get("Id", "")
        if rid.startswith("rId") and rid[3:].isdigit():
            max_num = max(max_num, int(rid[3:]))
    return max_num


def max_cnvpr_id(drawing_root: ET.Element) -> int:
    max_id = 0
    for elem in drawing_root.findall(".//xdr:cNvPr", NS):
        try:
            max_id = max(max_id, int(elem.attrib.get("id", "0")))
        except ValueError:
            pass
    return max_id


def chart_rows(z: ZipFile, drawing_path: str) -> dict[int, list[dict[str, str]]]:
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    drawing = read_xml(z, drawing_path)
    out: dict[int, list[dict[str, str]]] = {}
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or chart is None:
            continue
        row_text = frm.findtext("xdr:row", namespaces=NS)
        if row_text is None:
            continue
        rid = chart.attrib[f"{{{NS['officeRel']}}}id"]
        out.setdefault(int(row_text) + 1, []).append({"rid": rid, "chart_path": rel_map[rid]})
    return out


def find_template_anchor_and_chart(z: ZipFile, drawing_path: str, template_row: int) -> tuple[ET.Element, bytes]:
    drawing = read_xml(z, drawing_path)
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        if frm is None:
            continue
        row_text = frm.findtext("xdr:row", namespaces=NS)
        if row_text is None or int(row_text) + 1 != template_row:
            continue
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if chart is None:
            continue
        rid = chart.attrib[f"{{{NS['officeRel']}}}id"]
        return deepcopy(anchor), z.read(rel_map[rid])
    raise KeyError(f"Template chart row {template_row} not found")


def make_anchor(template_anchor: ET.Element, template_row: int, target_row: int, rid: str, cnvpr_id: int) -> ET.Element:
    anchor = deepcopy(template_anchor)
    delta = target_row - template_row
    for marker in ("from", "to"):
        elem = anchor.find(f"xdr:{marker}", NS)
        if elem is None:
            continue
        row_el = elem.find("xdr:row", NS)
        if row_el is not None and row_el.text is not None:
            row_el.text = str(int(row_el.text) + delta)

    cnvpr = anchor.find(".//xdr:cNvPr", NS)
    if cnvpr is not None:
        cnvpr.attrib["id"] = str(cnvpr_id)
        cnvpr.attrib["name"] = f"PE Trend {target_row}"

    chart = anchor.find(".//a:graphicData/c:chart", NS)
    if chart is None:
        raise KeyError("anchor chart rel")
    chart.attrib[f"{{{NS['officeRel']}}}id"] = rid
    return anchor


def update_chart_series(chart_xml: bytes, values: list[float]) -> bytes:
    root = ET.fromstring(chart_xml)
    literal = "{" + ",".join(format_number(v) for v in values) + "}"
    for formula in root.findall(".//c:lineChart/c:ser/c:val/c:numRef/c:f", NS):
        formula.text = literal

    for num_ref in root.findall(".//c:lineChart/c:ser/c:val/c:numRef", NS):
        num_cache = num_ref.find("c:numCache", NS)
        if num_cache is None:
            num_cache = ET.SubElement(num_ref, f"{{{NS['c']}}}numCache")
        for child in list(num_cache):
            tag = child.tag.rsplit("}", 1)[-1]
            if tag in {"ptCount", "pt"}:
                num_cache.remove(child)
        insert_at = 1 if num_cache.find("c:formatCode", NS) is not None else 0
        num_cache.insert(insert_at, ET.Element(f"{{{NS['c']}}}ptCount", {"val": str(len(values))}))
        for idx, value in enumerate(values):
            pt = ET.Element(f"{{{NS['c']}}}pt", {"idx": str(idx)})
            v = ET.SubElement(pt, f"{{{NS['c']}}}v")
            v.text = format_number(value)
            num_cache.append(pt)
    return write_xml(root)


def read_daily_pe(daily_path: Path) -> dict[int, dict[str, Any]]:
    wb = openpyxl.load_workbook(daily_path, data_only=True, read_only=False)
    front = wb["A股港股"]
    index = wb[DAILY_INDEX_SHEET]
    out: dict[int, dict[str, Any]] = {}

    front_by_code = {as_text(front.cell(row, 3).value).upper(): row for row in range(1, front.max_row + 1)}
    row_to_front_code = {
        33: "000001.SH",
        34: "399001.SZ",
        35: "000300.SH",
        36: "399006.SZ",
        37: "HSI.HI",
        38: "HSTECH.HI",
    }

    for tree_row, meta in TARGET_ROWS.items():
        daily_row = int(meta["daily_row"])
        code = row_to_front_code[daily_row]
        front_row = front_by_code.get(code)
        if front_row is None:
            raise KeyError(f"Daily front row not found for {code}")

        values = [
            float(v)
            for row in index.iter_rows(min_row=daily_row, max_row=daily_row, min_col=6, max_col=765, values_only=True)
            for v in row
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        ]
        if not values:
            raise RuntimeError(f"No PE trend values for index row {daily_row}")

        current = float(front.cell(front_row, 5).value)
        data_date = front.cell(front_row, 6).value
        change = float(front.cell(front_row, 7).value)
        less = sum(v < current for v in values)
        equal = sum(v == current for v in values)
        percentile = (less + 0.5 * equal) / len(values)
        out[tree_row] = {
            "name": meta["tree_name"],
            "code": meta["code"],
            "current": current,
            "date": data_date,
            "change": change,
            "trend": values,
            "percentile": percentile,
            "count": len(values),
            "source": f"A股港股!{front_row}/指数走势!{daily_row}",
        }
    return out


def pct_text(value: float) -> str:
    return f"{value * 100:.1f}%"


def fmt2(value: float) -> str:
    return f"{value:.2f}"


def row_judgement(name: str, current: float, change: float, percentile: float) -> str:
    position = "高位" if percentile >= 0.8 else "中高位置" if percentile >= 0.6 else "中位" if percentile >= 0.4 else "偏低位置"
    direction = "上行" if change > 0 else "回落" if change < 0 else "持平"
    change_text = f"{direction}{abs(change):.2f}倍" if change else "持平"
    return f"{name}为{fmt2(current)}倍，近三年分位约{pct_text(percentile)}，环比{change_text}，估值处于{position}。"


def face_conclusion(data: dict[int, dict[str, Any]]) -> str:
    hs300 = data[83]
    cyb = data[84]
    hsi = data[85]
    hstech = data[86]
    return (
        f"A股和港股估值数据已补齐：沪深300PE为{fmt2(hs300['current'])}倍、近三年分位{pct_text(hs300['percentile'])}，"
        f"创业板指PE为{fmt2(cyb['current'])}倍、近三年分位{pct_text(cyb['percentile'])}；"
        f"恒生指数PE为{fmt2(hsi['current'])}倍、近三年分位{pct_text(hsi['percentile'])}，"
        f"恒生科技PE为{fmt2(hstech['current'])}倍、近三年分位{pct_text(hstech['percentile'])}。"
        "整体看，A股核心和成长估值已处近三年高位，港股传统指数估值偏高、科技成长估值中高。"
    )


def update_tree_cells(sheet_root: ET.Element, data: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    updated = []
    style = {
        "judgement": cell_style(sheet_root, "F81", cell_style(sheet_root, "F75")),
        "tracking": cell_style(sheet_root, "M81", cell_style(sheet_root, "M75")),
        "code": cell_style(sheet_root, "N81", cell_style(sheet_root, "N75")),
        "current": cell_style(sheet_root, "O81", cell_style(sheet_root, "O75")),
        "date": cell_style(sheet_root, "P81", cell_style(sheet_root, "P75")),
        "change": cell_style(sheet_root, "S81", cell_style(sheet_root, "S75")),
        "dash": cell_style(sheet_root, "R81", cell_style(sheet_root, "R75")),
        "conclusion": cell_style(sheet_root, "D75"),
    }

    set_inline_string(find_cell(sheet_root, 75, 4), "估值偏高", style["conclusion"])
    set_inline_string(find_cell(sheet_root, 75, 6), face_conclusion(data), style["judgement"])

    for row, info in data.items():
        set_inline_string(
            find_cell(sheet_root, row, 6),
            row_judgement(as_text(info["name"]), info["current"], info["change"], info["percentile"]),
            style["judgement"],
        )
        set_inline_string(find_cell(sheet_root, row, 13), "当前PE+近三年分位", style["tracking"])
        set_inline_string(find_cell(sheet_root, row, 14), info["code"], style["code"])
        set_number(find_cell(sheet_root, row, 15), info["current"], style["current"])
        date_text = set_date(find_cell(sheet_root, row, 16), info["date"], style["date"])
        set_number(find_cell(sheet_root, row, 17), info["change"], style["change"])
        set_inline_string(find_cell(sheet_root, row, 18), "-", style["dash"])
        set_inline_string(find_cell(sheet_root, row, 19), "-", style["dash"])
        updated.append(
            {
                "row": row,
                "name": info["name"],
                "code": info["code"],
                "current": info["current"],
                "date": date_text,
                "change": info["change"],
                "percentile": info["percentile"],
                "trend_points": len(info["trend"]),
                "source": info["source"],
            }
        )
    return updated


def update_display_cells(sheet_root: ET.Element, data: dict[int, dict[str, Any]]) -> None:
    style = cell_style(sheet_root, "U51")
    text = (
        f"A股估值面：偏高；沪深300PE {fmt2(data[83]['current'])}倍、近三年分位{pct_text(data[83]['percentile'])}，"
        f"创业板指PE {fmt2(data[84]['current'])}倍、近三年分位{pct_text(data[84]['percentile'])}，"
        "核心资产和成长估值均处高位，指数点位需结合盈利和利率继续观察。"
    )
    set_inline_string(find_cell(sheet_root, 51, 21), text, style)


def update_tree_charts(tree_path: Path, replacements: dict[str, bytes], data: dict[int, dict[str, Any]]) -> tuple[dict[str, bytes], dict[str, bytes], dict[str, Any]]:
    new_files: dict[str, bytes] = {}
    meta: dict[str, Any] = {"charts_added": [], "charts_updated": []}
    with ZipFile(tree_path, "r") as z:
        names = z.namelist()
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        drawing_path = sheet_drawing_path(z, sheet_path)
        if drawing_path is None:
            raise RuntimeError("TREE V3.0 sheet has no drawing")
        drawing_rels = drawing_rels_path(drawing_path)
        drawing_root = read_xml(z, drawing_path)
        rels_root = read_xml(z, drawing_rels)
        content_types = read_xml(z, "[Content_Types].xml")
        existing = chart_rows(z, drawing_path)
        template_anchor, template_chart_xml = find_template_anchor_and_chart(z, drawing_path, TEMPLATE_CHART_ROW)

        next_chart_num = max_chart_number(names) + 1
        next_rid_num = max_rel_id(rels_root) + 1
        next_cnvpr = max_cnvpr_id(drawing_root) + 1
        rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"

        for row, info in data.items():
            values = [float(v) for v in info["trend"] if to_number(v) is not None]
            if row in existing and existing[row]:
                chart_path = existing[row][0]["chart_path"]
                replacements[chart_path] = update_chart_series(z.read(chart_path), values)
                meta["charts_updated"].append({"row": row, "points": len(values), "chart": chart_path})
                continue

            chart_path = f"xl/charts/chart{next_chart_num}.xml"
            rid = f"rId{next_rid_num}"
            new_files[chart_path] = update_chart_series(template_chart_xml, values)

            rel = ET.Element("Relationship", {"Id": rid, "Type": rel_type, "Target": "../charts/" + PurePosixPath(chart_path).name})
            rels_root.append(rel)
            drawing_root.append(make_anchor(template_anchor, TEMPLATE_CHART_ROW, row, rid, next_cnvpr))

            if not any(
                elem.attrib.get("PartName") == f"/{chart_path}"
                for elem in content_types
                if elem.tag.rsplit("}", 1)[-1] == "Override"
            ):
                content_types.append(
                    ET.Element(
                        f"{{http://schemas.openxmlformats.org/package/2006/content-types}}Override",
                        {
                            "PartName": f"/{chart_path}",
                            "ContentType": "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
                        },
                    )
                )

            meta["charts_added"].append({"row": row, "points": len(values), "chart": chart_path})
            next_chart_num += 1
            next_rid_num += 1
            next_cnvpr += 1

        replacements[drawing_path] = write_xml(drawing_root)
        replacements[drawing_rels] = write_xml(rels_root)
        replacements["[Content_Types].xml"] = write_xml(content_types)
    return replacements, new_files, meta


def write_workbook(tree_path: Path, output_path: Path, replacements: dict[str, bytes], new_files: dict[str, bytes]) -> None:
    with ZipFile(tree_path, "r") as zin, ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name in replacements:
                zout.writestr(name, replacements[name])
            else:
                zout.writestr(name, zin.read(name))
        for name, data in new_files.items():
            zout.writestr(name, data)


def verify_output(output_path: Path) -> dict[str, Any]:
    with ZipFile(output_path, "r") as z:
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        drawing_path = sheet_drawing_path(z, sheet_path)
        rows = chart_rows(z, drawing_path) if drawing_path else {}
        missing = [row for row in TARGET_ROWS if row not in rows]
        point_counts = {}
        for row in TARGET_ROWS:
            info = rows.get(row, [{}])[0]
            path = info.get("chart_path")
            if not path:
                continue
            chart_root = ET.fromstring(z.read(path))
            count = chart_root.find(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:ptCount", NS)
            point_counts[row] = int(count.attrib.get("val", "0")) if count is not None else 0
    return {"chart_missing": missing, "chart_point_counts": point_counts}


def sync(tree_path: Path, daily_path: Path, output_path: Path) -> dict[str, Any]:
    data = read_daily_pe(daily_path)
    replacements: dict[str, bytes] = {}

    with ZipFile(tree_path, "r") as z:
        tree_sheet_path = workbook_sheet_path(z, TREE_SHEET)
        tree_root = read_xml(z, tree_sheet_path)
        updated_rows = update_tree_cells(tree_root, data)
        replacements[tree_sheet_path] = write_xml(tree_root)

        display_sheet_path = workbook_sheet_path(z, DISPLAY_SHEET)
        display_root = read_xml(z, display_sheet_path)
        update_display_cells(display_root, data)
        replacements[display_sheet_path] = write_xml(display_root)

    replacements, new_files, chart_meta = update_tree_charts(tree_path, replacements, data)
    write_workbook(tree_path, output_path, replacements, new_files)
    verify = verify_output(output_path)

    return {
        "tree": str(tree_path),
        "daily": str(daily_path),
        "output": str(output_path),
        "updated_rows": updated_rows,
        **chart_meta,
        **verify,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync newly added index PE valuation rows into the latest TREE workbook.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    result = sync(Path(args.tree), Path(args.daily), Path(args.output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
