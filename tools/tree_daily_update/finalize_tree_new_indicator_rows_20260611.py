from __future__ import annotations

import argparse
import json
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))
WORKSPACE = TOOL_DIR.parents[1]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

import sync_tree_20260604 as base  # type: ignore  # noqa: E402
from sync_tree_new_indicators_from_daily import (  # noqa: E402
    NS,
    chart_rows,
    drawing_rels_path,
    find_cell,
    find_existing_cell,
    find_template_anchor_and_chart,
    format_number,
    make_anchor,
    max_chart_number,
    max_cnvpr_id,
    max_rel_id,
    read_xml,
    rel_target,
    set_date,
    set_inline_string,
    set_number,
    sheet_drawing_path,
    update_chart_series,
    workbook_sheet_path,
    write_xml,
)
from sync_tree_v30_full import build_sources, to_number  # noqa: E402


TREE_SHEET = "重点策略跟踪情况(V3.0)"
TEMPLATE_CHART_ROW = 20


@dataclass(frozen=True)
class RowMeta:
    weight: float
    reason: str
    definition: str
    usage: str
    freq: str
    value_kind: str  # raw or percent


METADATA: dict[str, RowMeta] = {
    "M0046166": RowMeta(
        0.016,
        "一般公共预算支出累计值直接反映财政资金实际投放规模，是观察财政托底力度的重要金额口径；但单看累计值受年度进度和季节性影响较大，需要结合支出同比、财政收入和赤字脉冲判断。",
        "一般公共预算支出累计值是政府一般公共预算从年初到当前月份已经发生的支出总额，覆盖教育、社保、卫生、城乡社区、农林水、债务付息等公共财政支出。它反映的是财政资金已经投向经济和公共服务领域的累计规模。",
        "支出累计值加快上行，通常说明财政资金投放节奏更快，对基建、公共服务和总需求形成托底；若收入偏弱但支出仍扩张，财政逆周期力度更强。使用时要和公共财政收入、支出同比、政府债发行及财政存款一起看。",
        "月度",
        "raw",
    ),
    "M0048264": RowMeta(
        0.012,
        "财政存款累计值反映财政资金在银行体系中的沉淀和变化，是判断财政资金投放节奏的辅助指标；但累计口径不如财政存款当月变化直接，因此权重低于财政存款边际变化和财政赤字脉冲。",
        "财政存款累计值是金融机构人民币存款统计中，财政部门存款从年初以来的累计新增规模。它衡量的是财政资金在银行体系中形成存款的累计变化，不等同于单月财政支出，也不是美国TGA余额这类国库现金余额。",
        "累计增加较快，通常说明财政资金更多沉淀在存款端，对实体投放和银行体系流动性的即时支持偏弱；累计增速放缓或减少，通常说明财政资金加快拨付和使用。更适合与财政存款当月值、财政支出和政府债发行配合判断。",
        "月度",
        "raw",
    ),
    "M0001427": RowMeta(
        0.020,
        "社零当月值反映消费市场的实际名义销售规模，是验证内需修复的重要金额口径；但当月值有春节、假期和价格因素扰动，方向判断仍应以社零同比和居民收入等指标为主。",
        "社会消费品零售总额当月值是企业和个体经营者当月向个人、社会集团销售消费品以及提供餐饮服务取得的名义金额总和。它反映一个月内国内消费市场的实际销售规模。",
        "当月值扩大说明消费交易规模上升，通常有利于消费链和服务业景气；若同比改善但金额仍偏弱，说明修复质量仍需验证。使用时要重点结合社零同比、CPI、居民收入和节假日错位影响。",
        "月度",
        "raw",
    ),
    "M0039354": RowMeta(
        0.0333333333333,
        "中国GDP不变价当季同比是宏观增长最核心的真实增长指标，能直接验证经济周期方向和政策成效；但季度发布且相对滞后，更适合作为增长模块的锚，而不是高频边际信号。",
        "中国GDP不变价当季同比是剔除价格因素后，本季度国内生产总值相对上年同期的实际增长率。它衡量真实产出的同比变化，比现价GDP更能反映经济活动本身的增长速度。",
        "读数回升通常说明真实经济动能改善，权益和顺周期资产的基本面支撑增强；读数下行说明增长压力加大，政策托底预期可能上升。需要和社零、固定资产投资、出口、PMI等高频指标交叉验证。",
        "季度",
        "percent",
    ),
    "M0000272": RowMeta(
        0.020,
        "固定资产投资累计值反映年内投资完成规模，是观察稳增长落地和实体资本开支的重要金额口径；但累计值受年度进度影响，边际判断应结合累计同比和制造业、基建、地产分项。",
        "固定资产投资累计值是从年初到当前月份，全社会在建筑安装工程、设备购置及其他固定资产建设活动中已经完成的名义投资总额。它覆盖基建、制造业、房地产等长期资产投入。",
        "累计值上行代表投资规模继续扩张，对总需求和未来供给能力都有支撑；如果累计值增长但同比走弱，说明投资节奏可能低于去年同期。分析时要拆分制造业、基建、地产和民间投资，避免只看总量。",
        "月度",
        "raw",
    ),
    "G1109077": RowMeta(
        0.020,
        "美国商业银行银行信贷反映银行资产端信用供给和资产负债表扩张情况，是美元实体流动性的核心余额指标之一；但它是存量指标，边际信号需结合周度变化和贷款分项。",
        "美国所有商业银行银行信贷是美联储H.8统计中的银行资产端口径，通常包括商业银行持有的证券以及贷款和租赁等信用资产，并经过季节调整。它衡量商业银行体系向经济和金融市场提供信用的总体规模。",
        "银行信贷扩张通常说明银行资产负债表扩张、信用供给改善，有利于实体融资和风险偏好；信贷收缩或增长放缓通常提示金融条件收紧。使用时应结合工商业贷款、房地产贷款、消费贷款和信用利差一起判断。",
        "周度",
        "raw",
    ),
    "G0000003": RowMeta(
        0.012,
        "美国现价GDP季调折年数用于观察美国名义经济规模和名义需求背景，可作为企业收入、财政赤字率和资产配置的分母锚；但它同时包含真实增长和通胀影响，边际周期信号弱于实际GDP和零售销售。",
        "美国GDP现价季调折年数是以当前价格计算、经过季节调整并折算成年率的美国国内生产总值。它反映美国经济在名义金额上的年度化规模，包含真实产出变化和价格变化两部分。",
        "名义GDP上行说明经济规模和名义需求扩张，但需要拆分真实增长和通胀贡献；如果名义GDP强而实际GDP弱，可能更多来自价格因素。它更适合做规模和分母校准，而不是单独判断经济动能。",
        "季度",
        "raw",
    ),
    "G1109245": RowMeta(
        0.026,
        "美国零售和食品服务销售额同比是美国消费需求的核心月度指标，能补充环比数据判断消费趋势；但它仍受价格和基数影响，需结合实际收入、消费者信心和通胀数据验证。",
        "美国零售和食品服务销售额季调当月同比是经季节调整后的零售商和食品服务企业销售额，相比上年同月的增长率。它覆盖商品零售和餐饮消费，是观察美国居民消费需求的重要月度指标。",
        "同比上行通常说明消费需求韧性增强，支持美国经济和企业盈利；同比下行说明消费动能减弱，可能强化降息或衰退预期。分析时要和零售环比、核心CPI、工资增速和消费者信心一起看。",
        "月度",
        "percent",
    ),
}


def json_default(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def cell_style(sheet_root: ET.Element, row: int, col: int, fallback: int | None = None) -> int | None:
    ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
    cell = find_existing_cell(sheet_root, ref)
    if cell is not None and "s" in cell.attrib:
        return int(cell.attrib["s"])
    return fallback


def source_values(source: Any, kind: str) -> tuple[Any, Any]:
    current = source.current
    change = source.change
    if kind == "percent":
        current_num = to_number(current)
        change_num = to_number(change)
        return (
            current_num / 100 if current_num is not None else current,
            change_num / 100 if change_num is not None else change,
        )
    return current, change


def update_sheet_xml(tree_path: Path, daily_path: Path, as_of: datetime) -> tuple[dict[str, bytes], dict[str, Any]]:
    by_code, _, best_trend = build_sources(daily_path, as_of.date())
    with ZipFile(tree_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)

        raw_current_style = cell_style(sheet_root, 36, 11)
        raw_change_style = cell_style(sheet_root, 36, 13)
        pct_current_style = cell_style(sheet_root, 47, 11)
        pct_change_style = cell_style(sheet_root, 47, 13)
        date_style = cell_style(sheet_root, 47, 12, cell_style(sheet_root, 36, 12))

        wb_tree = openpyxl.load_workbook(tree_path, data_only=False, read_only=False)
        ws = wb_tree[TREE_SHEET]
        rows: dict[str, int] = {}
        for row in range(6, ws.max_row + 1):
            code = str(ws.cell(row, 10).value or "").strip().upper()
            if code in METADATA:
                rows[code] = row
        wb_tree.close()

        updated = []
        missing_source = []
        for code, meta in METADATA.items():
            row = rows.get(code)
            source = by_code.get(code.upper())
            if row is None:
                updated.append({"code": code, "status": "missing_tree_row"})
                continue
            if source is None:
                missing_source.append({"row": row, "code": code})
                continue
            source = base.with_best_trend(source, best_trend)

            set_number(find_cell(sheet_root, row, 5), meta.weight, cell_style(sheet_root, row, 5))
            set_inline_string(find_cell(sheet_root, row, 6), meta.reason, cell_style(sheet_root, row, 6))
            set_inline_string(find_cell(sheet_root, row, 7), meta.definition, cell_style(sheet_root, row, 7))
            set_inline_string(find_cell(sheet_root, row, 8), meta.usage, cell_style(sheet_root, row, 8))
            set_inline_string(find_cell(sheet_root, row, 9), meta.freq, cell_style(sheet_root, row, 9))
            set_inline_string(find_cell(sheet_root, row, 10), code, cell_style(sheet_root, row, 10))

            current, change = source_values(source, meta.value_kind)
            current_style = pct_current_style if meta.value_kind == "percent" else raw_current_style
            change_style = pct_change_style if meta.value_kind == "percent" else raw_change_style
            set_number(find_cell(sheet_root, row, 11), current, current_style)
            date_text = set_date(find_cell(sheet_root, row, 12), source.data_date, date_style)
            set_number(find_cell(sheet_root, row, 13), change, change_style)
            updated.append(
                {
                    "row": row,
                    "code": code,
                    "source": f"{source.sheet}!{source.row}",
                    "freq": meta.freq,
                    "current": current,
                    "date": date_text,
                    "change": change,
                    "value_kind": meta.value_kind,
                    "trend_points": len(source.trend or []),
                }
            )

        return {sheet_path: write_xml(sheet_root)}, {
            "updated_text_and_values": updated,
            "missing_source": missing_source,
        }


def add_or_update_charts(tree_path: Path, replacements: dict[str, bytes], daily_path: Path, as_of: datetime) -> tuple[dict[str, bytes], dict[str, bytes], dict[str, Any]]:
    by_code, _, best_trend = build_sources(daily_path, as_of.date())
    new_files: dict[str, bytes] = {}
    meta: dict[str, Any] = {"charts_updated": [], "charts_added": [], "charts_skipped_no_trend": []}

    wb_tree = openpyxl.load_workbook(tree_path, data_only=True, read_only=False)
    ws = wb_tree[TREE_SHEET]
    target_rows: dict[int, str] = {}
    for row in range(6, ws.max_row + 1):
        code = str(ws.cell(row, 10).value or "").strip().upper()
        if code in METADATA:
            target_rows[row] = code
    wb_tree.close()

    with ZipFile(tree_path, "r") as zin:
        names = zin.namelist()
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        drawing_path = sheet_drawing_path(zin, sheet_path)
        if drawing_path is None:
            return replacements, new_files, meta | {"chart_error": "target sheet has no drawing"}

        drawing_rels = drawing_rels_path(drawing_path)
        drawing_root = read_xml(zin, drawing_path)
        rels_root = read_xml(zin, drawing_rels)
        content_types = read_xml(zin, "[Content_Types].xml")
        template_anchor, template_chart_xml = find_template_anchor_and_chart(zin, drawing_path, TEMPLATE_CHART_ROW)
        existing = chart_rows(zin, drawing_path)
        next_chart_num = max_chart_number(names) + 1
        next_rid_num = max_rel_id(rels_root) + 1
        next_cnvpr = max_cnvpr_id(drawing_root) + 1
        rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"

        for row, code in sorted(target_rows.items()):
            source = by_code.get(code.upper())
            if source is None:
                continue
            source = base.with_best_trend(source, best_trend)
            values = [float(v) for v in (source.trend or []) if to_number(v) is not None]
            if not values:
                meta["charts_skipped_no_trend"].append({"row": row, "code": code})
                continue

            if row in existing:
                chart_path = existing[row]["chart_path"]
                replacements[chart_path] = update_chart_series(zin.read(chart_path), values)
                meta["charts_updated"].append({"row": row, "code": code, "points": len(values)})
                continue

            chart_path = f"xl/charts/chart{next_chart_num}.xml"
            rid = f"rId{next_rid_num}"
            new_files[chart_path] = update_chart_series(template_chart_xml, values)
            rels_root.append(
                ET.Element("Relationship", {"Id": rid, "Type": rel_type, "Target": "../charts/" + PurePosixPath(chart_path).name})
            )
            drawing_root.append(make_anchor(template_anchor, row, rid, next_cnvpr))

            if not any(
                elem.attrib.get("PartName") == f"/{chart_path}"
                for elem in content_types
                if elem.tag.rsplit("}", 1)[-1] == "Override"
            ):
                content_types.append(
                    ET.Element(
                        "{http://schemas.openxmlformats.org/package/2006/content-types}Override",
                        {
                            "PartName": f"/{chart_path}",
                            "ContentType": "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
                        },
                    )
                )

            meta["charts_added"].append({"row": row, "code": code, "points": len(values), "chart": chart_path})
            next_chart_num += 1
            next_rid_num += 1
            next_cnvpr += 1

        replacements[drawing_path] = write_xml(drawing_root)
        replacements[drawing_rels] = write_xml(rels_root)
        replacements["[Content_Types].xml"] = write_xml(content_types)

    return replacements, new_files, meta


def write_workbook(input_path: Path, output_path: Path, replacements: dict[str, bytes], new_files: dict[str, bytes]) -> None:
    with ZipFile(input_path, "r") as zin, ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name in replacements:
                zout.writestr(name, replacements[name])
            else:
                zout.writestr(name, zin.read(name))
        for name, data in new_files.items():
            zout.writestr(name, data)


def verify(output_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(output_path, data_only=True, read_only=False)
    ws = wb[TREE_SHEET]
    rows = []
    with ZipFile(output_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        drawing_path = sheet_drawing_path(zin, sheet_path)
        chart_map = chart_rows(zin, drawing_path) if drawing_path else {}
        for row in range(6, ws.max_row + 1):
            code = str(ws.cell(row, 10).value or "").strip().upper()
            if code not in METADATA:
                continue
            chart_info = chart_map.get(row)
            chart_points = 0
            if chart_info:
                chart_xml = read_xml(zin, chart_info["chart_path"])
                chart_points = len(chart_xml.findall(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:pt", NS))
            rows.append(
                {
                    "row": row,
                    "name": ws.cell(row, 4).value,
                    "weight": ws.cell(row, 5).value,
                    "freq": ws.cell(row, 9).value,
                    "code": code,
                    "current": ws.cell(row, 11).value,
                    "date": ws.cell(row, 12).value,
                    "change": ws.cell(row, 13).value,
                    "chart": bool(chart_info),
                    "chart_points": chart_points,
                }
            )
    wb.close()
    missing = [
        item
        for item in rows
        if item["weight"] is None
        or not item["freq"]
        or item["current"] is None
        or item["date"] is None
        or item["change"] is None
        or not item["chart"]
    ]
    return {"rows": rows, "missing_after": missing}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default="2026-06-11")
    args = parser.parse_args()

    tree_path = Path(args.tree)
    daily_path = Path(args.daily)
    output_path = Path(args.output)
    as_of = datetime.strptime(args.as_of, "%Y-%m-%d")

    replacements, data_meta = update_sheet_xml(tree_path, daily_path, as_of)
    replacements, new_files, chart_meta = add_or_update_charts(tree_path, replacements, daily_path, as_of)
    write_workbook(tree_path, output_path, replacements, new_files)
    verify_meta = verify(output_path)

    result = {
        "tree": str(tree_path),
        "daily": str(daily_path),
        "output": str(output_path),
        **data_meta,
        **chart_meta,
        **verify_meta,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))
    if verify_meta["missing_after"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
