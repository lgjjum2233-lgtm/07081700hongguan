from __future__ import annotations

import argparse
import copy
import json
import math
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils.datetime import to_excel


TREE_SHEET = "重点策略跟踪情况(V3.0)"
FRONT_SHEETS = ["A股港股", "中债", "中国宏观", "海外数据"]
TEMPLATE_CHART_ROW = 20
RED = "FFFF0000"
GREEN = "FF008000"

ROW_CODE = {
    19: "M0017142",
    27: "M6404533",
    28: "M6404535",
    29: "M6404534",
    30: "M0096870",
    32: "M0325687",
    33: "G0000891",
    34: "M0325687-G0000891",
    36: "M0046168",
    37: "M0046167",
    38: "M0096886",
    39: "M5567950",
    40: "M5405502",
    41: "G1147446",
    45: "M5567876",
    47: "M6001128",
    48: "M6001129",
    49: "M6001130",
    51: "S0029657",
    52: "M0000357",
    53: "M5440435",
    57: "M0008499",
    58: "M0007911",
    59: "M0000610",
    67: "V6842305",
    68: "S0031525",
    79: "M0000561",
    80: "M0017126",
    81: "M0001227",
    82: "S0206721",
    138: "G1112986",
}

NO_DAILY_SOURCE_ROWS = {
    46: "GDP同比：中国",
    78: "A股:全部上市公司:ROE(TTM)",
    137: "GDP总量：美国",
}

NAME_CORRECTIONS = {
    36: "中国:公共财政收入:累计值",
    57: "中国:出口金额:美国:累计同比",
    58: "中国:出口金额:东南亚国家联盟:累计同比",
}

# TREE stores percentage cells as decimals, while front-sheet chart caches keep raw percent values.
PERCENT_DECIMAL_ROWS = {
    27,
    28,
    29,
    37,
    38,
    40,
    41,
    47,
    48,
    49,
    51,
    52,
    53,
    57,
    58,
    79,
    80,
    81,
    82,
    138,
}

NUMERIC_2_ROWS = {59}
GENERAL_NUMBER_ROWS = {36, 39, 45, 67, 68}

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}

for prefix, uri in [
    ("", NS["main"]),
    ("r", NS["officeRel"]),
    ("xdr", NS["xdr"]),
    ("a", NS["a"]),
    ("c", NS["c"]),
    ("rel", NS["rel"]),
]:
    ET.register_namespace(prefix, uri)


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_code(value: Any) -> str:
    return as_text(value).upper()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = as_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return datetime(parsed.year, parsed.month, parsed.day)
        except ValueError:
            pass
    return None


def date_text(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else as_text(value)


def scale_for_tree(row: int, value: Any) -> Any:
    number = to_number(value)
    if number is None:
        return value
    return number / 100 if row in PERCENT_DECIMAL_ROWS else number


def format_number(value: float) -> str:
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return "0"
    return format(float(value), ".12g")


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def write_xml(root: ET.Element, xml_declaration: bool = True) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=xml_declaration)


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def workbook_sheet_paths(z: ZipFile) -> dict[str, str]:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: dict[str, str] = {}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
        out[sheet.attrib["name"]] = rel_target("xl/workbook.xml", rel_map[rid])
    return out


def sheet_drawing_path(z: ZipFile, sheet_path: str) -> str | None:
    sheet = read_xml(z, sheet_path)
    drawing = sheet.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib[f"{{{NS['officeRel']}}}id"]
    rels_name = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
    rels = read_xml(z, rels_name)
    for rel in rels:
        if rel.attrib["Id"] == rid:
            return rel_target(sheet_path, rel.attrib["Target"])
    return None


def drawing_rels_path(drawing_path: str) -> str:
    p = PurePosixPath(drawing_path)
    return str(p.parent / "_rels" / (p.name + ".rels"))


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    rows = sheet_data.findall("main:row", NS)
    insert_at = len(rows)
    for idx, row in enumerate(rows):
        if int(row.attrib["r"]) > row_num:
            insert_at = idx
            break
    sheet_data.insert(insert_at, new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def find_existing_cell(sheet_root: ET.Element, ref: str) -> ET.Element | None:
    row_num = int("".join(ch for ch in ref if ch.isdigit()))
    row = find_row(sheet_root, row_num)
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    return None


def cell_style(sheet_root: ET.Element, ref: str, fallback: int = 0) -> int:
    cell = find_existing_cell(sheet_root, ref)
    if cell is None:
        return fallback
    return int(cell.attrib.get("s", str(fallback)))


def remove_children(cell: ET.Element) -> None:
    for child in list(cell):
        cell.remove(child)


def apply_style(cell: ET.Element, style_id: int | None) -> None:
    if style_id is not None:
        cell.attrib["s"] = str(style_id)


def set_inline_string(cell: ET.Element, value: Any, style_id: int | None = None) -> None:
    remove_children(cell)
    cell.attrib["t"] = "inlineStr"
    apply_style(cell, style_id)
    is_el = ET.SubElement(cell, q("is"))
    t = ET.SubElement(is_el, q("t"))
    t.text = as_text(value)


def set_blank(cell: ET.Element, style_id: int | None = None) -> None:
    remove_children(cell)
    cell.attrib.pop("t", None)
    apply_style(cell, style_id)


def set_number(cell: ET.Element, value: Any, style_id: int | None = None) -> None:
    number = to_number(value)
    if number is None:
        set_blank(cell, style_id)
        return
    remove_children(cell)
    cell.attrib.pop("t", None)
    apply_style(cell, style_id)
    v = ET.SubElement(cell, q("v"))
    v.text = format_number(number)


def set_date(cell: ET.Element, value: Any, style_id: int | None = None) -> str:
    parsed = parse_date(value)
    if parsed is None:
        set_blank(cell, style_id)
        return ""
    remove_children(cell)
    cell.attrib.pop("t", None)
    apply_style(cell, style_id)
    v = ET.SubElement(cell, q("v"))
    v.text = format_number(to_excel(parsed))
    return parsed.strftime("%Y-%m-%d")


def font_without_color(font: ET.Element) -> ET.Element:
    font = copy.deepcopy(font)
    for color in list(font.findall("main:color", NS)):
        font.remove(color)
    return font


def font_with_rgb(font: ET.Element, rgb: str) -> ET.Element:
    font = font_without_color(font)
    color_el = ET.Element(q("color"), {"rgb": rgb})
    insert_at = len(list(font))
    for idx, child in enumerate(list(font)):
        if child.tag.rsplit("}", 1)[-1] in {"sz", "u", "vertAlign", "scheme"}:
            insert_at = idx
            break
    font.insert(insert_at, color_el)
    return font


def ensure_font_style(
    styles_root: ET.Element,
    base_style_id: int,
    rgb: str | None,
    style_cache: dict[tuple[int, str | None], int],
) -> int:
    key = (base_style_id, rgb)
    if key in style_cache:
        return style_cache[key]

    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fonts is None or cell_xfs is None:
        raise KeyError("styles font/cellXfs")

    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]
    base_font_id = int(base_xf.attrib.get("fontId", "0"))
    font_nodes = fonts.findall("main:font", NS)
    base_font = font_nodes[base_font_id] if 0 <= base_font_id < len(font_nodes) else font_nodes[0]

    new_font = font_with_rgb(base_font, rgb) if rgb else font_without_color(base_font)
    fonts.append(new_font)
    new_font_id = len(font_nodes)
    fonts.attrib["count"] = str(len(font_nodes) + 1)

    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["applyFont"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(len(xfs) + 1)
    style_cache[key] = new_style_id
    return new_style_id


def chart_rows(z: ZipFile, drawing_path: str) -> dict[int, dict[str, str]]:
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    drawing = read_xml(z, drawing_path)
    out: dict[int, dict[str, str]] = {}
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or chart is None:
            continue
        row_text = frm.findtext("xdr:row", namespaces=NS)
        if row_text is None:
            continue
        rid = chart.attrib[f"{{{NS['officeRel']}}}id"]
        out[int(row_text) + 1] = {"rid": rid, "chart_path": rel_map[rid]}
    return out


def chart_values(chart_xml: bytes) -> list[float]:
    root = ET.fromstring(chart_xml)
    values: list[float] = []
    for pt in root.findall(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:pt", NS):
        number = to_number(pt.findtext("c:v", namespaces=NS))
        if number is not None:
            values.append(number)
    if values:
        return values
    for pt in root.findall(".//c:lineChart/c:ser/c:val/c:numLit/c:pt", NS):
        number = to_number(pt.findtext("c:v", namespaces=NS))
        if number is not None:
            values.append(number)
    if values:
        return values
    formula = root.findtext(".//c:lineChart/c:ser/c:val/c:numRef/c:f", namespaces=NS)
    if formula and formula.startswith("{") and formula.endswith("}"):
        for part in formula[1:-1].split(","):
            number = to_number(part)
            if number is not None:
                values.append(number)
    return values


def strip_chart_display_labels(root: ET.Element) -> dict[str, int]:
    removed = {"dLbls": 0, "legend": 0}
    for parent in list(root.iter()):
        for child in list(parent):
            tag = child.tag.rsplit("}", 1)[-1]
            if tag == "dLbls":
                parent.remove(child)
                removed["dLbls"] += 1
            elif tag == "legend":
                parent.remove(child)
                removed["legend"] += 1
    return removed


def update_chart_series(chart_xml: bytes, values: list[float]) -> bytes:
    root = ET.fromstring(chart_xml)
    strip_chart_display_labels(root)
    literal = "{" + ",".join(format_number(v) for v in values) + "}"
    formula = root.find(".//c:lineChart/c:ser/c:val/c:numRef/c:f", NS)
    if formula is not None:
        formula.text = literal

    num_ref = root.find(".//c:lineChart/c:ser/c:val/c:numRef", NS)
    if num_ref is None:
        raise KeyError("chart numRef")
    num_cache = num_ref.find("c:numCache", NS)
    if num_cache is None:
        num_cache = ET.SubElement(num_ref, f"{{{NS['c']}}}numCache")
    for child in list(num_cache):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag in {"ptCount", "pt"}:
            num_cache.remove(child)
    pt_count = ET.Element(f"{{{NS['c']}}}ptCount", {"val": str(len(values))})
    insert_at = 1 if num_cache.find("c:formatCode", NS) is not None else 0
    num_cache.insert(insert_at, pt_count)
    for idx, value in enumerate(values):
        pt = ET.Element(f"{{{NS['c']}}}pt", {"idx": str(idx)})
        v = ET.SubElement(pt, f"{{{NS['c']}}}v")
        v.text = format_number(value)
        num_cache.append(pt)
    return write_xml(root)


def find_template_anchor_and_chart(z: ZipFile, drawing_path: str, template_row: int) -> tuple[ET.Element, bytes]:
    drawing = read_xml(z, drawing_path)
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        if frm is None:
            continue
        row_text = frm.findtext("xdr:row", namespaces=NS)
        if row_text is None or int(row_text) + 1 != template_row:
            continue
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if chart is None:
            continue
        rid = chart.attrib[f"{{{NS['officeRel']}}}id"]
        return copy.deepcopy(anchor), z.read(rel_map[rid])
    raise KeyError(f"Template chart row {template_row} not found")


def existing_chart_rows(drawing_root: ET.Element) -> set[int]:
    rows = set()
    for anchor in list(drawing_root):
        frm = anchor.find("xdr:from", NS)
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or chart is None:
            continue
        row_text = frm.findtext("xdr:row", namespaces=NS)
        if row_text is not None:
            rows.add(int(row_text) + 1)
    return rows


def max_chart_number(names: list[str]) -> int:
    max_num = 0
    for name in names:
        match = re.fullmatch(r"xl/charts/chart(\d+)\.xml", name)
        if match:
            max_num = max(max_num, int(match.group(1)))
    return max_num


def max_rel_id(rels_root: ET.Element) -> int:
    max_num = 0
    for rel in rels_root:
        match = re.fullmatch(r"rId(\d+)", rel.attrib.get("Id", ""))
        if match:
            max_num = max(max_num, int(match.group(1)))
    return max_num


def max_cnvpr_id(drawing_root: ET.Element) -> int:
    max_id = 0
    for elem in drawing_root.findall(".//xdr:cNvPr", NS):
        try:
            max_id = max(max_id, int(elem.attrib.get("id", "0")))
        except ValueError:
            pass
    return max_id


def make_anchor(template_anchor: ET.Element, row: int, rid: str, cnvpr_id: int) -> ET.Element:
    anchor = copy.deepcopy(template_anchor)
    frm = anchor.find("xdr:from", NS)
    if frm is None:
        raise KeyError("anchor from")
    frm.find("xdr:col", NS).text = "13"
    frm.find("xdr:colOff", NS).text = "50000"
    frm.find("xdr:row", NS).text = str(row - 1)
    frm.find("xdr:rowOff", NS).text = "30000"

    cnvpr = anchor.find(".//xdr:cNvPr", NS)
    if cnvpr is not None:
        cnvpr.attrib["id"] = str(cnvpr_id)
        cnvpr.attrib["name"] = f"Trend {row}"

    chart = anchor.find(".//a:graphicData/c:chart", NS)
    if chart is None:
        raise KeyError("anchor chart rel")
    chart.attrib[f"{{{NS['officeRel']}}}id"] = rid
    return anchor


def build_daily_sources(daily_path: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(daily_path, data_only=True)
    sources: dict[str, dict[str, Any]] = {}
    for sheet_name in FRONT_SHEETS:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            code = norm_code(ws.cell(row, 3).value)
            if not code:
                continue
            sources.setdefault(
                code,
                {
                    "sheet": sheet_name,
                    "row": row,
                    "name": ws.cell(row, 1).value,
                    "freq": ws.cell(row, 2).value,
                    "unit": ws.cell(row, 4).value,
                    "current": ws.cell(row, 5).value,
                    "date": ws.cell(row, 6).value,
                    "change": ws.cell(row, 7).value,
                    "previous": ws.cell(row, 8).value,
                    "previous_date": ws.cell(row, 9).value,
                    "trend": [],
                },
            )

    with ZipFile(daily_path, "r") as z:
        sheet_paths = workbook_sheet_paths(z)
        for sheet_name in FRONT_SHEETS:
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path:
                continue
            drawing_path = sheet_drawing_path(z, sheet_path)
            if not drawing_path:
                continue
            rows = chart_rows(z, drawing_path)
            for code, source in sources.items():
                if source["sheet"] != sheet_name:
                    continue
                chart_info = rows.get(source["row"])
                if not chart_info:
                    continue
                source["trend"] = chart_values(z.read(chart_info["chart_path"]))
    return sources


def style_templates(sheet_root: ET.Element) -> dict[str, int]:
    return {
        "freq": cell_style(sheet_root, "I31", cell_style(sheet_root, "I20")),
        "code": cell_style(sheet_root, "J31", cell_style(sheet_root, "J20")),
        "date": cell_style(sheet_root, "L31", cell_style(sheet_root, "L20")),
        "percent_current": cell_style(sheet_root, "K25"),
        "percent_change": cell_style(sheet_root, "M25"),
        "numeric2_current": cell_style(sheet_root, "K31", cell_style(sheet_root, "K88")),
        "numeric2_change": cell_style(sheet_root, "M31", cell_style(sheet_root, "M88")),
        "general_current": cell_style(sheet_root, "K43"),
        "general_change": cell_style(sheet_root, "M43"),
    }


def style_for_value(row: int, col: int, sheet_root: ET.Element, templates: dict[str, int]) -> int | None:
    if col == 11:
        if row in PERCENT_DECIMAL_ROWS:
            return templates["percent_current"]
        if row in NUMERIC_2_ROWS:
            return templates["numeric2_current"]
        if row in GENERAL_NUMBER_ROWS:
            return templates["general_current"]
    if col == 13:
        if row in PERCENT_DECIMAL_ROWS:
            return templates["percent_change"]
        if row in NUMERIC_2_ROWS:
            return templates["numeric2_change"]
        if row in GENERAL_NUMBER_ROWS:
            return templates["general_change"]
    cell = find_existing_cell(sheet_root, f"{openpyxl.utils.get_column_letter(col)}{row}")
    return int(cell.attrib["s"]) if cell is not None and "s" in cell.attrib else None


def update_sheet_and_styles(
    tree_path: Path,
    daily_sources: dict[str, dict[str, Any]],
    today: date,
) -> tuple[dict[str, bytes], dict[str, Any]]:
    replacements: dict[str, bytes] = {}
    updated_rows: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    renamed: list[dict[str, Any]] = []
    style_fixes: list[int] = []
    color_rows: list[dict[str, Any]] = []

    with ZipFile(tree_path, "r") as z:
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        sheet_root = read_xml(z, sheet_path)
        styles_root = read_xml(z, "xl/styles.xml")
        templates = style_templates(sheet_root)
        style_cache: dict[tuple[int, str | None], int] = {}

        for row, code in ROW_CODE.items():
            source = daily_sources.get(code.upper())
            if source is None:
                missing.append({"row": row, "name": NO_DAILY_SOURCE_ROWS.get(row, ""), "code": code})
                continue

            if row in NAME_CORRECTIONS:
                set_inline_string(find_cell(sheet_root, row, 4), NAME_CORRECTIONS[row])
                renamed.append({"row": row, "name": NAME_CORRECTIONS[row]})

            set_inline_string(find_cell(sheet_root, row, 9), source["freq"], templates["freq"])
            set_inline_string(find_cell(sheet_root, row, 10), code, templates["code"])

            current_style = style_for_value(row, 11, sheet_root, templates)
            change_style = style_for_value(row, 13, sheet_root, templates)
            if row in PERCENT_DECIMAL_ROWS or row in NUMERIC_2_ROWS or row in GENERAL_NUMBER_ROWS:
                style_fixes.append(row)

            set_number(find_cell(sheet_root, row, 11), scale_for_tree(row, source["current"]), current_style)
            data_date_text = set_date(find_cell(sheet_root, row, 12), source["date"], templates["date"])
            change_value = scale_for_tree(row, source["change"])
            change_cell = find_cell(sheet_root, row, 13)
            set_number(change_cell, change_value, change_style)

            raw_change = to_number(source["change"])
            if data_date_text == today.strftime("%Y-%m-%d") and raw_change not in (None, 0):
                base_style = int(change_cell.attrib.get("s", str(change_style or 0)))
                rgb = RED if raw_change > 0 else GREEN
                change_cell.attrib["s"] = str(ensure_font_style(styles_root, base_style, rgb, style_cache))
                color_rows.append({"row": row, "change": source["change"], "rgb": rgb})

            updated_rows.append(
                {
                    "row": row,
                    "name": NAME_CORRECTIONS.get(row, source["name"]),
                    "code": code,
                    "source": f"{source['sheet']}!{source['row']}",
                    "freq": source["freq"],
                    "current_raw": source["current"],
                    "current_tree": scale_for_tree(row, source["current"]),
                    "date": data_date_text,
                    "change_raw": source["change"],
                    "change_tree": change_value,
                    "trend_points": len(source.get("trend") or []),
                }
            )

        replacements[sheet_path] = write_xml(sheet_root)
        replacements["xl/styles.xml"] = write_xml(styles_root)

    return replacements, {
        "updated_rows": updated_rows,
        "missing_rows": [{"row": row, "name": name} for row, name in NO_DAILY_SOURCE_ROWS.items()],
        "renamed_rows": renamed,
        "style_fixed_rows": sorted(set(style_fixes)),
        "colored_today_rows": color_rows,
    }


def update_charts(
    tree_path: Path,
    replacements: dict[str, bytes],
    daily_sources: dict[str, dict[str, Any]],
) -> tuple[dict[str, bytes], dict[str, bytes], dict[str, Any]]:
    new_files: dict[str, bytes] = {}
    meta: dict[str, Any] = {"charts_updated": [], "charts_added": [], "charts_skipped_no_trend": []}

    with ZipFile(tree_path, "r") as z:
        names = z.namelist()
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        drawing_path = sheet_drawing_path(z, sheet_path)
        if drawing_path is None:
            return replacements, new_files, meta | {"chart_error": "target sheet has no drawing"}
        drawing_rels = drawing_rels_path(drawing_path)
        drawing_root = read_xml(z, drawing_path)
        rels_root = read_xml(z, drawing_rels)
        content_types = read_xml(z, "[Content_Types].xml")
        template_anchor, template_chart_xml = find_template_anchor_and_chart(z, drawing_path, TEMPLATE_CHART_ROW)
        before_rows = existing_chart_rows(drawing_root)
        existing = chart_rows(z, drawing_path)

        next_chart_num = max_chart_number(names) + 1
        next_rid_num = max_rel_id(rels_root) + 1
        next_cnvpr = max_cnvpr_id(drawing_root) + 1
        rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"

        for row, code in ROW_CODE.items():
            source = daily_sources.get(code.upper())
            if source is None:
                continue
            values = [float(v) for v in (source.get("trend") or []) if to_number(v) is not None]
            if not values:
                meta["charts_skipped_no_trend"].append(row)
                continue

            if row in existing:
                chart_path = existing[row]["chart_path"]
                replacements[chart_path] = update_chart_series(z.read(chart_path), values)
                meta["charts_updated"].append(row)
                continue

            chart_path = f"xl/charts/chart{next_chart_num}.xml"
            rid = f"rId{next_rid_num}"
            new_files[chart_path] = update_chart_series(template_chart_xml, values)

            rel = ET.Element("Relationship", {"Id": rid, "Type": rel_type, "Target": "../charts/" + PurePosixPath(chart_path).name})
            rels_root.append(rel)
            drawing_root.append(make_anchor(template_anchor, row, rid, next_cnvpr))

            if not any(
                elem.attrib.get("PartName") == f"/{chart_path}"
                for elem in content_types
                if elem.tag.rsplit("}", 1)[-1] == "Override"
            ):
                override = ET.Element(
                    f"{{http://schemas.openxmlformats.org/package/2006/content-types}}Override",
                    {
                        "PartName": f"/{chart_path}",
                        "ContentType": "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
                    },
                )
                content_types.append(override)

            meta["charts_added"].append(row)
            next_chart_num += 1
            next_rid_num += 1
            next_cnvpr += 1

        replacements[drawing_path] = write_xml(drawing_root)
        replacements[drawing_rels] = write_xml(rels_root)
        replacements["[Content_Types].xml"] = write_xml(content_types)
        meta["chart_rows_before"] = len(before_rows)

    return replacements, new_files, meta


def write_workbook(tree_path: Path, output_path: Path, replacements: dict[str, bytes], new_files: dict[str, bytes]) -> None:
    with ZipFile(tree_path, "r") as zin, ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name in replacements:
                zout.writestr(name, replacements[name])
            else:
                zout.writestr(name, zin.read(name))
        for name, data in new_files.items():
            zout.writestr(name, data)


def verify_output(output_path: Path, daily_sources: dict[str, dict[str, Any]]) -> dict[str, Any]:
    with ZipFile(output_path, "r") as z:
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        drawing_path = sheet_drawing_path(z, sheet_path)
        chart_rows_after = chart_rows(z, drawing_path) if drawing_path else {}
        chart_missing = []
        chart_empty = []
        for row, code in ROW_CODE.items():
            source = daily_sources.get(code.upper())
            if source is None or not source.get("trend"):
                continue
            info = chart_rows_after.get(row)
            if not info:
                chart_missing.append(row)
                continue
            values = chart_values(z.read(info["chart_path"]))
            if not values:
                chart_empty.append(row)
        return {
            "chart_rows_after": len(chart_rows_after),
            "chart_missing_after": chart_missing,
            "chart_empty_after": chart_empty,
        }


def default_output_path(tree_path: Path) -> Path:
    base = tree_path.with_name(tree_path.stem + "_日报数据补充版.xlsx")
    if not base.exists():
        return base
    for idx in range(2, 30):
        candidate = tree_path.with_name(tree_path.stem + f"_日报数据补充版V{idx}.xlsx")
        if not candidate.exists():
            return candidate
    raise RuntimeError("No available output path.")


def sync(tree_path: Path, daily_path: Path, output_path: Path) -> dict[str, Any]:
    daily_sources = build_daily_sources(daily_path)
    replacements, data_meta = update_sheet_and_styles(tree_path, daily_sources, datetime.now().date())
    replacements, new_files, chart_meta = update_charts(tree_path, replacements, daily_sources)
    write_workbook(tree_path, output_path, replacements, new_files)
    verify_meta = verify_output(output_path, daily_sources)
    return {
        "tree": str(tree_path),
        "daily": str(daily_path),
        "output": str(output_path),
        **data_meta,
        **chart_meta,
        **verify_meta,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill newly inserted TREE V3.0 indicators from the refreshed daily workbook.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output")
    args = parser.parse_args()

    tree_path = Path(args.tree)
    daily_path = Path(args.daily)
    output_path = Path(args.output) if args.output else default_output_path(tree_path)
    result = sync(tree_path, daily_path, output_path)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
