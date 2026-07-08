from __future__ import annotations

import argparse
import json
import math
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl


TREE_SHEET = "重点策略跟踪情况(V2.5)"
INDEX_SHEET = "指数走势"
ROWS = {
    "SPX.GI": 114,
    "IXIC.GI": 115,
    "DJI.GI": 116,
}
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {
    "main": SHEET_NS,
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

ET.register_namespace("", SHEET_NS)
ET.register_namespace("r", NS["officeRel"])


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def q(tag: str) -> str:
    return f"{{{SHEET_NS}}}{tag}"


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib["name"] == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    sheet_data.append(new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def set_number(cell: ET.Element, value: float) -> None:
    style = cell.attrib.get("s")
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    if style is not None:
        cell.attrib["s"] = style
    v = ET.SubElement(cell, q("v"))
    v.text = format(float(value), ".15g")


def index_pct_changes(daily_path: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(daily_path, data_only=True)
    ws = wb[INDEX_SHEET]
    out: dict[str, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        code = str(ws.cell(row, 5).value or "").strip().upper()
        if code not in ROWS:
            continue
        points = []
        for col in range(6, ws.max_column + 1):
            value = to_number(ws.cell(row, col).value)
            if value is not None:
                points.append((ws.cell(1, col).value, value))
        if len(points) < 2:
            raise ValueError(f"Not enough points for {code}")
        prev_date, prev = points[-2]
        current_date, current = points[-1]
        if prev == 0:
            raise ValueError(f"Previous value is zero for {code}")
        out[code] = {
            "daily_row": row,
            "name": ws.cell(row, 4).value,
            "previous_date": prev_date,
            "previous": prev,
            "current_date": current_date,
            "current": current,
            "pct_change": (current - prev) / prev,
        }
    missing = sorted(set(ROWS) - set(out))
    if missing:
        raise KeyError(missing)
    return out


def fix_tree(daily_path: Path, tree_path: Path, output_path: Path) -> dict[str, Any]:
    changes = index_pct_changes(daily_path)
    with ZipFile(tree_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)
        for code, item in changes.items():
            row = ROWS[code]
            set_number(find_cell(sheet_root, row, 13), item["pct_change"])
        sheet_xml = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, sheet_xml)
                else:
                    zout.writestr(name, zin.read(name))
    return {
        "output": str(output_path),
        "fixed": [
            {"code": code, "tree_row": ROWS[code], **item}
            for code, item in changes.items()
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--daily", required=True)
    parser.add_argument("--tree", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    result = fix_tree(Path(args.daily), Path(args.tree), Path(args.output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
