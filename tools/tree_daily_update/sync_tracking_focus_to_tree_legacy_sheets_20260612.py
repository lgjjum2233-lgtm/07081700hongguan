from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "（6月12日V13）TREE宏观分析_重点跟踪项补全.xlsx"
OUTPUT = ROOT / "（6月12日V14）TREE宏观分析_重点跟踪项全表同步.xlsx"
AUDIT = ROOT / "20260612_TREE重点跟踪项全表同步复核.xlsx"
CANONICAL_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def normalize_name(value: Any) -> str:
    s = text(value)
    s = re.sub(r"[①②③④⑤⑥⑦⑧⑨⑩]", "", s)
    s = re.sub(r"[：:]", "", s)
    s = re.sub(r"\（.*?\）|\(.*?\)", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def compact(value: Any) -> str:
    return text(value).replace("\n", "").replace(" ", "")


def has_any(haystack: str, words: list[str]) -> bool:
    return any(word in haystack for word in words)


def is_macro(big: str, dim: str) -> bool:
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    if "产业" in dim:
        return False
    return True


def move_charts_right(ws, insert_col: int) -> None:
    zero_col = insert_col - 1
    for chart in ws._charts:
        anchor = chart.anchor
        if not hasattr(anchor, "_from"):
            continue
        if anchor._from.col >= zero_col:
            anchor._from.col += 1
        if hasattr(anchor, "to") and anchor.to.col >= zero_col:
            anchor.to.col += 1


def copy_style(source, target) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy.copy(source.alignment)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.protection = copy.copy(source.protection)


def build_canonical_maps(wb) -> tuple[dict[str, str], dict[str, str]]:
    ws = wb[CANONICAL_SHEET]
    by_code: dict[str, str] = {}
    by_name: dict[str, str] = {}
    big = dim = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, 1).value):
            big = text(ws.cell(row, 1).value)
        if text(ws.cell(row, 2).value):
            dim = text(ws.cell(row, 2).value)
        name = text(ws.cell(row, 4).value)
        if not name:
            continue
        focus = text(ws.cell(row, 17).value)
        if not focus or focus == "-":
            continue
        code = text(ws.cell(row, 10).value)
        if code and code not in {"—", "-"}:
            by_code[code] = focus
        by_name[normalize_name(name)] = focus
    return by_code, by_name


def decide_focus(name: str, freq: str, code: str, caliber_or_definition: str, by_code: dict[str, str], by_name: dict[str, str]) -> tuple[str, str]:
    if code in by_code:
        return by_code[code], "按V3.0同指标代码同步。"

    norm = normalize_name(name)
    if norm in by_name:
        return by_name[norm], "按V3.0同名指标同步。"

    t = compact(f"{name} {freq} {code} {caliber_or_definition}")
    if has_any(t, ["社会融资", "新增人民币贷款", "政府债券", "票据融资", "企业债券融资", "新增外币贷款", "同比多增"]):
        return "同比增量", "信用流量类指标，市场常用同比多增/少增判断信用扩张强弱。"
    if has_any(t, ["环比"]):
        return "环比", "名称或口径已体现环比，优先沿用该指标自身主读数。"
    if has_any(t, ["利率", "收益率", "利差", "DR001", "DR007", "R001", "R007", "SHIBOR", "SOFR", "IORB", "LPR", "TGA余额", "逆回购规模", "资产负债表"]):
        return "环比", "利率、利差和高频流动性余额类指标，市场通常关注边际变化。"
    if has_any(t, ["公开市场操作", "净投放", "总投放", "财政存款变动", "财政资金"]):
        return "环比", "流动性投放/回笼类指标本身是边际流量，更适合看上期变化。"
    if has_any(t, ["PMI", "ISM", "信心指数", "乐观指数", "新订单", "扩散指数"]):
        return "环比", "景气扩散指数通常看本期较上期改善或走弱。"
    if has_any(t, ["杠杆率", "财政赤字脉冲", "赤字率", "贡献率", "利润率"]):
        return "同比增量", "比例、脉冲和贡献率类指标用百分点变化更直观。"
    if has_any(t, ["货币政策", "政策报告", "关注货币政策"]):
        return "定性跟踪", "政策文本不是连续数值指标，重点跟踪政策取向和表述边际变化。"
    if has_any(t, ["同比", "累计同比", "当月同比", "增速", "M1", "M2", "GDP", "CPI", "PPI", "社零", "社会消费品", "固定资产投资", "出口", "工业增加值", "利润总额"]):
        return "同比", "增长、通胀和金额规模类指标通常优先看同比趋势。"
    if code in {"—", "-", ""}:
        return "定性跟踪", "无连续数值代码，作为定性跟踪项处理。"
    return "同比", "默认按宏观增长/规模类指标处理，优先观察同比趋势。"


def find_focus_col(ws) -> int | None:
    for col in range(1, ws.max_column + 1):
        if text(ws.cell(HEADER_ROW, col).value) == "重点跟踪项":
            return col
    return None


def insert_focus_col(ws, sheet_type: str) -> int:
    if sheet_type == "v3":
        insert_col = 18  # after Q 边际变化
        style_col = 17
    else:
        insert_col = 14  # after M 边际变化
        style_col = 13

    ws.insert_cols(insert_col)
    move_charts_right(ws, insert_col)
    for row in range(HEADER_ROW, ws.max_row + 1):
        copy_style(ws.cell(row, style_col), ws.cell(row, insert_col))
        ws.cell(row, insert_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if row > HEADER_ROW:
            ws.cell(row, insert_col).fill = PatternFill(fill_type=None)
    ws.cell(HEADER_ROW, insert_col).value = "重点跟踪项"
    ws.column_dimensions[get_column_letter(insert_col)].width = 13
    return insert_col


def sheet_type(ws) -> str | None:
    headers = {text(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    if headers.get("监测指标") == 9 and headers.get("边际变化") == 17:
        return "v3"
    if headers.get("监测指标") == 4 and headers.get("边际变化") == 13:
        return "v25"
    return None


def sync_sheet(ws, by_code: dict[str, str], by_name: dict[str, str]) -> list[dict[str, Any]]:
    stype = sheet_type(ws)
    if not stype:
        return []

    focus_col = find_focus_col(ws)
    inserted = False
    if focus_col is None:
        focus_col = insert_focus_col(ws, stype)
        inserted = True

    if stype == "v3":
        big_col, dim_col, sub_col, name_col, freq_col, code_col, caliber_col = 6, 7, 8, 9, 12, None, 10
    else:
        big_col, dim_col, sub_col, name_col, freq_col, code_col, caliber_col = 1, 2, 3, 4, 9, 10, 7

    rows: list[dict[str, Any]] = []
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, big_col).value):
            big = text(ws.cell(row, big_col).value)
        if text(ws.cell(row, dim_col).value):
            dim = text(ws.cell(row, dim_col).value)
        if text(ws.cell(row, sub_col).value):
            sub = text(ws.cell(row, sub_col).value)

        name = text(ws.cell(row, name_col).value)
        if not name or not is_macro(big, dim):
            continue

        freq = text(ws.cell(row, freq_col).value)
        code = text(ws.cell(row, code_col).value) if code_col else ""
        caliber = text(ws.cell(row, caliber_col).value)
        old_focus = text(ws.cell(row, focus_col).value)
        focus, reason = decide_focus(name, freq, code, caliber, by_code, by_name)
        ws.cell(row, focus_col).value = focus
        rows.append(
            {
                "sheet": ws.title,
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "code": code,
                "old_focus": old_focus,
                "new_focus": focus,
                "inserted": inserted,
                "reason": reason,
            }
        )
    return rows


def build_audit(rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重点跟踪项全表同步"
    headers = ["表名", "行号", "一级分类", "维度", "子维度", "指标名称", "指标代码", "原重点跟踪项", "新重点跟踪项", "判断依据"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in ["sheet", "row", "big", "dim", "sub", "name", "code", "old_focus", "new_focus", "reason"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [28, 8, 14, 14, 20, 34, 24, 14, 14, 58]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(AUDIT)


def main() -> None:
    wb = openpyxl.load_workbook(INPUT)
    by_code, by_name = build_canonical_maps(wb)

    audit_rows: list[dict[str, Any]] = []
    for sheet_name in [s for s in wb.sheetnames if "重点策略跟踪情况" in s and s != CANONICAL_SHEET]:
        audit_rows.extend(sync_sheet(wb[sheet_name], by_code, by_name))

    wb.save(OUTPUT)
    build_audit(audit_rows)
    print(f"saved: {OUTPUT}")
    print(f"audit: {AUDIT}")
    print(f"synced rows: {len(audit_rows)}")
    for row in audit_rows:
        if "社会融资规模" in row["name"] or "新增人民币贷款" in row["name"]:
            print(f'{row["sheet"]} R{row["row"]} {row["name"]} -> {row["new_focus"]}')


if __name__ == "__main__":
    main()
