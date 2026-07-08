from __future__ import annotations

import argparse
import copy
import json
import math
from bisect import bisect_right
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import add_tree_yoy_column_20260611 as yoy_base
import apply_macro_change_history_colors as color_base
import apply_tree_current_value_history_colors_20260611 as current_color


TREE_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5
CURRENT_COL = 11  # K
DATE_COL = 12  # L
YOY_COL = 13  # M, user-renamed "同比"
INSERT_COL = 14  # N, insert "同比增量" before user-renamed "环比"
CODE_COL = 10
CALIBER_COL = 16  # P before insertion
LOOKBACK_DAYS = 365
MIN_HISTORY_VALUES = 4
MIN_PERCENTILE_TO_FILL = 0.55


def to_number(value: Any) -> float | None:
    return yoy_base.to_number(value)


def parse_date(value: Any) -> date | None:
    return yoy_base.parse_date(value)


def display_scale(current_display: Any, current_raw: float | None) -> float:
    display = to_number(current_display)
    if display is None or current_raw is None or abs(current_raw) <= 1e-12:
        return 1.0
    ratio = display / current_raw
    candidates = [1.0, 0.01, 0.0001, 100.0]
    best = min(candidates, key=lambda item: abs(math.log(abs(ratio / item))) if ratio and item else float("inf"))
    # Only apply a scale when it is clearly a unit conversion, not a small date/value mismatch.
    if best == 1.0:
        return 1.0
    if abs(ratio / best - 1.0) <= 0.20:
        return best
    return 1.0


def yearly_increment(
    series: yoy_base.MacroSeries,
    current_dt: date,
    current_display: Any,
) -> tuple[Any, dict[str, Any]]:
    current_point = yoy_base.closest_current_period(series.points, current_dt)
    if current_point is None:
        return "-", {"basis": "底层宏观数据找不到当前日期值"}
    current_point_dt, current_raw = current_point
    prior = yoy_base.closest_prior_period(series.points, current_dt)
    if prior is None:
        return "-", {
            "source_current_date": current_point_dt.isoformat(),
            "source_current_value": current_raw,
            "basis": "底层宏观数据找不到去年同期值",
        }
    prior_dt, prior_raw = prior
    scale = display_scale(current_display, current_raw)
    value = (current_raw - prior_raw) * scale
    return value, {
        "source_current_date": current_point_dt.isoformat(),
        "source_current_value": current_raw,
        "prior_date": prior_dt.isoformat(),
        "prior_value": prior_raw,
        "scale": scale,
        "basis": "同比增量=今年同期值-去年同期值，并按TREE展示单位缩放",
    }


def increment_history(
    series: yoy_base.MacroSeries,
    current_dt: date,
    current_display: Any,
) -> tuple[list[tuple[date, float]], float]:
    current_point = yoy_base.closest_current_period(series.points, current_dt)
    scale = display_scale(current_display, current_point[1] if current_point else None)
    start = current_dt - timedelta(days=LOOKBACK_DAYS)
    out: list[tuple[date, float]] = []
    for d, value in sorted(series.points, key=lambda item: item[0]):
        if not (start <= d <= current_dt):
            continue
        prior = yoy_base.closest_prior_period(series.points, d)
        if prior is None:
            continue
        _, prior_value = prior
        inc = (value - prior_value) * scale
        if math.isfinite(inc):
            out.append((d, inc))
    return out, scale


def color_for_increment(value: Any, history: list[tuple[date, float]]) -> tuple[str | None, str, dict[str, Any]]:
    latest = to_number(value)
    values = [item[1] for item in history if math.isfinite(item[1])]
    detail = {
        "history_count": len(values),
        "history_start": history[0][0].isoformat() if history else "",
        "history_end": history[-1][0].isoformat() if history else "",
        "percentile": None,
        "intensity": None,
    }
    if latest is None or not math.isfinite(latest):
        return None, color_base.BLACK, detail
    if abs(latest) <= 1e-12:
        detail["percentile"] = 0.0
        detail["intensity"] = 0.0
        return None, color_base.BLACK, detail
    if len(values) < MIN_HISTORY_VALUES:
        return None, color_base.BLACK, detail
    pct = bisect_right(sorted(abs(item) for item in values), abs(latest)) / len(values)
    detail["percentile"] = pct
    if pct < MIN_PERCENTILE_TO_FILL:
        detail["intensity"] = 0.0
        return None, color_base.BLACK, detail
    intensity = max(0.15, min(1.0, (pct - MIN_PERCENTILE_TO_FILL) / (1.0 - MIN_PERCENTILE_TO_FILL)))
    detail["intensity"] = intensity
    fill = (
        color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
        if latest > 0
        else color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
    )
    return fill, color_base.text_color_for_fill(fill), detail


def apply_style(cell, fill_rgb: str | None, font_rgb: str) -> None:
    font = copy.copy(cell.font)
    font.color = font_rgb
    cell.font = font
    if fill_rgb:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_rgb)
    else:
        cell.fill = PatternFill(fill_type=None)


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.protection = copy.copy(src.protection)


def move_charts_right(ws, insert_col: int) -> None:
    zero_col = insert_col - 1
    for chart in ws._charts:
        anchor = getattr(chart, "anchor", None)
        if not hasattr(anchor, "_from"):
            continue
        if anchor._from.col >= zero_col:
            anchor._from.col += 1
        if hasattr(anchor, "to") and anchor.to.col >= zero_col:
            anchor.to.col += 1


def write_audit(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "同比增量颜色复核"
    headers = [
        "TREE行号",
        "大分类",
        "细分维度",
        "子维度",
        "指标名称",
        "指标代码",
        "数据日期",
        "当前值",
        "日报当前日期",
        "日报当前值",
        "去年同期日期",
        "去年同期值",
        "同比增量",
        "缩放比例",
        "近一年样本数",
        "历史起点",
        "历史终点",
        "历史分位",
        "填充色",
        "字体色",
        "备注",
    ]
    ws.append(headers)
    for item in rows:
        ws.append([item.get(header, "") for header in headers])
    widths = [10, 18, 18, 18, 34, 22, 14, 14, 14, 14, 14, 14, 14, 10, 12, 14, 14, 12, 14, 12, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def add_column(input_tree: Path, daily_path: Path, output_tree: Path, audit_output: Path) -> dict[str, Any]:
    macro_series = yoy_base.read_macro_series(daily_path)
    wb = openpyxl.load_workbook(input_tree, data_only=False, read_only=False)
    ws = wb[TREE_SHEET]

    # Collect row metadata before inserting the column because "数据口径" moves right after insertion.
    rows = current_color.iter_tree_rows(ws)
    for row in rows:
        row["caliber"] = ws.cell(row["row"], CALIBER_COL).value

    ws.insert_cols(INSERT_COL, 1)
    move_charts_right(ws, INSERT_COL)
    ws.cell(HEADER_ROW, INSERT_COL).value = "同比增量"
    copy_cell_style(ws.cell(HEADER_ROW, INSERT_COL + 1), ws.cell(HEADER_ROW, INSERT_COL))

    styled = 0
    positive_styled = 0
    negative_styled = 0
    dash_rows = 0
    no_history = 0
    unmatched = 0
    nonmacro = 0
    audit_rows: list[dict[str, Any]] = []

    for row in rows:
        r = row["row"]
        cell = ws.cell(r, INSERT_COL)
        copy_cell_style(ws.cell(r, INSERT_COL + 1), cell)
        note = ""
        value: Any = "-"
        detail: dict[str, Any] = {}
        fill_rgb: str | None = None
        font_rgb = color_base.BLACK
        color_detail: dict[str, Any] = {"history_count": 0, "history_start": "", "history_end": "", "percentile": None}

        if not yoy_base.is_macro_tree_row(row["big"], row["dim"]):
            nonmacro += 1
            note = "非宏观计算区，按要求用-"
        else:
            series = macro_series.get(row["code"])
            current_dt = parse_date(row["current_date"])
            if series is None or current_dt is None:
                unmatched += 1
                note = "未匹配到底层宏观数据或日期为空"
            else:
                value, detail = yearly_increment(series, current_dt, row["current"])
                if value == "-":
                    note = detail.get("basis", "无法计算")
                else:
                    history, _ = increment_history(series, current_dt, row["current"])
                    fill_rgb, font_rgb, color_detail = color_for_increment(value, history)
                    if fill_rgb:
                        styled += 1
                        if value > 0:
                            positive_styled += 1
                        else:
                            negative_styled += 1
                    elif color_detail.get("history_count", 0) < MIN_HISTORY_VALUES and abs(float(value)) > 1e-12:
                        no_history += 1
                        note = "近一年同比增量历史样本不足"
                    else:
                        note = "历史分位未达到填色阈值或数值接近0"

        if value == "-":
            cell.value = "-"
            dash_rows += 1
        else:
            cell.value = value
        apply_style(cell, fill_rgb, font_rgb)

        audit_rows.append(
            {
                "TREE行号": r,
                "大分类": row["big"],
                "细分维度": row["dim"],
                "子维度": row["sub"],
                "指标名称": row["name"],
                "指标代码": row["code"],
                "数据日期": row["current_date"],
                "当前值": row["current"],
                "日报当前日期": detail.get("source_current_date", ""),
                "日报当前值": detail.get("source_current_value", ""),
                "去年同期日期": detail.get("prior_date", ""),
                "去年同期值": detail.get("prior_value", ""),
                "同比增量": value,
                "缩放比例": detail.get("scale", ""),
                "近一年样本数": color_detail.get("history_count", 0),
                "历史起点": color_detail.get("history_start", ""),
                "历史终点": color_detail.get("history_end", ""),
                "历史分位": color_detail.get("percentile"),
                "填充色": fill_rgb or "",
                "字体色": font_rgb,
                "备注": note,
            }
        )

    wb.save(output_tree)
    wb.close()
    write_audit(audit_output, audit_rows)
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        "rows": len(rows),
        "styled_rows": styled,
        "positive_styled_rows": positive_styled,
        "negative_styled_rows": negative_styled,
        "dash_rows": dash_rows,
        "nonmacro_rows": nonmacro,
        "unmatched_rows": unmatched,
        "no_history_rows": no_history,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()
    result = add_column(Path(args.tree), Path(args.daily), Path(args.output_tree), Path(args.audit_output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
