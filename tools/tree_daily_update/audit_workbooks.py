from __future__ import annotations

import argparse
import json
import math
import re
import subprocess
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl

from postprocess_daily import (
    NS,
    as_text,
    close_enough,
    load_config,
    norm,
    read_xml,
    rel_target,
    to_number,
    values_from_range_formula,
    workbook_sheet_paths,
    sheet_drawing_path,
    drawing_rels_path,
)


ERROR_TERMS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A")


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        dt = datetime(1899, 12, 30) + timedelta(days=float(value))
        return dt.strftime("%Y-%m-%d")
    return "" if value is None else str(value)


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def header_row_for(ws, row: int) -> int:
    for header in range(row, 0, -1):
        if as_text(ws.cell(header, 5).value).lower() in {"date", "日期"}:
            return header
    return 1


def last_points(ws, row: int) -> list[tuple[Any, float]]:
    header = header_row_for(ws, row)
    points = []
    for col in range(6, ws.max_column + 1):
        n = to_number(ws.cell(row, col).value)
        if n is not None:
            points.append((ws.cell(header, col).value, n))
    return points


def parse_wsd_coverage(path: Path, sheet_name: str) -> dict[int, dict[str, Any]]:
    coverage: dict[int, dict[str, Any]] = {}
    with ZipFile(path) as z:
        sheet_paths = workbook_sheet_paths(z)
        if sheet_name not in sheet_paths:
            return coverage
        root = read_xml(z, sheet_paths[sheet_name])
        for cell in root.findall(".//main:c", NS):
            f = cell.find("main:f", NS)
            if f is None or not f.text or "wsd(" not in f.text.lower():
                continue
            text = f.text
            for match in re.finditer(r"E(\d+):E(\d+)", text, flags=re.IGNORECASE):
                start = int(match.group(1))
                end = int(match.group(2))
                for row in range(start, end + 1):
                    coverage[row] = {"anchor": cell.attrib.get("r"), "formula": text[:240]}
    return coverage


def find_base_source(wb, daily_path: Path, source_sheet: str, code: str, name: str) -> dict[str, Any] | None:
    if source_sheet not in wb.sheetnames:
        return None
    ws = wb[source_sheet]
    wsd_coverage = parse_wsd_coverage(daily_path, source_sheet)
    code_key = code.upper()
    best = None
    for row in range(2, ws.max_row + 1):
        row_code = as_text(ws.cell(row, 5).value).upper()
        row_name = as_text(ws.cell(row, 4).value)
        if row_code != code_key and norm(row_name) != norm(name):
            continue
        points = last_points(ws, row)
        if not points:
            continue
        latest = points[-1]
        previous = points[-2] if len(points) >= 2 else (None, None)
        item = {
            "sheet": source_sheet,
            "row": row,
            "name": row_name,
            "code": row_code,
            "current": latest[1],
            "date": latest[0],
            "previous": previous[1],
            "previous_date": previous[0],
            "points": len(points),
            "wind_formula_covered": row in wsd_coverage,
            "wind_formula_anchor": wsd_coverage.get(row, {}).get("anchor"),
        }
        if best is None or item["points"] > best["points"]:
            best = item
    return best


def find_front_source(wb, source_sheet: str, code: str, name: str) -> dict[str, Any] | None:
    if source_sheet not in wb.sheetnames:
        return None
    ws = wb[source_sheet]
    code_key = code.upper()
    for row in range(2, ws.max_row + 1):
        row_code = as_text(ws.cell(row, 3).value).upper()
        row_name = as_text(ws.cell(row, 1).value)
        if row_code == code_key or norm(row_name) == norm(name):
            return {
                "sheet": source_sheet,
                "row": row,
                "name": row_name,
                "code": row_code,
                "current": ws.cell(row, 5).value,
                "date": ws.cell(row, 6).value,
                "previous": ws.cell(row, 8).value,
                "previous_date": ws.cell(row, 9).value,
                "points": None,
                "wind_formula_covered": True,
                "wind_formula_anchor": "front_or_dataset",
            }
    return None


def mandatory_sources(daily_path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(daily_path, data_only=True)
    out = []
    base_sheets = set(config["base_sheets"])
    for item in config["mandatory_sources"]:
        source = (
            find_base_source(wb, daily_path, item["source_sheet"], item["code"], item["name"])
            if item["source_sheet"] in base_sheets
            else find_front_source(wb, item["source_sheet"], item["code"], item["name"])
        )
        status = {
            "name": item["name"],
            "code": item["code"],
            "expected_sheet": item["source_sheet"],
            "found": source is not None,
            "tree_row": item.get("tree_row"),
        }
        if source:
            status |= {
                "source_sheet": source["sheet"],
                "source_row": source["row"],
                "date": date_text(source["date"]),
                "current": source["current"],
                "previous": source.get("previous"),
                "previous_date": date_text(source.get("previous_date")),
                "points": source["points"],
                "wind_formula_covered": source["wind_formula_covered"],
                "wind_formula_anchor": source["wind_formula_anchor"],
            }
        out.append(status)
    return out


def chart_rows(z: ZipFile, drawing_path: str) -> list[dict[str, Any]]:
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    drawing = read_xml(z, drawing_path)
    rows = []
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        cref = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or cref is None:
            continue
        rid = cref.attrib[f"{{{NS['officeRel']}}}id"]
        rows.append({"row": int(frm.findtext("xdr:row", namespaces=NS)) + 1, "chart_path": rel_map[rid]})
    return rows


def chart_formula_and_values(z: ZipFile, chart_path: str, wb_values) -> tuple[str, list[float], int]:
    root = read_xml(z, chart_path)
    f = root.find(".//c:lineChart/c:ser/c:val/c:numRef/c:f", NS)
    formula = f.text if f is not None and f.text else ""
    values = [to_number(v) for v in values_from_range_formula(wb_values, formula)]
    numeric = [float(v) for v in values if v is not None]
    error_like_count = sum(1 for v in values_from_range_formula(wb_values, formula) if isinstance(v, str) and v.startswith("#"))
    return formula, numeric, error_like_count


def audit_daily_charts(daily_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    wb_values = openpyxl.load_workbook(daily_path, data_only=True)
    mismatches = []
    bad_errors = []
    chart_count = 0
    with ZipFile(daily_path) as z:
        sheet_paths = workbook_sheet_paths(z)
        for sheet_name in config["front_sheets"]:
            if sheet_name not in sheet_paths:
                continue
            drawing_path = sheet_drawing_path(z, sheet_paths[sheet_name])
            if drawing_path is None:
                continue
            ws = wb_values[sheet_name]
            for item in chart_rows(z, drawing_path):
                chart_count += 1
                formula, values, formula_error_count = chart_formula_and_values(z, item["chart_path"], wb_values)
                current = ws.cell(item["row"], 5).value
                if formula_error_count:
                    bad_errors.append({"sheet": sheet_name, "row": item["row"], "name": ws.cell(item["row"], 1).value, "formula": formula})
                if values and to_number(current) is not None and not close_enough(values[-1], current):
                    mismatches.append(
                        {
                            "sheet": sheet_name,
                            "row": item["row"],
                            "name": ws.cell(item["row"], 1).value,
                            "current": current,
                            "chart_last": values[-1],
                        }
                    )
    return {
        "chart_count": chart_count,
        "last_mismatch_count": len(mismatches),
        "formula_error_chart_count": len(bad_errors),
        "mismatch_sample": mismatches[:20],
        "formula_error_sample": bad_errors[:20],
    }


def whitelisted_error(sheet: str, value: str, config: dict[str, Any]) -> bool:
    for item in config.get("formula_error_whitelist", []):
        if item.get("sheet") != sheet:
            continue
        if any(term in value for term in item.get("terms", [])):
            return True
    return False


def audit_formula_errors(path: Path, config: dict[str, Any], tree_sheet: str | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    raw = []
    non_whitelist = []
    tree_sheet_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(term in value for term in ERROR_TERMS):
                    hit = {"sheet": ws.title, "cell": cell.coordinate, "value": value}
                    raw.append(hit)
                    if tree_sheet and ws.title == tree_sheet:
                        tree_sheet_errors.append(hit)
                    if not whitelisted_error(ws.title, value, config):
                        non_whitelist.append(hit)
    return {
        "raw_count": len(raw),
        "raw_by_sheet": dict(Counter(item["sheet"] for item in raw)),
        "non_whitelist_count": len(non_whitelist),
        "non_whitelist_sample": non_whitelist[:20],
        "tree_sheet_error_count": len(tree_sheet_errors),
        "tree_sheet_error_sample": tree_sheet_errors[:20],
    }


def audit_tree_key_rows(tree_path: Path, daily_status: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(tree_path, data_only=True)
    ws = wb[config["tree_sheet"]]
    cols = config["tree_columns"]
    by_code = {item["code"].upper(): item for item in daily_status if item.get("found")}
    pct_change_codes = {str(code).upper() for code in config.get("stock_index_pct_change_codes", [])}
    mismatches = []
    checked = 0
    for item in config["mandatory_sources"]:
        row = item.get("tree_row")
        if not row:
            continue
        source = by_code.get(item["code"].upper())
        if not source:
            mismatches.append({"row": row, "name": item["name"], "issue": "daily_source_missing"})
            continue
        checked += 1
        tree_current = ws.cell(row, cols["current"]).value
        tree_date = ws.cell(row, cols["date"]).value
        expected_current = to_number(source["current"])
        if expected_current is not None:
            expected_current *= float(item.get("tree_scale", 1))
        issue = {}
        if expected_current is not None and not close_enough(tree_current, expected_current):
            issue["current"] = {"tree": tree_current, "daily": expected_current}
        if date_text(tree_date) != source["date"]:
            issue["date"] = {"tree": date_text(tree_date), "daily": source["date"]}
        if item["code"].upper() in pct_change_codes:
            previous = to_number(source.get("previous"))
            current = to_number(source.get("current"))
            tree_change = ws.cell(row, cols["change"]).value
            if previous not in (None, 0) and current is not None:
                expected_change = (current - previous) / previous
                if not close_enough(tree_change, expected_change):
                    issue["change"] = {"tree": tree_change, "daily_pct_change": expected_change}
        if issue:
            mismatches.append({"row": row, "name": item["name"], "code": item["code"], "issue": issue})
    return {"checked": checked, "mismatch_count": len(mismatches), "mismatch_sample": mismatches[:20]}


def run_deep_audit(
    tree_path: Path,
    daily_path: Path,
    config: dict[str, Any],
    skip: bool,
    previous_tree: Path | None = None,
) -> dict[str, Any] | None:
    if skip:
        return None
    workspace = Path(config["workspace"])
    script = workspace / config.get("deep_audit_script", "")
    if not script.exists():
        return {"available": False, "reason": f"not found: {script}"}
    cmd = [sys.executable, str(script), "--tree", str(tree_path), "--daily", str(daily_path)]
    if previous_tree:
        cmd.extend(["--previous-tree", str(previous_tree)])
    try:
        proc = subprocess.run(cmd, cwd=workspace, text=True, capture_output=True, timeout=240, check=False)
    except Exception as exc:
        return {"available": True, "error": str(exc)}
    if proc.returncode != 0:
        return {"available": True, "returncode": proc.returncode, "stderr": proc.stderr[-1000:]}
    try:
        data = json.loads(proc.stdout)
    except json.JSONDecodeError:
        return {"available": True, "parse_error": True, "stdout_tail": proc.stdout[-1000:]}
    return {
        "available": True,
        "dates_stale_count": data.get("dates", {}).get("stale_count"),
        "dates_checked_rows": data.get("dates", {}).get("checked_rows"),
        "colors_bad_count": data.get("colors", {}).get("bad_color_count"),
        "colors_extra_count": data.get("colors", {}).get("extra_colored_count"),
        "tree_chart_style_issue_count": data.get("charts", {}).get("style_issue_count"),
        "tree_key_chart_mismatch_count": data.get("charts", {}).get("key_chart_mismatch_count"),
        "tree_sheet_error_count": data.get("formulas", {}).get("tree_sheet_error_count"),
        "unmatched_count": data.get("dates", {}).get("unmatched_count"),
    }


def audit(
    daily_path: Path,
    tree_path: Path | None,
    config: dict[str, Any],
    skip_deep: bool,
    previous_tree: Path | None = None,
) -> dict[str, Any]:
    daily_sources = mandatory_sources(daily_path, config)
    missing_sources = [item for item in daily_sources if not item["found"]]
    uncovered_wind = [
        item
        for item in daily_sources
        if item["found"] and item["expected_sheet"] in config["base_sheets"] and not item.get("wind_formula_covered")
    ]
    daily_charts = audit_daily_charts(daily_path, config)
    daily_errors = audit_formula_errors(daily_path, config)

    result: dict[str, Any] = {
        "daily": {
            "path": str(daily_path),
            "mandatory_source_count": len(daily_sources),
            "missing_source_count": len(missing_sources),
            "uncovered_wind_source_count": len(uncovered_wind),
            "source_sample": daily_sources[:20],
            "missing_sources": missing_sources,
            "uncovered_wind_sources": uncovered_wind,
            "charts": daily_charts,
            "formula_errors": daily_errors,
        }
    }
    pass_checks = (
        len(missing_sources) == 0
        and len(uncovered_wind) == 0
        and daily_charts["last_mismatch_count"] == 0
        and daily_charts["formula_error_chart_count"] == 0
        and daily_errors["non_whitelist_count"] == 0
    )

    if tree_path:
        tree_errors = audit_formula_errors(tree_path, config, tree_sheet=config["tree_sheet"])
        key_rows = audit_tree_key_rows(tree_path, daily_sources, config)
        deep = run_deep_audit(tree_path, daily_path, config, skip=skip_deep, previous_tree=previous_tree)
        result["tree"] = {
            "path": str(tree_path),
            "key_rows": key_rows,
            "formula_errors": tree_errors,
            "deep_audit": deep,
        }
        pass_checks = pass_checks and key_rows["mismatch_count"] == 0 and tree_errors["tree_sheet_error_count"] == 0
        if deep and deep.get("available"):
            pass_checks = pass_checks and all(
                (deep.get(key) in (0, None))
                for key in [
                    "dates_stale_count",
                    "colors_bad_count",
                    "colors_extra_count",
                    "tree_chart_style_issue_count",
                    "tree_key_chart_mismatch_count",
                    "tree_sheet_error_count",
                ]
            )

    result["summary"] = {"pass": pass_checks}
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Compact TREE/daily workbook audit.")
    parser.add_argument("--daily", required=True)
    parser.add_argument("--tree")
    parser.add_argument("--config", default=str(Path(__file__).with_name("config.json")))
    parser.add_argument("--skip-deep", action="store_true")
    parser.add_argument("--previous-tree")
    parser.add_argument("--output")
    args = parser.parse_args()

    config = load_config(Path(args.config))
    result = audit(
        Path(args.daily),
        Path(args.tree) if args.tree else None,
        config,
        args.skip_deep,
        previous_tree=Path(args.previous_tree) if args.previous_tree else None,
    )
    text = json.dumps(result, ensure_ascii=False, indent=2, default=json_default)
    if args.output:
        Path(args.output).write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
