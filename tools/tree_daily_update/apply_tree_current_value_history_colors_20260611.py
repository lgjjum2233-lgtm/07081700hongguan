from __future__ import annotations

import argparse
import json
import math
import re
import sys
import xml.etree.ElementTree as ET
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils import get_column_letter

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import add_tree_yoy_column_20260611 as yoy_base  # noqa: E402
import apply_macro_change_history_colors as color_base  # noqa: E402


TREE_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5
CURRENT_COL = 11
CODE_COL = 10
DATE_COL = 12
CALIBER_COL = 16
LOOKBACK_DAYS = 365
LOW_THRESHOLD = 0.30
HIGH_THRESHOLD = 0.70
MIN_HISTORY_VALUES = 4


@dataclass(frozen=True)
class DirectionRule:
    direction: str
    reason: str
    confidence: str


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    return yoy_base.to_number(value)


def parse_date(value: Any) -> date | None:
    return yoy_base.parse_date(value)


def norm_text(value: str) -> str:
    return re.sub(r"\s+", "", value or "").lower()


def contains_any(text: str, keywords: tuple[str, ...]) -> str | None:
    for keyword in keywords:
        if keyword.lower() in text:
            return keyword
    return None


def direction_for_indicator(row: dict[str, Any], macro_name: str) -> DirectionRule:
    text = norm_text(" ".join(
        [
            row.get("big", ""),
            row.get("dim", ""),
            row.get("sub", ""),
            row.get("name", ""),
            row.get("caliber", ""),
            macro_name,
            row.get("code", ""),
        ]
    ))

    low_good_keywords = (
        "shibor",
        "lpr",
        "dr001",
        "dr007",
        "r001",
        "r007",
        "sofr",
        "iorb",
        "effr",
        "联邦基金利率",
        "逆回购利率",
        "利率水平",
        "收益率",
        "国债收益率",
        "美债收益率",
        "存款准备金率",
        "准备金率",
        "tga",
        "财政部一般账户",
        "财政存款",
        "通胀",
        "cpi",
        "ppi",
        "pce",
        "原材料购价",
        "购进价格",
        "出厂价格",
        "猪肉",
        "布伦特原油",
        "原油",
        "失业率",
        "杠杆率",
        "债务余额",
        "债务",
        "地缘政治风险",
    )
    hit = contains_any(text, low_good_keywords)
    if hit:
        return DirectionRule("low_good", f"命中低位偏好关键词：{hit}", "规则")

    high_good_keywords = (
        "净投放",
        "净流入",
        "结售汇差额",
        "社会融资",
        "社融",
        "新增人民币贷款",
        "人民币贷款",
        "贷款",
        "融资",
        "m1",
        "m2",
        "货币供应量",
        "银行信贷",
        "总资产",
        "gdp",
        "pmi",
        "工业增加值",
        "社会消费品",
        "零售",
        "固定资产投资",
        "房地产开发投资",
        "制造业",
        "基础设施",
        "出口",
        "进口",
        "贸易差额",
        "财政支出",
        "财政收入",
        "公共预算收入",
        "公共预算支出",
        "土地使用权出让收入",
        "房屋销售",
        "住宅价格",
        "消费者信心",
        "ism",
        "新订单",
        "新建住房",
        "建造支出",
        "就业人数",
        "adp",
        "nfib",
        "中小企业乐观",
        "利润总额",
        "利润率",
        "净利润",
        "贡献率",
    )
    hit = contains_any(text, high_good_keywords)
    if hit:
        return DirectionRule("high_good", f"命中高位偏好关键词：{hit}", "规则")

    if "利差" in text:
        return DirectionRule("neutral", "利差类指标高低含义依场景变化，暂不做方向填色", "谨慎")
    if "赤字脉冲" in text or "赤字率" in text:
        return DirectionRule("neutral", "财政赤字/脉冲类高低含义需要结合口径和政策目标，暂不做方向填色", "谨慎")
    return DirectionRule("neutral", "未命中方向规则，避免误判好坏", "未判定")


def percentile_rank(sorted_values: list[float], value: float) -> float:
    return bisect_right(sorted_values, value) / len(sorted_values)


def current_history(
    series: yoy_base.MacroSeries,
    current_date: date,
) -> tuple[list[tuple[date, float]], tuple[date, float] | None]:
    start_date = current_date - timedelta(days=LOOKBACK_DAYS)
    points = sorted((d, v) for d, v in series.points if d <= current_date)
    history = [(d, v) for d, v in points if start_date <= d <= current_date and math.isfinite(v)]
    current_point = yoy_base.closest_current_period(points, current_date)
    return history, current_point


def color_for_position(
    current_raw: float,
    percentile: float,
    direction: str,
) -> tuple[str | None, str, float | None, str]:
    side = "中位"
    if percentile >= HIGH_THRESHOLD:
        side = "高位"
        distance = (percentile - HIGH_THRESHOLD) / (1.0 - HIGH_THRESHOLD)
        favorable = direction == "high_good"
    elif percentile <= LOW_THRESHOLD:
        side = "低位"
        distance = (LOW_THRESHOLD - percentile) / LOW_THRESHOLD
        favorable = direction == "low_good"
    else:
        return None, color_base.BLACK, 0.0, side

    if direction == "neutral":
        return None, color_base.BLACK, None, side

    intensity = max(0.15, min(1.0, distance))
    if favorable:
        fill = color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
    else:
        fill = color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
    return fill, color_base.text_color_for_fill(fill), intensity, side


def iter_tree_rows(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    big = ""
    dim = ""
    sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        big_value = as_text(ws.cell(row, 1).value)
        dim_value = as_text(ws.cell(row, 2).value)
        sub_value = as_text(ws.cell(row, 3).value)
        if big_value:
            big = big_value
            dim = ""
            sub = ""
        if dim_value:
            dim = dim_value
            sub = ""
        if sub_value:
            sub = sub_value
        name = as_text(ws.cell(row, 4).value)
        code = yoy_base.norm_code(ws.cell(row, CODE_COL).value)
        if not name and not code:
            continue
        rows.append(
            {
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "code": code,
                "current": ws.cell(row, CURRENT_COL).value,
                "current_date": ws.cell(row, DATE_COL).value,
                "caliber": as_text(ws.cell(row, CALIBER_COL).value),
            }
        )
    return rows


def read_xml(zf: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(zf.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def write_audit(path: Path, audit_rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "当前值历史分位颜色复核"
    headers = [
        "TREE行号",
        "大分类",
        "细分维度",
        "子维度",
        "指标名称",
        "指标代码",
        "数据日期",
        "当前值",
        "日报当前日期",
        "日报当前值",
        "近一年样本数",
        "历史起点",
        "历史终点",
        "历史分位",
        "位置判断",
        "方向规则",
        "方向依据",
        "填充色",
        "字体色",
        "颜色强度",
        "备注",
    ]
    ws.append(headers)
    for item in audit_rows:
        ws.append([item.get(header, "") for header in headers])
    widths = [10, 18, 18, 18, 34, 22, 14, 14, 14, 14, 12, 14, 14, 12, 12, 12, 34, 14, 12, 12, 44]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def apply_colors(input_tree: Path, daily_path: Path, output_tree: Path, audit_output: Path) -> dict[str, Any]:
    macro_series = yoy_base.read_macro_series(daily_path)
    wb = openpyxl.load_workbook(input_tree, data_only=True, read_only=False)
    ws = wb[TREE_SHEET]
    rows = iter_tree_rows(ws)
    wb.close()

    summaries = {
        "rows": len(rows),
        "macro_rows": 0,
        "styled_rows": 0,
        "red_rows": 0,
        "green_rows": 0,
        "neutral_position_rows": 0,
        "neutral_direction_rows": 0,
        "no_history_rows": 0,
        "unmatched_rows": 0,
        "nonmacro_rows": 0,
    }
    audit_rows: list[dict[str, Any]] = []
    style_cache: dict[tuple[int, str, str], int] = {}
    replacements: dict[str, bytes] = {}

    with ZipFile(input_tree, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        sheet_path = color_base.workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)

        for row in rows:
            cell = color_base.find_cell(sheet_root, row["row"], CURRENT_COL)
            base_style_id = int(cell.attrib.get("s", "0"))
            fill_rgb: str | None = None
            font_rgb = color_base.BLACK
            note = ""
            hist_count = 0
            hist_start = ""
            hist_end = ""
            pct: float | None = None
            position = ""
            source_date = ""
            source_current: Any = ""
            intensity: float | None = None
            direction_rule = DirectionRule("neutral", "未处理", "未判定")

            if not yoy_base.is_macro_tree_row(row["big"], row["dim"]):
                summaries["nonmacro_rows"] += 1
                note = "非宏观计算区，不填色"
            else:
                summaries["macro_rows"] += 1
                series = macro_series.get(row["code"])
                current_dt = parse_date(row["current_date"])
                if series is None or current_dt is None:
                    summaries["unmatched_rows"] += 1
                    note = "未匹配到底层宏观数据或日期为空"
                else:
                    history, current_point = current_history(series, current_dt)
                    hist_count = len(history)
                    if history:
                        hist_start = history[0][0].isoformat()
                        hist_end = history[-1][0].isoformat()
                    if current_point is not None:
                        source_date = current_point[0].isoformat()
                        source_current = current_point[1]

                    if hist_count < MIN_HISTORY_VALUES or current_point is None:
                        summaries["no_history_rows"] += 1
                        note = "近一年当前值历史样本不足"
                    else:
                        current_raw = current_point[1]
                        values = sorted(value for _, value in history)
                        pct = percentile_rank(values, current_raw)
                        direction_rule = direction_for_indicator(row, series.name)
                        fill_rgb, font_rgb, intensity, position = color_for_position(
                            current_raw,
                            pct,
                            direction_rule.direction,
                        )
                        if fill_rgb:
                            summaries["styled_rows"] += 1
                            if direction_rule.direction == "high_good":
                                red_fill = pct >= HIGH_THRESHOLD
                            else:
                                red_fill = pct <= LOW_THRESHOLD
                            if red_fill:
                                summaries["red_rows"] += 1
                            else:
                                summaries["green_rows"] += 1
                        else:
                            if direction_rule.direction == "neutral":
                                summaries["neutral_direction_rows"] += 1
                                note = direction_rule.reason
                            else:
                                summaries["neutral_position_rows"] += 1
                                note = "当前值处于近一年常态区间，未填色"

            cell.attrib["s"] = str(
                color_base.ensure_cell_style(styles_root, base_style_id, fill_rgb, font_rgb, style_cache)
            )
            audit_rows.append(
                {
                    "TREE行号": row["row"],
                    "大分类": row["big"],
                    "细分维度": row["dim"],
                    "子维度": row["sub"],
                    "指标名称": row["name"],
                    "指标代码": row["code"],
                    "数据日期": row["current_date"],
                    "当前值": row["current"],
                    "日报当前日期": source_date,
                    "日报当前值": source_current,
                    "近一年样本数": hist_count,
                    "历史起点": hist_start,
                    "历史终点": hist_end,
                    "历史分位": pct,
                    "位置判断": position,
                    "方向规则": direction_rule.direction,
                    "方向依据": direction_rule.reason,
                    "填充色": fill_rgb or "",
                    "字体色": font_rgb,
                    "颜色强度": intensity,
                    "备注": note,
                }
            )

        replacements[sheet_path] = write_xml(sheet_root)
        replacements["xl/styles.xml"] = write_xml(styles_root)

        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, replacements[name] if name in replacements else zin.read(name))

    write_audit(audit_output, audit_rows)
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        "lookback_days": LOOKBACK_DAYS,
        "low_threshold": LOW_THRESHOLD,
        "high_threshold": HIGH_THRESHOLD,
        "min_history_values": MIN_HISTORY_VALUES,
        **summaries,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()
    result = apply_colors(
        Path(args.tree),
        Path(args.daily),
        Path(args.output_tree),
        Path(args.audit_output),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
