from __future__ import annotations

import argparse
import copy
import json
import math
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from sync_tree_new_indicators_from_daily import (
    NS,
    find_cell,
    read_xml,
    workbook_sheet_path,
    write_xml,
)


RED = "FFFF0000"
BLACK = "FF000000"


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def as_text(value) -> str:
    return "" if value is None else str(value).strip()


def to_number(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "-", "—", "None"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if "%" in str(value) else number


def find_indicator_weight_columns(ws) -> tuple[int, int, int] | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        indicator_col = None
        weight_col = None
        for col in range(1, ws.max_column + 1):
            text = as_text(ws.cell(row, col).value)
            if text == "监测指标":
                indicator_col = col
            elif text == "权重占比":
                weight_col = col
        if indicator_col and weight_col:
            return row, indicator_col, weight_col
    return None


def font_without_color_and_bold(font: ET.Element) -> ET.Element:
    font = copy.deepcopy(font)
    for child in list(font):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag in {"color", "b"}:
            font.remove(child)
    return font


def font_with_style(font: ET.Element, rgb: str, bold: bool) -> ET.Element:
    font = font_without_color_and_bold(font)
    insert_at = 0
    if bold:
        font.insert(insert_at, ET.Element(q("b")))
        insert_at += 1
    color_el = ET.Element(q("color"), {"rgb": rgb})
    font.insert(insert_at, color_el)
    return font


def ensure_font_style(
    styles_root: ET.Element,
    base_style_id: int,
    rgb: str,
    bold: bool,
    cache: dict[tuple[int, str, bool], int],
) -> int:
    key = (base_style_id, rgb, bold)
    if key in cache:
        return cache[key]

    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fonts is None or cell_xfs is None:
        raise KeyError("styles font/cellXfs")

    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]
    font_nodes = fonts.findall("main:font", NS)
    base_font_id = int(base_xf.attrib.get("fontId", "0"))
    base_font = font_nodes[base_font_id] if 0 <= base_font_id < len(font_nodes) else font_nodes[0]

    new_font = font_with_style(base_font, rgb, bold)
    fonts.append(new_font)
    new_font_id = len(font_nodes)
    fonts.attrib["count"] = str(len(font_nodes) + 1)

    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["applyFont"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(len(xfs) + 1)
    cache[key] = new_style_id
    return new_style_id


def apply_style(input_path: Path, output_path: Path, threshold: float = 0.02) -> dict:
    wb_values = openpyxl.load_workbook(input_path, data_only=True)
    targets = []
    for sheet_name in wb_values.sheetnames:
        if not sheet_name.startswith("重点策略跟踪情况"):
            continue
        ws = wb_values[sheet_name]
        found = find_indicator_weight_columns(ws)
        if found is None:
            continue
        header_row, indicator_col, weight_col = found
        targets.append((sheet_name, header_row, indicator_col, weight_col, ws.max_row))

    replacements: dict[str, bytes] = {}
    styled_summary = []
    style_cache: dict[tuple[int, str, bool], int] = {}

    with ZipFile(input_path, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        for sheet_name, header_row, indicator_col, weight_col, max_row in targets:
            sheet_path = workbook_sheet_path(zin, sheet_name)
            sheet_root = read_xml(zin, sheet_path)
            ws = wb_values[sheet_name]
            red_rows = []
            black_rows = []
            for row in range(header_row + 1, max_row + 1):
                weight = to_number(ws.cell(row, weight_col).value)
                use_red = weight is not None and weight >= threshold
                rgb = RED if use_red else BLACK
                bold = bool(use_red)
                for col in (indicator_col, weight_col):
                    cell = find_cell(sheet_root, row, col)
                    base_style = int(cell.attrib.get("s", "0"))
                    cell.attrib["s"] = str(ensure_font_style(styles_root, base_style, rgb, bold, style_cache))
                if use_red:
                    red_rows.append(row)
                else:
                    black_rows.append(row)
            replacements[sheet_path] = write_xml(sheet_root)
            styled_summary.append(
                {
                    "sheet": sheet_name,
                    "indicator_col": indicator_col,
                    "weight_col": weight_col,
                    "red_bold_rows": red_rows,
                    "black_regular_count": len(black_rows),
                }
            )

        replacements["xl/styles.xml"] = write_xml(styles_root)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    return {"output": str(output_path), "styled_sheets": styled_summary}


def main() -> None:
    parser = argparse.ArgumentParser(description="Bold red indicator name and weight when TREE weight is at least 2%; black regular otherwise.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--threshold", type=float, default=0.02)
    args = parser.parse_args()
    result = apply_style(Path(args.input), Path(args.output), args.threshold)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
