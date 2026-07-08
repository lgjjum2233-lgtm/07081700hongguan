from __future__ import annotations

import argparse
import copy
import json
import math
import sys
import xml.etree.ElementTree as ET
from bisect import bisect_right
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS = {"main": SHEET_NS, "officeRel": REL_NS}

ET.register_namespace("", SHEET_NS)
ET.register_namespace("r", REL_NS)

WORKSPACE = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
for path in (SCRIPT_DIR, LEGACY_TOOLS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

import sync_tree_20260604 as base  # noqa: E402
from sync_tree_v30_full import build_sources  # noqa: E402


BLACK = "FF000000"
WHITE = "FFFFFFFF"
RED_LIGHT = "FFFFE5E5"
RED_DARK = "FFC00000"
GREEN_LIGHT = "FFE2F0D9"
GREEN_DARK = "FF006100"
MIN_HISTORY_DIFFS = 4
MIN_PERCENTILE_TO_FILL = 0.55


@dataclass(frozen=True)
class SheetConfig:
    sheet: str
    header_row: int
    big_col: int
    dim_col: int
    sub_col: int
    indicator_col: int
    code_col: int | None
    change_col: int


SHEET_CONFIGS = {
    "重点策略跟踪情况(V3.0)": SheetConfig(
        sheet="重点策略跟踪情况(V3.0)",
        header_row=5,
        big_col=1,
        dim_col=2,
        sub_col=3,
        indicator_col=4,
        code_col=10,
        change_col=13,
    ),
    "重点策略跟踪情况(V3)": SheetConfig(
        sheet="重点策略跟踪情况(V3)",
        header_row=5,
        big_col=6,
        dim_col=7,
        sub_col=8,
        indicator_col=9,
        code_col=None,
        change_col=17,
    ),
    "重点策略跟踪情况(V3)(1)": SheetConfig(
        sheet="重点策略跟踪情况(V3)(1)",
        header_row=5,
        big_col=6,
        dim_col=7,
        sub_col=8,
        indicator_col=9,
        code_col=None,
        change_col=17,
    ),
}


def q(tag: str) -> str:
    return f"{{{SHEET_NS}}}{tag}"


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "None", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}:
        return None
    is_percent = "%" in text
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if is_percent else number


def read_xml(zf: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(zf.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def rel_target(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base_path).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(zf: ZipFile, sheet_name: str) -> str:
    workbook = read_xml(zf, "xl/workbook.xml")
    rels = read_xml(zf, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in workbook.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{REL_NS}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib.get("r", "0")) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    rows = sheet_data.findall("main:row", NS)
    insert_at = len(rows)
    for idx, row in enumerate(rows):
        if int(row.attrib.get("r", "0")) > row_num:
            insert_at = idx
            break
    sheet_data.insert(insert_at, new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def hex_to_rgb(rgb: str) -> tuple[int, int, int]:
    rgb = rgb[-6:]
    return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)


def rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"FF{r:02X}{g:02X}{b:02X}"


def interpolate_color(light: str, dark: str, norm: float) -> str:
    norm = max(0.0, min(1.0, norm))
    lr, lg, lb = hex_to_rgb(light)
    dr, dg, db = hex_to_rgb(dark)
    return rgb_to_hex(
        round(lr + (dr - lr) * norm),
        round(lg + (dg - lg) * norm),
        round(lb + (db - lb) * norm),
    )


def text_color_for_fill(fill_rgb: str) -> str:
    r, g, b = hex_to_rgb(fill_rgb)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return WHITE if luminance < 135 else BLACK


def font_with_rgb(font: ET.Element, rgb: str) -> ET.Element:
    new_font = copy.deepcopy(font)
    for color in list(new_font.findall("main:color", NS)):
        new_font.remove(color)
    color_el = ET.Element(q("color"), {"rgb": rgb})
    insert_at = len(list(new_font))
    for idx, child in enumerate(list(new_font)):
        if child.tag.rsplit("}", 1)[-1] in {"sz", "u", "vertAlign", "scheme"}:
            insert_at = idx
            break
    new_font.insert(insert_at, color_el)
    return new_font


def fill_with_rgb(rgb: str) -> ET.Element:
    fill = ET.Element(q("fill"))
    pattern = ET.SubElement(fill, q("patternFill"), {"patternType": "solid"})
    ET.SubElement(pattern, q("fgColor"), {"rgb": rgb})
    ET.SubElement(pattern, q("bgColor"), {"indexed": "64"})
    return fill


def ensure_cell_style(
    styles_root: ET.Element,
    base_style_id: int,
    fill_rgb: str | None,
    font_rgb: str,
    cache: dict[tuple[int, str, str], int],
) -> int:
    fill_key = fill_rgb or "NO_FILL"
    key = (base_style_id, fill_key, font_rgb)
    if key in cache:
        return cache[key]

    fills = styles_root.find("main:fills", NS)
    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fills is None or fonts is None or cell_xfs is None:
        raise KeyError("styles fills/fonts/cellXfs")

    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]

    fonts_list = fonts.findall("main:font", NS)
    base_font_id = int(base_xf.attrib.get("fontId", "0"))
    base_font = fonts_list[base_font_id] if 0 <= base_font_id < len(fonts_list) else fonts_list[0]
    new_font_id = len(fonts_list)
    fonts.append(font_with_rgb(base_font, font_rgb))
    fonts.attrib["count"] = str(new_font_id + 1)

    if fill_rgb is None:
        new_fill_id = 0
    else:
        fills_list = fills.findall("main:fill", NS)
        new_fill_id = len(fills_list)
        fills.append(fill_with_rgb(fill_rgb))
        fills.attrib["count"] = str(new_fill_id + 1)

    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["fillId"] = str(new_fill_id)
    new_xf.attrib["applyFont"] = "1"
    new_xf.attrib["applyFill"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(new_style_id + 1)
    cache[key] = new_style_id
    return new_style_id


def is_asset_or_industry(big: str, dim: str, sub: str) -> bool:
    joined = f"{big} {dim} {sub}"
    asset_keys = ("A股", "港股", "美股", "短债", "中长债", "美债", "黄金")
    return any(key in joined for key in asset_keys) or "产业" in joined


def include_macro_scope(big: str, dim: str, sub: str) -> bool:
    if is_asset_or_industry(big, dim, sub):
        return False
    joined = f"{dim} {sub}"
    if big.startswith("中国基本面"):
        return any(key in joined for key in ("流动性", "政策", "经济面", "增长", "价格利润"))
    if big.startswith("美国经济基本面"):
        return any(key in joined for key in ("流动性", "经济基本面"))
    return False


def iter_indicator_rows(ws: openpyxl.worksheet.worksheet.Worksheet, config: SheetConfig) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    last_big = ""
    last_dim = ""
    last_sub = ""
    for row in range(config.header_row + 1, ws.max_row + 1):
        big = as_text(ws.cell(row, config.big_col).value)
        dim = as_text(ws.cell(row, config.dim_col).value)
        sub = as_text(ws.cell(row, config.sub_col).value)
        if big:
            last_big = big
            last_dim = ""
            last_sub = ""
        if dim:
            last_dim = dim
            last_sub = ""
        if sub:
            last_sub = sub

        name = as_text(ws.cell(row, config.indicator_col).value)
        if not name:
            continue
        code = as_text(ws.cell(row, config.code_col).value) if config.code_col else ""
        rows.append(
            {
                "row": row,
                "big": last_big,
                "dim": last_dim,
                "sub": last_sub,
                "name": name,
                "code": code,
            }
        )
    return rows


def history_intensity(source: Any) -> tuple[float | None, float | None, int, float | None]:
    values = [to_number(value) for value in (source.trend or [])]
    values = [value for value in values if value is not None]
    diffs = [values[idx] - values[idx - 1] for idx in range(1, len(values))]
    diffs = [diff for diff in diffs if math.isfinite(diff)]

    latest = to_number(source.change)
    if latest is None and len(values) >= 2:
        latest = values[-1] - values[-2]
    if latest is None or not math.isfinite(latest):
        return latest, None, len(diffs), None
    if abs(latest) <= 1e-12:
        return latest, 0.0, len(diffs), 0.0
    if len(diffs) < MIN_HISTORY_DIFFS:
        return latest, None, len(diffs), None

    history_abs = sorted(abs(diff) for diff in diffs)
    percentile = bisect_right(history_abs, abs(latest)) / len(history_abs)
    if percentile < MIN_PERCENTILE_TO_FILL:
        return latest, 0.0, len(diffs), percentile
    intensity = (percentile - MIN_PERCENTILE_TO_FILL) / (1.0 - MIN_PERCENTILE_TO_FILL)
    return latest, max(0.15, min(1.0, intensity)), len(diffs), percentile


def style_for_source(source: Any) -> tuple[str | None, str, dict[str, Any]]:
    latest, intensity, history_count, percentile = history_intensity(source)
    detail = {
        "latest_change": latest,
        "history_diff_count": history_count,
        "percentile": percentile,
    }
    if latest is None or intensity is None or intensity <= 0:
        return None, BLACK, detail
    if latest > 0:
        fill_rgb = interpolate_color(RED_LIGHT, RED_DARK, intensity)
    else:
        fill_rgb = interpolate_color(GREEN_LIGHT, GREEN_DARK, intensity)
    return fill_rgb, text_color_for_fill(fill_rgb), detail


def resolve_row_source(
    row_info: dict[str, Any],
    by_code: dict[str, Any],
    by_name: dict[str, list[Any]],
    best_trend: dict[str, Any],
) -> Any | None:
    source = base.resolve_source(row_info["name"], row_info["code"], by_code, by_name)
    if source is None:
        return None
    return base.with_best_trend(source, best_trend)


def apply_history_colors(
    input_tree: Path,
    daily_path: Path,
    output_tree: Path,
    as_of: datetime,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    by_code, by_name, best_trend = build_sources(daily_path, as_of.date())
    wb = openpyxl.load_workbook(input_tree, data_only=True, read_only=False)
    selected_names = sheet_names or [name for name in wb.sheetnames if name in SHEET_CONFIGS]
    selected_configs = [SHEET_CONFIGS[name] for name in selected_names if name in wb.sheetnames]
    if not selected_configs:
        raise RuntimeError("No supported TREE sheet was found.")

    replacements: dict[str, bytes] = {}
    style_cache: dict[tuple[int, str, str], int] = {}
    summaries: list[dict[str, Any]] = []

    with ZipFile(input_tree, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")

        for config in selected_configs:
            ws = wb[config.sheet]
            sheet_path = workbook_sheet_path(zin, config.sheet)
            sheet_root = read_xml(zin, sheet_path)
            rows = iter_indicator_rows(ws, config)

            summary = {
                "sheet": config.sheet,
                "indicator_rows": len(rows),
                "macro_rows": 0,
                "styled_rows": 0,
                "cleared_rows": 0,
                "small_or_zero_rows": 0,
                "no_history_rows": 0,
                "unmatched_rows": 0,
                "positive_styled_rows": 0,
                "negative_styled_rows": 0,
                "styled_examples": [],
                "unmatched_examples": [],
                "no_history_examples": [],
            }

            for row_info in rows:
                row = row_info["row"]
                target_scope = include_macro_scope(row_info["big"], row_info["dim"], row_info["sub"])
                cell = find_cell(sheet_root, row, config.change_col)
                base_style_id = int(cell.attrib.get("s", "0"))

                fill_rgb: str | None = None
                font_rgb = BLACK
                style_detail: dict[str, Any] = {}

                if target_scope:
                    summary["macro_rows"] += 1
                    source = resolve_row_source(row_info, by_code, by_name, best_trend)
                    if source is None:
                        summary["unmatched_rows"] += 1
                        if len(summary["unmatched_examples"]) < 12:
                            summary["unmatched_examples"].append(
                                {"row": row, "name": row_info["name"], "code": row_info["code"]}
                            )
                    else:
                        fill_rgb, font_rgb, style_detail = style_for_source(source)
                        if fill_rgb is None:
                            if style_detail.get("percentile") is None and style_detail.get("latest_change") not in (None, 0.0):
                                summary["no_history_rows"] += 1
                                if len(summary["no_history_examples"]) < 12:
                                    summary["no_history_examples"].append(
                                        {
                                            "row": row,
                                            "name": row_info["name"],
                                            "code": row_info["code"] or source.code,
                                            "history_diff_count": style_detail.get("history_diff_count"),
                                        }
                                    )
                            else:
                                summary["small_or_zero_rows"] += 1
                        else:
                            summary["styled_rows"] += 1
                            if style_detail.get("latest_change", 0) > 0:
                                summary["positive_styled_rows"] += 1
                            else:
                                summary["negative_styled_rows"] += 1
                            if len(summary["styled_examples"]) < 20:
                                summary["styled_examples"].append(
                                    {
                                        "row": row,
                                        "name": row_info["name"],
                                        "code": row_info["code"] or source.code,
                                        "change": style_detail.get("latest_change"),
                                        "percentile": style_detail.get("percentile"),
                                        "fill": fill_rgb,
                                    }
                                )

                if fill_rgb is None:
                    summary["cleared_rows"] += 1
                cell.attrib["s"] = str(ensure_cell_style(styles_root, base_style_id, fill_rgb, font_rgb, style_cache))

            replacements[sheet_path] = write_xml(sheet_root)
            summaries.append(summary)

        replacements["xl/styles.xml"] = write_xml(styles_root)
        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    wb.close()
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "as_of": as_of.strftime("%Y-%m-%d"),
        "min_history_diffs": MIN_HISTORY_DIFFS,
        "min_percentile_to_fill": MIN_PERCENTILE_TO_FILL,
        "sheets": summaries,
    }


def json_default(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Color TREE macro marginal changes by each indicator's own historical change percentile.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default=datetime.today().strftime("%Y-%m-%d"))
    parser.add_argument("--sheet", action="append", help="Supported sheet to process. Defaults to supported TREE sheets.")
    args = parser.parse_args()

    result = apply_history_colors(
        Path(args.tree),
        Path(args.daily),
        Path(args.output),
        datetime.strptime(args.as_of, "%Y-%m-%d"),
        args.sheet,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
