from __future__ import annotations

import copy
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "（6月12日V16）TREE宏观分析_当前数据红高绿低_平值修正.xlsx"
OUTPUT = ROOT / "（6月12日V17）TREE宏观分析_权重重估.xlsx"
AUDIT = ROOT / "20260612_TREE权重重估复核.xlsx"
SHEET_MAIN = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5

RED_FONT = "FFFF0000"
BLACK_FONT = "FF000000"
WHITE_FONT = "FFFFFFFF"
WEIGHT_LIGHT = "FFFFF7E6"
WEIGHT_DARK = "FFE2A87B"
WEIGHT_THRESHOLD = 0.02


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def normalize_name(value: Any) -> str:
    s = text(value)
    s = re.sub(r"[①②③④⑤⑥⑦⑧⑨⑩]", "", s)
    s = re.sub(r"[：:]", "", s)
    s = re.sub(r"\（.*?\）|\(.*?\)", "", s)
    s = re.sub(r"\[.*?\]", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def norm_code(value: Any) -> str:
    return text(value).upper()


def is_macro(big: str, dim: str) -> bool:
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    # 延续之前处理边界：产业和资本市场不重算本轮宏观权重。
    if "产业" in dim:
        return False
    return True


def interpolate_color(start: str, end: str, ratio: float) -> str:
    ratio = max(0.0, min(1.0, ratio))
    s = start[-6:]
    e = end[-6:]
    vals = []
    for idx in range(0, 6, 2):
        a = int(s[idx : idx + 2], 16)
        b = int(e[idx : idx + 2], 16)
        vals.append(round(a + (b - a) * ratio))
    return "FF" + "".join(f"{value:02X}" for value in vals)


def text_color_for_fill(fill: str) -> str:
    r = int(fill[2:4], 16)
    g = int(fill[4:6], 16)
    b = int(fill[6:8], 16)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return WHITE_FONT if luminance < 150 else BLACK_FONT


def tier(weight: float) -> str:
    if weight >= 0.08:
        return "核心权重"
    if weight >= 0.05:
        return "高权重"
    if weight >= 0.03:
        return "中高权重"
    if weight >= 0.02:
        return "中等权重"
    return "辅助权重"


def reason(role: str, why: str, weight: float) -> str:
    return f"{role}。本轮新增指标后，将其定位为{tier(weight)}：{why}"


RAW_META: dict[int, tuple[float, str, str]] = {
    # 中国流动性面：央行20%，银行间15%，实体50%，海外15%。
    6: (0.05, "7天逆回购利率是中国短端政策利率锚", "它直接决定银行间资金成本中枢，是判断货币政策松紧最权威的价格信号。"),
    7: (0.02, "法定存款准备金率代表长期基础货币约束", "降准影响大但频率低，更多用于确认长期流动性环境。"),
    8: (0.04, "日度公开市场净投放体现央行短期流动性态度", "它能捕捉短期投放和回笼节奏，但单日噪音较大，需要配合周度口径。"),
    9: (0.02, "周度公开市场净投放用于平滑日度扰动", "它比单日数据更适合观察阶段性投放方向，但仍是辅助信号。"),
    10: (0.03, "中央银行贷款总投放汇总结构性工具变化", "它能观察中长期工具合力，但月度频率和结构性属性使其弱于政策利率和信用数据。"),
    11: (0.005, "SLF是压力阶段的应急流动性工具", "只有在资金压力明显时信号才强，常态下解释力较弱。"),
    12: (0.015, "MLF影响银行中期负债成本和续作预期", "当前政策利率和LPR更直接，MLF主要作为中期资金补充信号。"),
    13: (0.01, "PSL是定向结构性资金工具", "主要影响特定领域，不能代表总量流动性。"),
    14: (0.01, "其他结构性工具用于补充观察定向投放", "透明度和连续性弱于常规工具，因此只保留低权重。"),
    15: (0.02, "DR001是隔夜资金温度计", "它高频敏感但波动较大，适合识别短期资金面异常。"),
    16: (0.05, "DR007是银行间资金面的核心市场利率", "它最贴近7天政策利率传导，是判断资金面合理充裕的重要标尺。"),
    17: (0.015, "R001反映非银隔夜融资压力", "它用于识别银行与非银之间是否存在流动性分层。"),
    18: (0.02, "R007补充观察非银7天融资成本", "它比DR007更能反映广义机构资金压力，但稳定性略弱。"),
    19: (0.045, "3个月SHIBOR反映银行中期同业负债成本", "它能观察资金压力是否从隔夜和7天期限传导到银行负债端。"),
    20: (0.09, "社会融资规模是实体融资最宽口径", "它代表信用周期总量方向，是中国流动性面最核心的实体信用指标。"),
    21: (0.08, "新增人民币贷款是社融中最重要的银行信用来源", "它最直接反映金融体系对实体的信贷支持。"),
    22: (0.05, "政府债券融资反映财政发力对社融的贡献", "它直接对应稳增长节奏和财政扩张强度。"),
    23: (0.02, "票据融资用于判断银行冲量和真实融资需求质量", "它对社融结构解释重要，但不能单独代表信用扩张。"),
    24: (0.03, "企业债券融资反映直接融资环境和信用风险偏好", "它补充银行信贷之外的企业融资渠道。"),
    25: (0.06, "M1反映资金活化程度", "它比M2更能捕捉企业经营活跃度和居民交易性需求。"),
    26: (0.04, "M2代表广义货币总水位", "它适合观察总量环境，但对增长拐点的领先性弱于社融和M1。"),
    27: (0.025, "居民部门杠杆率反映居民资产负债表扩张空间", "它影响消费和地产信用，但季度更新且属于慢变量。"),
    28: (0.02, "政府部门杠杆率衡量财政加杠杆空间", "它用于约束性判断，边际信号弱于政府债和财政支出。"),
    29: (0.035, "非金融企业部门杠杆率反映企业资产负债表压力", "它关系到投资扩张和信用风险，是实体信用质量的重要背景指标。"),
    30: (0.05, "1年LPR是实体贷款定价核心基准", "它直接影响企业和居民融资成本，是货币政策向信用端传导的重要指标。"),
    31: (0.03, "银行结售汇差额反映境内外汇供求和跨境资金意愿", "它对人民币流动性和汇率预期有辅助判断价值。"),
    32: (0.035, "中国10年期国债收益率是国内长期无风险利率锚", "它反映增长预期、通胀预期和配置需求变化。"),
    33: (0.035, "美国10年期国债收益率是全球美元长期利率锚", "它影响人民币资产估值、汇率压力和跨境资金偏好。"),
    34: (0.05, "中美10年期国债利差衡量中美长期利率相对吸引力", "它比单独看两国利率更适合判断外部流动性约束和汇率压力。"),
    # 中国政策面：货币25%，财政75%。
    35: (0.25, "货币政策报告和政策表述决定流动性框架边界", "虽然不定期，但权威性最高，能校准后续利率、数量工具和信用政策方向。"),
    36: (0.10, "一般公共预算收入累计值反映财政收入资源基础", "它能观察经济活动和税收修复，但收入不等同于财政发力。"),
    37: (0.18, "一般公共预算支出累计值反映财政资金实际投放", "它比收入端更直接影响当期需求和稳增长力度。"),
    38: (0.10, "土地出让收入累计同比反映地方财政和地产链景气", "它对地方政府性基金收入、基建配套和地产政策空间有解释力。"),
    39: (0.08, "中央政府债务余额衡量中央财政加杠杆空间", "它是财政扩张的背景约束，边际信号弱于支出和赤字脉冲。"),
    40: (0.17, "财政赤字脉冲反映财政边际扩张方向", "它能直接观察财政政策是边际发力还是收缩，是政策面核心量化指标。"),
    41: (0.12, "财政存款累计值反映财政资金沉淀和投放节奏", "它影响银行体系流动性，但累计口径需要结合支出和赤字脉冲判断。"),
    # 中国经济面：增长65%，价格利润35%。
    42: (0.03, "社零当月值反映消费市场名义销售规模", "它验证内需修复，但受价格和节假日扰动，应弱于同比口径。"),
    43: (0.05, "社零同比是消费需求的核心月度代理", "消费是中国内需的重要组成，社零同比对经济修复方向更敏感。"),
    44: (0.025, "中国GDP总量校准经济规模和产出基础", "它是总量锚，但季度发布且偏滞后。"),
    45: (0.045, "中国GDP不变价当季同比是宏观真实增长核心指标", "它直接验证经济周期方向和政策成效，但频率低于月度指标。"),
    46: (0.015, "最终消费支出贡献率拆解GDP增长来源", "它用于确认增长是否由内需驱动，属于结构验证指标。"),
    47: (0.015, "资本形成总额贡献率拆解投资对GDP的拉动", "它验证投资链条结论，但季度更新、边际性较弱。"),
    48: (0.01, "货物和服务净出口贡献率衡量外需对GDP的贡献", "它受进口、价格和汇率共同影响，权重低于直接出口数据。"),
    49: (0.03, "固定资产投资累计值反映年内投资完成规模", "它验证稳增长落地，但累计值受年度进度影响。"),
    50: (0.05, "固定资产投资累计同比连接制造业、基建和地产投资", "它是稳增长和实体资本开支的核心月度指标。"),
    51: (0.04, "房地产开发投资累计同比反映地产链景气", "地产链对信用、地方财政和周期行业影响仍大。"),
    52: (0.03, "制造业投资累计同比反映企业资本开支意愿", "它用于判断实体投资是否具备内生动能。"),
    53: (0.03, "基建投资累计同比反映政策性投资发力", "它是财政政策向实体需求传导的重要观察口径。"),
    54: (0.035, "房屋销售累计同比领先地产投资和居民预期", "它能更早反映地产链需求变化。"),
    55: (0.015, "70城二手住宅价格同比验证地产真实供需", "二手房价格有参考意义，但影响范围窄于销售和投资。"),
    56: (0.045, "出口金额当月同比反映外需和制造业竞争力", "它是增长的重要边际变量，对制造业和汇率都有影响。"),
    57: (0.015, "对美国出口累计同比反映外需结构中的美国方向", "它是出口总指标的地区拆解，权重低于总出口。"),
    58: (0.015, "对东盟出口累计同比反映区域贸易和产业链转移", "它补充出口结构判断，但不替代总出口。"),
    59: (0.015, "贸易差额反映净出口对外汇收入和总需求的影响", "它受进口、价格和汇率影响，需要与出口分项结合。"),
    60: (0.035, "制造业PMI综合是月初最早的景气风向标", "它能快速捕捉制造业扩张或收缩。"),
    61: (0.045, "PMI新订单领先生产和库存", "它是需求转折的关键领先分项。"),
    62: (0.015, "PMI生产验证供给端实际产出", "它偏同步确认，重要性低于新订单。"),
    63: (0.02, "非制造业PMI补充服务业和建筑业景气", "它能覆盖制造业PMI以外的经济活动。"),
    64: (0.025, "工业增加值是生产侧硬数据", "它反映工业部门真实产出，但公布滞后于PMI。"),
    65: (0.07, "PPI与工业盈利周期高度相关", "它是价格利润链条的核心价格指标。"),
    66: (0.04, "CPI衡量居民终端通胀和政策约束", "它影响货币政策空间，但对工业利润传导弱于PPI。"),
    67: (0.015, "猪肉批发价格指数是CPI食品项的重要高频线索", "它能提前观察居民通胀压力，但品类单一。"),
    68: (0.025, "布伦特原油价格代表全球能源成本和输入性通胀", "它影响PPI、CPI和企业成本，但受供给和地缘扰动大。"),
    69: (0.04, "原材料购价领先PPI和成本压力", "它是利润率和价格传导的前瞻信号。"),
    70: (0.09, "工业企业利润总额累计同比连接宏观景气和企业盈利", "它最直接体现利润周期，是价格利润模块核心指标。"),
    71: (0.07, "工业企业营业收入利润率衡量盈利质量", "它能区分增收与真实增利，是利润修复质量的核心指标。"),
    # 美国流动性面：联邦40%，银行间25%，实体35%。
    113: (0.13, "联邦基金利率是美元政策立场和资金成本核心锚", "它直接影响全球折现率、美元流动性和风险资产估值。"),
    114: (0.10, "美联储资产负债表反映QE/QT总量影响", "它决定美元基础流动性的方向，是联邦流动性核心数量指标。"),
    115: (0.04, "IORB约束银行准备金收益和利率走廊", "它用于判断准备金市场和政策利率传导。"),
    116: (0.01, "隔夜逆回购利率影响现金停放收益", "该行缺少连续底层代码，作为低权重定性辅助项。"),
    117: (0.05, "TGA余额会抽离或释放银行体系准备金", "它对短期美元流动性有直接扰动。"),
    118: (0.07, "隔夜逆回购规模反映过剩流动性吸收和释放空间", "它能观察货币市场闲置流动性变化。"),
    119: (0.13, "SOFR是美元回购融资成本的高频核心指标", "它最直接反映美元担保融资市场的松紧。"),
    120: (0.12, "SOFR-IORB利差衡量回购市场相对准备金收益的紧张程度", "它比单看SOFR更能识别美元资金压力。"),
    121: (0.07, "AAA信用利差反映企业融资环境和风险偏好", "它是实体金融条件的重要价格信号。"),
    122: (0.07, "商业银行银行信贷反映银行资产端信用供给", "它是美元实体流动性的核心余额指标之一。"),
    123: (0.04, "银行贷款同比反映信用供给趋势", "它能确认银行信贷是否持续扩张。"),
    124: (0.05, "工商业贷款反映企业投资和经营融资需求", "它是美国实体信用的重要企业端分项。"),
    125: (0.03, "居民消费贷款反映家庭消费信用需求", "它补充观察居民部门信用扩张。"),
    126: (0.03, "居民房地产贷款反映住房信用传导", "它变化较慢，但能观察高利率对地产信用的影响。"),
    127: (0.02, "美国M1反映交易性货币需求", "美国市场更重视价格型金融条件，因此M1为辅助指标。"),
    128: (0.04, "美国M2反映广义货币背景", "它能补充美元流动性总量判断，但领先性弱于利率和信用利差。"),
    # 美国经济基本面：供给20%，需求30%，价格50%。
    129: (0.04, "制造业ISM是美国制造业景气核心先行指标", "它能较早反映制造业订单、生产和库存变化。"),
    130: (0.05, "非制造业ISM覆盖美国服务业景气", "服务业占美国经济比重高，因此权重高于制造业ISM。"),
    131: (0.04, "ISM新订单领先生产变化", "它是判断美国制造业拐点的关键分项。"),
    132: (0.03, "非农新增就业反映企业用工需求", "就业数据影响消费和政策预期，但月度波动较大。"),
    133: (0.02, "NFIB中小企业乐观指数反映企业经营信心", "它补充ISM和就业数据，但调查属性较强。"),
    134: (0.01, "失业率衡量劳动力闲置程度", "它重要但偏滞后，因此在供给模块中权重较低。"),
    135: (0.01, "工业生产同比是供给侧硬数据", "美国经济服务占比较高，工业生产代表性低于ISM和就业。"),
    136: (0.03, "密歇根消费者信心指数反映消费意愿", "它是美国需求的前瞻信号。"),
    137: (0.02, "美国名义GDP总量提供经济规模和名义需求背景", "它是分母锚，但边际周期信号弱于实际增长和零售。"),
    138: (0.035, "美国GDP同比衡量经济增长总方向", "它验证软着陆或衰退风险，但季度发布滞后。"),
    139: (0.045, "GDP年化增速是市场最常用的美国增长节奏指标", "它更贴近交易和政策讨论口径。"),
    140: (0.07, "零售销售环比是美国消费的高频核心代理", "美国消费权重高，该指标对需求边际变化最敏感。"),
    141: (0.04, "零售和食品服务销售额同比补充消费趋势判断", "它能平滑环比波动，但受价格和基数影响。"),
    142: (0.02, "房屋销售同比反映利率向地产需求的传导", "它是需求链条的辅助观察。"),
    143: (0.02, "时薪同比影响消费能力和服务通胀压力", "它连接收入、消费和通胀。"),
    144: (0.02, "个人可支配收入同比是消费能力来源", "它能验证消费是否有收入支撑。"),
    145: (0.15, "美国CPI同比是终端通胀核心指标", "它直接影响美联储政策预期、利率和估值。"),
    146: (0.15, "核心CPI同比剔除食品和能源扰动", "它更能反映美国内生通胀粘性，是政策路径核心指标。"),
    147: (0.08, "PPI同比反映上游价格和成本传导", "它能提前观察通胀链条和利润压力。"),
    148: (0.12, "核心PCE同比是美联储更关注的通胀口径", "它对降息预期和美元利率判断具有核心影响。"),
}

META = {
    row: {"weight": weight, "reason": reason(role, why, weight)}
    for row, (weight, role, why) in RAW_META.items()
}

SUB_LABELS_V30 = {
    6: "央行流动性\n（20%）",
    15: "银行间流动性\n（15%）",
    20: "实体流动性\n（50%）",
    31: "海外流动性\n（15%）",
    35: "货币政策\n（25%）",
    36: "财政政策\n（75%）",
    42: "增长\n（65%）",
    65: "价格利润\n（35%）",
    113: "联邦流动性\n（40%）",
    119: "银行间流动性\n（25%）",
    121: "实体流动性\n（35%）",
    129: "供给\n（20%）",
    136: "需求\n（30%）",
    145: "价格\n（50%）",
}


def style_weight_cell(cell, weight: float | None) -> None:
    if weight is None:
        return
    intensity = min(1.0, max(0.0, weight / 0.10))
    fill = interpolate_color(WEIGHT_LIGHT, WEIGHT_DARK, intensity)
    cell.fill = PatternFill(fill_type="solid", fgColor=fill)
    font = copy.copy(cell.font)
    font.color = RED_FONT if weight > WEIGHT_THRESHOLD else BLACK_FONT
    font.bold = weight > WEIGHT_THRESHOLD
    cell.font = font


def style_indicator_cell(cell, weight: float | None) -> None:
    if weight is None:
        return
    font = copy.copy(cell.font)
    font.color = RED_FONT if weight > WEIGHT_THRESHOLD else BLACK_FONT
    font.bold = weight > WEIGHT_THRESHOLD
    cell.font = font


def copy_style(source, target) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy.copy(source.alignment)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.protection = copy.copy(source.protection)


def build_main_maps(ws):
    by_code: dict[str, dict[str, Any]] = {}
    by_name: dict[tuple[str, str], dict[str, Any]] = {}
    big = dim = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, 1).value):
            big = text(ws.cell(row, 1).value)
            dim = ""
        if text(ws.cell(row, 2).value):
            dim = text(ws.cell(row, 2).value)
        name = text(ws.cell(row, 4).value)
        if row not in META or not name:
            continue
        item = {
            "weight": META[row]["weight"],
            "reason": META[row]["reason"],
            "big": big,
            "dim": dim,
            "name": name,
            "row": row,
        }
        code = norm_code(ws.cell(row, 10).value)
        if code and code not in {"—", "-"}:
            by_code[code] = item
        by_name[(big, normalize_name(name))] = item
    return by_code, by_name


def fallback_match(big: str, name: str, by_name: dict[tuple[str, str], dict[str, Any]]) -> dict[str, Any] | None:
    norm = normalize_name(name)
    if not norm:
        return None
    best = None
    best_score = 0.0
    for (candidate_big, candidate_norm), item in by_name.items():
        if candidate_big != big:
            continue
        if norm in candidate_norm or candidate_norm in norm:
            score = min(len(norm), len(candidate_norm)) / max(len(norm), len(candidate_norm))
            if score > best_score:
                best_score = score
                best = item
    return best if best_score >= 0.35 else None


def sheet_layout(ws) -> dict[str, int] | None:
    headers = {text(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    if headers.get("监测指标") == 4 and headers.get("权重占比") == 5:
        return {"big": 1, "dim": 2, "sub": 3, "name": 4, "weight": 5, "reason": 6, "code": 10}
    if headers.get("监测指标") == 9 and headers.get("权重占比") == 13:
        return {"big": 6, "dim": 7, "sub": 8, "name": 9, "weight": 13, "reason": 14, "code": 0}
    return None


def update_main_sheet(ws) -> list[dict[str, Any]]:
    rows = []
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if row in SUB_LABELS_V30:
            ws.cell(row, 3).value = SUB_LABELS_V30[row]
        if text(ws.cell(row, 1).value):
            big = text(ws.cell(row, 1).value)
            dim = ""
            sub = ""
        if text(ws.cell(row, 2).value):
            dim = text(ws.cell(row, 2).value)
            sub = ""
        if text(ws.cell(row, 3).value):
            sub = text(ws.cell(row, 3).value)
        name = text(ws.cell(row, 4).value)
        if row not in META or not name:
            continue
        old_weight = ws.cell(row, 5).value
        old_reason = ws.cell(row, 6).value
        weight = META[row]["weight"]
        ws.cell(row, 5).value = weight
        ws.cell(row, 5).number_format = "0.00%"
        ws.cell(row, 6).value = META[row]["reason"]
        ws.cell(row, 6).alignment = Alignment(wrap_text=True, vertical="center")
        style_weight_cell(ws.cell(row, 5), weight)
        style_indicator_cell(ws.cell(row, 4), weight)
        rows.append(
            {
                "sheet": ws.title,
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "old_weight": old_weight,
                "new_weight": weight,
                "old_reason": old_reason,
                "new_reason": META[row]["reason"],
                "match": "V3.0行号",
            }
        )
    return rows


def update_legacy_sheet(ws, by_code, by_name) -> list[dict[str, Any]]:
    layout = sheet_layout(ws)
    if not layout:
        return []
    rows = []
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, layout["big"]).value):
            big = text(ws.cell(row, layout["big"]).value)
            dim = ""
            sub = ""
        if text(ws.cell(row, layout["dim"]).value):
            dim = text(ws.cell(row, layout["dim"]).value)
            sub = ""
        if text(ws.cell(row, layout["sub"]).value):
            sub = text(ws.cell(row, layout["sub"]).value)
        name = text(ws.cell(row, layout["name"]).value)
        if not name or not is_macro(big, dim):
            continue
        code = norm_code(ws.cell(row, layout["code"]).value) if layout["code"] else ""
        item = by_code.get(code) if code else None
        match = "指标代码"
        if item is None:
            item = by_name.get((big, normalize_name(name)))
            match = "同名"
        if item is None:
            item = fallback_match(big, name, by_name)
            match = "相近名称"
        if item is None:
            continue
        old_weight = ws.cell(row, layout["weight"]).value
        old_reason = ws.cell(row, layout["reason"]).value
        weight = item["weight"]
        weight_written = not isinstance(ws.cell(row, layout["weight"]), MergedCell)
        reason_written = not isinstance(ws.cell(row, layout["reason"]), MergedCell)
        if weight_written:
            ws.cell(row, layout["weight"]).value = weight
            ws.cell(row, layout["weight"]).number_format = "0.00%"
            style_weight_cell(ws.cell(row, layout["weight"]), weight)
        if reason_written:
            ws.cell(row, layout["reason"]).value = item["reason"]
            ws.cell(row, layout["reason"]).alignment = Alignment(wrap_text=True, vertical="center")
        style_indicator_cell(ws.cell(row, layout["name"]), weight)
        merge_note = ""
        if not weight_written and not reason_written:
            merge_note = "；权重和理由位于合并单元格非左上角，未写入"
        elif not weight_written:
            merge_note = "；权重位于合并单元格非左上角，未写入"
        elif not reason_written:
            merge_note = "；理由位于合并单元格非左上角，未写入"
        rows.append(
            {
                "sheet": ws.title,
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "old_weight": old_weight,
                "new_weight": weight if weight_written else "",
                "old_reason": old_reason,
                "new_reason": item["reason"] if reason_written else merge_note.lstrip("；"),
                "match": match + merge_note,
            }
        )
    return rows


def collect_group_sums(ws) -> list[dict[str, Any]]:
    out = []
    big = dim = sub = ""
    sums = defaultdict(float)
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, 1).value):
            big = text(ws.cell(row, 1).value)
            dim = ""
            sub = ""
        if text(ws.cell(row, 2).value):
            dim = text(ws.cell(row, 2).value)
            sub = ""
        if text(ws.cell(row, 3).value):
            sub = text(ws.cell(row, 3).value)
        name = text(ws.cell(row, 4).value)
        weight = ws.cell(row, 5).value
        if not name or not is_macro(big, dim) or not isinstance(weight, (int, float)):
            continue
        sums[(big, dim)] += weight
        sums[(big, dim, sub)] += weight
    for key, value in sorted(sums.items(), key=lambda item: str(item[0])):
        out.append({"层级": "面" if len(key) == 2 else "子维度", "分类": " / ".join(key), "权重合计": value})
    return out


def write_audit(changes: list[dict[str, Any]], sums: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "权重明细"
    headers = ["表名", "行号", "一级分类", "维度", "子维度", "指标名称", "原权重", "新权重", "匹配方式", "新权重理由"]
    ws.append(headers)
    for item in changes:
        ws.append(
            [
                item["sheet"],
                item["row"],
                item["big"],
                item["dim"],
                item["sub"],
                item["name"],
                item["old_weight"],
                item["new_weight"],
                item["match"],
                item["new_reason"],
            ]
        )
    ws2 = wb.create_sheet("权重合计复核")
    ws2.append(["层级", "分类", "权重合计"])
    for item in sums:
        ws2.append([item["层级"], item["分类"], item["权重合计"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws, ws2]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = [28, 8, 14, 14, 22, 36, 12, 12, 12, 72]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate([12, 60, 14], start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = width
    wb.save(AUDIT)


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, data_only=False)
    ws_main = wb[SHEET_MAIN]
    changes = update_main_sheet(ws_main)
    by_code, by_name = build_main_maps(ws_main)
    for sheet_name in [s for s in wb.sheetnames if "重点策略跟踪情况" in s and s != SHEET_MAIN]:
        changes.extend(update_legacy_sheet(wb[sheet_name], by_code, by_name))
    sums = collect_group_sums(ws_main)
    wb.save(OUTPUT)
    wb.close()
    write_audit(changes, sums)
    print({"output": str(OUTPUT), "audit": str(AUDIT), "changed_rows": len(changes), "sum_rows": len(sums)})


if __name__ == "__main__":
    main()
