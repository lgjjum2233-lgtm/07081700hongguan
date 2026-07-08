from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils.cell import range_boundaries


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}

for prefix, uri in [
    ("", NS["main"]),
    ("r", NS["officeRel"]),
    ("xdr", NS["xdr"]),
    ("a", NS["a"]),
    ("c", NS["c"]),
]:
    ET.register_namespace(prefix, uri)


def load_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    text = as_text(value).lower()
    text = re.sub(r"\[[^\]]*\]", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[（）():：;；\s_\-—]+", "", text)
    return text


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def close_enough(a: Any, b: Any) -> bool:
    na = to_number(a)
    nb = to_number(b)
    if na is None or nb is None:
        return False
    return abs(na - nb) <= max(1e-8, abs(na) * 1e-7, abs(nb) * 1e-7)


def numeric_values(values: list[Any]) -> list[float]:
    out = []
    for value in values:
        n = to_number(value)
        if n is not None:
            out.append(float(n))
    return out


def frequency_count(freq: Any, fallback: int = 60) -> int:
    text = as_text(freq)
    if "日" in text:
        return 60
    if "周" in text:
        return 12
    if "季" in text:
        return 8
    if "月" in text:
        return 6
    return fallback or 60


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_paths(z: ZipFile) -> dict[str, str]:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    paths = {}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
        paths[sheet.attrib["name"]] = rel_target("xl/workbook.xml", rel_map[rid])
    return paths


def sheet_drawing_path(z: ZipFile, sheet_path: str) -> str | None:
    sheet = read_xml(z, sheet_path)
    drawing = sheet.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib[f"{{{NS['officeRel']}}}id"]
    rels_name = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
    rels = read_xml(z, rels_name)
    for rel in rels:
        if rel.attrib["Id"] == rid:
            return rel_target(sheet_path, rel.attrib["Target"])
    return None


def drawing_rels_path(drawing_path: str) -> str:
    p = PurePosixPath(drawing_path)
    return str(p.parent / "_rels" / (p.name + ".rels"))


def chart_rows(z: ZipFile, drawing_path: str) -> list[dict[str, Any]]:
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    drawing = read_xml(z, drawing_path)
    out = []
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        chart_ref = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or chart_ref is None:
            continue
        rid = chart_ref.attrib[f"{{{NS['officeRel']}}}id"]
        out.append({"row": int(frm.findtext("xdr:row", namespaces=NS)) + 1, "chart_path": rel_map[rid]})
    return out


def split_sheet_range(formula: str) -> tuple[str, str] | None:
    if "!" not in formula:
        return None
    sheet, cell_range = formula.rsplit("!", 1)
    return sheet.strip().strip("'"), cell_range.replace("$", "")


def values_from_range_formula(wb_values, formula: str) -> list[Any]:
    formula = as_text(formula)
    if formula.startswith("{") and formula.endswith("}"):
        return [to_number(part) for part in formula[1:-1].split(",")]
    split = split_sheet_range(formula)
    if split is None:
        return []
    sheet, cell_range = split
    if sheet not in wb_values.sheetnames:
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return []
    ws = wb_values[sheet]
    values = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            values.append(ws.cell(row, col).value)
    return values


def chart_count_and_formula(chart_xml: bytes) -> tuple[int, str]:
    root = ET.fromstring(chart_xml)
    formula = ""
    f = root.find(".//c:lineChart/c:ser/c:val/c:numRef/c:f", NS)
    if f is not None and f.text:
        formula = f.text
    pt_count = root.find(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:ptCount", NS)
    count = int(pt_count.attrib.get("val", "0")) if pt_count is not None else 0
    return count, formula


def helper_index(wb_values, helper_sheet: str) -> dict[str, list[float]]:
    if helper_sheet not in wb_values.sheetnames:
        return {}
    ws = wb_values[helper_sheet]
    out: dict[str, list[float]] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not label:
            continue
        values = numeric_values([ws.cell(row, col).value for col in range(2, ws.max_column + 1)])
        if values:
            out[norm(label)] = values
    return out


def helper_values_for_name(name: Any, helper: dict[str, list[float]]) -> list[float] | None:
    nname = norm(name)
    if not nname:
        return None
    if nname in helper:
        return helper[nname]
    candidates = []
    for key, values in helper.items():
        if len(key) >= 3 and (key in nname or nname in key):
            candidates.append((abs(len(key) - len(nname)), len(values), values))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], -item[1]))
    return candidates[0][2]


def divisor_from_formula(formula: Any) -> float:
    text = as_text(formula).replace(" ", "")
    for divisor in ["1000000000000", "100000000", "10000", "100"]:
        if f"/{divisor}" in text:
            return float(divisor)
    return 1.0


RANGE_RE = re.compile(r"('?[^'!]+(?:'?)|[A-Za-z0-9_\u4e00-\u9fff]+)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")


def formula_source_values(wb_values, formula: Any) -> list[float] | None:
    text = as_text(formula)
    matches = RANGE_RE.findall(text)
    if not matches:
        return None
    divisor = divisor_from_formula(formula)
    best: list[float] = []
    for sheet_raw, start_col, start_row, end_col, end_row in matches:
        sheet = sheet_raw.strip("'")
        if sheet not in wb_values.sheetnames:
            continue
        ws = wb_values[sheet]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(f"{start_col}{start_row}:{end_col}{end_row}")
        except ValueError:
            continue
        values = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, min(max_col, ws.max_column) + 1):
                n = to_number(ws.cell(row, col).value)
                if n is not None:
                    values.append(float(n) / divisor)
        if len(values) > len(best):
            best = values
    return best or None


def header_row_for(ws, row: int) -> int:
    for header in range(row, 0, -1):
        if as_text(ws.cell(header, 5).value).lower() in {"date", "日期"}:
            return header
    return 1


def base_series_by_code(wb_values, base_sheets: list[str]) -> dict[str, list[list[float]]]:
    out: dict[str, list[list[float]]] = {}
    for sheet_name in base_sheets:
        if sheet_name not in wb_values.sheetnames:
            continue
        ws = wb_values[sheet_name]
        for row in range(2, ws.max_row + 1):
            code = as_text(ws.cell(row, 5).value).upper()
            if not code:
                continue
            header = header_row_for(ws, row)
            values = []
            for col in range(6, ws.max_column + 1):
                n = to_number(ws.cell(row, col).value)
                if n is not None and ws.cell(header, col).value is not None:
                    values.append(float(n))
            if values:
                out.setdefault(code, []).append(values)
    return out


def code_values_for_current(code: Any, current: Any, base: dict[str, list[list[float]]]) -> list[float] | None:
    key = as_text(code).upper()
    if not key or key not in base:
        return None
    series_list = base[key]
    exact = [series for series in series_list if series and close_enough(series[-1], current)]
    if exact:
        return max(exact, key=len)
    scaled = []
    cur = to_number(current)
    if cur is not None:
        for series in series_list:
            if not series:
                continue
            for divisor in (100.0, 10000.0, 100000000.0, 1000000000000.0):
                if close_enough(series[-1] / divisor, cur):
                    scaled.append([v / divisor for v in series])
                    break
    if scaled:
        return max(scaled, key=len)
    return max(series_list, key=len)


def choose_trend_values(wb_values, wb_formula, helper, base_by_code, sheet_name: str, row: int, chart_formula: str, existing_count: int) -> tuple[list[float] | None, str]:
    ws_v = wb_values[sheet_name]
    ws_f = wb_formula[sheet_name]
    name = ws_v.cell(row, 1).value
    code = ws_v.cell(row, 3).value
    current = ws_v.cell(row, 5).value
    desired = frequency_count(ws_v.cell(row, 2).value, existing_count or 60)

    candidates: list[tuple[int, str, list[float]]] = []

    chart_range_values = numeric_values(values_from_range_formula(wb_values, chart_formula))
    if chart_range_values:
        score = 100 if close_enough(chart_range_values[-1], current) else 20
        candidates.append((score, "chart_range", chart_range_values))

    formula_values = formula_source_values(wb_values, ws_f.cell(row, 5).value)
    if formula_values:
        score = 90 if close_enough(formula_values[-1], current) else 15
        candidates.append((score, "front_formula", formula_values))

    code_series = code_values_for_current(code, current, base_by_code)
    if code_series:
        score = 85 if close_enough(code_series[-1], current) else 10
        candidates.append((score, "code_series", code_series))

    helper_values = helper_values_for_name(name, helper)
    if helper_values:
        score = 95 if close_enough(helper_values[-1], current) else 25
        candidates.append((score, "helper", helper_values))

    if not candidates:
        return None, "no_source"
    candidates.sort(key=lambda item: (item[0], len(item[2])), reverse=True)
    _score, source_name, values = candidates[0]
    selected = values[-desired:] if desired else values
    current_number = to_number(current)
    if selected and current_number is not None and not close_enough(selected[-1], current_number):
        selected = [*selected]
        selected[-1] = current_number
        source_name = f"{source_name}+front_current_anchor"
    return selected, source_name


def literal(values: list[float]) -> str:
    return "{" + ",".join(format(float(v), ".12g") for v in values) + "}"


def update_chart_xml(chart_xml: bytes, values: list[float]) -> bytes:
    root = ET.fromstring(chart_xml)
    num_ref = root.find(".//c:lineChart/c:ser/c:val/c:numRef", NS)
    if num_ref is None:
        return chart_xml
    f = num_ref.find("c:f", NS)
    if f is None:
        f = ET.SubElement(num_ref, f"{{{NS['c']}}}f")
    f.text = literal(values)

    num_cache = num_ref.find("c:numCache", NS)
    if num_cache is None:
        num_cache = ET.SubElement(num_ref, f"{{{NS['c']}}}numCache")
    for child in list(num_cache):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag in {"ptCount", "pt"}:
            num_cache.remove(child)
    fmt = num_cache.find("c:formatCode", NS)
    insert_at = 1 if fmt is not None else 0
    num_cache.insert(insert_at, ET.Element(f"{{{NS['c']}}}ptCount", {"val": str(len(values))}))
    for idx, value in enumerate(values):
        pt = ET.Element(f"{{{NS['c']}}}pt", {"idx": str(idx)})
        v = ET.SubElement(pt, f"{{{NS['c']}}}v")
        v.text = format(float(value), ".12g")
        num_cache.append(pt)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def clear_stray_error_cells(z: ZipFile, sheet_paths: dict[str, str], config: dict[str, Any], replacements: dict[str, bytes]) -> int:
    cleared = 0
    by_sheet: dict[str, list[str]] = {}
    for item in config.get("stray_error_cells", []):
        if item.get("action") == "clear":
            by_sheet.setdefault(item["sheet"], []).append(item["cell"])
    for sheet_name, cells in by_sheet.items():
        if sheet_name not in sheet_paths:
            continue
        sheet_path = sheet_paths[sheet_name]
        root = ET.fromstring(replacements[sheet_path]) if sheet_path in replacements else read_xml(z, sheet_path)
        for cell_ref in cells:
            for cell in root.findall(f".//main:c[@r='{cell_ref}']", NS):
                for child in list(cell):
                    cell.remove(child)
                cell.attrib.pop("t", None)
                cleared += 1
        replacements[sheet_path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return cleared


def postprocess_daily(input_path: Path, output_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    wb_values = openpyxl.load_workbook(input_path, data_only=True)
    wb_formula = openpyxl.load_workbook(input_path, data_only=False)
    helper = helper_index(wb_values, config["helper_sheet"])
    base_by_code = base_series_by_code(wb_values, config["base_sheets"])
    replacements: dict[str, bytes] = {}
    reports = []

    with ZipFile(input_path, "r") as zin:
        sheet_paths = workbook_sheet_paths(zin)
        for sheet_name in config["front_sheets"]:
            if sheet_name not in sheet_paths:
                continue
            drawing_path = sheet_drawing_path(zin, sheet_paths[sheet_name])
            if drawing_path is None:
                continue
            for item in chart_rows(zin, drawing_path):
                chart_path = item["chart_path"]
                chart_xml = zin.read(chart_path)
                existing_count, chart_formula = chart_count_and_formula(chart_xml)
                values, source_name = choose_trend_values(
                    wb_values, wb_formula, helper, base_by_code, sheet_name, item["row"], chart_formula, existing_count
                )
                if not values:
                    reports.append({"sheet": sheet_name, "row": item["row"], "chart": chart_path, "source": source_name, "updated": False})
                    continue
                replacements[chart_path] = update_chart_xml(chart_xml, values)
                reports.append(
                    {
                        "sheet": sheet_name,
                        "row": item["row"],
                        "name": wb_values[sheet_name].cell(item["row"], 1).value,
                        "chart": chart_path,
                        "source": source_name,
                        "points": len(values),
                        "last": values[-1],
                        "current": wb_values[sheet_name].cell(item["row"], 5).value,
                        "updated": True,
                    }
                )

        stray_cleared = clear_stray_error_cells(zin, sheet_paths, config, replacements)

        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, replacements.get(name, zin.read(name)))

    mismatches = [r for r in reports if r.get("updated") and not close_enough(r.get("last"), r.get("current"))]
    return {
        "input": str(input_path),
        "output": str(output_path),
        "charts_seen": len(reports),
        "charts_updated": sum(1 for r in reports if r.get("updated")),
        "charts_not_updated": sum(1 for r in reports if not r.get("updated")),
        "chart_last_current_mismatch_count": len(mismatches),
        "stray_error_cells_cleared": stray_cleared,
        "mismatch_sample": mismatches[:20],
    }


def default_output_path(input_path: Path, config: dict[str, Any]) -> Path:
    suffix = config.get("output_name_patterns", {}).get("daily_postprocessed_suffix", "_图表缓存修正.xlsx")
    stem = input_path.stem
    if stem.endswith("_图表缓存修正") or "图表缓存修正" in stem:
        stem = stem.replace("_图表缓存修正", "")
    return input_path.with_name(stem + suffix)


def main() -> None:
    parser = argparse.ArgumentParser(description="Fix daily workbook front-chart series caches.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output")
    parser.add_argument("--config", default=str(Path(__file__).with_name("config.json")))
    args = parser.parse_args()

    config = load_config(Path(args.config))
    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else default_output_path(input_path, config)
    result = postprocess_daily(input_path, output_path, config)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
