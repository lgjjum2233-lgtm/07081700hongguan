from __future__ import annotations

import argparse
import copy
import json
import math
import re
import sys
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from pathlib import Path, PurePosixPath
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from postprocess_daily import load_config


RED = "FFFF0000"
GREEN = "FF008000"
TREE_SHEET_DEFAULT = "重点策略跟踪情况(V2.5)"
NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

ET.register_namespace("", NS["main"])
ET.register_namespace("r", NS["officeRel"])


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def date_key(value: Any) -> tuple[int, Any]:
    if isinstance(value, datetime):
        return 2, datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return 2, datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        return 2, datetime(1899, 12, 30) + timedelta(days=float(value))
    if value is None:
        return 0, ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return 2, datetime.strptime(text, fmt)
        except ValueError:
            pass
    return (1 if text else 0), text


def date_text(value: Any) -> str:
    kind, key = date_key(value)
    if kind == 2:
        return key.strftime("%Y-%m-%d")
    return str(key) if key else ""


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib["name"] == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    rows = sheet_data.findall("main:row", NS)
    insert_at = len(rows)
    for idx, row in enumerate(rows):
        if int(row.attrib["r"]) > row_num:
            insert_at = idx
            break
    sheet_data.insert(insert_at, new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def font_without_color(font: ET.Element) -> ET.Element:
    font = copy.deepcopy(font)
    for color in list(font.findall("main:color", NS)):
        font.remove(color)
    return font


def font_with_rgb(font: ET.Element, rgb: str) -> ET.Element:
    font = font_without_color(font)
    color_el = ET.Element(q("color"), {"rgb": rgb})
    insert_at = len(list(font))
    for idx, child in enumerate(list(font)):
        if child.tag.rsplit("}", 1)[-1] in {"sz", "u", "vertAlign", "scheme"}:
            insert_at = idx
            break
    font.insert(insert_at, color_el)
    return font


def ensure_font_style(
    styles_root: ET.Element,
    base_style_id: int,
    rgb: str | None,
    style_cache: dict[tuple[int, str | None], int],
) -> int:
    key = (base_style_id, rgb)
    if key in style_cache:
        return style_cache[key]

    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fonts is None or cell_xfs is None:
        raise KeyError("styles font/cellXfs")

    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]
    base_font_id = int(base_xf.attrib.get("fontId", "0"))
    font_nodes = fonts.findall("main:font", NS)
    base_font = font_nodes[base_font_id] if 0 <= base_font_id < len(font_nodes) else font_nodes[0]

    new_font = font_with_rgb(base_font, rgb) if rgb else font_without_color(base_font)
    fonts.append(new_font)
    new_font_id = len(font_nodes)
    fonts.attrib["count"] = str(len(font_nodes) + 1)

    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["applyFont"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(len(xfs) + 1)
    style_cache[key] = new_style_id
    return new_style_id


def font_rgb(cell) -> str | None:
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return f"{color.type}:{color.value}"


def normalize_name(value: Any) -> str:
    text = str(value or "").lower()
    text = re.sub(r"\[[^\]]+\]", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[：:;\s_\-—]+", "", text)
    return text


def row_identity(ws, row: int) -> tuple[str, str] | None:
    code = str(ws.cell(row, 10).value or "").strip().upper()
    if code and code not in {"-", "—"}:
        return "code", code
    normalized = normalize_name(ws.cell(row, 4).value)
    if normalized:
        return "name", normalized
    return None


def previous_row_index(previous_tree: Path, tree_sheet: str) -> dict[tuple[str, str], dict[str, Any]]:
    wb = openpyxl.load_workbook(previous_tree, data_only=True)
    ws = wb[tree_sheet]
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for row in range(6, ws.max_row + 1):
        identity = row_identity(ws, row)
        if identity is None:
            continue
        date_value = ws.cell(row, 12).value
        if not date_text(date_value):
            continue
        index[identity] = {
            "row": row,
            "name": ws.cell(row, 4).value,
            "date": date_value,
        }
    return index


def expected_color_rows(previous_tree: Path, current_tree: Path, tree_sheet: str) -> list[dict[str, Any]]:
    prev_index = previous_row_index(previous_tree, tree_sheet)
    curr_wb = openpyxl.load_workbook(current_tree, data_only=True)
    curr_ws = curr_wb[tree_sheet]
    rows_by_row: dict[int, dict[str, Any]] = {}
    for row in range(6, curr_ws.max_row + 1):
        name = curr_ws.cell(row, 4).value
        if not name:
            continue
        identity = row_identity(curr_ws, row)
        if identity is None or identity not in prev_index:
            continue
        curr_date = curr_ws.cell(row, 12).value
        prev_date = prev_index[identity]["date"]
        change_num = to_number(curr_ws.cell(row, 13).value)
        if date_key(curr_date) > date_key(prev_date) and change_num is not None and change_num != 0:
            rows_by_row[row] = {
                "row": row,
                "name": name,
                "identity": identity,
                "previous_row": prev_index[identity]["row"],
                "code": curr_ws.cell(row, 10).value,
                "previous_date": date_text(prev_date),
                "current_date": date_text(curr_date),
                "change": curr_ws.cell(row, 13).value,
                "change_num": change_num,
                "expected_rgb": RED if change_num > 0 else GREEN,
                "reason": "date_advanced",
            }

    latest_key: tuple[int, Any] = (0, "")
    latest_text = ""
    for row in range(6, curr_ws.max_row + 1):
        change_num = to_number(curr_ws.cell(row, 13).value)
        if change_num is None or change_num == 0:
            continue
        key = date_key(curr_ws.cell(row, 12).value)
        if key > latest_key:
            latest_key = key
            latest_text = date_text(curr_ws.cell(row, 12).value)
    if latest_text:
        for row in range(6, curr_ws.max_row + 1):
            name = curr_ws.cell(row, 4).value
            change_num = to_number(curr_ws.cell(row, 13).value)
            if not name or change_num is None or change_num == 0:
                continue
            if date_text(curr_ws.cell(row, 12).value) != latest_text:
                continue
            if row not in rows_by_row:
                rows_by_row[row] = {
                    "row": row,
                    "name": name,
                    "identity": row_identity(curr_ws, row),
                    "previous_row": None,
                    "code": curr_ws.cell(row, 10).value,
                    "previous_date": "",
                    "current_date": latest_text,
                    "change": curr_ws.cell(row, 13).value,
                    "change_num": change_num,
                    "expected_rgb": RED if change_num > 0 else GREEN,
                    "reason": "latest_nonzero_date",
                }
    return [rows_by_row[row] for row in sorted(rows_by_row)]


def apply_change_colors(
    current_tree: Path,
    previous_tree: Path,
    output_tree: Path,
    tree_sheet: str,
) -> dict[str, Any]:
    rows = expected_color_rows(previous_tree, current_tree, tree_sheet)
    targets = {item["row"]: item for item in rows}
    style_cache: dict[tuple[int, str | None], int] = {}

    with ZipFile(current_tree, "r") as zin:
        sheet_path = workbook_sheet_path(zin, tree_sheet)
        sheet_root = read_xml(zin, sheet_path)
        styles_root = read_xml(zin, "xl/styles.xml")

        for row_num in range(6, openpyxl.load_workbook(current_tree, read_only=True)[tree_sheet].max_row + 1):
            cell = find_cell(sheet_root, row_num, 13)
            base_style_id = int(cell.attrib.get("s", "0"))
            uncolored_style = ensure_font_style(styles_root, base_style_id, None, style_cache)
            cell.attrib["s"] = str(uncolored_style)

        for row_num, item in targets.items():
            cell = find_cell(sheet_root, row_num, 13)
            base_style_id = int(cell.attrib.get("s", "0"))
            rgb = RED if item["change_num"] > 0 else GREEN
            cell.attrib["s"] = str(ensure_font_style(styles_root, base_style_id, rgb, style_cache))

        sheet_xml = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
        styles_xml = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)

        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, sheet_xml)
                elif name == "xl/styles.xml":
                    zout.writestr(name, styles_xml)
                else:
                    zout.writestr(name, zin.read(name))

    check_wb = openpyxl.load_workbook(output_tree, data_only=True)
    check_ws = check_wb[tree_sheet]
    bad = []
    extras = []
    colored = []
    for row_num in range(6, check_ws.max_row + 1):
        rgb = font_rgb(check_ws.cell(row_num, 13))
        if rgb in {RED, GREEN}:
            colored.append(row_num)
            expected = targets.get(row_num)
            if expected is None:
                extras.append({"row": row_num, "name": check_ws.cell(row_num, 4).value, "rgb": rgb})
            elif rgb != expected["expected_rgb"]:
                bad.append(expected | {"actual_rgb": rgb})
    for row_num, expected in targets.items():
        rgb = font_rgb(check_ws.cell(row_num, 13))
        if rgb != expected["expected_rgb"]:
            bad.append(expected | {"actual_rgb": rgb})

    return {
        "previous_tree": str(previous_tree),
        "colored_rows": len(rows),
        "actual_colored_rows": len(colored),
        "bad_color_count": len(bad),
        "extra_color_count": len(extras),
        "red_rows": sum(1 for row_num in colored if font_rgb(check_ws.cell(row_num, 13)) == RED),
        "green_rows": sum(1 for row_num in colored if font_rgb(check_ws.cell(row_num, 13)) == GREEN),
        "colored_sample": rows[:20],
        "bad_color_sample": bad[:20],
        "extra_color_sample": extras[:20],
    }


def default_output_path(tree_path: Path) -> Path:
    stem = tree_path.stem
    if "_交付版" in stem:
        stem = stem.replace("_交付版", "")
    for suffix in ["_V1", "_V2", "_V3", "_V4", "_V5", "_V6"]:
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
            break
    for idx in range(7, 50):
        candidate = tree_path.with_name(f"{stem}_V{idx}_新流程更新复核版.xlsx")
        if not candidate.exists():
            return candidate
    raise RuntimeError("No available TREE output version.")


def sync_tree(
    daily_path: Path,
    tree_path: Path,
    output_path: Path,
    previous_tree_for_colors: Path,
    config: dict[str, Any],
) -> dict[str, Any]:
    workspace = Path(config["workspace"])
    legacy_tools = workspace / "codex_tmp" / "20260604_tree_update"
    refresh_tools = workspace / "codex_tmp" / "20260605_tree_refresh"
    if not legacy_tools.exists():
        raise FileNotFoundError(f"Missing legacy TREE tools: {legacy_tools}")
    if str(legacy_tools) not in sys.path:
        sys.path.insert(0, str(legacy_tools))
    if str(refresh_tools) not in sys.path and refresh_tools.exists():
        sys.path.insert(0, str(refresh_tools))

    import sync_tree_20260604 as base  # type: ignore
    import sync_tree_20260604_legacy_charts as legacy  # type: ignore
    import sync_tree_20260604_preserve_chart_xml as preserve  # type: ignore
    import sync_tree_20260605_refreshed_daily as refreshed  # type: ignore

    pct_change_codes = {str(code).upper() for code in config.get("stock_index_pct_change_codes", [])}
    original_stock_index_change_row = base.stock_index_change_row

    def stock_index_pct_change_row(row_name: str, row_code: str) -> bool:
        return original_stock_index_change_row(row_name, row_code) or str(row_code or "").strip().upper() in pct_change_codes

    base.stock_index_change_row = stock_index_pct_change_row
    if hasattr(preserve, "stock_index_change_row"):
        preserve.stock_index_change_row = stock_index_pct_change_row
    if hasattr(legacy, "stock_index_change_row"):
        legacy.stock_index_change_row = stock_index_pct_change_row

    tree_sheet = config.get("tree_sheet", TREE_SHEET_DEFAULT)
    for module in (base, legacy, preserve, refreshed):
        if hasattr(module, "ROOT"):
            module.ROOT = workspace
        if hasattr(module, "DAILY_PATH"):
            module.DAILY_PATH = daily_path
        if hasattr(module, "TREE_PATH"):
            module.TREE_PATH = tree_path
        if hasattr(module, "TREE_SHEET"):
            module.TREE_SHEET = tree_sheet

    row_sources, trends, source_meta = legacy.build_sources()
    sheet_xml, data_meta = preserve.update_sheet_xml(row_sources)

    with TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / "tree_data_updated.xlsx"
        chart_tmp = Path(tmp_dir) / "tree_charts_updated.xlsx"
        chart_complete_tmp = Path(tmp_dir) / "tree_charts_complete.xlsx"
        preserve.write_data_only_xlsx(sheet_xml, tmp_path)
        chart_meta = refreshed.update_existing_tree_charts(tmp_path, chart_tmp, trends)
        required_chart_rows = list(config.get("required_tree_chart_rows", []))
        missing_chart_meta: dict[str, Any] = {"charts_added": []}
        if required_chart_rows:
            old_target_rows = getattr(legacy, "TARGET_ROWS", [])
            try:
                legacy.TARGET_ROWS = required_chart_rows
                missing_chart_meta = legacy.append_legacy_charts(chart_tmp, chart_complete_tmp, trends)
            finally:
                legacy.TARGET_ROWS = old_target_rows
        else:
            chart_complete_tmp = chart_tmp
        color_meta = apply_change_colors(chart_complete_tmp, previous_tree_for_colors, output_path, tree_sheet)

    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[tree_sheet]
    mandatory_rows = []
    for item in config.get("mandatory_sources", []):
        row = item.get("tree_row")
        if not row:
            continue
        mandatory_rows.append(
            {
                "row": row,
                "name": ws.cell(row, 4).value,
                "code": ws.cell(row, 10).value,
                "current": ws.cell(row, 11).value,
                "date": date_text(ws.cell(row, 12).value),
                "change": ws.cell(row, 13).value,
                "source": f"{row_sources[row].sheet}!{row_sources[row].row}" if row in row_sources else None,
                "trend_points": len(trends.get(row, [])),
            }
        )

    return {
        "input_daily": str(daily_path),
        "input_tree": str(tree_path),
        "output": str(output_path),
        **source_meta,
        **data_meta,
        **chart_meta,
        "required_chart_rows": required_chart_rows,
        "missing_chart_fix": missing_chart_meta,
        "colors": color_meta,
        "mandatory_rows": mandatory_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Synchronize TREE from the refreshed daily workbook.")
    parser.add_argument("--daily", required=True)
    parser.add_argument("--tree", required=True)
    parser.add_argument("--output")
    parser.add_argument("--previous-tree")
    parser.add_argument("--config", default=str(Path(__file__).with_name("config.json")))
    args = parser.parse_args()

    config = load_config(Path(args.config))
    daily_path = Path(args.daily)
    tree_path = Path(args.tree)
    output_path = Path(args.output) if args.output else default_output_path(tree_path)
    previous_tree = Path(args.previous_tree) if args.previous_tree else tree_path
    result = sync_tree(daily_path, tree_path, output_path, previous_tree, config)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
