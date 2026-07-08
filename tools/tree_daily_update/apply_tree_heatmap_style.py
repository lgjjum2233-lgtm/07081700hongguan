from __future__ import annotations

import argparse
import copy
import json
import math
import xml.etree.ElementTree as ET
from bisect import bisect_right
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS = {
    "main": SHEET_NS,
    "officeRel": REL_NS,
}

ET.register_namespace("", SHEET_NS)
ET.register_namespace("r", REL_NS)

BLACK = "FF000000"
WHITE = "FFFFFFFF"
WEIGHT_LIGHT = "FFFFF7E6"
WEIGHT_DARK = "FFC65911"
CHANGE_LIGHT = "FFEAF2F8"
CHANGE_DARK = "FF1F4E79"


def q(tag: str) -> str:
    return f"{{{SHEET_NS}}}{tag}"


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "None", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}:
        return None
    is_percent = "%" in text
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if is_percent else number


def hex_to_rgb(rgb: str) -> tuple[int, int, int]:
    rgb = rgb[-6:]
    return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)


def rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"FF{r:02X}{g:02X}{b:02X}"


def interpolate_color(light: str, dark: str, norm: float) -> str:
    norm = max(0.0, min(1.0, norm))
    lr, lg, lb = hex_to_rgb(light)
    dr, dg, db = hex_to_rgb(dark)
    r = round(lr + (dr - lr) * norm)
    g = round(lg + (dg - lg) * norm)
    b = round(lb + (db - lb) * norm)
    return rgb_to_hex(r, g, b)


def text_color_for_fill(fill_rgb: str) -> str:
    r, g, b = hex_to_rgb(fill_rgb)
    # Perceived luminance is enough here; keep dark heat cells readable.
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return WHITE if luminance < 135 else BLACK


def normalized_rank(sorted_values: list[float], value: float) -> float:
    if not sorted_values:
        return 0.0
    return bisect_right(sorted_values, value) / len(sorted_values)


def weight_norm(value: float, max_value: float, sorted_values: list[float]) -> float:
    if max_value <= 0:
        return 0.0
    value_norm = math.sqrt(max(0.0, value) / max_value)
    rank_norm = normalized_rank(sorted_values, value)
    return 0.75 * value_norm + 0.25 * rank_norm


def change_norm(value_abs: float, max_value_abs: float, sorted_abs_values: list[float]) -> float:
    if max_value_abs <= 0:
        return 0.0
    log_norm = math.log1p(value_abs) / math.log1p(max_value_abs)
    rank_norm = normalized_rank(sorted_abs_values, value_abs)
    return 0.35 * log_norm + 0.65 * rank_norm


def read_xml(zf: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(zf.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(zf: ZipFile, sheet_name: str) -> str:
    workbook = read_xml(zf, "xl/workbook.xml")
    rels = read_xml(zf, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in workbook.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{REL_NS}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib.get("r", "0")) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    rows = sheet_data.findall("main:row", NS)
    insert_at = len(rows)
    for idx, row in enumerate(rows):
        if int(row.attrib.get("r", "0")) > row_num:
            insert_at = idx
            break
    sheet_data.insert(insert_at, new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def split_col_attrs(attrs: dict[str, str], min_col: int, max_col: int) -> dict[str, str]:
    new_attrs = dict(attrs)
    new_attrs["min"] = str(min_col)
    new_attrs["max"] = str(max_col)
    return new_attrs


def unhide_single_column(sheet_root: ET.Element, col_num: int, width: float | None = None) -> bool:
    cols = sheet_root.find("main:cols", NS)
    if cols is None:
        worksheet_children = list(sheet_root)
        insert_at = 0
        for idx, child in enumerate(worksheet_children):
            if child.tag.rsplit("}", 1)[-1] in {"sheetPr", "dimension", "sheetViews", "sheetFormatPr"}:
                insert_at = idx + 1
        cols = ET.Element(q("cols"))
        sheet_root.insert(insert_at, cols)

    changed = False
    new_col_nodes: list[ET.Element] = []
    handled = False
    for col in list(cols.findall("main:col", NS)):
        min_col = int(col.attrib["min"])
        max_col = int(col.attrib["max"])
        if min_col <= col_num <= max_col:
            handled = True
            attrs = dict(col.attrib)
            if min_col <= col_num - 1:
                new_col_nodes.append(ET.Element(q("col"), split_col_attrs(attrs, min_col, col_num - 1)))

            target_attrs = split_col_attrs(attrs, col_num, col_num)
            for attr in ("hidden", "collapsed"):
                target_attrs.pop(attr, None)
            if width is not None:
                target_attrs["width"] = f"{width:.2f}"
                target_attrs["customWidth"] = "1"
            new_col_nodes.append(ET.Element(q("col"), target_attrs))

            if col_num + 1 <= max_col:
                new_col_nodes.append(ET.Element(q("col"), split_col_attrs(attrs, col_num + 1, max_col)))
            if col.attrib.get("hidden") == "1" or "hidden" in col.attrib:
                changed = True
        else:
            new_col_nodes.append(copy.deepcopy(col))

    if not handled:
        attrs = {"min": str(col_num), "max": str(col_num)}
        if width is not None:
            attrs["width"] = f"{width:.2f}"
            attrs["customWidth"] = "1"
        new_col_nodes.append(ET.Element(q("col"), attrs))
        changed = True

    for col in list(cols.findall("main:col", NS)):
        cols.remove(col)
    new_col_nodes.sort(key=lambda node: (int(node.attrib["min"]), int(node.attrib["max"])))
    for col in new_col_nodes:
        cols.append(col)

    return changed


def font_with_rgb(font: ET.Element, rgb: str) -> ET.Element:
    new_font = copy.deepcopy(font)
    for color in list(new_font.findall("main:color", NS)):
        new_font.remove(color)
    color_el = ET.Element(q("color"), {"rgb": rgb})
    insert_at = len(list(new_font))
    for idx, child in enumerate(list(new_font)):
        if child.tag.rsplit("}", 1)[-1] in {"sz", "u", "vertAlign", "scheme"}:
            insert_at = idx
            break
    new_font.insert(insert_at, color_el)
    return new_font


def fill_with_rgb(rgb: str) -> ET.Element:
    fill = ET.Element(q("fill"))
    pattern = ET.SubElement(fill, q("patternFill"), {"patternType": "solid"})
    ET.SubElement(pattern, q("fgColor"), {"rgb": rgb})
    ET.SubElement(pattern, q("bgColor"), {"indexed": "64"})
    return fill


def ensure_heatmap_style(
    styles_root: ET.Element,
    base_style_id: int,
    fill_rgb: str,
    font_rgb: str,
    cache: dict[tuple[int, str, str], int],
) -> int:
    key = (base_style_id, fill_rgb, font_rgb)
    if key in cache:
        return cache[key]

    fills = styles_root.find("main:fills", NS)
    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fills is None or fonts is None or cell_xfs is None:
        raise KeyError("styles fills/fonts/cellXfs")

    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]

    fonts_list = fonts.findall("main:font", NS)
    base_font_id = int(base_xf.attrib.get("fontId", "0"))
    base_font = fonts_list[base_font_id] if 0 <= base_font_id < len(fonts_list) else fonts_list[0]
    fonts.append(font_with_rgb(base_font, font_rgb))
    new_font_id = len(fonts_list)
    fonts.attrib["count"] = str(len(fonts_list) + 1)

    fills_list = fills.findall("main:fill", NS)
    fills.append(fill_with_rgb(fill_rgb))
    new_fill_id = len(fills_list)
    fills.attrib["count"] = str(len(fills_list) + 1)

    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["fillId"] = str(new_fill_id)
    new_xf.attrib["applyFont"] = "1"
    new_xf.attrib["applyFill"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(len(xfs) + 1)

    cache[key] = new_style_id
    return new_style_id


def find_tree_header(ws) -> dict[str, int] | None:
    required = {"indicator": "监测指标", "weight": "权重占比", "change": "边际变化"}
    for row in range(1, min(ws.max_row, 15) + 1):
        found: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            text = as_text(ws.cell(row, col).value)
            for key, label in required.items():
                if text == label:
                    found[key] = col
        if set(found) == set(required):
            found["header_row"] = row
            return found
    return None


def collect_targets(wb, sheet_names: list[str] | None, include_hidden: bool) -> list[dict[str, Any]]:
    targets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if sheet_names is not None and ws.title not in sheet_names:
            continue
        if sheet_names is None and not ws.title.startswith("重点策略跟踪情况"):
            continue
        if not include_hidden and ws.sheet_state != "visible":
            continue
        header = find_tree_header(ws)
        if header is None:
            continue
        targets.append(
            {
                "sheet": ws.title,
                "header_row": header["header_row"],
                "indicator_col": header["indicator"],
                "weight_col": header["weight"],
                "change_col": header["change"],
                "max_row": ws.max_row,
            }
        )
    return targets


def apply_heatmap(input_path: Path, output_path: Path, sheet_names: list[str] | None, include_hidden: bool) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    targets = collect_targets(wb, sheet_names, include_hidden)
    if not targets:
        raise RuntimeError("No visible TREE sheet with headers: 监测指标 / 权重占比 / 边际变化")

    replacements: dict[str, bytes] = {}
    style_cache: dict[tuple[int, str, str], int] = {}
    summaries: list[dict[str, Any]] = []

    with ZipFile(input_path, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        for target in targets:
            ws = wb[target["sheet"]]
            sheet_path = workbook_sheet_path(zin, target["sheet"])
            sheet_root = read_xml(zin, sheet_path)

            weight_rows: list[tuple[int, float]] = []
            change_rows: list[tuple[int, float]] = []
            for row in range(target["header_row"] + 1, target["max_row"] + 1):
                indicator = as_text(ws.cell(row, target["indicator_col"]).value)
                if not indicator:
                    continue
                weight = to_number(ws.cell(row, target["weight_col"]).value)
                if weight is not None and weight >= 0:
                    weight_rows.append((row, weight))
                change = to_number(ws.cell(row, target["change_col"]).value)
                if change is not None and abs(change) > 0:
                    change_rows.append((row, abs(change)))

            weights = sorted(value for _, value in weight_rows)
            changes = sorted(value for _, value in change_rows)
            max_weight = max(weights) if weights else 0.0
            max_change = max(changes) if changes else 0.0

            weight_fills = 0
            change_fills = 0

            for row, value in weight_rows:
                norm = weight_norm(value, max_weight, weights)
                fill_rgb = interpolate_color(WEIGHT_LIGHT, WEIGHT_DARK, norm)
                font_rgb = text_color_for_fill(fill_rgb)
                for col in (target["indicator_col"], target["weight_col"]):
                    cell = find_cell(sheet_root, row, col)
                    base_style_id = int(cell.attrib.get("s", "0"))
                    cell.attrib["s"] = str(ensure_heatmap_style(styles_root, base_style_id, fill_rgb, font_rgb, style_cache))
                    weight_fills += 1

            for row, value_abs in change_rows:
                norm = change_norm(value_abs, max_change, changes)
                fill_rgb = interpolate_color(CHANGE_LIGHT, CHANGE_DARK, norm)
                font_rgb = text_color_for_fill(fill_rgb)
                cell = find_cell(sheet_root, row, target["change_col"])
                base_style_id = int(cell.attrib.get("s", "0"))
                cell.attrib["s"] = str(ensure_heatmap_style(styles_root, base_style_id, fill_rgb, font_rgb, style_cache))
                change_fills += 1

            weight_column_unhidden = unhide_single_column(sheet_root, target["weight_col"])
            replacements[sheet_path] = write_xml(sheet_root)
            summaries.append(
                {
                    "sheet": target["sheet"],
                    "header_row": target["header_row"],
                    "indicator_col": target["indicator_col"],
                    "weight_col": target["weight_col"],
                    "change_col": target["change_col"],
                    "weight_rows": len(weight_rows),
                    "nonzero_change_rows": len(change_rows),
                    "max_weight": max_weight,
                    "max_abs_change": max_change,
                    "styled_weight_cells": weight_fills,
                    "styled_change_cells": change_fills,
                    "weight_column_unhidden": weight_column_unhidden,
                }
            )

        replacements["xl/styles.xml"] = write_xml(styles_root)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    return {"input": str(input_path), "output": str(output_path), "styled_sheets": summaries}


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply weight and marginal-change heatmap fills to TREE workbook display sheets.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--sheet", action="append", help="Sheet name to style. Can be repeated. Defaults to visible TREE sheets.")
    parser.add_argument("--include-hidden", action="store_true", help="Also style hidden TREE sheets when --sheet is not provided.")
    args = parser.parse_args()

    result = apply_heatmap(Path(args.input), Path(args.output), args.sheet, args.include_hidden)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
