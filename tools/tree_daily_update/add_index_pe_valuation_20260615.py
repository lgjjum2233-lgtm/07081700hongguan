from __future__ import annotations

import argparse
import copy
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


INDEX_SHEET = "指数走势"
FRONT_SHEET = "A股港股"
HELPER_SHEET = "辅助数据"

PE_FIELD = "pe_ttm"
PE_BLOCK_TITLE = "市盈率(TTM)"
PE_ROWS = [
    ("上证指数PE(TTM)", "000001.SH"),
    ("深证成指PE(TTM)", "399001.SZ"),
    ("沪深300PE(TTM)", "000300.SH"),
    ("创业板指PE(TTM)", "399006.SZ"),
    ("恒生指数PE(TTM)", "HSI.HI"),
    ("恒生科技PE(TTM)", "HSTECH.HI"),
]

PE_COLS = 760
HELPER_POINTS = 760
HELPER_START_COL = 2


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.protection = copy.copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    ws.row_dimensions[dst_row].hidden = ws.row_dimensions[src_row].hidden
    ws.row_dimensions[dst_row].outlineLevel = ws.row_dimensions[src_row].outlineLevel


def clear_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def norm(value: Any) -> str:
    return str(value or "").strip().upper()


def find_existing_pe_rows(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        title = str(ws.cell(row, 3).value or "")
        code = norm(ws.cell(row, 5).value)
        if title.lower() == PE_FIELD and code:
            out[code] = row
    return out


def next_blank_row(ws) -> int:
    row = ws.max_row + 1
    while row > 1:
        if any(ws.cell(row - 1, col).value not in (None, "") for col in range(1, 6)):
            return row
        row -= 1
    return ws.max_row + 1


def front_formulas(front_row: int, source_row: int, header_row: int, end_col: int) -> dict[int, str]:
    end_letter = get_column_letter(end_col)
    data_rng = f"{INDEX_SHEET}!$F${source_row}:${end_letter}${source_row}"
    date_rng = f"{INDEX_SHEET}!$F${header_row}:${end_letter}${header_row}"
    return {
        5: f'=IFERROR(LOOKUP(9.99E+307,{data_rng}),"")',
        6: f'=IFERROR(LOOKUP(2,1/({data_rng}<>""),{date_rng}),"")',
        7: f'=IFERROR(E{front_row}-H{front_row},"")',
        8: (
            f'=IFERROR(LOOKUP(9.99E+307,OFFSET({INDEX_SHEET}!$F${source_row},0,0,1,'
            f'LOOKUP(2,1/({data_rng}<>""),COLUMN({data_rng}))-COLUMN({INDEX_SHEET}!$F${source_row}))),"")'
        ),
        9: f'=IFERROR(LOOKUP(2,1/({data_rng}<>"")/({date_rng}<F{front_row}),{date_rng}),"")',
    }


def helper_formula(source_row: int, end_col: int, point: int) -> str:
    end_letter = get_column_letter(end_col)
    data_rng = f"{INDEX_SHEET}!$F${source_row}:${end_letter}${source_row}"
    last_idx = f'LOOKUP(2,1/({data_rng}<>""),COLUMN({data_rng}))-COLUMN({INDEX_SHEET}!$F${source_row})+1'
    idx = f"{last_idx}-{HELPER_POINTS}+{point}"
    value = f"INDEX({data_rng},1,{idx})"
    return f"=IF(ISERROR({value}),NA(),IF({value}=0,NA(),{value}))"


def remove_charts_at_rows(ws, rows: set[int]) -> int:
    before = len(ws._charts)
    kept = []
    for chart in ws._charts:
        anchor = getattr(chart, "anchor", None)
        chart_row = None
        if hasattr(anchor, "_from"):
            chart_row = anchor._from.row + 1
        if chart_row in rows:
            continue
        kept.append(chart)
    ws._charts = kept
    return before - len(kept)


def add_trend_chart(front_ws, helper_ws, front_row: int, helper_row: int) -> None:
    chart = LineChart()
    chart.title = None
    chart.legend = None
    chart.x_axis.delete = True
    chart.y_axis.delete = True
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.numFmt = "General"
    chart.y_axis.numFmt = "General"

    data = Reference(
        helper_ws,
        min_col=HELPER_START_COL,
        max_col=HELPER_START_COL + HELPER_POINTS - 1,
        min_row=helper_row,
        max_row=helper_row,
    )
    chart.add_data(data, from_rows=True, titles_from_data=False)
    for series in chart.series:
        series.graphicalProperties.line.solidFill = "C00000"
        series.graphicalProperties.line.width = 15120
        series.marker.symbol = "none"

    chart.anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=9, colOff=0, row=front_row - 1, rowOff=0),
        to=AnchorMarker(col=10, colOff=559080, row=front_row, rowOff=286200),
    )
    front_ws.add_chart(chart)


def ensure_index_pe_block(wb) -> tuple[int, dict[str, int], int]:
    ws = wb[INDEX_SHEET]
    existing = find_existing_pe_rows(ws)
    if all(code in existing for _name, code in PE_ROWS):
        header_row = min(existing.values()) - 1
        return header_row, existing, 6 + PE_COLS - 1

    header_row = next_blank_row(ws) + 1
    first_data_row = header_row + 1
    end_col = 6 + PE_COLS - 1

    # Leave one blank row before the PE block.
    copy_row_style(ws, 19, header_row, end_col)
    clear_row(ws, header_row, end_col)
    ws.cell(header_row, 1).value = "开始日期"
    ws.cell(header_row, 2).value = "截止日期"
    ws.cell(header_row, 3).value = PE_BLOCK_TITLE
    ws.cell(header_row, 4).value = "日期"
    ws.cell(header_row, 5).value = "Date"
    for col in range(6, end_col + 1):
        copy_cell_style(ws.cell(19, min(col, ws.max_column)), ws.cell(header_row, col))

    rows_by_code: dict[str, int] = {}
    for offset, (_label, code) in enumerate(PE_ROWS):
        row = first_data_row + offset
        copy_row_style(ws, 20, row, end_col)
        clear_row(ws, row, end_col)
        ws.cell(row, 4).value = f"=[1]!s_info_name(E{row})"
        ws.cell(row, 5).value = code
        rows_by_code[code] = row
        for col in range(6, end_col + 1):
            ws.cell(row, col).number_format = "0.00"

    ws.cell(first_data_row, 1).value = '=TEXT(EDATE(TODAY(),-36),"YYYYMMDD")'
    ws.cell(first_data_row, 2).value = '=TEXT(TODAY(),"YYYYMMDD")'
    ws.cell(first_data_row, 3).value = PE_FIELD
    formula = (
        f'=[1]!wsd(E{first_data_row}:E{first_data_row + len(PE_ROWS) - 1},'
        f'C{first_data_row},A{first_data_row},B{first_data_row},'
        '"TradingCalendar=SSE","rptType=1","Direction=V","Version=1","ShowParams=Y",'
        f'"cols={PE_COLS};rows={len(PE_ROWS)}")'
    )
    ws.cell(first_data_row, 6).value = ArrayFormula(f"F{first_data_row}", formula)

    return header_row, rows_by_code, end_col


def ensure_front_rows(wb, rows_by_code: dict[str, int], header_row: int, end_col: int) -> dict[str, int]:
    ws = wb[FRONT_SHEET]
    existing_labels = {str(ws.cell(row, 1).value or "").strip(): row for row in range(1, ws.max_row + 1)}
    rows: dict[str, int] = {}
    for label, code in PE_ROWS:
        if label in existing_labels:
            front_row = existing_labels[label]
        else:
            front_row = ws.max_row + 1
            copy_row_style(ws, 2, front_row, 10)
            for col in range(1, 11):
                ws.cell(front_row, col).value = None
            ws.cell(front_row, 1).value = label
            ws.cell(front_row, 2).value = "日度"
            ws.cell(front_row, 3).value = code
            ws.cell(front_row, 4).value = "倍"
            ws.cell(front_row, 5).number_format = "0.00"
            ws.cell(front_row, 7).number_format = "0.00"
            ws.cell(front_row, 8).number_format = "0.00"
        source_row = rows_by_code[code]
        for col, formula in front_formulas(front_row, source_row, header_row, end_col).items():
            ws.cell(front_row, col).value = formula
        rows[label] = front_row
    return rows


def ensure_helper_rows_and_charts(wb, front_rows: dict[str, int], rows_by_code: dict[str, int], end_col: int) -> dict[str, Any]:
    helper = wb[HELPER_SHEET]
    front = wb[FRONT_SHEET]
    existing = {str(helper.cell(row, 1).value or "").strip(): row for row in range(1, helper.max_row + 1)}
    helper_rows: dict[str, int] = {}

    for label, code in PE_ROWS:
        helper_row = existing.get(label)
        if helper_row is None:
            helper_row = helper.max_row + 1
        copy_row_style(helper, 1, helper_row, HELPER_START_COL + HELPER_POINTS - 1)
        helper.cell(helper_row, 1).value = label
        for point, col in enumerate(range(HELPER_START_COL, HELPER_START_COL + HELPER_POINTS), start=1):
            helper.cell(helper_row, col).value = helper_formula(rows_by_code[code], end_col, point)
        helper_rows[label] = helper_row

    removed = remove_charts_at_rows(front, set(front_rows.values()))
    for label, _code in PE_ROWS:
        add_trend_chart(front, helper, front_rows[label], helper_rows[label])

    return {"helper_rows": helper_rows, "removed_existing_charts": removed}


def apply_basic_styles(wb, front_rows: dict[str, int]) -> None:
    front = wb[FRONT_SHEET]
    for row in front_rows.values():
        for col in range(1, 11):
            copy_cell_style(front.cell(2, col), front.cell(row, col))
        front.cell(row, 1).font = copy.copy(front.cell(2, 1).font)
        front.cell(row, 4).value = "倍"
        for col in (5, 7, 8):
            front.cell(row, col).number_format = "0.00"
        for col in (6, 9):
            front.cell(row, col).number_format = "yyyy-mm-dd"

    # Keep the new labels plain black like the surrounding A/H equity rows.
    for row in front_rows.values():
        front.cell(row, 1).font = Font(name=front.cell(2, 1).font.name, size=front.cell(2, 1).font.sz, color="000000")
        for col in range(1, 11):
            front.cell(row, col).fill = copy.copy(front.cell(2, col).fill)
    if isinstance(front.cell(2, 1).fill, PatternFill):
        pass


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    wb = openpyxl.load_workbook(in_path, data_only=False)
    header_row, rows_by_code, end_col = ensure_index_pe_block(wb)
    front_rows = ensure_front_rows(wb, rows_by_code, header_row, end_col)
    helper_info = ensure_helper_rows_and_charts(wb, front_rows, rows_by_code, end_col)
    apply_basic_styles(wb, front_rows)

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    wb.save(out_path)

    verify = openpyxl.load_workbook(out_path, data_only=False)
    vfront = verify[FRONT_SHEET]
    vindex = verify[INDEX_SHEET]
    summary = {
        "output": str(out_path),
        "pe_field": PE_FIELD,
        "pe_block_header_row": header_row,
        "pe_source_rows": rows_by_code,
        "pe_end_column": get_column_letter(end_col),
        "front_rows": front_rows,
        "helper": helper_info,
        "front_chart_count": len(vfront._charts),
        "index_array_formula": getattr(vindex.cell(min(rows_by_code.values()), 6).value, "text", vindex.cell(min(rows_by_code.values()), 6).value),
        "front_samples": {
            label: [vfront.cell(row, col).value for col in range(1, 10)]
            for label, row in front_rows.items()
        },
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
