from __future__ import annotations

import argparse
import copy
import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


DATE_LAST_COL = 534  # TN
MACRO_SHEET_INDEX = 4
CHINA_SHEET_INDEX = 2
OVERSEAS_SHEET_INDEX = 3

NS = {
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}

for prefix, uri in NS.items():
    ET.register_namespace(prefix, uri)


@dataclass(frozen=True)
class Indicator:
    name: str
    code: str
    sheet_index: int
    freq: str
    unit: str


INDICATORS = [
    Indicator("中国:一般公共预算支出:累计值", "M0046166", CHINA_SHEET_INDEX, "月度", "亿元"),
    Indicator("中国:固定资产投资累计值", "M0000272", CHINA_SHEET_INDEX, "月度", "亿元"),
    Indicator("美国:所有商业银行:资产:银行信贷:季调", "G1109077", OVERSEAS_SHEET_INDEX, "月度", "亿美元"),
    Indicator("美国:GDP:现价:季调:折年数", "G0000003", OVERSEAS_SHEET_INDEX, "季度", "十亿美元"),
    Indicator("美国:零售和食品服务销售额:季调:当月同比", "G1109245", OVERSEAS_SHEET_INDEX, "月度", "%"),
]


def norm(value: Any) -> str:
    return str(value or "").strip().upper()


def formula_text(value: Any) -> str:
    return value.text if isinstance(value, ArrayFormula) else str(value or "")


def copy_cell(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.protection = copy.copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        copy_cell(ws.cell(src_row, col), ws.cell(dst_row, col))
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    ws.row_dimensions[dst_row].hidden = ws.row_dimensions[src_row].hidden
    ws.row_dimensions[dst_row].outlineLevel = ws.row_dimensions[src_row].outlineLevel


def copy_row_all(ws, src_row: int, dst_row: int, max_col: int) -> None:
    copy_row_style(ws, src_row, dst_row, max_col)
    for col in range(1, max_col + 1):
        ws.cell(dst_row, col).value = ws.cell(src_row, col).value


def clear_row_values(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def find_raw_end(macro) -> int:
    text = formula_text(macro["F2"].value)
    match = re.search(r"E2:E(\d+)", text)
    if not match:
        raise RuntimeError("Cannot parse WSD range from macro!F2")
    return int(match.group(1))


def rows_by_code(ws, code_col: int) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    for row in range(1, ws.max_row + 1):
        code = norm(ws.cell(row, code_col).value)
        if code:
            rows.setdefault(code, []).append(row)
    return rows


def first_rows_by_code(ws, code_col: int) -> dict[str, int]:
    return {code: row_list[0] for code, row_list in rows_by_code(ws, code_col).items()}


def update_wsd_formula(macro, raw_end_row: int) -> str:
    cell = macro["F2"]
    text = formula_text(cell.value)
    text = re.sub(r"E2:E\d+", f"E2:E{raw_end_row}", text)
    text = re.sub(r"rows=\d+", f"rows={raw_end_row - 1}", text)
    if isinstance(cell.value, ArrayFormula):
        cell.value.text = text
        cell.value.ref = "F2"
    else:
        cell.value = text
    return text


def replace_macro_row_refs(wb, macro_title: str, moved_rows: dict[int, int]) -> int:
    replaced = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                new_value = value
                for old_row, new_row in moved_rows.items():
                    patterns = [
                        (f"{macro_title}!$F${old_row}:$TN${old_row}", f"{macro_title}!$F${new_row}:$TN${new_row}"),
                        (f"{macro_title}!$F${old_row},", f"{macro_title}!$F${new_row},"),
                        (f"COLUMN({macro_title}!$F${old_row})", f"COLUMN({macro_title}!$F${new_row})"),
                        (f"COLUMN({macro_title}!$F${old_row}:$TN${old_row})", f"COLUMN({macro_title}!$F${new_row}:$TN${new_row})"),
                        (f"{macro_title}!$F${old_row}:$TR${old_row}", f"{macro_title}!$F${new_row}:$TR${new_row}"),
                        (f"COLUMN({macro_title}!$F${old_row}:$TR${old_row})", f"COLUMN({macro_title}!$F${new_row}:$TR${new_row})"),
                    ]
                    for src, dst in patterns:
                        new_value = new_value.replace(src, dst)
                if new_value != value:
                    cell.value = new_value
                    replaced += 1
    return replaced


def front_formulas(front_row: int, src_row: int, macro_title: str) -> dict[int, str]:
    return {
        5: f'=IFERROR(LOOKUP(9.99E+307,{macro_title}!$F${src_row}:$TN${src_row}),"")',
        6: f'=IFERROR(LOOKUP(2,1/({macro_title}!$F${src_row}:$TN${src_row}<>""),{macro_title}!$F$1:$TN$1),"")',
        7: f'=IFERROR(E{front_row}-H{front_row},"")',
        8: (
            f'=IFERROR(LOOKUP(9.99E+307,OFFSET({macro_title}!$F${src_row},0,0,1,'
            f'LOOKUP(2,1/({macro_title}!$F${src_row}:$TN${src_row}<>""),'
            f'COLUMN({macro_title}!$F${src_row}:$TN${src_row}))-COLUMN({macro_title}!$F${src_row}))),"")'
        ),
        9: (
            f'=IFERROR(LOOKUP(2,1/({macro_title}!$F${src_row}:$TN${src_row}<>"")/'
            f'({macro_title}!$F$1:$TN$1<F{front_row}),{macro_title}!$F$1:$TN$1),"")'
        ),
    }


def chart_by_row(ws, row: int):
    for chart in ws._charts:
        anchor = getattr(chart, "anchor", None)
        if hasattr(anchor, "_from") and anchor._from.row + 1 == row:
            return chart
    return None


def clone_template_chart(ws, front_row: int, src_row: int, macro_title: str) -> bool:
    template = chart_by_row(ws, front_row - 1)
    if template is None:
        for chart in reversed(ws._charts):
            anchor = getattr(chart, "anchor", None)
            if hasattr(anchor, "_from") and anchor._from.row + 1 < front_row:
                template = chart
                break
    if template is None:
        return False

    chart = copy.deepcopy(template)
    if hasattr(chart.anchor, "_from"):
        old_from = template.anchor._from
        old_to = template.anchor.to
        row_delta = front_row - (old_from.row + 1)
        chart.anchor._from.row = old_from.row + row_delta
        chart.anchor.to.row = old_to.row + row_delta

    cats_ref = f"'{macro_title}'!$F$1:$TN$1"
    vals_ref = f"'{macro_title}'!$F${src_row}:$TN${src_row}"
    for series in chart.series:
        try:
            series.val.numRef.f = vals_ref
        except Exception:
            pass
        try:
            series.cat.numRef.f = cats_ref
        except Exception:
            try:
                series.cat.strRef.f = cats_ref
            except Exception:
                pass
        try:
            series.graphicalProperties.line.solidFill = "C00000"
            series.graphicalProperties.line.width = 15120
            series.marker.symbol = "none"
        except Exception:
            pass
    try:
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
    except Exception:
        pass
    ws.add_chart(chart)
    return True


def normalize_charts(ws) -> None:
    for chart in ws._charts:
        try:
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
        except Exception:
            pass
        try:
            chart.graphical_properties.noFill = True
            chart.graphical_properties.line.noFill = True
        except Exception:
            pass
        try:
            chart.plot_area.graphicalProperties.noFill = True
            chart.plot_area.graphicalProperties.line.noFill = True
        except Exception:
            pass
        for series in chart.series:
            try:
                series.graphicalProperties.line.solidFill = "C00000"
                series.graphicalProperties.line.width = 15120
                series.marker.symbol = "none"
            except Exception:
                pass


def remove_children(parent: ET.Element, names: set[str]) -> None:
    for child in list(parent):
        if child.tag in names:
            parent.remove(child)


def set_shape_no_fill(sppr: ET.Element) -> None:
    fills = {
        f"{{{NS['a']}}}solidFill",
        f"{{{NS['a']}}}gradFill",
        f"{{{NS['a']}}}blipFill",
        f"{{{NS['a']}}}pattFill",
        f"{{{NS['a']}}}grpFill",
    }
    remove_children(sppr, fills)
    if sppr.find("a:noFill", NS) is None:
        sppr.insert(0, ET.Element(f"{{{NS['a']}}}noFill"))
    line = sppr.find("a:ln", NS)
    if line is None:
        line = ET.SubElement(sppr, f"{{{NS['a']}}}ln")
    remove_children(line, fills)
    if line.find("a:noFill", NS) is None:
        line.insert(0, ET.Element(f"{{{NS['a']}}}noFill"))


def set_series_line(line: ET.Element) -> None:
    line.attrib["w"] = "15120"
    fills = {
        f"{{{NS['a']}}}noFill",
        f"{{{NS['a']}}}solidFill",
        f"{{{NS['a']}}}gradFill",
        f"{{{NS['a']}}}blipFill",
        f"{{{NS['a']}}}pattFill",
        f"{{{NS['a']}}}grpFill",
    }
    remove_children(line, fills)
    solid = ET.SubElement(line, f"{{{NS['a']}}}solidFill")
    ET.SubElement(solid, f"{{{NS['a']}}}srgbClr", {"val": "C00000"})


def standardize_chart_xmls(path: Path) -> None:
    tmp = path.with_suffix(".tmp.xlsx")
    with ZipFile(path, "r") as zin, ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.startswith("xl/charts/chart") and info.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for parent in root.findall(".//c:valAx", NS) + root.findall(".//c:catAx", NS):
                    for grid in parent.findall("c:majorGridlines", NS):
                        parent.remove(grid)
                chart_space_sppr = root.find("c:spPr", NS)
                if chart_space_sppr is None:
                    chart_space_sppr = ET.SubElement(root, f"{{{NS['c']}}}spPr")
                set_shape_no_fill(chart_space_sppr)
                plot_area = root.find(".//c:plotArea", NS)
                if plot_area is not None:
                    plot_sppr = plot_area.find("c:spPr", NS)
                    if plot_sppr is None:
                        plot_sppr = ET.SubElement(plot_area, f"{{{NS['c']}}}spPr")
                    set_shape_no_fill(plot_sppr)
                for line in root.findall(".//c:ser/c:spPr/a:ln", NS):
                    set_series_line(line)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            zout.writestr(info, data)
    tmp.replace(path)


def add_front_row(ws, item: Indicator, src_row: int, macro_title: str) -> tuple[int, bool]:
    front_row = ws.max_row + 1
    copy_row_style(ws, front_row - 1, front_row, ws.max_column)
    for col in range(1, ws.max_column + 1):
        ws.cell(front_row, col).value = None
    ws.cell(front_row, 1).value = item.name
    ws.cell(front_row, 2).value = item.freq
    ws.cell(front_row, 3).value = item.code
    ws.cell(front_row, 4).value = item.unit
    for col, formula in front_formulas(front_row, src_row, macro_title).items():
        ws.cell(front_row, col).value = formula
    chart_added = clone_template_chart(ws, front_row, src_row, macro_title)
    return front_row, chart_added


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    report_path = Path(args.report)

    wb = openpyxl.load_workbook(input_path, data_only=False)
    macro = wb.worksheets[MACRO_SHEET_INDEX]
    macro_title = macro.title
    macro_max_col = max(macro.max_column, DATE_LAST_COL)
    old_raw_end = find_raw_end(macro)

    macro_rows = first_rows_by_code(macro, 5)
    items_to_add_macro = [item for item in INDICATORS if norm(item.code) not in macro_rows]

    moved_rows: dict[int, int] = {}
    added_macro: list[dict[str, Any]] = []
    skipped_macro: list[dict[str, Any]] = []

    if items_to_add_macro:
        add_count = len(items_to_add_macro)
        formula_start = old_raw_end + 1
        old_max_row = macro.max_row
        for src_row in range(old_max_row, formula_start - 1, -1):
            dst_row = src_row + add_count
            copy_row_all(macro, src_row, dst_row, macro_max_col)
            moved_rows[src_row] = dst_row
        for row in range(formula_start, formula_start + add_count):
            copy_row_style(macro, old_raw_end, row, macro_max_col)
            clear_row_values(macro, row, macro_max_col)
        for offset, item in enumerate(items_to_add_macro):
            row = formula_start + offset
            macro.cell(row, 4).value = item.name
            macro.cell(row, 5).value = item.code
            macro_rows[norm(item.code)] = row
            added_macro.append({"code": item.code, "name": item.name, "row": row})

    for item in INDICATORS:
        if norm(item.code) not in {entry["code"] for entry in added_macro} and norm(item.code) in macro_rows:
            skipped_macro.append({"code": item.code, "name": item.name, "row": macro_rows[norm(item.code)]})

    new_raw_end = old_raw_end + len(items_to_add_macro)
    f2_formula = update_wsd_formula(macro, new_raw_end)
    replaced_refs = replace_macro_row_refs(wb, macro_title, moved_rows) if moved_rows else 0
    macro_rows = first_rows_by_code(macro, 5)

    added_front: list[dict[str, Any]] = []
    skipped_front: list[dict[str, Any]] = []
    for item in INDICATORS:
        front_ws = wb.worksheets[item.sheet_index]
        front_rows = first_rows_by_code(front_ws, 3)
        if norm(item.code) in front_rows:
            skipped_front.append(
                {
                    "code": item.code,
                    "name": item.name,
                    "sheet": front_ws.title,
                    "row": front_rows[norm(item.code)],
                }
            )
            continue
        src_row = macro_rows.get(norm(item.code))
        if src_row is None:
            raise RuntimeError(f"No macro source row for {item.code}")
        front_row, chart_added = add_front_row(front_ws, item, src_row, macro_title)
        added_front.append(
            {
                "code": item.code,
                "name": item.name,
                "sheet": front_ws.title,
                "row": front_row,
                "macro_row": src_row,
                "chart_added": chart_added,
            }
        )

    normalize_charts(wb.worksheets[CHINA_SHEET_INDEX])
    normalize_charts(wb.worksheets[OVERSEAS_SHEET_INDEX])
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    standardize_chart_xmls(output_path)

    verify = openpyxl.load_workbook(output_path, data_only=False)
    vmacro = verify.worksheets[MACRO_SHEET_INDEX]
    errors: list[str] = []
    f2_after = formula_text(vmacro["F2"].value)
    if f"E2:E{new_raw_end}" not in f2_after or f"rows={new_raw_end - 1}" not in f2_after:
        errors.append("macro!F2 WSD range mismatch")
    vmacro_rows = first_rows_by_code(vmacro, 5)
    for item in INDICATORS:
        if norm(item.code) not in vmacro_rows:
            errors.append(f"missing macro code {item.code}")
        front_ws = verify.worksheets[item.sheet_index]
        vfront_rows = first_rows_by_code(front_ws, 3)
        if norm(item.code) not in vfront_rows:
            errors.append(f"missing front code {item.code}")
        else:
            front_row = vfront_rows[norm(item.code)]
            if chart_by_row(front_ws, front_row) is None:
                errors.append(f"missing chart {item.code} {front_ws.title}!{front_row}")
    report = {
        "input": str(input_path),
        "output": str(output_path),
        "macro_sheet": vmacro.title,
        "old_raw_end": old_raw_end,
        "new_raw_end": new_raw_end,
        "f2_formula": f2_after,
        "moved_rows": moved_rows,
        "formula_refs_replaced": replaced_refs,
        "added_macro": added_macro,
        "skipped_existing_macro": skipped_macro,
        "added_front": added_front,
        "skipped_existing_front": skipped_front,
        "sheet_rows": {
            verify.worksheets[CHINA_SHEET_INDEX].title: verify.worksheets[CHINA_SHEET_INDEX].max_row,
            verify.worksheets[OVERSEAS_SHEET_INDEX].title: verify.worksheets[OVERSEAS_SHEET_INDEX].max_row,
        },
        "chart_counts": {
            verify.worksheets[CHINA_SHEET_INDEX].title: len(verify.worksheets[CHINA_SHEET_INDEX]._charts),
            verify.worksheets[OVERSEAS_SHEET_INDEX].title: len(verify.worksheets[OVERSEAS_SHEET_INDEX]._charts),
        },
        "errors": errors,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
