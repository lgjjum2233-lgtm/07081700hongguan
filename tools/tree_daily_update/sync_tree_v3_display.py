from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from sync_tree_new_indicators_from_daily import (
    find_cell,
    read_xml,
    set_date,
    set_number,
    workbook_sheet_path,
    write_xml,
)
from sync_tree_v30_full import (
    build_sources,
    is_stock_index_row,
    scale_for_cell,
    to_number,
)


WORKSPACE = Path(__file__).resolve().parents[2]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

import sync_tree_20260604 as base  # noqa: E402


TREE_SHEET = "重点策略跟踪情况(V3)"
HEADER_ROW = 5
NAME_COL = 9
FREQ_COL = 12
CURRENT_COL = 15
DATE_COL = 16
CHANGE_COL = 17


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_blank(value: Any) -> bool:
    text = as_text(value)
    return text == "" or text in {"-", "—", "None"}


def update_display_sheet(input_tree: Path, daily_path: Path, output_tree: Path, as_of: date) -> dict[str, Any]:
    by_code, by_name, best_trend = build_sources(daily_path, as_of)
    wb_tree = openpyxl.load_workbook(input_tree, data_only=False)
    ws = wb_tree[TREE_SHEET]

    with ZipFile(input_tree, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)

        updated_rows = []
        unmatched = []

        for row in range(HEADER_ROW + 1, ws.max_row + 1):
            row_name = ws.cell(row, NAME_COL).value
            if is_blank(row_name):
                continue

            source = base.resolve_source(as_text(row_name), "", by_code, by_name)
            if source is None:
                unmatched.append({"row": row, "name": as_text(row_name)})
                continue
            source = base.with_best_trend(source, best_trend)

            if is_blank(ws.cell(row, FREQ_COL).value) and source.freq:
                # The display sheet normally already has a frequency, but fill blanks if the source is clear.
                from sync_tree_new_indicators_from_daily import set_inline_string

                set_inline_string(find_cell(sheet_root, row, FREQ_COL), source.freq)

            current_format = ws.cell(row, CURRENT_COL).number_format
            change_format = ws.cell(row, CHANGE_COL).number_format
            set_number(find_cell(sheet_root, row, CURRENT_COL), scale_for_cell(source.current, current_format))
            set_date(find_cell(sheet_root, row, DATE_COL), source.data_date)

            change_value = source.change
            already_percent_decimal = False
            if is_stock_index_row(row_name, source.code):
                current = to_number(source.current)
                previous = to_number(source.previous)
                if current is not None and previous not in (None, 0):
                    change_value = (current - previous) / previous
                    already_percent_decimal = True
            set_number(
                find_cell(sheet_root, row, CHANGE_COL),
                scale_for_cell(change_value, change_format, already_percent_decimal),
            )

            updated_rows.append(
                {
                    "row": row,
                    "name": as_text(row_name),
                    "code": source.code,
                    "source": f"{source.sheet}!{source.row}" if source.sheet else source.code,
                    "date": source.data_date,
                    "current": source.current,
                    "change": change_value,
                }
            )

        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, write_xml(sheet_root))
                else:
                    zout.writestr(name, zin.read(name))

    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output": str(output_tree),
        "as_of": as_of.isoformat(),
        "updated_row_count": len(updated_rows),
        "updated_rows_sample": updated_rows[:30],
        "unmatched_count": len(unmatched),
        "unmatched_sample": unmatched[:30],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Synchronize TREE V3 display-sheet current data from the refreshed daily workbook.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default=date.today().isoformat())
    args = parser.parse_args()
    as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
    result = update_display_sheet(Path(args.input), Path(args.daily), Path(args.output), as_of)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
