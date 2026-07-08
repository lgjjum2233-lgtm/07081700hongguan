from __future__ import annotations

import argparse
import json
from pathlib import Path, PurePosixPath
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl


TREE_SHEET = "重点策略跟踪情况(V3.0)"
ROWS = [
    19,
    27,
    28,
    29,
    30,
    32,
    33,
    34,
    36,
    37,
    38,
    39,
    40,
    41,
    45,
    46,
    47,
    48,
    49,
    51,
    52,
    53,
    57,
    58,
    59,
    67,
    68,
    78,
    79,
    80,
    81,
    82,
    137,
    138,
]

EXPECTED_BLANK_ROWS = {46, 78, 137}

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


def read_xml(z, name):
    return ET.fromstring(z.read(name))


def rel_target(base, target):
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts = []
    for part in (base_dir + "/" + target).split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z, sheet_name):
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib["name"] == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def sheet_drawing_path(z, sheet_path):
    sheet = read_xml(z, sheet_path)
    drawing = sheet.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib[f"{{{NS['officeRel']}}}id"]
    rels_name = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
    rels = read_xml(z, rels_name)
    for rel in rels:
        if rel.attrib["Id"] == rid:
            return rel_target(sheet_path, rel.attrib["Target"])
    return None


def drawing_rels_path(drawing_path):
    p = PurePosixPath(drawing_path)
    return str(p.parent / "_rels" / (p.name + ".rels"))


def chart_rows(z, drawing_path):
    rels = read_xml(z, drawing_rels_path(drawing_path))
    rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
    drawing = read_xml(z, drawing_path)
    out = {}
    for anchor in list(drawing):
        frm = anchor.find("xdr:from", NS)
        chart = anchor.find(".//a:graphicData/c:chart", NS)
        if frm is None or chart is None:
            continue
        rid = chart.attrib[f"{{{NS['officeRel']}}}id"]
        out[int(frm.findtext("xdr:row", namespaces=NS)) + 1] = rel_map[rid]
    return out


def chart_count(z, chart_path):
    root = read_xml(z, chart_path)
    pts = root.findall(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:pt", NS)
    return len(pts)


def font_rgb(cell):
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return f"{color.type}:{color.value}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    args = parser.parse_args()
    path = Path(args.workbook)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[TREE_SHEET]
    row_checks = []
    unexpected_blank = []
    for row in ROWS:
        item = {
            "row": row,
            "name": ws.cell(row, 4).value,
            "freq": ws.cell(row, 9).value,
            "code": ws.cell(row, 10).value,
            "current": ws.cell(row, 11).value,
            "date": ws.cell(row, 12).value,
            "change": ws.cell(row, 13).value,
            "fmt_current": ws.cell(row, 11).number_format,
            "fmt_change": ws.cell(row, 13).number_format,
            "change_font": font_rgb(ws.cell(row, 13)),
        }
        if row not in EXPECTED_BLANK_ROWS and (item["freq"] is None or item["code"] is None or item["current"] is None or item["date"] is None):
            unexpected_blank.append(item)
        row_checks.append(item)

    with ZipFile(path) as z:
        sheet_path = workbook_sheet_path(z, TREE_SHEET)
        drawing_path = sheet_drawing_path(z, sheet_path)
        charts = chart_rows(z, drawing_path) if drawing_path else {}
        missing_charts = [row for row in ROWS if row not in EXPECTED_BLANK_ROWS and row not in charts]
        empty_charts = [row for row in ROWS if row in charts and chart_count(z, charts[row]) == 0]

    result = {
        "pass": not unexpected_blank and not missing_charts and not empty_charts,
        "unexpected_blank": unexpected_blank,
        "missing_charts": missing_charts,
        "empty_charts": empty_charts,
        "chart_count": len(charts),
        "rows": row_checks,
    }
    print(json.dumps(result, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
