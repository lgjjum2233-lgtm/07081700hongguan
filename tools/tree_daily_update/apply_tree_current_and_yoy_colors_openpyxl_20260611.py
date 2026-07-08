from __future__ import annotations

import argparse
import copy
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import add_tree_yoy_column_20260611 as yoy_base
import apply_tree_current_value_history_colors_20260611 as current_color
import apply_tree_yoy_history_colors_20260611 as yoy_color


TREE_SHEET = "重点策略跟踪情况(V3.0)"
CURRENT_COL = 11
YOY_COL = 14


def to_number(value: Any) -> float | None:
    return yoy_base.to_number(value)


def apply_style(cell, fill_rgb: str | None, font_rgb: str) -> None:
    font = copy.copy(cell.font)
    font.color = font_rgb
    cell.font = font
    if fill_rgb:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_rgb)
    else:
        cell.fill = PatternFill(fill_type=None)


def current_decision(row: dict[str, Any], macro_series: dict[str, yoy_base.MacroSeries]) -> dict[str, Any]:
    out = {
        "fill": None,
        "font": "FF000000",
        "history_count": 0,
        "history_start": "",
        "history_end": "",
        "percentile": None,
        "position": "",
        "direction": "neutral",
        "direction_reason": "",
        "source_date": "",
        "source_current": "",
        "intensity": None,
        "note": "",
    }
    if not yoy_base.is_macro_tree_row(row["big"], row["dim"]):
        out["note"] = "非宏观计算区，不填色"
        return out
    series = macro_series.get(row["code"])
    current_dt = current_color.parse_date(row["current_date"])
    if series is None or current_dt is None:
        out["note"] = "未匹配到底层宏观数据或日期为空"
        return out
    history, current_point = current_color.current_history(series, current_dt)
    out["history_count"] = len(history)
    if history:
        out["history_start"] = history[0][0].isoformat()
        out["history_end"] = history[-1][0].isoformat()
    if current_point is None or len(history) < current_color.MIN_HISTORY_VALUES:
        out["note"] = "近一年当前值历史样本不足"
        return out
    out["source_date"] = current_point[0].isoformat()
    out["source_current"] = current_point[1]
    values = sorted(value for _, value in history)
    pct = current_color.percentile_rank(values, current_point[1])
    rule = current_color.direction_for_indicator(row, series.name)
    fill, font, intensity, position = current_color.color_for_position(current_point[1], pct, rule.direction)
    out.update(
        {
            "fill": fill,
            "font": font,
            "percentile": pct,
            "position": position,
            "direction": rule.direction,
            "direction_reason": rule.reason,
            "intensity": intensity,
        }
    )
    if not fill:
        out["note"] = rule.reason if rule.direction == "neutral" else "当前值处于近一年常态区间，未填色"
    return out


def yoy_decision(row: dict[str, Any], macro_series: dict[str, yoy_base.MacroSeries]) -> dict[str, Any]:
    out = {
        "fill": None,
        "font": "FF000000",
        "history_count": 0,
        "history_start": "",
        "history_end": "",
        "percentile": None,
        "intensity": None,
        "method": "",
        "note": "",
    }
    if row["yoy"] == "-":
        out["note"] = "同比变化率为-，不填色"
        return out
    if not yoy_base.is_macro_tree_row(row["big"], row["dim"]):
        out["note"] = "非宏观计算区，不填色"
        return out
    series = macro_series.get(row["code"])
    current_dt = yoy_base.parse_date(row["current_date"])
    if series is None or current_dt is None:
        out["note"] = "未匹配到底层宏观数据或日期为空"
        return out
    history, method = yoy_color.build_yoy_history(series, current_dt, row["yoy"], row["caliber"])
    fill, font, detail = yoy_color.color_for_history(row["yoy"], history)
    out.update(
        {
            "fill": fill,
            "font": font,
            "history_count": detail.get("history_count", 0),
            "history_start": detail.get("history_start", ""),
            "history_end": detail.get("history_end", ""),
            "percentile": detail.get("percentile"),
            "intensity": detail.get("intensity"),
            "method": method,
        }
    )
    if not fill:
        if detail.get("history_count", 0) < yoy_color.MIN_HISTORY_VALUES and to_number(row["yoy"]) not in (None, 0.0):
            out["note"] = "近一年同比变化率历史样本不足"
        else:
            out["note"] = "历史分位未达到填色阈值或数值接近0"
    return out


def write_audit(path: Path, audit_rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "当前值与同比颜色复核"
    headers = [
        "TREE行号",
        "大分类",
        "细分维度",
        "子维度",
        "指标名称",
        "指标代码",
        "数据日期",
        "当前值",
        "当前值分位",
        "当前值位置",
        "当前值方向规则",
        "当前值填充色",
        "当前值备注",
        "同比变化率",
        "同比历史样本数",
        "同比历史分位",
        "同比填充色",
        "同比备注",
    ]
    ws.append(headers)
    for item in audit_rows:
        ws.append([item.get(header, "") for header in headers])
    widths = [10, 18, 18, 18, 34, 22, 14, 14, 12, 12, 18, 14, 42, 14, 14, 12, 14, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def apply_colors(input_tree: Path, daily_path: Path, output_tree: Path, audit_output: Path) -> dict[str, Any]:
    macro_series = yoy_base.read_macro_series(daily_path)
    wb = openpyxl.load_workbook(input_tree, data_only=False, read_only=False)
    ws = wb[TREE_SHEET]
    rows = current_color.iter_tree_rows(ws)
    for row in rows:
        row["yoy"] = ws.cell(row["row"], YOY_COL).value

    current_filled = 0
    yoy_filled = 0
    nonmacro_current_filled = 0
    nonmacro_yoy_filled = 0
    audit_rows: list[dict[str, Any]] = []

    for row in rows:
        c_decision = current_decision(row, macro_series)
        y_decision = yoy_decision(row, macro_series)
        apply_style(ws.cell(row["row"], CURRENT_COL), c_decision["fill"], c_decision["font"])
        apply_style(ws.cell(row["row"], YOY_COL), y_decision["fill"], y_decision["font"])

        is_macro = yoy_base.is_macro_tree_row(row["big"], row["dim"])
        if c_decision["fill"]:
            if is_macro:
                current_filled += 1
            else:
                nonmacro_current_filled += 1
        if y_decision["fill"]:
            if is_macro:
                yoy_filled += 1
            else:
                nonmacro_yoy_filled += 1

        audit_rows.append(
            {
                "TREE行号": row["row"],
                "大分类": row["big"],
                "细分维度": row["dim"],
                "子维度": row["sub"],
                "指标名称": row["name"],
                "指标代码": row["code"],
                "数据日期": row["current_date"],
                "当前值": row["current"],
                "当前值分位": c_decision["percentile"],
                "当前值位置": c_decision["position"],
                "当前值方向规则": f'{c_decision["direction"]}: {c_decision["direction_reason"]}',
                "当前值填充色": c_decision["fill"] or "",
                "当前值备注": c_decision["note"],
                "同比变化率": row["yoy"],
                "同比历史样本数": y_decision["history_count"],
                "同比历史分位": y_decision["percentile"],
                "同比填充色": y_decision["fill"] or "",
                "同比备注": y_decision["note"],
            }
        )

    wb.save(output_tree)
    wb.close()
    write_audit(audit_output, audit_rows)
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        "current_filled": current_filled,
        "yoy_filled": yoy_filled,
        "nonmacro_current_filled": nonmacro_current_filled,
        "nonmacro_yoy_filled": nonmacro_yoy_filled,
        "rows": len(rows),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()
    result = apply_colors(
        Path(args.tree),
        Path(args.daily),
        Path(args.output_tree),
        Path(args.audit_output),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
