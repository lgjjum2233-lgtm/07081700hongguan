from __future__ import annotations

import argparse
import copy
import json
import math
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))
WORKSPACE = TOOL_DIR.parents[1]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

from sync_tree_new_indicators_from_daily import (  # noqa: E402
    NS,
    find_cell,
    find_existing_cell,
    read_xml,
    set_inline_string,
    set_number,
    sheet_drawing_path,
    workbook_sheet_path,
    write_xml,
)


TREE_SHEET = "重点策略跟踪情况(V3.0)"
MACRO_SHEET = "宏观数据"
HEADER_ROW = 5
CODE_COL = 10
CURRENT_COL = 11
DATE_COL = 12
CHANGE_COL = 13
INSERT_COL = 14  # N, immediately after marginal change.
CHART_COL = 14
CALIBER_COL = 15  # O before insertion.


@dataclass
class MacroSeries:
    name: str
    code: str
    unit: str
    points: list[tuple[date, float]]


@dataclass
class YoyResult:
    value: Any
    style_kind: str
    method: str
    basis: str
    source_current_date: str = ""
    source_current_value: Any = None
    prior_date: str = ""
    prior_value: Any = None
    macro_name: str = ""


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_code(value: Any) -> str:
    return as_text(value).upper()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = as_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def prior_year_date(current: date) -> date:
    try:
        return current.replace(year=current.year - 1)
    except ValueError:
        return current.replace(year=current.year - 1, day=28)


def closest_prior_period(points: list[tuple[date, float]], current_date: date) -> tuple[date, float] | None:
    target = prior_year_date(current_date)
    if not points:
        return None
    # Prefer exact same calendar date; otherwise use the nearest observation around the same period.
    by_date = {d: v for d, v in points}
    if target in by_date:
        return target, by_date[target]
    candidates = [(abs((d - target).days), d, v) for d, v in points if abs((d - target).days) <= 45]
    if not candidates:
        return None
    _, d, v = min(candidates, key=lambda item: (item[0], item[1]))
    return d, v


def closest_current_period(points: list[tuple[date, float]], current_date: date) -> tuple[date, float] | None:
    if not points:
        return None
    by_date = {d: v for d, v in points}
    if current_date in by_date:
        return current_date, by_date[current_date]
    candidates = [(abs((d - current_date).days), d, v) for d, v in points if abs((d - current_date).days) <= 45]
    if not candidates:
        return None
    _, d, v = min(candidates, key=lambda item: (item[0], item[1]))
    return d, v


def read_macro_series(daily_path: Path) -> dict[str, MacroSeries]:
    wb = openpyxl.load_workbook(daily_path, data_only=True, read_only=False)
    ws = wb[MACRO_SHEET]
    dates: dict[int, date] = {}
    for col in range(6, ws.max_column + 1):
        parsed = parse_date(ws.cell(1, col).value)
        if parsed:
            dates[col] = parsed
    out: dict[str, MacroSeries] = {}
    for row in range(2, ws.max_row + 1):
        code = norm_code(ws.cell(row, 5).value)
        if not code or code in out:
            continue
        points = []
        for col, d in dates.items():
            value = to_number(ws.cell(row, col).value)
            if value is not None:
                points.append((d, value))
        if points:
            out[code] = MacroSeries(
                name=as_text(ws.cell(row, 4).value),
                code=code,
                unit="",
                points=points,
            )
    wb.close()
    return out


def is_yoy_caliber(caliber: str, macro_name: str) -> bool:
    text = f"{caliber} {macro_name}"
    return any(key in text for key in ["同比", "累计同比", "当月同比", "同比多增", "同比少增"])


def is_level_diff_caliber(caliber: str) -> bool:
    return any(
        key in caliber
        for key in [
            "利率",
            "收益率",
            "利差",
            "比例",
            "扩散指数",
            "估值倍数",
            "盈利比例",
            "净投放",
            "净流量",
            "差额",
            "净额",
            "衍生",
            "合成流量",
        ]
    )


def is_macro_tree_row(big_category: str, dimension: str) -> bool:
    big = as_text(big_category)
    dim = as_text(dimension)
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    # Process macro rows only; China industry rows are excluded by the user's rule.
    if big == "中国基本面" and dim == "产业":
        return False
    return True


def calculate_yoy(
    code: str,
    current: Any,
    current_date: Any,
    caliber: str,
    macro_series: dict[str, MacroSeries],
) -> YoyResult:
    code = norm_code(code)
    current_number = to_number(current)
    current_dt = parse_date(current_date)
    series = macro_series.get(code)

    if not code or code in {"—", "-"}:
        return YoyResult("-", "text", "未计算", "无有效Wind代码")
    if series is None:
        return YoyResult("-", "text", "未计算", "未匹配到底层宏观数据，资本市场或定性项用-")
    if current_number is None or current_dt is None:
        return YoyResult("-", "text", "未计算", "当前值或日期为空", macro_name=series.name)

    if is_yoy_caliber(caliber, series.name):
        return YoyResult(
            current_number,
            "same_as_current",
            "原值",
            "底层宏观数据标准名称/数据口径已是同比类，直接取当前值",
            macro_name=series.name,
        )

    current_point = closest_current_period(series.points, current_dt)
    if current_point is None:
        return YoyResult("-", "text", "未计算", "底层宏观数据找不到当前日期值", macro_name=series.name)
    current_point_dt, current_raw = current_point
    prior = closest_prior_period(series.points, current_dt)
    if prior is None:
        return YoyResult("-", "text", "未计算", "底层宏观数据找不到去年同期值", source_current_date=current_point_dt.isoformat(), source_current_value=current_raw, macro_name=series.name)
    prior_dt, prior_value = prior

    if prior_value == 0:
        return YoyResult("-", "text", "未计算", "去年同期值为0，无法计算同比变化率", prior_date=prior_dt.isoformat(), prior_value=prior_value, macro_name=series.name)
    value = (current_raw - prior_value) / abs(prior_value)
    return YoyResult(
        value,
        "percent",
        "同比变化率",
        "统一口径：(今年同期值-去年同期值)/ABS(去年同期值)，避免去年同期为负时方向被分母符号扰乱",
        source_current_date=current_point_dt.isoformat(),
        source_current_value=current_raw,
        prior_date=prior_dt.isoformat(),
        prior_value=prior_value,
        macro_name=series.name,
    )


def split_cell_ref(ref: str) -> tuple[str, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(ref)
    return match.group(1), int(match.group(2))


def shift_formula_refs(formula: str, min_col: int = INSERT_COL) -> str:
    def repl(match: re.Match[str]) -> str:
        prefix, col_letters, row = match.groups()
        col = column_index_from_string(col_letters)
        if col >= min_col:
            col_letters = get_column_letter(col + 1)
        return f"{prefix}{col_letters}{row}"

    return re.sub(r"(\$?)([A-Z]{1,3})(\$?\d+)", repl, formula)


def shift_cell_right(cell: ET.Element, min_col: int = INSERT_COL) -> None:
    ref = cell.attrib.get("r")
    if not ref:
        return
    col_letters, row = split_cell_ref(ref)
    col = column_index_from_string(col_letters)
    if col >= min_col:
        cell.attrib["r"] = f"{get_column_letter(col + 1)}{row}"
    for formula in cell.findall("main:f", NS):
        if formula.text:
            formula.text = shift_formula_refs(formula.text, min_col)


def shift_sheet_cells(sheet_root: ET.Element, min_col: int = INSERT_COL) -> None:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        return
    for row in sheet_data.findall("main:row", NS):
        for cell in row.findall("main:c", NS):
            shift_cell_right(cell, min_col)
        row[:] = sorted(
            list(row),
            key=lambda c: column_index_from_string(re.match(r"([A-Z]+)", c.attrib.get("r", "A1")).group(1)),
        )


def shift_dimension(sheet_root: ET.Element) -> None:
    dim = sheet_root.find("main:dimension", NS)
    if dim is not None:
        dim.attrib["ref"] = "A1:P238"


def shift_cols(sheet_root: ET.Element, min_col: int = INSERT_COL) -> None:
    cols = sheet_root.find("main:cols", NS)
    if cols is None:
        return
    for col in cols.findall("main:col", NS):
        min_v = int(float(col.attrib.get("min", "0")))
        max_v = int(float(col.attrib.get("max", "0")))
        if min_v >= min_col:
            col.attrib["min"] = str(min_v + 1)
            col.attrib["max"] = str(max_v + 1)
        elif max_v >= min_col:
            col.attrib["max"] = str(max_v + 1)
    # Set the new YoY column width.
    cols.append(ET.Element(f"{{{NS['main']}}}col", {"min": str(INSERT_COL), "max": str(INSERT_COL), "width": "12", "customWidth": "1"}))


def shift_merged_ranges(sheet_root: ET.Element, min_col: int = INSERT_COL) -> None:
    merge_cells = sheet_root.find("main:mergeCells", NS)
    if merge_cells is None:
        return
    for merge in merge_cells.findall("main:mergeCell", NS):
        ref = merge.attrib.get("ref", "")
        if ":" not in ref:
            continue
        start, end = ref.split(":", 1)
        s_col, s_row = split_cell_ref(start)
        e_col, e_row = split_cell_ref(end)
        s_idx = column_index_from_string(s_col)
        e_idx = column_index_from_string(e_col)
        if s_idx >= min_col:
            s_idx += 1
        if e_idx >= min_col:
            e_idx += 1
        merge.attrib["ref"] = f"{get_column_letter(s_idx)}{s_row}:{get_column_letter(e_idx)}{e_row}"


def shift_drawing_anchors(drawing_root: ET.Element, min_zero_col: int = INSERT_COL - 1) -> None:
    for anchor in list(drawing_root):
        for tag in ("xdr:from", "xdr:to"):
            node = anchor.find(tag, NS)
            if node is None:
                continue
            col = node.find("xdr:col", NS)
            if col is not None and col.text is not None:
                value = int(col.text)
                if value >= min_zero_col:
                    col.text = str(value + 1)


def cell_style(sheet_root: ET.Element, row: int, col: int, fallback: int | None = None) -> int | None:
    cell = find_existing_cell(sheet_root, f"{get_column_letter(col)}{row}")
    if cell is not None and "s" in cell.attrib:
        return int(cell.attrib["s"])
    return fallback


def write_yoy_column(tree_path: Path, output_path: Path, records: list[dict[str, Any]]) -> None:
    record_by_row = {int(item["row"]): item for item in records}
    replacements: dict[str, bytes] = {}

    with ZipFile(tree_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)
        drawing_path = sheet_drawing_path(zin, sheet_path)
        drawing_root = read_xml(zin, drawing_path) if drawing_path else None

        shift_sheet_cells(sheet_root)
        shift_cols(sheet_root)
        shift_merged_ranges(sheet_root)
        shift_dimension(sheet_root)
        if drawing_root is not None:
            shift_drawing_anchors(drawing_root)

        header_style = cell_style(sheet_root, HEADER_ROW, CHANGE_COL, cell_style(sheet_root, HEADER_ROW, INSERT_COL + 1))
        text_style = cell_style(sheet_root, HEADER_ROW + 1, INSERT_COL + 2, cell_style(sheet_root, HEADER_ROW + 1, CHANGE_COL))
        percent_style = cell_style(sheet_root, 45, CURRENT_COL, cell_style(sheet_root, 47, CURRENT_COL))
        general_style = cell_style(sheet_root, 37, CURRENT_COL, cell_style(sheet_root, HEADER_ROW + 1, CURRENT_COL))

        set_inline_string(find_cell(sheet_root, HEADER_ROW, INSERT_COL), "同比变化率", header_style)
        for row, item in record_by_row.items():
            cell = find_cell(sheet_root, row, INSERT_COL)
            if item["value"] == "-":
                set_inline_string(cell, "-", text_style)
            elif item["style_kind"] == "same_as_current":
                set_number(cell, item["value"], cell_style(sheet_root, row, CURRENT_COL, general_style))
            elif item["style_kind"] == "percent":
                set_number(cell, item["value"], percent_style)
            else:
                set_number(cell, item["value"], general_style)

        replacements[sheet_path] = write_xml(sheet_root)
        if drawing_path and drawing_root is not None:
            replacements[drawing_path] = write_xml(drawing_root)

        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, replacements[name] if name in replacements else zin.read(name))


def build_records(tree_path: Path, daily_path: Path) -> list[dict[str, Any]]:
    macro_series = read_macro_series(daily_path)
    wb = openpyxl.load_workbook(tree_path, data_only=True, read_only=False)
    ws = wb[TREE_SHEET]
    records = []
    current_big = ""
    current_dim = ""
    current_sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        big_value = as_text(ws.cell(row, 1).value)
        dim_value = as_text(ws.cell(row, 2).value)
        sub_value = as_text(ws.cell(row, 3).value)
        if big_value:
            current_big = big_value
        if dim_value:
            current_dim = dim_value
        if sub_value:
            current_sub = sub_value
        name = as_text(ws.cell(row, 4).value)
        code = norm_code(ws.cell(row, CODE_COL).value)
        if not name and not code:
            continue
        caliber = as_text(ws.cell(row, CALIBER_COL).value)
        if is_macro_tree_row(current_big, current_dim):
            yoy = calculate_yoy(code, ws.cell(row, CURRENT_COL).value, ws.cell(row, DATE_COL).value, caliber, macro_series)
        else:
            yoy = YoyResult("-", "text", "未计算", "非宏观数据区，按要求用-")
        records.append(
            {
                "row": row,
                "big_category": current_big,
                "dimension": current_dim,
                "sub_dimension": current_sub,
                "tree_name": name,
                "code": code,
                "caliber": caliber,
                "current": ws.cell(row, CURRENT_COL).value,
                "current_date": ws.cell(row, DATE_COL).value,
                "value": yoy.value,
                "style_kind": yoy.style_kind,
                "method": yoy.method,
                "basis": yoy.basis,
                "source_current_date": yoy.source_current_date,
                "source_current_value": yoy.source_current_value,
                "prior_date": yoy.prior_date,
                "prior_value": yoy.prior_value,
                "macro_name": yoy.macro_name,
            }
        )
    wb.close()
    return records


def write_audit(audit_path: Path, records: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "同比变化率计算复核"
    headers = [
        "TREE行号",
        "大分类",
        "细分维度",
        "子维度",
        "TREE名称",
        "指标代码",
        "数据口径",
        "当前数据",
        "当前日期",
        "同比变化率",
        "计算方法",
        "日报当前日期",
        "日报当前值",
        "去年同期日期",
        "去年同期值",
        "日报宏观标准名称",
        "判断依据",
    ]
    ws.append(headers)
    for item in records:
        ws.append(
            [
                item["row"],
                item["big_category"],
                item["dimension"],
                item["sub_dimension"],
                item["tree_name"],
                item["code"],
                item["caliber"],
                item["current"],
                item["current_date"],
                item["value"],
                item["method"],
                item["source_current_date"],
                item["source_current_value"],
                item["prior_date"],
                item["prior_value"],
                item["macro_name"],
                item["basis"],
            ]
        )
    for idx, width in enumerate([10, 18, 18, 18, 34, 24, 16, 14, 16, 14, 16, 14, 16, 14, 42, 48], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(audit_path)


def verify(output_path: Path, records: list[dict[str, Any]]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(output_path, data_only=True, read_only=False)
    ws = wb[TREE_SHEET]
    chart_rows = [
        (getattr(chart.anchor, "_from", None).row + 1, getattr(chart.anchor, "_from", None).col + 1)
        for chart in ws._charts
        if hasattr(chart.anchor, "_from")
    ]
    sample_rows = {item["row"]: item for item in records if item["value"] != "-"}
    mismatches = []
    for row, item in list(sample_rows.items())[:80]:
        actual = ws.cell(row, INSERT_COL).value
        expected = item["value"]
        if isinstance(expected, float) and isinstance(actual, (int, float)):
            if abs(float(actual) - expected) > 1e-9:
                mismatches.append({"row": row, "actual": actual, "expected": expected})
        elif actual != expected:
            mismatches.append({"row": row, "actual": actual, "expected": expected})
    result = {
        "header": ws.cell(HEADER_ROW, INSERT_COL).value,
        "max_col": ws.max_column,
        "chart_count": len(ws._charts),
        "last_chart_cols": chart_rows[-12:],
        "mismatch_sample": mismatches[:10],
        "dash_count": sum(1 for item in records if item["value"] == "-"),
        "numeric_count": sum(1 for item in records if item["value"] != "-"),
    }
    wb.close()
    return result


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()

    tree_path = Path(args.tree)
    daily_path = Path(args.daily)
    output_tree = Path(args.output_tree)
    audit_output = Path(args.audit_output)

    records = build_records(tree_path, daily_path)
    write_yoy_column(tree_path, output_tree, records)
    write_audit(audit_output, records)
    check = verify(output_tree, records)
    methods: dict[str, int] = {}
    for item in records:
        methods[item["method"]] = methods.get(item["method"], 0) + 1
    result = {
        "tree": str(tree_path),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        **check,
        "methods": methods,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))
    if check["header"] != "同比变化率" or check["mismatch_sample"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
