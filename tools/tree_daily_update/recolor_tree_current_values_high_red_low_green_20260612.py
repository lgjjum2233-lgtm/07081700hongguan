from __future__ import annotations

import copy
import math
import re
import sys
from bisect import bisect_left, bisect_right
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import add_tree_yoy_column_20260611 as macro_base  # noqa: E402
import apply_macro_change_history_colors as color_base  # noqa: E402


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "（6月12日V15）TREE宏观分析_当前数据红高绿低.xlsx"
DAILY = ROOT / "20260611月报宏观数据_扩展2024历史_Wind刷新.xlsx"
OUTPUT = ROOT / "（6月12日V16）TREE宏观分析_当前数据红高绿低_平值修正.xlsx"
AUDIT = ROOT / "20260612_TREE当前数据红高绿低复核_平值修正.xlsx"

HEADER_ROW = 5
LOOKBACK_DAYS = 365
LOW_THRESHOLD = 0.30
HIGH_THRESHOLD = 0.70
MIN_HISTORY_VALUES = 4


SHEET_LAYOUTS = {
    "v30": {
        "big_col": 1,
        "dim_col": 2,
        "sub_col": 3,
        "name_col": 4,
        "code_col": 10,
        "caliber_col": 11,
        "current_col": 12,
        "date_col": 13,
    },
    "v3": {
        "big_col": 6,
        "dim_col": 7,
        "sub_col": 8,
        "name_col": 9,
        "code_col": None,
        "caliber_col": 10,
        "current_col": 15,
        "date_col": 16,
    },
    "v25": {
        "big_col": 1,
        "dim_col": 2,
        "sub_col": 3,
        "name_col": 4,
        "code_col": 10,
        "caliber_col": 7,
        "current_col": 11,
        "date_col": 12,
    },
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def norm_code(value: Any) -> str:
    return macro_base.norm_code(value)


def parse_date(value: Any) -> date | None:
    parsed = macro_base.parse_date(value)
    if parsed:
        return parsed
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None
    return None


def to_number(value: Any) -> float | None:
    return macro_base.to_number(value)


def normalize_name(value: Any) -> str:
    s = text(value)
    s = re.sub(r"[①②③④⑤⑥⑦⑧⑨⑩]", "", s)
    s = re.sub(r"[：:]", "", s)
    s = re.sub(r"\（.*?\）|\(.*?\)", "", s)
    s = re.sub(r"\[.*?\]", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def is_macro(big: str, dim: str) -> bool:
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    if "产业" in dim:
        return False
    return True


def detect_layout(ws) -> str | None:
    headers = {text(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    if headers.get("监测指标") == 4 and headers.get("当前数据") == 12:
        return "v30"
    if headers.get("监测指标") == 9 and headers.get("当前数据") == 15:
        return "v3"
    if headers.get("监测指标") == 4 and headers.get("当前数据") == 11:
        return "v25"
    return None


def iter_rows(ws, layout_key: str):
    cfg = SHEET_LAYOUTS[layout_key]
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row, cfg["big_col"]).value):
            big = text(ws.cell(row, cfg["big_col"]).value)
            dim = ""
            sub = ""
        if text(ws.cell(row, cfg["dim_col"]).value):
            dim = text(ws.cell(row, cfg["dim_col"]).value)
            sub = ""
        if text(ws.cell(row, cfg["sub_col"]).value):
            sub = text(ws.cell(row, cfg["sub_col"]).value)

        name = text(ws.cell(row, cfg["name_col"]).value)
        if not name:
            continue
        code_col = cfg["code_col"]
        yield {
            "row": row,
            "big": big,
            "dim": dim,
            "sub": sub,
            "name": name,
            "code": norm_code(ws.cell(row, code_col).value) if code_col else "",
            "caliber": text(ws.cell(row, cfg["caliber_col"]).value),
            "current": ws.cell(row, cfg["current_col"]).value,
            "current_date": ws.cell(row, cfg["date_col"]).value,
            "current_col": cfg["current_col"],
        }


def build_canonical_code_map(wb) -> tuple[dict[tuple[str, str], str], dict[str, list[tuple[str, str, str]]]]:
    ws = wb["重点策略跟踪情况(V3.0)"]
    exact: dict[tuple[str, str], str] = {}
    by_big: dict[str, list[tuple[str, str, str]]] = {}
    for row in iter_rows(ws, "v30"):
        if not is_macro(row["big"], row["dim"]):
            continue
        code = row["code"]
        if not code or code in {"—", "-"}:
            continue
        norm = normalize_name(row["name"])
        exact[(row["big"], norm)] = code
        by_big.setdefault(row["big"], []).append((norm, code, row["name"]))
    return exact, by_big


def lookup_code(row: dict[str, Any], exact: dict[tuple[str, str], str], by_big: dict[str, list[tuple[str, str, str]]]) -> tuple[str, str]:
    if row["code"] and row["code"] not in {"—", "-"}:
        return row["code"], "本表指标代码"

    norm = normalize_name(row["name"])
    exact_code = exact.get((row["big"], norm))
    if exact_code:
        return exact_code, "按V3.0同名指标补充代码"

    candidates = []
    for canon_norm, code, canon_name in by_big.get(row["big"], []):
        if len(norm) >= 2 and (norm in canon_norm or canon_norm in norm):
            score = min(len(norm), len(canon_norm)) / max(len(norm), len(canon_norm))
            candidates.append((score, code, canon_name))
    if candidates:
        candidates.sort(reverse=True, key=lambda item: item[0])
        score, code, canon_name = candidates[0]
        if score >= 0.35:
            return code, f"按V3.0相近指标“{canon_name}”补充代码"
    return "", "未找到可匹配代码"


def current_history(series: macro_base.MacroSeries, current_date: date):
    start_date = current_date - timedelta(days=LOOKBACK_DAYS)
    points = sorted((d, v) for d, v in series.points if d <= current_date and math.isfinite(v))
    history = [(d, v) for d, v in points if start_date <= d <= current_date]
    current_point = macro_base.closest_current_period(points, current_date)
    return history, current_point


def percentile_rank(values: list[float], value: float) -> float:
    sorted_values = sorted(values)
    if not sorted_values:
        return 0.5
    if math.isclose(sorted_values[0], sorted_values[-1], rel_tol=0.0, abs_tol=1e-12):
        return 0.5
    left = bisect_left(sorted_values, value)
    right = bisect_right(sorted_values, value)
    return ((left + right) / 2) / len(sorted_values)


def color_for_percentile(percentile: float) -> tuple[str | None, str, float | None, str]:
    if percentile >= HIGH_THRESHOLD:
        distance = (percentile - HIGH_THRESHOLD) / (1 - HIGH_THRESHOLD)
        intensity = max(0.15, min(1.0, distance))
        fill = color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
        return fill, color_base.text_color_for_fill(fill), intensity, "高位"
    if percentile <= LOW_THRESHOLD:
        distance = (LOW_THRESHOLD - percentile) / LOW_THRESHOLD
        intensity = max(0.15, min(1.0, distance))
        fill = color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
        return fill, color_base.text_color_for_fill(fill), intensity, "低位"
    return None, color_base.BLACK, None, "中位"


def apply_style(cell, fill_rgb: str | None, font_rgb: str) -> None:
    font = copy.copy(cell.font)
    font.color = font_rgb
    cell.font = font
    if fill_rgb:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_rgb)
    else:
        cell.fill = PatternFill(fill_type=None)


def write_audit(rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "当前数据红高绿低复核"
    headers = [
        "表名",
        "TREE行号",
        "一级分类",
        "维度",
        "子维度",
        "指标名称",
        "匹配代码",
        "代码来源",
        "TREE数据日期",
        "TREE当前数据",
        "日报当前日期",
        "日报当前值",
        "近一年样本数",
        "历史起点",
        "历史终点",
        "历史分位",
        "位置判断",
        "填充颜色",
        "字体颜色",
        "颜色强度",
        "备注",
    ]
    ws.append(headers)
    for item in rows:
        ws.append([item.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [28, 10, 14, 14, 20, 36, 24, 30, 14, 14, 14, 14, 12, 14, 14, 12, 12, 14, 12, 12, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    wb.save(AUDIT)


def main() -> None:
    series_map = macro_base.read_macro_series(DAILY)
    wb = openpyxl.load_workbook(INPUT, data_only=False, read_only=False)
    exact_codes, codes_by_big = build_canonical_code_map(wb)

    audit_rows: list[dict[str, Any]] = []
    summary = {
        "macro_rows": 0,
        "red_rows": 0,
        "green_rows": 0,
        "cleared_rows": 0,
        "unmatched_rows": 0,
        "no_history_rows": 0,
        "nonmacro_rows_skipped": 0,
    }

    for sheet_name in [s for s in wb.sheetnames if "重点策略跟踪情况" in s]:
        ws = wb[sheet_name]
        layout = detect_layout(ws)
        if not layout:
            continue
        for row in iter_rows(ws, layout):
            if not is_macro(row["big"], row["dim"]):
                summary["nonmacro_rows_skipped"] += 1
                continue

            summary["macro_rows"] += 1
            cell = ws.cell(row["row"], row["current_col"])
            code, code_source = lookup_code(row, exact_codes, codes_by_big)
            current_dt = parse_date(row["current_date"])
            source_date = ""
            source_current: Any = ""
            hist_count = 0
            hist_start = ""
            hist_end = ""
            pct: float | None = None
            position = ""
            fill_rgb: str | None = None
            font_rgb = color_base.BLACK
            intensity: float | None = None
            note = ""

            series = series_map.get(code)
            if not code or series is None or current_dt is None:
                summary["unmatched_rows"] += 1
                note = "未匹配到底层宏观数据或日期为空，清除旧颜色"
            else:
                history, current_point = current_history(series, current_dt)
                hist_count = len(history)
                if history:
                    hist_start = history[0][0].isoformat()
                    hist_end = history[-1][0].isoformat()
                if current_point:
                    source_date = current_point[0].isoformat()
                    source_current = current_point[1]

                if hist_count < MIN_HISTORY_VALUES or current_point is None:
                    summary["no_history_rows"] += 1
                    note = "近一年历史样本不足，清除旧颜色"
                else:
                    pct = percentile_rank([value for _, value in history], current_point[1])
                    fill_rgb, font_rgb, intensity, position = color_for_percentile(pct)
                    if fill_rgb:
                        if position == "高位":
                            summary["red_rows"] += 1
                            note = "近一年高位，统一填红"
                        elif position == "低位":
                            summary["green_rows"] += 1
                            note = "近一年低位，统一填绿"
                    else:
                        summary["cleared_rows"] += 1
                        note = "近一年中位区间，清除旧颜色"

            apply_style(cell, fill_rgb, font_rgb)
            audit_rows.append(
                {
                    "表名": sheet_name,
                    "TREE行号": row["row"],
                    "一级分类": row["big"],
                    "维度": row["dim"],
                    "子维度": row["sub"],
                    "指标名称": row["name"],
                    "匹配代码": code,
                    "代码来源": code_source,
                    "TREE数据日期": row["current_date"],
                    "TREE当前数据": row["current"],
                    "日报当前日期": source_date,
                    "日报当前值": source_current,
                    "近一年样本数": hist_count,
                    "历史起点": hist_start,
                    "历史终点": hist_end,
                    "历史分位": pct,
                    "位置判断": position,
                    "填充颜色": fill_rgb or "",
                    "字体颜色": font_rgb,
                    "颜色强度": intensity,
                    "备注": note,
                }
            )

    wb.save(OUTPUT)
    wb.close()
    write_audit(audit_rows)
    print(
        {
            "input": str(INPUT),
            "daily": str(DAILY),
            "output": str(OUTPUT),
            "audit": str(AUDIT),
            **summary,
        }
    )


if __name__ == "__main__":
    main()
