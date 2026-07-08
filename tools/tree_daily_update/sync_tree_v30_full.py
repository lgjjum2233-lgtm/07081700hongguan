from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils.datetime import from_excel

from sync_tree import apply_change_colors
from sync_tree_new_indicators_from_daily import (
    TREE_SHEET,
    chart_rows,
    find_cell,
    read_xml,
    set_date,
    set_inline_string,
    set_number,
    sheet_drawing_path,
    update_chart_series,
    workbook_sheet_path,
    write_xml,
)


WORKSPACE = Path(__file__).resolve().parents[2]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

import sync_tree_20260604 as base  # noqa: E402


STOCK_INDEX_PCT_CHANGE_CODES = {
    "000001.SH",
    "399001.SZ",
    "000300.SH",
    "399006.SZ",
    "HSI.HI",
    "HSTECH.HI",
    "SPX.GI",
    "IXIC.GI",
    "DJI.GI",
}


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def is_blank(value: Any) -> bool:
    text = as_text(value)
    return text == "" or text in {"-", "—", "None"}


def is_percent_format(number_format: Any) -> bool:
    return "%" in as_text(number_format)


def is_stock_index_row(row_name: Any, row_code: Any) -> bool:
    code = as_text(row_code).upper()
    return code in STOCK_INDEX_PCT_CHANGE_CODES or base.stock_index_change_row(as_text(row_name), as_text(row_code))


def scale_for_cell(value: Any, number_format: Any, already_percent_decimal: bool = False) -> Any:
    number = to_number(value)
    if number is None:
        return value
    if is_percent_format(number_format) and not already_percent_decimal:
        return number / 100
    return number


def date_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        try:
            parsed = from_excel(value)
            return datetime(parsed.year, parsed.month, parsed.day)
        except Exception:
            return None
    text = as_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return datetime(parsed.year, parsed.month, parsed.day)
        except ValueError:
            pass
    return None


def is_future_date(value: Any, as_of: date) -> bool:
    parsed = date_value(value)
    return parsed is not None and parsed.date() > as_of


def latest_series_values_asof(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, as_of: date) -> tuple[Any, Any, Any, list[float]]:
    points: list[tuple[int, Any, float]] = []
    for col in range(6, ws.max_column + 1):
        header = ws.cell(1, col).value
        parsed = date_value(header)
        if parsed is not None and parsed.date() > as_of:
            continue
        number = to_number(ws.cell(row, col).value)
        if number is not None:
            points.append((col, header, number))
    if not points:
        return None, None, None, []
    trend = [point[2] for point in points[-60:]]
    current = points[-1][2]
    current_date = points[-1][1]
    previous = points[-2][2] if len(points) >= 2 else None
    change = current - previous if previous is not None else None
    return current, current_date, change, trend


def build_series_sources_asof(wb_values: openpyxl.Workbook, as_of: date) -> list[Any]:
    sources = []
    for sheet in base.SERIES_SHEETS:
        if sheet not in wb_values.sheetnames:
            continue
        ws = wb_values[sheet]
        for row in range(2, ws.max_row + 1):
            name = as_text(ws.cell(row, 4).value)
            code = as_text(ws.cell(row, 5).value)
            if not name and not code:
                continue
            current, data_date, change, trend = latest_series_values_asof(ws, row, as_of)
            if current is None:
                continue
            previous = current - change if change is not None else None
            sources.append(
                base.Source(
                    name=name,
                    code=code,
                    freq="",
                    unit="",
                    current=current,
                    data_date=data_date,
                    change=change,
                    previous=previous,
                    trend=trend,
                    sheet=sheet,
                    row=row,
                )
            )
    return sources


def build_sources(daily_path: Path, as_of: date):
    wb_daily_formula = openpyxl.load_workbook(daily_path, data_only=False)
    wb_daily_values = openpyxl.load_workbook(daily_path, data_only=True)
    sources = [
        source
        for source in base.build_front_sources(wb_daily_formula, wb_daily_values)
        if not is_future_date(source.data_date, as_of)
    ]
    sources.extend(build_series_sources_asof(wb_daily_values, as_of))
    by_code, by_name = base.build_indexes(sources)
    best_trend = base.build_best_trend_index(sources)
    return by_code, by_name, best_trend


def update_data_and_charts(input_tree: Path, daily_path: Path, data_output: Path, as_of: date) -> dict[str, Any]:
    by_code, by_name, best_trend = build_sources(daily_path, as_of)
    wb_tree = openpyxl.load_workbook(input_tree, data_only=False)
    ws = wb_tree[TREE_SHEET]

    with ZipFile(input_tree, "r") as zin:
        names = zin.namelist()
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)
        drawing_path = sheet_drawing_path(zin, sheet_path)
        chart_map = chart_rows(zin, drawing_path) if drawing_path else {}

        replacements: dict[str, bytes] = {}
        row_sources: dict[int, Any] = {}
        updated_rows = []
        unmatched = []
        charts_updated = []
        chart_rows_without_source = []

        for row in range(6, ws.max_row + 1):
            row_name = ws.cell(row, 4).value
            row_code = ws.cell(row, 10).value
            if is_blank(row_name) and is_blank(row_code):
                continue

            source = base.resolve_source(as_text(row_name), as_text(row_code), by_code, by_name)
            if source is None:
                if not is_blank(row_name):
                    unmatched.append({"row": row, "name": as_text(row_name), "code": as_text(row_code)})
                continue
            source = base.with_best_trend(source, best_trend)
            row_sources[row] = source

            if is_blank(ws.cell(row, 9).value) and source.freq:
                set_inline_string(find_cell(sheet_root, row, 9), source.freq)
            if is_blank(ws.cell(row, 10).value) and source.code:
                set_inline_string(find_cell(sheet_root, row, 10), source.code)

            current_format = ws.cell(row, 11).number_format
            change_format = ws.cell(row, 13).number_format
            set_number(find_cell(sheet_root, row, 11), scale_for_cell(source.current, current_format))
            set_date(find_cell(sheet_root, row, 12), source.data_date)

            change_value = source.change
            already_percent_decimal = False
            if is_stock_index_row(row_name, row_code):
                current = to_number(source.current)
                previous = to_number(source.previous)
                if current is not None and previous not in (None, 0):
                    change_value = (current - previous) / previous
                    already_percent_decimal = True
            set_number(find_cell(sheet_root, row, 13), scale_for_cell(change_value, change_format, already_percent_decimal))

            updated_rows.append(
                {
                    "row": row,
                    "name": as_text(row_name),
                    "code": as_text(row_code) or source.code,
                    "source": f"{source.sheet}!{source.row}" if source.sheet else source.code,
                    "date": source.data_date,
                    "current": source.current,
                    "change": change_value,
                }
            )

            chart_info = chart_map.get(row)
            if chart_info:
                take = base.frequency_take_count(ws, row, source)
                values = base.trim_trend_values(source, take)
                if values:
                    replacements[chart_info["chart_path"]] = update_chart_series(zin.read(chart_info["chart_path"]), values)
                    charts_updated.append({"row": row, "points": len(values), "chart": chart_info["chart_path"]})

        for row, info in sorted(chart_map.items()):
            if row >= 6 and row <= ws.max_row and row not in row_sources:
                chart_rows_without_source.append(row)

        replacements[sheet_path] = write_xml(sheet_root)
        with ZipFile(data_output, "w", ZIP_DEFLATED) as zout:
            for name in names:
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "as_of": as_of.isoformat(),
        "data_output": str(data_output),
        "updated_row_count": len(updated_rows),
        "updated_rows_sample": updated_rows[:30],
        "unmatched_count": len(unmatched),
        "unmatched_sample": unmatched[:30],
        "charts_updated_count": len(charts_updated),
        "charts_updated_sample": charts_updated[:30],
        "chart_rows_without_source_count": len(chart_rows_without_source),
        "chart_rows_without_source_sample": chart_rows_without_source[:30],
    }


def sync(input_tree: Path, daily_path: Path, output_tree: Path, previous_tree: Path | None, as_of: date) -> dict[str, Any]:
    with TemporaryDirectory() as tmp_dir:
        data_output = Path(tmp_dir) / "tree_v30_data_updated.xlsx"
        data_meta = update_data_and_charts(input_tree, daily_path, data_output, as_of)
        if previous_tree is not None:
            color_meta = apply_change_colors(data_output, previous_tree, output_tree, TREE_SHEET)
        else:
            output_tree.write_bytes(data_output.read_bytes())
            color_meta = {}
    return {**data_meta, "output": str(output_tree), "colors": color_meta}


def main() -> None:
    parser = argparse.ArgumentParser(description="Synchronize the full TREE V3.0 sheet from a refreshed daily workbook.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--previous-tree")
    parser.add_argument("--as-of", default=date.today().isoformat(), help="Only use source data dated on or before this day.")
    args = parser.parse_args()

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
    result = sync(
        Path(args.tree),
        Path(args.daily),
        Path(args.output),
        Path(args.previous_tree) if args.previous_tree else None,
        as_of,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
