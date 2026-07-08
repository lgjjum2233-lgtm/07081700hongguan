from __future__ import annotations

import argparse
import copy
import json
import math
import re
import sys
import xml.etree.ElementTree as ET
from bisect import bisect_left, bisect_right
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import add_tree_yoy_column_20260611 as yoy_base  # noqa: E402
import apply_macro_change_history_colors as color_base  # noqa: E402
import sync_tree_v30_full as v30  # noqa: E402
from sync_tree_new_indicators_from_daily import (  # noqa: E402
    NS,
    chart_rows,
    find_cell,
    read_xml,
    set_date,
    set_inline_string,
    set_number,
    sheet_drawing_path,
    update_chart_series,
    workbook_sheet_path,
    write_xml,
)


YOY_WORD = "\u540c\u6bd4"
INDUSTRY_WORD = "\u4ea7\u4e1a"
DAILY_WORD = "\u65e5"
WEEKLY_WORD = "\u5468"
MONTHLY_WORD = "\u6708"
QUARTERLY_WORD = "\u5b63"
PE_WORDS = ("\u5e02\u76c8\u7387", "\u4f30\u503c", "PE")
BLACK = "FF000000"
MARGIN_BALANCE_WORD = "\u878d\u8d44\u878d\u5238"
YOY_GROWTH_NOT_MEANINGFUL_WORDS = (
    "\u51c0\u6295\u653e",
    "\u5229\u5dee",
    "\u5229\u7387",
    "\u6536\u76ca\u7387",
    "PMI",
    "\u51c6\u5907\u91d1\u7387",
    "\u6760\u6746\u7387",
    "\u8d64\u5b57\u7387",
    "\u6bd4\u7387",
)
NON_MACRO_SCOPE_WORDS = (
    "A\u80a1",
    "\u6e2f\u80a1",
    "\u7f8e\u80a1",
    "\u60c5\u7eea\u9762",
    "\u4f30\u503c\u9762",
    "\u4ea7\u4e1a\u57fa\u672c\u9762",
    "\u4ea7\u4e1a\u6307\u6807",
)
STOCK_INDEX_CODES = {
    "000001.SH",
    "399001.SZ",
    "000300.SH",
    "399006.SZ",
    "HSI.HI",
    "HSTECH.HI",
    "SPX.GI",
    "IXIC.GI",
    "DJI.GI",
}
SIMPLE_WIND_CODE_RE = re.compile(r"^[A-Z][A-Z0-9.]*$")
DISPLAY_SOURCE_OVERRIDES = (
    ("\u6807\u666e500\u52a8\u6001PE", "G0011297"),
    ("VIX", "G0003892"),
)
YIELD_CODE_OVERRIDES = {
    "M0325687": "S0059749",
    "M0325687-G0000891": "S0059749-G0000891",
}
USD_MILLION_BALANCE_CODES = {
    "G1109077",
    "G1109086",
    "G1109091",
    "G1109087",
}
USD_BILLION_BALANCE_CODES = {
    "G0003382",
    "G0003383",
}
MASTER_CHART_COL = 20
DISPLAY_CHART_COL = 19
CHART_COL_OFF = "50000"
CHART_ROW_OFF = "30000"
MASTER_CHART_EXT = ("2500000", "850000")

MASTER_COLS = {
    "big": 1,
    "asset": 2,
    "face": 3,
    "framework": 5,
    "judgement": 6,
    "name": 7,
    "freq": 12,
    "track": 13,
    "code": 14,
    "current": 15,
    "date": 16,
    "change": 17,
    "yoy": 18,
    "yoy_inc": 19,
}

DISPLAY_COLS = {
    "big": 6,
    "dim": 7,
    "sub": 8,
    "name": 9,
    "freq": 12,
    "current": 15,
    "date": 16,
    "change": 17,
}

CHANGE_THRESHOLD = 0.65
POSITION_LOW = 0.35
POSITION_HIGH = 0.65
MIN_HISTORY = 4
BUILTIN_NUM_FMT_IDS = {
    "General": "0",
    "0.00": "2",
    "0.00%": "10",
}


@dataclass
class Series:
    name: str
    code: str
    points: list[tuple[date, float]]
    kind: str = "macro"


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    return v30.to_number(value)


def parse_date(value: Any) -> date | None:
    parsed = v30.date_value(value)
    if parsed is not None:
        return parsed.date()
    return yoy_base.parse_date(value)


def norm_code(value: Any) -> str:
    return as_text(value).upper()


def canonical_tree_code(value: Any) -> str:
    code = norm_code(value)
    return YIELD_CODE_OVERRIDES.get(code, code)


def is_blank(value: Any) -> bool:
    text = as_text(value)
    return text == "" or text in {"-", "None"}


def find_sheet_name(wb: openpyxl.Workbook, suffix: str) -> str:
    for name in wb.sheetnames:
        if name.endswith(suffix):
            return name
    raise KeyError(f"No sheet ending with {suffix!r}")


def is_percent_format(fmt: Any) -> bool:
    return "%" in as_text(fmt)


def closest_point(points: list[tuple[date, float]], current_date: date) -> tuple[date, float] | None:
    by_date = {d: v for d, v in points}
    if current_date in by_date:
        return current_date, by_date[current_date]
    candidates = [(abs((d - current_date).days), d, v) for d, v in points if abs((d - current_date).days) <= 45]
    if not candidates:
        return None
    _, d, v = min(candidates, key=lambda item: (item[0], item[1]))
    return d, v


def previous_observation(points: list[tuple[date, float]], current_date: date) -> tuple[date, float] | None:
    prior = [(d, v) for d, v in points if d < current_date]
    return prior[-1] if prior else None


def raw_display_scale(display_value: Any, raw_value: float | None) -> float:
    display = to_number(display_value)
    if display is None or raw_value is None or abs(raw_value) <= 1e-12:
        return 1.0
    ratio = display / raw_value
    candidates = [1.0, 0.01, 0.0001, 100.0]
    best = min(candidates, key=lambda item: abs(math.log(abs(ratio / item))) if ratio and item else float("inf"))
    if best != 1.0 and abs(ratio / best - 1.0) <= 0.20:
        return best
    return 1.0


def read_full_series(daily_path: Path, as_of: date) -> dict[str, Series]:
    wb = openpyxl.load_workbook(daily_path, data_only=True, read_only=True)
    out: dict[str, Series] = {}
    source_sheet_names = {"\u5b8f\u89c2\u6570\u636e", "\u6307\u6570\u8d70\u52bf"}
    for sheet_name in wb.sheetnames:
        if sheet_name not in source_sheet_names:
            continue
        ws = wb[sheet_name]
        global_header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        def parsed_dates(header_values: tuple[Any, ...]) -> list[tuple[int, date]]:
            dates: list[tuple[int, date]] = []
            for idx, value in enumerate(header_values, start=1):
                if idx < 6:
                    continue
                parsed = parse_date(value)
                if parsed is not None and parsed <= as_of:
                    dates.append((idx - 1, parsed))
            return dates

        global_dates = parsed_dates(global_header)
        if not global_dates:
            continue
        current_dates = global_dates
        previous_row_values: tuple[Any, ...] | None = None
        special_tags = {"TURN", "PE_EST", "PE_TTM"}
        current_tag = ""
        sheet_kind = "macro" if sheet_name == "\u5b8f\u89c2\u6570\u636e" else "index"
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            name = row_values[3] if len(row_values) > 3 else None
            code = norm_code(row_values[4] if len(row_values) > 4 else None)
            tag = as_text(row_values[2] if len(row_values) > 2 else None).upper()
            if tag:
                current_tag = tag
                if tag in special_tags and previous_row_values is not None:
                    block_dates = parsed_dates(previous_row_values)
                    if block_dates:
                        current_dates = block_dates
                else:
                    current_dates = global_dates
            if not code or code in out:
                if code and current_tag in special_tags:
                    pass
                else:
                    previous_row_values = row_values
                    continue
            key = code
            if current_tag in special_tags:
                key = f"{code}|{current_tag}"
            if key in out:
                previous_row_values = row_values
                continue
            points: list[tuple[date, float]] = []
            for value_idx, d in current_dates:
                if value_idx >= len(row_values):
                    continue
                number = to_number(row_values[value_idx])
                if number is not None:
                    points.append((d, number))
            if points:
                points.sort(key=lambda item: item[0])
                out[key] = Series(name=as_text(name), code=key, points=points, kind=sheet_kind)
            previous_row_values = row_values
    for key, series in list(out.items()):
        if not key.endswith("|PE_TTM"):
            continue
        base_code = key.split("|", 1)[0]
        price_series = out.get(base_code)
        if not price_series or not price_series.points:
            continue
        price_date = price_series.points[-1][0]
        trimmed = [point for point in series.points if point[0] <= price_date]
        if trimmed:
            series.points = trimmed
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_column < 9:
            continue
        for row_values in ws.iter_rows(values_only=True):
            name = as_text(row_values[0] if len(row_values) > 0 else None)
            if "PE(TTM)" not in name:
                continue
            code = norm_code(row_values[2] if len(row_values) > 2 else None)
            if not code:
                continue
            points: list[tuple[date, float]] = []
            previous_date = parse_date(row_values[8] if len(row_values) > 8 else None)
            previous_value = to_number(row_values[7] if len(row_values) > 7 else None)
            current_date = parse_date(row_values[5] if len(row_values) > 5 else None)
            current_value = to_number(row_values[4] if len(row_values) > 4 else None)
            if previous_date is not None and previous_value is not None:
                points.append((previous_date, previous_value))
            if current_date is not None and current_value is not None:
                points.append((current_date, current_value))
            if points:
                points.sort(key=lambda item: item[0])
                out[f"{code}|PE_TTM_FRONT"] = Series(name=name, code=f"{code}|PE_TTM_FRONT", points=points, kind="front")
    wb.close()
    return out


CODE_TOKEN_RE = re.compile(r"[A-Z][A-Z0-9]*(?:\.[A-Z0-9]+)?")


def expression_series(code: str, full_series: dict[str, Series]) -> Series | None:
    expr = norm_code(code).replace(" ", "")
    if not any(op in expr for op in "+-*/"):
        return None
    if "*" in expr or "/" in expr:
        return None
    tokens = []
    for token in CODE_TOKEN_RE.findall(expr):
        if token not in full_series:
            return None
        if token not in tokens:
            tokens.append(token)
    if not tokens:
        return None
    common: set[date] | None = None
    maps: dict[str, dict[date, float]] = {}
    for token in tokens:
        values = {d: v for d, v in full_series[token].points}
        maps[token] = values
        common = set(values) if common is None else common & set(values)
    if not common:
        return None
    replaced = expr
    token_names: dict[str, str] = {}
    for idx, token in enumerate(sorted(tokens, key=len, reverse=True)):
        name = f"V{idx}"
        token_names[token] = name
        replaced = re.sub(rf"(?<![A-Z0-9.]){re.escape(token)}(?![A-Z0-9.])", name, replaced)
    points: list[tuple[date, float]] = []
    for d in sorted(common):
        env = {name: maps[token][d] for token, name in token_names.items()}
        try:
            value = eval(replaced, {"__builtins__": {}}, env)
        except Exception:
            continue
        if isinstance(value, (int, float)) and math.isfinite(value):
            points.append((d, float(value)))
    kinds = {full_series[token].kind for token in tokens}
    kind = "macro" if kinds == {"macro"} else "index"
    return Series(name=code, code=code, points=points, kind=kind) if points else None


def get_series(code: Any, full_series: dict[str, Series]) -> Series | None:
    key = norm_code(code)
    if not key or key in {"-", "NONE"}:
        return None
    return full_series.get(key) or expression_series(key, full_series)


def is_macro_row(code: Any, row_text: str, full_series: dict[str, Series]) -> bool:
    if INDUSTRY_WORD in row_text or any(word in row_text for word in NON_MACRO_SCOPE_WORDS):
        return False
    series = get_series(code, full_series)
    return series is not None and series.kind == "macro"


def percentile_rank(sorted_values: list[float], value: float) -> float:
    if not sorted_values:
        return 0.0
    left = bisect_left(sorted_values, value)
    right = bisect_right(sorted_values, value)
    return ((left + right) / 2) / len(sorted_values)


def color_by_position(points: list[tuple[date, float]], current_date: date) -> tuple[str | None, str, dict[str, Any]]:
    current = closest_point(points, current_date)
    if current is None:
        return None, BLACK, {"reason": "no_current"}
    start = current_date - timedelta(days=365)
    window = [v for d, v in points if start <= d <= current_date and math.isfinite(v)]
    if len(window) < MIN_HISTORY:
        return None, BLACK, {"reason": "insufficient", "history": len(window)}
    if max(window) - min(window) <= max(1e-12, abs(current[1]) * 1e-10):
        return None, BLACK, {"reason": "flat", "history": len(window)}
    pct = percentile_rank(sorted(window), current[1])
    if pct >= POSITION_HIGH:
        intensity = max(0.15, min(1.0, (pct - POSITION_HIGH) / (1.0 - POSITION_HIGH)))
        fill = color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
    elif pct <= POSITION_LOW:
        intensity = max(0.15, min(1.0, (POSITION_LOW - pct) / POSITION_LOW))
        fill = color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
    else:
        return None, BLACK, {"reason": "middle", "percentile": pct}
    return fill, color_base.text_color_for_fill(fill), {"percentile": pct, "intensity": intensity}


def color_by_abs_change(series_values: list[tuple[date, float]], latest_value: Any, latest_date: date) -> tuple[str | None, str, dict[str, Any]]:
    latest = to_number(latest_value)
    if latest is None or abs(latest) <= 1e-12:
        return None, BLACK, {"reason": "zero_or_blank"}
    start = latest_date - timedelta(days=365)
    values = [(d, v) for d, v in series_values if start <= d <= latest_date]
    if len(values) < MIN_HISTORY + 1:
        return None, BLACK, {"reason": "insufficient", "history": len(values)}
    diffs = [values[idx][1] - values[idx - 1][1] for idx in range(1, len(values))]
    prior_abs = sorted(abs(x) for x in diffs[:-1] if math.isfinite(x))
    if len(prior_abs) < MIN_HISTORY:
        return None, BLACK, {"reason": "insufficient_diffs", "history": len(prior_abs)}
    pct = percentile_rank(prior_abs, abs(latest))
    if pct < CHANGE_THRESHOLD:
        return None, BLACK, {"reason": "small", "percentile": pct}
    intensity = max(0.15, min(1.0, (pct - CHANGE_THRESHOLD) / (1.0 - CHANGE_THRESHOLD)))
    fill = (
        color_base.interpolate_color(color_base.RED_LIGHT, color_base.RED_DARK, intensity)
        if latest > 0
        else color_base.interpolate_color(color_base.GREEN_LIGHT, color_base.GREEN_DARK, intensity)
    )
    return fill, color_base.text_color_for_fill(fill), {"percentile": pct, "intensity": intensity}


def is_yoy_input(row_name: Any, series: Series) -> bool:
    text = f"{as_text(row_name)} {series.name}"
    return YOY_WORD in text


def yoy_growth_applicable(row_name: Any, series: Series) -> bool:
    if is_yoy_input(row_name, series):
        return True
    text = f"{as_text(row_name)} {series.name}".upper()
    return not any(word.upper() in text for word in YOY_GROWTH_NOT_MEANINGFUL_WORDS)


def yoy_value_and_history(series: Series, current_date: date, current_display: Any, row_name: Any) -> tuple[Any, list[tuple[date, float]]]:
    current = closest_point(series.points, current_date)
    if current is None:
        return "-", []
    if is_yoy_input(row_name, series):
        scale = raw_display_scale(current_display, current[1])
        start = current_date - timedelta(days=365)
        hist = [(d, v * scale) for d, v in series.points if start <= d <= current_date]
        return to_number(current_display) if to_number(current_display) is not None else current[1] * scale, hist
    if not yoy_growth_applicable(row_name, series):
        return "-", []
    prior = yoy_base.closest_prior_period(series.points, current_date)
    if prior is None or abs(prior[1]) <= 1e-12:
        return "-", []
    latest = (current[1] - prior[1]) / abs(prior[1])
    start = current_date - timedelta(days=365)
    hist: list[tuple[date, float]] = []
    for d, value in series.points:
        if not (start <= d <= current_date):
            continue
        old = yoy_base.closest_prior_period(series.points, d)
        if old is None or abs(old[1]) <= 1e-12:
            continue
        hist.append((d, (value - old[1]) / abs(old[1])))
    return latest, hist


def yoy_increment_and_history(series: Series, current_date: date, current_display: Any) -> tuple[Any, list[tuple[date, float]]]:
    current = closest_point(series.points, current_date)
    if current is None:
        return "-", []
    prior = yoy_base.closest_prior_period(series.points, current_date)
    if prior is None:
        return "-", []
    scale = raw_display_scale(current_display, current[1])
    latest = (current[1] - prior[1]) * scale
    start = current_date - timedelta(days=365)
    hist: list[tuple[date, float]] = []
    for d, value in series.points:
        if not (start <= d <= current_date):
            continue
        old = yoy_base.closest_prior_period(series.points, d)
        if old is None:
            continue
        hist.append((d, (value - old[1]) * scale))
    return latest, hist


def set_style(sheet_root: Any, styles_root: Any, row: int, col: int, fill: str | None, font: str, cache: dict[tuple[int, str, str], int]) -> None:
    cell = find_cell(sheet_root, row, col)
    base_style = int(cell.attrib.get("s", "0"))
    cell.attrib["s"] = str(color_base.ensure_cell_style(styles_root, base_style, fill, font, cache))


def ensure_num_fmt_id(styles_root: Any, fmt_code: str) -> str:
    if fmt_code in BUILTIN_NUM_FMT_IDS:
        return BUILTIN_NUM_FMT_IDS[fmt_code]
    num_fmts = styles_root.find("main:numFmts", NS)
    if num_fmts is None:
        num_fmts = ET.Element(f"{{{NS['main']}}}numFmts", {"count": "0"})
        styles_root.insert(0, num_fmts)
    for num_fmt in num_fmts.findall("main:numFmt", NS):
        if num_fmt.attrib.get("formatCode") == fmt_code:
            return num_fmt.attrib["numFmtId"]
    existing = [int(num_fmt.attrib.get("numFmtId", "163")) for num_fmt in num_fmts.findall("main:numFmt", NS)]
    next_id = max([163, *existing]) + 1
    num_fmts.append(ET.Element(f"{{{NS['main']}}}numFmt", {"numFmtId": str(next_id), "formatCode": fmt_code}))
    num_fmts.attrib["count"] = str(len(num_fmts.findall("main:numFmt", NS)))
    return str(next_id)


def ensure_number_style(styles_root: Any, base_style_id: int, fmt_code: str, cache: dict[tuple[int, str], int]) -> int:
    key = (base_style_id, fmt_code)
    if key in cache:
        return cache[key]
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if cell_xfs is None:
        raise KeyError("styles cellXfs")
    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]
    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["numFmtId"] = ensure_num_fmt_id(styles_root, fmt_code)
    new_xf.attrib["applyNumberFormat"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(new_style_id + 1)
    cache[key] = new_style_id
    return new_style_id


def set_number_format(
    sheet_root: Any,
    styles_root: Any,
    row: int,
    col: int,
    fmt_code: str,
    cache: dict[tuple[int, str], int],
) -> None:
    cell = find_cell(sheet_root, row, col)
    base_style = int(cell.attrib.get("s", "0"))
    cell.attrib["s"] = str(ensure_number_style(styles_root, base_style, fmt_code, cache))


def metric_number_formats(current_format: Any, code: Any = None) -> tuple[str, str, str, str]:
    # Display raw source values without scaling or text-unit suffixes.
    # Percentage semantics remain percentage-formatted; other values use
    # two visible decimal places while retaining full stored precision.
    current_fmt = "0.00%" if v30.is_percent_format(current_format) else "0.00"
    return current_fmt, current_fmt, "0.00%", current_fmt


def chart_take_count(name: Any, freq: Any, values_count: int) -> int:
    text = f"{as_text(name)} {as_text(freq)}"
    if any(word in text for word in PE_WORDS):
        return min(values_count, 750)
    if DAILY_WORD in text or WEEKLY_WORD in text:
        return min(values_count, 60)
    if QUARTERLY_WORD in text:
        return min(values_count, 8)
    if MONTHLY_WORD in text:
        return min(values_count, 12)
    return min(values_count, 12)


def match_key(value: Any) -> str:
    key = v30.base.norm(value)
    key = re.sub(r"^[\u2460-\u2473\u2776-\u277f0-9]+", "", key)
    for token in [
        "\u540c\u6bd4\u591a\u589e",
        "\u540c\u6bd4\u5c11\u589e",
        "\u5355\u4f4d\u4ebf\u5143",
        "\u4ebf\u5143",
        "\u4ebf\u6e2f\u5143",
        "\u7d2f\u8ba1\u503c",
        "\u5f53\u6708\u503c",
        "\u53d8\u52a8",
    ]:
        key = key.replace(token, "")
    return key


def build_master_index(master_cache: dict[int, dict[str, Any]]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = {}
    for row, item in master_cache.items():
        key = match_key(item["name"])
        if len(key) >= 2:
            index.setdefault(key, []).append(row)
    return index


def is_pe_like(value: Any) -> bool:
    text = as_text(value).upper()
    return any(word.upper() in text for word in PE_WORDS)


def find_date_style_id(ws: openpyxl.worksheet.worksheet.Worksheet, col: int) -> int | None:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, col)
        number_format = as_text(cell.number_format).lower()
        if "y" in number_format and "d" in number_format:
            return getattr(cell, "style_id", getattr(cell, "_style_id", None))
    return None


def display_override_code(name: Any) -> str | None:
    text = as_text(name).upper().replace(":", "").replace("\uff1a", "")
    for token, code in DISPLAY_SOURCE_OVERRIDES:
        if token.upper() in text:
            return code
    if "\u6807\u666e500" in text and ("PE" in text or "\u5e02\u76c8" in text):
        return "G0011297"
    return None


def is_suspicious_display_source(name: Any, source: Any) -> bool:
    name_text = as_text(name).upper()
    source_name = as_text(getattr(source, "name", "")).upper()
    if "\u76c8\u5229\u589e\u957f" in name_text and not any(token in source_name for token in ("\u76c8\u5229", "\u5229\u6da6", "\u51c0\u5229\u6da6")):
        return True
    if is_pe_like(name_text) and not any(token in source_name for token in ("PE", "\u5e02\u76c8")):
        return True
    return False


def protect_margin_zero(name: Any, new_value: Any, old_value: Any) -> bool:
    if MARGIN_BALANCE_WORD not in as_text(name):
        return False
    new_number = to_number(new_value)
    old_number = to_number(old_value)
    return new_number == 0 and old_number not in (None, 0)


def source_from_series(series: Series, name: Any, code: Any) -> Any | None:
    if not series.points:
        return None
    current_date, current = series.points[-1]
    previous = series.points[-2][1] if len(series.points) >= 2 else None
    change = current - previous if previous is not None else None
    return v30.base.Source(
        name=as_text(name) or series.name,
        code=as_text(code) or series.code,
        freq="",
        unit="",
        current=current,
        data_date=current_date,
        change=change,
        previous=previous,
        trend=[value for _d, value in series.points],
        sheet="full_series",
        row=0,
    )


def preferred_source(name: Any, code: Any, full_series: dict[str, Series]) -> Any | None:
    key = norm_code(code)
    if not key:
        return None
    if "|PE_TTM" in key:
        series = full_series.get(key)
        return source_from_series(series, name, code) if series else None
    if key in STOCK_INDEX_CODES and not is_pe_like(name):
        series = full_series.get(key)
        return source_from_series(series, name, code) if series else None
    return None


def resolve_master_source(
    name: Any,
    code: Any,
    full_series: dict[str, Series],
    by_code: dict[str, Any],
    by_name: dict[str, list[Any]],
) -> Any | None:
    source = preferred_source(name, code, full_series)
    if source is not None:
        return source

    code_text = norm_code(code)
    if SIMPLE_WIND_CODE_RE.fullmatch(code_text):
        # A single Wind code is authoritative. If this source workbook does
        # not contain it, preserve the prior monthly/supplemental overlay
        # instead of falling back to a similarly named but different series.
        return by_code.get(v30.base.code_key(code_text))

    return v30.base.resolve_source(as_text(name), as_text(code), by_code, by_name)


def match_master(display_name: Any, master_cache: dict[int, dict[str, Any]], master_index: dict[str, list[int]]) -> dict[str, Any] | None:
    key = match_key(display_name)
    display_pe = is_pe_like(display_name)
    if key in master_index:
        for row in master_index[key]:
            item = master_cache[row]
            if is_pe_like(item["name"]) == display_pe:
                return item
    candidates: list[tuple[int, int]] = []
    for mkey, rows in master_index.items():
        if len(mkey) < 3 or len(key) < 3:
            continue
        filtered_rows = [row for row in rows if is_pe_like(master_cache[row]["name"]) == display_pe]
        if not filtered_rows:
            continue
        if key in mkey or mkey in key:
            candidates.append((abs(len(mkey) - len(key)), filtered_rows[0]))
    if candidates:
        _score, row = min(candidates)
        return master_cache[row]
    return None


def values_for_chart(code: Any, source: Any, full_series: dict[str, Series], take: int) -> list[float]:
    series = get_series(code or getattr(source, "code", ""), full_series)
    if series is not None and series.points:
        return [v for _d, v in series.points[-take:]]
    raw = [to_number(value) for value in (getattr(source, "trend", None) or [])]
    vals = [float(value) for value in raw if value is not None]
    return vals[-take:]


def chart_anchor_row(anchor: ET.Element) -> int | None:
    if anchor.find(".//a:graphicData/c:chart", NS) is None:
        return None
    frm = anchor.find("xdr:from", NS)
    if frm is None:
        return None
    row_text = frm.findtext("xdr:row", namespaces=NS)
    if row_text is None:
        return None
    return int(row_text) + 1


def set_marker(marker: ET.Element | None, col: int, row: int, col_off: str, row_off: str) -> None:
    if marker is None:
        return
    values = {
        "col": str(col - 1),
        "colOff": col_off,
        "row": str(row - 1),
        "rowOff": row_off,
    }
    for tag, value in values.items():
        child = marker.find(f"xdr:{tag}", NS)
        if child is not None:
            child.text = value


def normalize_chart_anchor(anchor: ET.Element, row: int, target_col: int, default_ext: tuple[str, str] | None) -> None:
    set_marker(anchor.find("xdr:from", NS), target_col, row, CHART_COL_OFF, CHART_ROW_OFF)

    ext = anchor.find("xdr:ext", NS)
    xfrm_ext = anchor.find(".//xdr:graphicFrame/xdr:xfrm/a:ext", NS)
    if default_ext is not None:
        cx, cy = default_ext
        if ext is not None:
            ext.attrib["cx"] = cx
            ext.attrib["cy"] = cy
        if xfrm_ext is not None:
            xfrm_ext.attrib["cx"] = cx
            xfrm_ext.attrib["cy"] = cy

    to_marker = anchor.find("xdr:to", NS)
    if to_marker is not None:
        if default_ext is not None:
            set_marker(to_marker, target_col, row, default_ext[0], default_ext[1])
        else:
            to_col_off = to_marker.findtext("xdr:colOff", namespaces=NS) or CHART_COL_OFF
            to_row_off = to_marker.findtext("xdr:rowOff", namespaces=NS) or CHART_ROW_OFF
            set_marker(to_marker, target_col, row, to_col_off, to_row_off)


def normalize_chart_anchors(
    zin: ZipFile,
    drawing_path: str | None,
    target_col: int,
    default_ext: tuple[str, str] | None = None,
) -> tuple[bytes | None, dict[str, Any]]:
    if not drawing_path:
        return None, {"charts": 0, "unique_rows": 0, "duplicates_removed": 0}
    drawing_root = read_xml(zin, drawing_path)
    entries: list[tuple[int, int, ET.Element]] = []
    for idx, anchor in enumerate(list(drawing_root)):
        row = chart_anchor_row(anchor)
        if row is not None:
            entries.append((idx, row, anchor))

    keep_idx_by_row: dict[int, int] = {}
    for idx, row, _anchor in entries:
        keep_idx_by_row[row] = idx

    removed = 0
    for idx, row, anchor in reversed(entries):
        if keep_idx_by_row[row] != idx:
            drawing_root.remove(anchor)
            removed += 1

    for idx, row, anchor in entries:
        if keep_idx_by_row[row] == idx:
            normalize_chart_anchor(anchor, row, target_col, default_ext)

    return write_xml(drawing_root), {
        "charts": len(entries) - removed,
        "unique_rows": len(keep_idx_by_row),
        "duplicates_removed": removed,
    }


def sync_master_sheet(
    wb_values: openpyxl.Workbook,
    zin: ZipFile,
    sheet_name: str,
    daily_path: Path,
    as_of: date,
    full_series: dict[str, Series],
    replacements: dict[str, bytes],
    styles_root: Any,
    style_cache: dict[tuple[int, str, str], int],
    master_cache: dict[int, dict[str, Any]],
    preserve_layout: bool,
    write_canonical_codes: bool,
) -> dict[str, Any]:
    by_code, by_name, best_trend = v30.build_sources(daily_path, as_of)
    ws = wb_values[sheet_name]
    sheet_path = workbook_sheet_path(zin, sheet_name)
    sheet_root = read_xml(zin, sheet_path)
    drawing_path = sheet_drawing_path(zin, sheet_path)
    chart_map = chart_rows(zin, drawing_path) if drawing_path else {}

    updated = []
    unmatched = []
    charts = []
    colored = {"current": 0, "change": 0, "yoy": 0, "yoy_inc": 0}
    protected = []
    date_style_id = find_date_style_id(ws, MASTER_COLS["date"])
    num_fmt_cache: dict[tuple[int, str], int] = {}
    row_context = {col: "" for col in (1, 2, 3, 5)}

    for row in range(2, ws.max_row + 1):
        if row % 50 == 0:
            print(f"master row {row}/{ws.max_row}", file=sys.stderr, flush=True)
        name = ws.cell(row, MASTER_COLS["name"]).value
        original_code = ws.cell(row, MASTER_COLS["code"]).value
        code = canonical_tree_code(original_code)
        if is_blank(name) and is_blank(code):
            continue
        if write_canonical_codes and code and code != norm_code(original_code):
            set_inline_string(find_cell(sheet_root, row, MASTER_COLS["code"]), code)
        for col in row_context:
            value = as_text(ws.cell(row, col).value)
            if value:
                row_context[col] = value
        row_text = " ".join([*(row_context[col] for col in (1, 2, 3, 5)), as_text(name)])
        source = resolve_master_source(name, code, full_series, by_code, by_name)
        if source is None:
            if not is_blank(name):
                unmatched.append({"row": row, "name": as_text(name), "code": as_text(code)})
            continue
        if getattr(source, "sheet", "") != "full_series":
            source = v30.base.with_best_trend(source, best_trend)

        current_format = ws.cell(row, MASTER_COLS["current"]).number_format
        current_num_fmt, change_num_fmt, yoy_num_fmt, yoy_inc_num_fmt = metric_number_formats(
            current_format, code or source.code
        )
        current_display = v30.scale_for_cell(source.current, current_num_fmt)
        change_value = source.change
        already_percent_decimal = False
        if v30.is_stock_index_row(name, code or source.code):
            current_number = to_number(source.current)
            previous_number = to_number(source.previous)
            if current_number is not None and previous_number not in (None, 0):
                change_value = (current_number - previous_number) / previous_number
                already_percent_decimal = True
        change_display = v30.scale_for_cell(change_value, change_num_fmt, already_percent_decimal)

        if protect_margin_zero(name, current_display, ws.cell(row, MASTER_COLS["current"]).value):
            old_current = ws.cell(row, MASTER_COLS["current"]).value
            old_date_raw = ws.cell(row, MASTER_COLS["date"]).value
            old_date = parse_date(old_date_raw) or old_date_raw
            old_change = ws.cell(row, MASTER_COLS["change"]).value

            set_number(find_cell(sheet_root, row, MASTER_COLS["current"]), old_current)
            set_date(find_cell(sheet_root, row, MASTER_COLS["date"]), old_date, date_style_id)
            set_number(find_cell(sheet_root, row, MASTER_COLS["change"]), old_change)
            set_number_format(sheet_root, styles_root, row, MASTER_COLS["current"], current_num_fmt, num_fmt_cache)
            set_number_format(sheet_root, styles_root, row, MASTER_COLS["change"], change_num_fmt, num_fmt_cache)
            for col in (MASTER_COLS["current"], MASTER_COLS["change"], MASTER_COLS["yoy"], MASTER_COLS["yoy_inc"]):
                set_style(sheet_root, styles_root, row, col, None, BLACK, style_cache)

            master_cache[row] = {
                "row": row,
                "name": as_text(name),
                "code": as_text(code or source.code),
                "current": old_current,
                "date": old_date,
                "change": old_change,
                "freq": ws.cell(row, MASTER_COLS["freq"]).value or source.freq,
                "chart_values": [],
            }
            protected.append({"row": row, "name": as_text(name), "old": old_current, "date": as_text(old_date)})
            continue

        set_number(find_cell(sheet_root, row, MASTER_COLS["current"]), current_display)
        set_date(find_cell(sheet_root, row, MASTER_COLS["date"]), source.data_date, date_style_id)
        set_number(find_cell(sheet_root, row, MASTER_COLS["change"]), change_display)
        set_number_format(sheet_root, styles_root, row, MASTER_COLS["current"], current_num_fmt, num_fmt_cache)
        set_number_format(sheet_root, styles_root, row, MASTER_COLS["change"], change_num_fmt, num_fmt_cache)

        current_date = parse_date(source.data_date)
        series = get_series(code or source.code, full_series)
        macro_scope = current_date is not None and is_macro_row(code or source.code, row_text, full_series)
        if macro_scope and series is not None:
            current_point = closest_point(series.points, current_date)
            display_scale = raw_display_scale(current_display, current_point[1]) if current_point else 1.0
            yoy_value, yoy_hist = yoy_value_and_history(series, current_date, current_display, name)
            yoy_inc_value, yoy_inc_hist = yoy_increment_and_history(series, current_date, current_display)
            if yoy_value == "-":
                set_inline_string(find_cell(sheet_root, row, MASTER_COLS["yoy"]), "-")
            else:
                set_number(find_cell(sheet_root, row, MASTER_COLS["yoy"]), yoy_value)
                set_number_format(sheet_root, styles_root, row, MASTER_COLS["yoy"], yoy_num_fmt, num_fmt_cache)
            if yoy_inc_value == "-":
                set_inline_string(find_cell(sheet_root, row, MASTER_COLS["yoy_inc"]), "-")
            else:
                set_number(find_cell(sheet_root, row, MASTER_COLS["yoy_inc"]), yoy_inc_value)
                set_number_format(sheet_root, styles_root, row, MASTER_COLS["yoy_inc"], yoy_inc_num_fmt, num_fmt_cache)

            fill, font, _ = color_by_position(series.points, current_date)
            if fill:
                colored["current"] += 1
            set_style(sheet_root, styles_root, row, MASTER_COLS["current"], fill, font, style_cache)

            change_series = [(d, v * display_scale) for d, v in series.points if d <= current_date]
            fill, font, _ = color_by_abs_change(change_series, change_display, current_date)
            if fill:
                colored["change"] += 1
            set_style(sheet_root, styles_root, row, MASTER_COLS["change"], fill, font, style_cache)

            fill, font, _ = color_by_position(yoy_hist, current_date)
            if fill:
                colored["yoy"] += 1
            set_style(sheet_root, styles_root, row, MASTER_COLS["yoy"], fill, font, style_cache)

            fill, font, _ = color_by_abs_change(yoy_inc_hist, yoy_inc_value, current_date)
            if fill:
                colored["yoy_inc"] += 1
            set_style(sheet_root, styles_root, row, MASTER_COLS["yoy_inc"], fill, font, style_cache)
        else:
            for col in (MASTER_COLS["current"], MASTER_COLS["change"], MASTER_COLS["yoy"], MASTER_COLS["yoy_inc"]):
                set_style(sheet_root, styles_root, row, col, None, BLACK, style_cache)

        chart_info = chart_map.get(row)
        if chart_info:
            freq = ws.cell(row, MASTER_COLS["freq"]).value or source.freq
            series = get_series(code or source.code, full_series)
            count_base = len(series.points) if series is not None else len(source.trend or [])
            take = chart_take_count(name, freq, count_base)
            values = values_for_chart(code or source.code, source, full_series, take)
            if values:
                replacements[chart_info["chart_path"]] = update_chart_series(zin.read(chart_info["chart_path"]), values)
                charts.append({"row": row, "points": len(values)})
        else:
            values = []

        master_cache[row] = {
            "row": row,
            "name": as_text(name),
            "code": as_text(code or source.code),
            "current": current_display,
            "date": source.data_date,
            "change": change_display,
            "freq": ws.cell(row, MASTER_COLS["freq"]).value or source.freq,
            "chart_values": values,
        }
        updated.append({"row": row, "name": as_text(name), "code": as_text(code or source.code), "date": source.data_date})

    # Reapply macro metric colors after all number-format/style mutations.
    # Some source cells inherit a previously colored base style; a final pass
    # ensures the fill intensity is determined only by the current history.
    for row, item in master_cache.items():
        current_date = parse_date(item["date"])
        series = get_series(item["code"], full_series)
        row_text = " ".join(as_text(ws.cell(row, col).value) for col in (1, 2, 3, 5, 7))
        if current_date is None or series is None or not is_macro_row(item["code"], row_text, full_series):
            continue

        current_display = item["current"]
        change_display = item["change"]
        current_point = closest_point(series.points, current_date)
        display_scale = raw_display_scale(current_display, current_point[1]) if current_point else 1.0
        yoy_value, yoy_hist = yoy_value_and_history(series, current_date, current_display, item["name"])
        yoy_inc_value, yoy_inc_hist = yoy_increment_and_history(series, current_date, current_display)

        fill, font, _ = color_by_position(series.points, current_date)
        set_style(sheet_root, styles_root, row, MASTER_COLS["current"], fill, font, style_cache)

        change_series = [(d, v * display_scale) for d, v in series.points if d <= current_date]
        fill, font, _ = color_by_abs_change(change_series, change_display, current_date)
        set_style(sheet_root, styles_root, row, MASTER_COLS["change"], fill, font, style_cache)

        fill, font, _ = color_by_position(yoy_hist, current_date)
        set_style(sheet_root, styles_root, row, MASTER_COLS["yoy"], fill, font, style_cache)

        fill, font, _ = color_by_abs_change(yoy_inc_hist, yoy_inc_value, current_date)
        set_style(sheet_root, styles_root, row, MASTER_COLS["yoy_inc"], fill, font, style_cache)

    anchor_fix = {"charts": 0, "unique_rows": 0, "duplicates_removed": 0}
    if drawing_path and not preserve_layout:
        drawing_bytes, anchor_fix = normalize_chart_anchors(zin, drawing_path, MASTER_CHART_COL, MASTER_CHART_EXT)
        if drawing_bytes is not None:
            replacements[drawing_path] = drawing_bytes

    replacements[sheet_path] = write_xml(sheet_root)
    return {
        "sheet": sheet_name,
        "updated_rows": len(updated),
        "unmatched_rows": len(unmatched),
        "charts_updated": len(charts),
        "chart_anchor_fix": anchor_fix,
        "colored": colored,
        "protected_zero_rows": protected,
        "updated_sample": updated[:20],
        "unmatched_sample": unmatched[:20],
        "chart_sample": charts[:20],
    }


def sync_display_sheet(
    wb_values: openpyxl.Workbook,
    zin: ZipFile,
    sheet_name: str,
    daily_path: Path,
    as_of: date,
    full_series: dict[str, Series],
    replacements: dict[str, bytes],
    styles_root: Any,
    style_cache: dict[tuple[int, str, str], int],
    master_cache: dict[int, dict[str, Any]],
    preserve_layout: bool,
) -> dict[str, Any]:
    by_code, by_name, best_trend = v30.build_sources(daily_path, as_of)
    ws = wb_values[sheet_name]
    sheet_path = workbook_sheet_path(zin, sheet_name)
    sheet_root = read_xml(zin, sheet_path)
    drawing_path = sheet_drawing_path(zin, sheet_path)
    chart_map = chart_rows(zin, drawing_path) if drawing_path else {}
    master_index = build_master_index(master_cache)
    date_style_id = find_date_style_id(ws, DISPLAY_COLS["date"])

    updated = []
    unmatched = []
    charts = []
    colored = {"current": 0, "change": 0}
    num_fmt_cache: dict[tuple[int, str], int] = {}
    row_context = {col: "" for col in (6, 7, 8)}
    for row in range(6, ws.max_row + 1):
        if row % 40 == 0:
            print(f"display row {row}/{ws.max_row}", file=sys.stderr, flush=True)
        name = ws.cell(row, DISPLAY_COLS["name"]).value
        if is_blank(name):
            continue
        for col in row_context:
            value = as_text(ws.cell(row, col).value)
            if value:
                row_context[col] = value
        row_text = " ".join([*(row_context[col] for col in (6, 7, 8)), as_text(name)])
        source = None
        master_item = match_master(name, master_cache, master_index)
        if master_item is not None:
            current_display = master_item["current"]
            change_display = master_item["change"]
            source_code = master_item["code"]
            source_date = master_item["date"]
            chart_values = master_item.get("chart_values") or []
            current_num_fmt, change_num_fmt, _yoy_num_fmt, _yoy_inc_num_fmt = metric_number_formats(
                ws.cell(row, DISPLAY_COLS["current"]).number_format, source_code
            )
        else:
            override_code = display_override_code(name)
            if is_pe_like(name) and not override_code:
                unmatched.append({"row": row, "name": as_text(name), "reason": "no_pe_master_match"})
                continue
            source = v30.base.resolve_source(as_text(name), override_code or "", by_code, by_name)
            if source is not None and is_suspicious_display_source(name, source):
                for col in (DISPLAY_COLS["current"], DISPLAY_COLS["date"], DISPLAY_COLS["change"]):
                    set_inline_string(find_cell(sheet_root, row, col), "\u2014")
                for col in (DISPLAY_COLS["current"], DISPLAY_COLS["change"]):
                    set_style(sheet_root, styles_root, row, col, None, BLACK, style_cache)
                unmatched.append({"row": row, "name": as_text(name), "reason": "suspicious_source", "source": as_text(source.name)})
                continue
            if source is None:
                unmatched.append({"row": row, "name": as_text(name)})
                continue
            if getattr(source, "sheet", "") != "full_series":
                source = v30.base.with_best_trend(source, best_trend)
            current_format = ws.cell(row, DISPLAY_COLS["current"]).number_format
            current_num_fmt, change_num_fmt, _yoy_num_fmt, _yoy_inc_num_fmt = metric_number_formats(
                current_format, source.code
            )
            current_display = v30.scale_for_cell(source.current, current_num_fmt)
            change_value = source.change
            already_percent_decimal = False
            if v30.is_stock_index_row(name, source.code):
                current_number = to_number(source.current)
                previous_number = to_number(source.previous)
                if current_number is not None and previous_number not in (None, 0):
                    change_value = (current_number - previous_number) / previous_number
                    already_percent_decimal = True
            change_display = v30.scale_for_cell(change_value, change_num_fmt, already_percent_decimal)
            source_code = source.code
            source_date = source.data_date
            freq = ws.cell(row, DISPLAY_COLS["freq"]).value or source.freq
            series = get_series(source.code, full_series)
            count_base = len(series.points) if series is not None else len(source.trend or [])
            take = chart_take_count(name, freq, count_base)
            chart_values = values_for_chart(source.code, source, full_series, take)
        set_number(find_cell(sheet_root, row, DISPLAY_COLS["current"]), current_display)
        set_date(find_cell(sheet_root, row, DISPLAY_COLS["date"]), source_date, date_style_id)
        set_number(find_cell(sheet_root, row, DISPLAY_COLS["change"]), change_display)
        set_number_format(sheet_root, styles_root, row, DISPLAY_COLS["current"], current_num_fmt, num_fmt_cache)
        set_number_format(sheet_root, styles_root, row, DISPLAY_COLS["change"], change_num_fmt, num_fmt_cache)

        current_date = parse_date(source_date)
        series = get_series(source_code, full_series)
        macro_scope = current_date is not None and is_macro_row(source_code, row_text, full_series)
        if macro_scope and series is not None:
            current_point = closest_point(series.points, current_date)
            display_scale = raw_display_scale(current_display, current_point[1]) if current_point else 1.0
            fill, font, _ = color_by_position(series.points, current_date)
            if fill:
                colored["current"] += 1
            set_style(sheet_root, styles_root, row, DISPLAY_COLS["current"], fill, font, style_cache)

            change_series = [(d, v * display_scale) for d, v in series.points if d <= current_date]
            fill, font, _ = color_by_abs_change(change_series, change_display, current_date)
            if fill:
                colored["change"] += 1
            set_style(sheet_root, styles_root, row, DISPLAY_COLS["change"], fill, font, style_cache)
        else:
            for col in (DISPLAY_COLS["current"], DISPLAY_COLS["change"]):
                set_style(sheet_root, styles_root, row, col, None, BLACK, style_cache)

        chart_info = chart_map.get(row)
        if chart_info:
            if chart_values:
                replacements[chart_info["chart_path"]] = update_chart_series(zin.read(chart_info["chart_path"]), chart_values)
                charts.append({"row": row, "points": len(chart_values)})

        updated.append({"row": row, "name": as_text(name), "code": source_code, "date": source_date})

    anchor_fix = {"charts": 0, "unique_rows": 0, "duplicates_removed": 0}
    if drawing_path and not preserve_layout:
        drawing_bytes, anchor_fix = normalize_chart_anchors(zin, drawing_path, DISPLAY_CHART_COL)
        if drawing_bytes is not None:
            replacements[drawing_path] = drawing_bytes

    replacements[sheet_path] = write_xml(sheet_root)
    return {
        "sheet": sheet_name,
        "updated_rows": len(updated),
        "unmatched_rows": len(unmatched),
        "charts_updated": len(charts),
        "chart_anchor_fix": anchor_fix,
        "colored": colored,
        "updated_sample": updated[:20],
        "unmatched_sample": unmatched[:20],
        "chart_sample": charts[:20],
    }


def normalize_output_macro_colors(output_path: Path, full_series: dict[str, Series]) -> dict[str, int]:
    wb = openpyxl.load_workbook(output_path, data_only=True, read_only=False)
    master_sheet = (
        "重点策略跟踪情况"
        if "重点策略跟踪情况" in wb.sheetnames
        else find_sheet_name(wb, "(V3.0)")
    )
    ws = wb[master_sheet]
    replacements: dict[str, bytes] = {}
    style_cache: dict[tuple[int, str, str], int] = {}
    normalized = 0

    with ZipFile(output_path, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        sheet_path = workbook_sheet_path(zin, master_sheet)
        sheet_root = read_xml(zin, sheet_path)

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, MASTER_COLS["name"]).value
            code = ws.cell(row, MASTER_COLS["code"]).value
            current_date = parse_date(ws.cell(row, MASTER_COLS["date"]).value)
            series = get_series(code, full_series)
            row_text = " ".join(as_text(ws.cell(row, col).value) for col in (1, 2, 3, 5, 7))
            if current_date is None or series is None or not is_macro_row(code, row_text, full_series):
                continue

            current_display = ws.cell(row, MASTER_COLS["current"]).value
            change_display = ws.cell(row, MASTER_COLS["change"]).value
            current_point = closest_point(series.points, current_date)
            display_scale = raw_display_scale(current_display, current_point[1]) if current_point else 1.0
            yoy_value, yoy_hist = yoy_value_and_history(series, current_date, current_display, name)
            yoy_inc_value, yoy_inc_hist = yoy_increment_and_history(series, current_date, current_display)

            fill, font, _ = color_by_position(series.points, current_date)
            set_style(sheet_root, styles_root, row, MASTER_COLS["current"], fill, font, style_cache)

            change_series = [(d, v * display_scale) for d, v in series.points if d <= current_date]
            fill, font, _ = color_by_abs_change(change_series, change_display, current_date)
            set_style(sheet_root, styles_root, row, MASTER_COLS["change"], fill, font, style_cache)

            fill, font, _ = color_by_position(yoy_hist, current_date)
            set_style(sheet_root, styles_root, row, MASTER_COLS["yoy"], fill, font, style_cache)

            fill, font, _ = color_by_abs_change(yoy_inc_hist, yoy_inc_value, current_date)
            set_style(sheet_root, styles_root, row, MASTER_COLS["yoy_inc"], fill, font, style_cache)
            normalized += 1

        replacements["xl/styles.xml"] = write_xml(styles_root)
        replacements[sheet_path] = write_xml(sheet_root)
        temp_path = output_path.with_name(output_path.stem + ".color-normalized.xlsx")
        with ZipFile(temp_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    wb.close()
    temp_path.replace(output_path)
    return {"rows": normalized}


def run(
    tree_path: Path,
    daily_path: Path,
    output_path: Path,
    as_of: date,
    preserve_layout: bool = False,
    write_canonical_codes: bool = False,
) -> dict[str, Any]:
    print("reading full series", file=sys.stderr, flush=True)
    full_series = read_full_series(daily_path, as_of)
    print(f"full series: {len(full_series)}", file=sys.stderr, flush=True)
    print("opening tree workbook", file=sys.stderr, flush=True)
    wb_values = openpyxl.load_workbook(tree_path, data_only=False, read_only=True)
    display_sheet = (
        "重点策略跟踪情况(V3)"
        if "重点策略跟踪情况(V3)" in wb_values.sheetnames
        else find_sheet_name(wb_values, "(V3)")
    )
    master_sheet = (
        "重点策略跟踪情况"
        if "重点策略跟踪情况" in wb_values.sheetnames
        else find_sheet_name(wb_values, "(V3.0)")
    )
    replacements: dict[str, bytes] = {}
    style_cache: dict[tuple[int, str, str], int] = {}
    master_cache: dict[int, dict[str, Any]] = {}
    with ZipFile(tree_path, "r") as zin:
        styles_root = read_xml(zin, "xl/styles.xml")
        print("syncing master", file=sys.stderr, flush=True)
        master_result = sync_master_sheet(
            wb_values,
            zin,
            master_sheet,
            daily_path,
            as_of,
            full_series,
            replacements,
            styles_root,
            style_cache,
            master_cache,
            preserve_layout,
            write_canonical_codes,
        )
        print("syncing display", file=sys.stderr, flush=True)
        display_result = sync_display_sheet(
            wb_values,
            zin,
            display_sheet,
            daily_path,
            as_of,
            full_series,
            replacements,
            styles_root,
            style_cache,
            master_cache,
            preserve_layout,
        )
        replacements["xl/styles.xml"] = write_xml(styles_root)
        print("writing output", file=sys.stderr, flush=True)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))
    wb_values.close()
    color_normalization = normalize_output_macro_colors(output_path, full_series)
    return {
        "input_tree": str(tree_path),
        "daily": str(daily_path),
        "output": str(output_path),
        "as_of": as_of.isoformat(),
        "preserve_layout": preserve_layout,
        "write_canonical_codes": write_canonical_codes,
        "full_series_count": len(full_series),
        "color_normalization": color_normalization,
        "master": master_result,
        "display": display_result,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Update latest V3/V3.0 TREE layout from refreshed daily workbook.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default=date.today().isoformat())
    parser.add_argument("--preserve-layout", action="store_true")
    parser.add_argument("--write-canonical-codes", action="store_true")
    args = parser.parse_args()
    result = run(
        Path(args.tree),
        Path(args.daily),
        Path(args.output),
        datetime.strptime(args.as_of, "%Y-%m-%d").date(),
        preserve_layout=args.preserve_layout,
        write_canonical_codes=args.write_canonical_codes,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
