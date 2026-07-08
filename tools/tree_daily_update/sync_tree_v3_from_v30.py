from __future__ import annotations

import argparse
import copy
import json
import math
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from sync_tree import GREEN, RED, date_text, ensure_font_style
from sync_tree_new_indicators_from_daily import (
    NS,
    chart_rows,
    find_cell,
    read_xml,
    set_date,
    set_inline_string,
    set_number,
    sheet_drawing_path,
    update_chart_series,
    workbook_sheet_path,
    write_xml,
)


DISPLAY_SHEET = "重点策略跟踪情况(V3)"
MASTER_SHEET = "重点策略跟踪情况(V3.0)"

DISPLAY_HEADER_ROW = 5
DISPLAY_NAME_COL = 9
DISPLAY_FREQ_COL = 12
DISPLAY_CURRENT_COL = 15
DISPLAY_DATE_COL = 16
DISPLAY_CHANGE_COL = 17

MASTER_NAME_COL = 4
MASTER_FREQ_COL = 9
MASTER_CODE_COL = 10
MASTER_CURRENT_COL = 11
MASTER_DATE_COL = 12
MASTER_CHANGE_COL = 13


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = as_text(value).replace(",", "").replace("%", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def is_blank(value: Any) -> bool:
    return as_text(value) in {"", "-", "—", "None"}


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def strip_parentheses(text: str) -> str:
    text = re.sub(r"（[^）]*）", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    return text


def norm_name(value: Any, *, remove_parentheses: bool = False) -> str:
    text = as_text(value)
    if remove_parentheses:
        text = strip_parentheses(text)
    text = text.replace("\n", "")
    text = re.sub(r"^[①②③④⑤⑥⑦⑧⑨⑩]+", "", text)
    text = re.sub(r"[①②③④⑤⑥⑦⑧⑨⑩：:（）()\[\]【】\s、/，,。；;\-+]", "", text)
    for token in ["同比多增", "亿港元", "亿元", "单位", "担保隔夜融资利率"]:
        text = text.replace(token, "")
    return text.lower()


def keys_for_name(value: Any) -> list[str]:
    keys: list[str] = []
    for remove_parentheses in (False, True):
        key = norm_name(value, remove_parentheses=remove_parentheses)
        if key and key not in keys:
            keys.append(key)
    return keys


def is_percent_format(number_format: Any) -> bool:
    return "%" in as_text(number_format)


def convert_between_formats(value: Any, source_format: Any, target_format: Any) -> Any:
    number = to_number(value)
    if number is None:
        return value
    source_is_percent = is_percent_format(source_format)
    target_is_percent = is_percent_format(target_format)
    if source_is_percent and not target_is_percent:
        return number * 100
    if target_is_percent and not source_is_percent:
        return number / 100
    return number


def chart_values(chart_xml: bytes) -> list[float]:
    root = ET.fromstring(chart_xml)
    values: list[float] = []
    for pt in root.findall(".//c:lineChart/c:ser/c:val/c:numRef/c:numCache/c:pt", NS):
        value_el = pt.find("c:v", NS)
        if value_el is None:
            continue
        number = to_number(value_el.text)
        if number is not None:
            values.append(number)
    return values


def master_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[int, dict[str, Any]]:
    rows: dict[int, dict[str, Any]] = {}
    for row in range(6, ws.max_row + 1):
        name = ws.cell(row, MASTER_NAME_COL).value
        if is_blank(name):
            continue
        rows[row] = {
            "row": row,
            "name": as_text(name),
            "freq": ws.cell(row, MASTER_FREQ_COL).value,
            "code": ws.cell(row, MASTER_CODE_COL).value,
            "current": ws.cell(row, MASTER_CURRENT_COL).value,
            "date": ws.cell(row, MASTER_DATE_COL).value,
            "change": ws.cell(row, MASTER_CHANGE_COL).value,
            "current_format": ws.cell(row, MASTER_CURRENT_COL).number_format,
            "change_format": ws.cell(row, MASTER_CHANGE_COL).number_format,
        }
    return rows


def build_master_index(master: dict[int, dict[str, Any]]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = defaultdict(list)
    for row, item in master.items():
        for key in keys_for_name(item["name"]):
            if len(key) >= 2:
                index[key].append(row)
    for key in index:
        index[key].sort()
    return index


def match_master_row(display_name: Any, master_index: dict[str, list[int]], usage: dict[str, int]) -> tuple[int | None, str | None]:
    keys = keys_for_name(display_name)
    for key in keys:
        rows = master_index.get(key)
        if not rows:
            continue
        ordinal = usage[key]
        usage[key] += 1
        return rows[min(ordinal, len(rows) - 1)], key

    manual_targets = {
        norm_name("财政存款变动", remove_parentheses=True): norm_name("财政存款", remove_parentheses=True),
        norm_name(" M1（贷款的钱是在M1还是M2）", remove_parentheses=True): norm_name("M1", remove_parentheses=True),
        norm_name("M2（怎么辨别市场流动性的占比，多少是央行流出来的，多少是贷款倍数后的数）", remove_parentheses=True): norm_name("M2", remove_parentheses=True),
    }
    for key in keys:
        target = manual_targets.get(key)
        if target and target in master_index:
            rows = master_index[target]
            ordinal = usage[target]
            usage[target] += 1
            return rows[min(ordinal, len(rows) - 1)], target
    return None, None


def set_dash(cell: ET.Element) -> None:
    set_inline_string(cell, "—")


def previous_dates(previous_tree: Path | None) -> dict[int, str]:
    if previous_tree is None or not previous_tree.exists():
        return {}
    wb = openpyxl.load_workbook(previous_tree, data_only=True, read_only=True)
    if DISPLAY_SHEET not in wb.sheetnames:
        return {}
    ws = wb[DISPLAY_SHEET]
    return {row: date_text(ws.cell(row, DISPLAY_DATE_COL).value) for row in range(DISPLAY_HEADER_ROW + 1, ws.max_row + 1)}


def sync_display_from_master(input_tree: Path, output_tree: Path, previous_tree: Path | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_tree, data_only=True, read_only=True)
    wb_formats = openpyxl.load_workbook(input_tree, data_only=False, read_only=True)
    display_ws = wb[DISPLAY_SHEET]
    display_formats = wb_formats[DISPLAY_SHEET]
    master_ws = wb_formats[MASTER_SHEET]
    master = master_rows(master_ws)
    master_index = build_master_index(master)
    usage: dict[str, int] = defaultdict(int)
    old_dates = previous_dates(previous_tree)

    updated_rows = []
    unmatched_rows = []
    blanked_rows = []
    charts_updated = []
    red_rows = []
    green_rows = []

    with ZipFile(input_tree, "r") as zin:
        display_sheet_path = workbook_sheet_path(zin, DISPLAY_SHEET)
        master_sheet_path = workbook_sheet_path(zin, MASTER_SHEET)
        display_root = read_xml(zin, display_sheet_path)
        styles_root = read_xml(zin, "xl/styles.xml")

        display_drawing = sheet_drawing_path(zin, display_sheet_path)
        master_drawing = sheet_drawing_path(zin, master_sheet_path)
        display_chart_map = chart_rows(zin, display_drawing) if display_drawing else {}
        master_chart_map = chart_rows(zin, master_drawing) if master_drawing else {}

        style_cache: dict[tuple[int, str | None], int] = {}
        for row in range(DISPLAY_HEADER_ROW + 1, display_ws.max_row + 1):
            change_cell = find_cell(display_root, row, DISPLAY_CHANGE_COL)
            base_style = int(change_cell.attrib.get("s", "0"))
            change_cell.attrib["s"] = str(ensure_font_style(styles_root, base_style, None, style_cache))

        for row in range(DISPLAY_HEADER_ROW + 1, display_ws.max_row + 1):
            display_name = display_ws.cell(row, DISPLAY_NAME_COL).value
            if is_blank(display_name):
                continue

            master_row, match_key = match_master_row(display_name, master_index, usage)
            if master_row is None:
                unmatched_rows.append({"row": row, "name": as_text(display_name)})
                continue

            item = master[master_row]
            if is_blank(item["current"]) and is_blank(item["date"]) and is_blank(item["change"]):
                set_dash(find_cell(display_root, row, DISPLAY_CURRENT_COL))
                set_dash(find_cell(display_root, row, DISPLAY_DATE_COL))
                set_dash(find_cell(display_root, row, DISPLAY_CHANGE_COL))
                blanked_rows.append({"row": row, "name": as_text(display_name), "master_row": master_row})
            else:
                target_current_format = display_formats.cell(row, DISPLAY_CURRENT_COL).number_format
                target_change_format = display_formats.cell(row, DISPLAY_CHANGE_COL).number_format
                current_value = convert_between_formats(item["current"], item["current_format"], target_current_format)
                change_value = convert_between_formats(item["change"], item["change_format"], target_change_format)

                set_number(find_cell(display_root, row, DISPLAY_CURRENT_COL), current_value)
                set_date(find_cell(display_root, row, DISPLAY_DATE_COL), item["date"])
                set_number(find_cell(display_root, row, DISPLAY_CHANGE_COL), change_value)

                prev_date = old_dates.get(row, "")
                curr_date = date_text(item["date"])
                change_number = to_number(change_value)
                if previous_tree is not None and curr_date and curr_date != prev_date and change_number is not None:
                    color = RED if change_number > 0 else GREEN if change_number < 0 else None
                    if color:
                        change_cell = find_cell(display_root, row, DISPLAY_CHANGE_COL)
                        base_style = int(change_cell.attrib.get("s", "0"))
                        change_cell.attrib["s"] = str(ensure_font_style(styles_root, base_style, color, style_cache))
                        if color == RED:
                            red_rows.append(row)
                        else:
                            green_rows.append(row)

                updated_rows.append(
                    {
                        "row": row,
                        "name": as_text(display_name),
                        "master_row": master_row,
                        "master_name": item["name"],
                        "code": item["code"],
                        "date": item["date"],
                        "current": current_value,
                        "change": change_value,
                        "match_key": match_key,
                    }
                )

            display_chart = display_chart_map.get(row)
            master_chart = master_chart_map.get(master_row)
            if display_chart and master_chart:
                values = chart_values(zin.read(master_chart["chart_path"]))
                if values:
                    chart_xml = update_chart_series(zin.read(display_chart["chart_path"]), values)
                    charts_updated.append({"row": row, "master_row": master_row, "points": len(values)})
                else:
                    chart_xml = None
                if chart_xml is not None:
                    master.setdefault("_chart_replacements", {})[display_chart["chart_path"]] = chart_xml

        replacements: dict[str, bytes] = {
            display_sheet_path: write_xml(display_root),
            "xl/styles.xml": write_xml(styles_root),
        }
        for path, chart_xml in master.get("_chart_replacements", {}).items():
            replacements[path] = chart_xml

        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))

    return {
        "input": str(input_tree),
        "output": str(output_tree),
        "previous_tree": str(previous_tree) if previous_tree else None,
        "updated_row_count": len(updated_rows),
        "updated_rows_sample": updated_rows[:35],
        "blanked_row_count": len(blanked_rows),
        "blanked_rows_sample": blanked_rows[:20],
        "unmatched_count": len(unmatched_rows),
        "unmatched_sample": unmatched_rows[:35],
        "charts_updated_count": len(charts_updated),
        "charts_updated_sample": charts_updated[:20],
        "red_change_rows": red_rows,
        "green_change_rows": green_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Synchronize TREE V3 display sheet from the already-refreshed V3.0 master sheet.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--previous-tree")
    args = parser.parse_args()
    result = sync_display_from_master(
        Path(args.input),
        Path(args.output),
        Path(args.previous_tree) if args.previous_tree else None,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
