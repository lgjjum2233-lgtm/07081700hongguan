from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


MACRO_SHEET = "宏观数据"
START_DATE = datetime(2024, 1, 1)
NEW_WSD_COLS = 950
DATE_FIRST_COL = 6
DERIVED_SPREAD_ROW = 179
DERIVED_FISCAL_ROW = 180


def formula_text(value: Any) -> str:
    return value.text if isinstance(value, ArrayFormula) else str(value or "")


def set_formula_text(cell, text: str) -> None:
    if isinstance(cell.value, ArrayFormula):
        cell.value.text = text
        cell.value.ref = cell.coordinate
    else:
        cell.value = text


def update_wsd_formula(ws, new_cols: int) -> str:
    cell = ws["F2"]
    text = formula_text(cell.value)
    if not text:
        raise RuntimeError("宏观数据!F2 没有找到 Wind WSD 公式")
    text = re.sub(r"cols=\d+", f"cols={new_cols}", text)
    # Preserve the current raw-code row range. Added derived rows are outside WSD.
    match = re.search(r"E2:E(\d+)", text)
    if not match:
        raise RuntimeError("无法从宏观数据!F2 解析 E2:E 行范围")
    raw_end = int(match.group(1))
    text = re.sub(r"rows=\d+", f"rows={raw_end - 1}", text)
    set_formula_text(cell, text)
    return text


def update_macro_range_references(wb, macro_name: str, new_last_col: str) -> int:
    # Expand formulas that look at 宏观数据!$F$...:$TN$... / :$SZ$... / etc.
    escaped = re.escape(macro_name)
    pattern = re.compile(
        rf"({escaped}!\$F\$(\d+):\$)([A-Z]+)(\$(\d+))"
    )
    replaced = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=") and f"{macro_name}!$F$" in value):
                    continue

                def repl(match: re.Match[str]) -> str:
                    start_row = match.group(2)
                    end_row = match.group(5)
                    # Only expand same-row horizontal series, including date row.
                    if start_row != end_row:
                        return match.group(0)
                    return f"{match.group(1)}{new_last_col}{match.group(4)}"

                new_value = pattern.sub(repl, value)
                if new_value != value:
                    cell.value = new_value
                    replaced += 1
    return replaced


def fill_derived_rows(ws, new_last_col: int) -> None:
    for col in range(DATE_FIRST_COL, new_last_col + 1):
        letter = get_column_letter(col)
        ws.cell(DERIVED_SPREAD_ROW, col).value = (
            f'=IFERROR(IF(OR({letter}156="",{letter}58=""),"",{letter}156-{letter}58),"")'
        )
        ws.cell(DERIVED_FISCAL_ROW, col).value = (
            f'=IFERROR(IF(OR({letter}144="",{letter}145=""),"",'
            f'({letter}144-{letter}145)/'
            f'LOOKUP(2,1/((YEAR($F$1:{letter}$1)=YEAR({letter}$1)-1)*'
            f'($F$146:{letter}$146<>"")),$F$146:{letter}$146)*100),"")'
        )


def copy_column_dimensions(ws, source_col: int, last_col: int) -> None:
    src_letter = get_column_letter(source_col)
    src_dim = ws.column_dimensions[src_letter]
    for col in range(source_col + 1, last_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = src_dim.width
        ws.column_dimensions[letter].hidden = src_dim.hidden
        ws.column_dimensions[letter].outlineLevel = src_dim.outlineLevel


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--cols", type=int, default=NEW_WSD_COLS)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    wb = openpyxl.load_workbook(input_path, data_only=False, read_only=False)
    ws = wb[MACRO_SHEET]

    ws["A2"].value = START_DATE
    wsd_formula = update_wsd_formula(ws, args.cols)
    new_last_col_idx = DATE_FIRST_COL + args.cols - 1
    new_last_col = get_column_letter(new_last_col_idx)

    replaced = update_macro_range_references(wb, MACRO_SHEET, new_last_col)
    fill_derived_rows(ws, new_last_col_idx)
    copy_column_dimensions(ws, ws.max_column, new_last_col_idx)

    wb.save(output_path)
    wb.close()
    print(
        {
            "output": str(output_path),
            "start_date": START_DATE.strftime("%Y-%m-%d"),
            "wsd_cols": args.cols,
            "new_last_col": new_last_col,
            "formula_replacements": replaced,
            "wsd_formula": wsd_formula,
        }
    )


if __name__ == "__main__":
    main()
