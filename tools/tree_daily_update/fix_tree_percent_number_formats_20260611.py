from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import get_column_letter

import add_tree_yoy_column_20260611 as yoy_base


TREE_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5
CODE_COL = 10
CURRENT_COL = 11
DATE_COL = 12
YOY_COL = 13
YOY_INCREMENT_COL = 14
QOQ_COL = 15
CALIBER_COL = 17

PCT_FORMAT = "0.00%"
LITERAL_PCT_FORMAT = '0.00"%"'
GENERAL_FORMAT = "0.00"


@dataclass(frozen=True)
class FormatDecision:
    is_percent: bool
    number_format: str
    reason: str
    scale: float | None = None
    source_current: float | None = None


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    return yoy_base.to_number(value)


def parse_date(value: Any) -> date | None:
    return yoy_base.parse_date(value)


def norm_text(*parts: Any) -> str:
    return re.sub(r"\s+", "", " ".join(as_text(part) for part in parts)).lower()


def is_percent_metric(name: str, caliber: str) -> tuple[bool, str]:
    text = norm_text(name, caliber)
    non_percent_keys = (
        "pmi",
        "扩散指数",
        "指数水平",
        "价格指数",
        "估值倍数",
        "市盈率",
        "pe",
        "余额",
        "资产规模",
        "名义现值",
        "累计值",
        "当月值",
        "净投放",
        "净流入",
        "差额",
        "贸易差额",
        "点",
    )
    strong_percent_keys = (
        "同比",
        "环比",
        "累计同比",
        "当月同比",
        "增速",
        "比例",
        "利率",
        "收益率",
        "利差",
        "杠杆率",
        "赤字率",
        "贡献率",
        "换手率",
        "利润率",
        "roe",
        "roa",
        "cpi",
        "ppi",
        "pce",
        "shibor",
        "lpr",
        "sofr",
        "iorb",
        "effr",
    )
    for key in strong_percent_keys:
        if key in text:
            return True, f"命中百分/比例关键词：{key}"
    for key in non_percent_keys:
        if key in text:
            return False, f"命中非百分关键词：{key}"
    return False, "未命中百分/比例关键词"


def source_current_value(
    series: yoy_base.MacroSeries | None,
    current_date: Any,
) -> float | None:
    if series is None:
        return None
    current_dt = parse_date(current_date)
    if current_dt is None:
        return None
    point = yoy_base.closest_current_period(series.points, current_dt)
    return point[1] if point is not None else None


def detect_scale(display_value: Any, raw_value: float | None, name: str, caliber: str) -> tuple[float | None, str]:
    display = to_number(display_value)
    if display is None or raw_value is None or abs(raw_value) <= 1e-12:
        text = norm_text(name, caliber)
        if any(key in text for key in ("利率", "收益率", "shibor", "lpr")) and display is not None and abs(display) > 0.5:
            return 1.0, "无法匹配底层值，按利率/收益率百分数值显示"
        return None, "无法用底层值判断缩放"
    ratio = display / raw_value
    candidates = [1.0, 0.01, 0.0001, 100.0]
    best = min(candidates, key=lambda item: abs(math.log(abs(ratio / item))) if ratio and item else float("inf"))
    if abs(ratio / best - 1.0) <= 0.25:
        return best, f"TREE值/日报原值≈{best:g}"
    return None, f"TREE值/日报原值={ratio:g}，未匹配常见缩放"


def percent_format_for_metric(
    name: str,
    caliber: str,
    display_value: Any,
    raw_value: float | None,
) -> tuple[str, float | None, str]:
    scale, reason = detect_scale(display_value, raw_value, name, caliber)
    if scale is not None and abs(scale - 1.0) <= 1e-9:
        return LITERAL_PCT_FORMAT, scale, reason
    return PCT_FORMAT, scale, reason


def iter_tree_rows(ws) -> list[dict[str, Any]]:
    rows = []
    big = dim = sub = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row, 1).value:
            big = as_text(ws.cell(row, 1).value)
            dim = ""
            sub = ""
        if ws.cell(row, 2).value:
            dim = as_text(ws.cell(row, 2).value)
            sub = ""
        if ws.cell(row, 3).value:
            sub = as_text(ws.cell(row, 3).value)
        name = as_text(ws.cell(row, 4).value)
        code = yoy_base.norm_code(ws.cell(row, CODE_COL).value)
        if not name and not code:
            continue
        rows.append(
            {
                "row": row,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "code": code,
                "current_date": ws.cell(row, DATE_COL).value,
                "caliber": as_text(ws.cell(row, CALIBER_COL).value),
            }
        )
    return rows


def is_numeric_cell(cell) -> bool:
    return isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)


def set_cell_format(cell, fmt: str) -> None:
    cell.number_format = fmt


def write_audit(path: Path, audit_rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "百分号格式复核"
    headers = [
        "TREE行号",
        "指标名称",
        "指标代码",
        "数据口径",
        "是否百分指标",
        "日报当前值",
        "缩放判断",
        "K当前数据格式",
        "M同比格式",
        "N同比增量格式",
        "O环比格式",
        "备注",
    ]
    ws.append(headers)
    for item in audit_rows:
        ws.append([item.get(header, "") for header in headers])
    for idx, width in enumerate([10, 34, 22, 16, 12, 14, 28, 16, 16, 18, 16, 48], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def fix_formats(input_tree: Path, daily_path: Path, output_tree: Path, audit_output: Path) -> dict[str, Any]:
    macro_series = yoy_base.read_macro_series(daily_path)
    wb = openpyxl.load_workbook(input_tree, data_only=False, read_only=False)
    ws = wb[TREE_SHEET]
    rows = iter_tree_rows(ws)

    changed = 0
    pct_metric_rows = 0
    audit_rows = []
    for item in rows:
        row = item["row"]
        name = item["name"]
        caliber = item["caliber"]
        series = macro_series.get(item["code"])
        raw_current = source_current_value(series, item["current_date"])
        is_pct, pct_reason = is_percent_metric(name, caliber)
        if is_pct:
            pct_metric_rows += 1
        metric_fmt, scale, scale_reason = (
            percent_format_for_metric(name, caliber, ws.cell(row, CURRENT_COL).value, raw_current)
            if is_pct
            else (GENERAL_FORMAT, None, pct_reason)
        )

        before = {
            "K": ws.cell(row, CURRENT_COL).number_format,
            "M": ws.cell(row, YOY_COL).number_format,
            "N": ws.cell(row, YOY_INCREMENT_COL).number_format,
            "O": ws.cell(row, QOQ_COL).number_format,
        }

        if is_numeric_cell(ws.cell(row, CURRENT_COL)):
            set_cell_format(ws.cell(row, CURRENT_COL), metric_fmt if is_pct else GENERAL_FORMAT)

        if is_numeric_cell(ws.cell(row, YOY_COL)):
            # M is the year-over-year change rate column, so numeric values are ratios.
            set_cell_format(ws.cell(row, YOY_COL), PCT_FORMAT)
        else:
            set_cell_format(ws.cell(row, YOY_COL), GENERAL_FORMAT)

        if is_numeric_cell(ws.cell(row, YOY_INCREMENT_COL)):
            set_cell_format(ws.cell(row, YOY_INCREMENT_COL), metric_fmt if is_pct else GENERAL_FORMAT)
        else:
            set_cell_format(ws.cell(row, YOY_INCREMENT_COL), GENERAL_FORMAT)

        if is_numeric_cell(ws.cell(row, QOQ_COL)):
            # O follows the original marginal/period change unit. If it was already a rate
            # change, keep percent formatting; otherwise use the metric unit.
            old_o = before["O"]
            if "%" in old_o and not is_pct:
                set_cell_format(ws.cell(row, QOQ_COL), PCT_FORMAT)
            else:
                set_cell_format(ws.cell(row, QOQ_COL), metric_fmt if is_pct else GENERAL_FORMAT)
        else:
            set_cell_format(ws.cell(row, QOQ_COL), GENERAL_FORMAT)

        after = {
            "K": ws.cell(row, CURRENT_COL).number_format,
            "M": ws.cell(row, YOY_COL).number_format,
            "N": ws.cell(row, YOY_INCREMENT_COL).number_format,
            "O": ws.cell(row, QOQ_COL).number_format,
        }
        changed += sum(1 for key in before if before[key] != after[key])

        audit_rows.append(
            {
                "TREE行号": row,
                "指标名称": name,
                "指标代码": item["code"],
                "数据口径": caliber,
                "是否百分指标": "是" if is_pct else "否",
                "日报当前值": raw_current if raw_current is not None else "",
                "缩放判断": scale_reason,
                "K当前数据格式": after["K"],
                "M同比格式": after["M"],
                "N同比增量格式": after["N"],
                "O环比格式": after["O"],
                "备注": pct_reason,
            }
        )

    wb.save(output_tree)
    wb.close()
    write_audit(audit_output, audit_rows)
    return {
        "input_tree": str(input_tree),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        "rows": len(rows),
        "percent_metric_rows": pct_metric_rows,
        "format_cells_changed": changed,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()
    result = fix_formats(Path(args.tree), Path(args.daily), Path(args.output_tree), Path(args.audit_output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
