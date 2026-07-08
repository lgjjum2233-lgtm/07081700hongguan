from __future__ import annotations

import argparse
import json
import math
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

for prefix, uri in [("", NS["main"]), ("r", NS["officeRel"])]:
    ET.register_namespace(prefix, uri)


PE_CODES = {
    "000001.SH",
    "399001.SZ",
    "000300.SH",
    "399006.SZ",
    "HSI.HI",
    "HSTECH.HI",
}


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_code(value: Any) -> str:
    return as_text(value).upper()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = as_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def date_str(value: date | None) -> str:
    return value.isoformat() if value else ""


def excel_serial(value: date) -> float:
    return float(to_excel(datetime(value.year, value.month, value.day)))


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def write_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    row_el = None
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            row_el = row
            break
    if row_el is None:
        row_el = ET.Element(q("row"), {"r": str(row_num)})
        sheet_data.append(row_el)
    ref = f"{get_column_letter(col_num)}{row_num}"
    for cell in row_el.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    cell = ET.Element(q("c"), {"r": ref})
    row_el.append(cell)
    return cell


def format_number(value: float) -> str:
    return format(float(value), ".12g")


def set_formula_and_cache(cell: ET.Element, formula: str, value: float | date | None) -> None:
    cell.attrib.pop("t", None)
    formula_el = cell.find("main:f", NS)
    if formula_el is None:
        formula_el = ET.Element(q("f"))
        cell.insert(0, formula_el)
    formula_el.text = formula.lstrip("=")
    for child in list(cell):
        if child.tag in {q("v"), q("is")}:
            cell.remove(child)
    v = ET.Element(q("v"))
    if isinstance(value, date):
        v.text = format_number(excel_serial(value))
    elif value is None:
        v.text = ""
    else:
        v.text = format_number(float(value))
    children = list(cell)
    insert_at = len(children)
    for idx, child in enumerate(children):
        if child.tag == q("f"):
            insert_at = idx + 1
            break
    cell.insert(insert_at, v)


def header_row_for(ws, row: int) -> int:
    for header in range(row, 0, -1):
        marker = as_text(ws.cell(header, 5).value).lower()
        if marker in {"date", "日期"}:
            return header
    return 1


def row_points(ws, row: int, header: int) -> list[tuple[date, float]]:
    points: list[tuple[date, float]] = []
    for col in range(6, ws.max_column + 1):
        d = parse_date(ws.cell(header, col).value)
        v = to_number(ws.cell(row, col).value)
        if d is not None and v is not None:
            points.append((d, v))
    points.sort(key=lambda item: item[0])
    return points


def discover_index_rows(ws) -> tuple[dict[str, int], dict[str, int]]:
    price_rows: dict[str, int] = {}
    pe_rows: dict[str, int] = {}
    current_tag = ""
    for row in range(1, ws.max_row + 1):
        tag = as_text(ws.cell(row, 3).value).lower()
        if tag:
            current_tag = tag
        code = norm_code(ws.cell(row, 5).value)
        if code not in PE_CODES:
            continue
        if current_tag == "close3" and code not in price_rows:
            price_rows[code] = row
        if current_tag == "pe_ttm" and code not in pe_rows:
            pe_rows[code] = row
    return price_rows, pe_rows


def discover_front_rows(ws) -> tuple[dict[str, int], dict[str, int]]:
    price_rows: dict[str, int] = {}
    pe_rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        name = as_text(ws.cell(row, 1).value)
        code = norm_code(ws.cell(row, 3).value)
        if code not in PE_CODES:
            continue
        if "PE(TTM)" in name:
            pe_rows[code] = row
        elif code not in price_rows:
            price_rows[code] = row
    return price_rows, pe_rows


def select_current(points: list[tuple[date, float]], cutoff: date | None) -> tuple[date, float] | None:
    candidates = [point for point in points if cutoff is None or point[0] <= cutoff]
    return candidates[-1] if candidates else None


def select_previous(points: list[tuple[date, float]], current_date: date) -> tuple[date, float] | None:
    candidates = [point for point in points if point[0] < current_date]
    return candidates[-1] if candidates else None


def update_daily_pe_front(input_path: Path, output_path: Path) -> dict[str, Any]:
    wb_values = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    wb_formula = openpyxl.load_workbook(input_path, data_only=False, read_only=False)
    front_name = "A股港股"
    index_name = "指数走势"
    front_values = wb_values[front_name]
    index_values = wb_values[index_name]
    front_formula = wb_formula[front_name]
    price_index_rows, pe_index_rows = discover_index_rows(index_values)
    price_front_rows, pe_front_rows = discover_front_rows(front_values)
    max_col = index_values.max_column
    updates: dict[str, dict[str, Any]] = {}
    for code in sorted(PE_CODES):
        price_row = price_index_rows.get(code)
        pe_row = pe_index_rows.get(code)
        front_row = pe_front_rows.get(code)
        front_price_row = price_front_rows.get(code)
        if not price_row or not pe_row or not front_row or not front_price_row:
            continue
        price_points = row_points(index_values, price_row, header_row_for(index_values, price_row))
        price_current = price_points[-1] if price_points else None
        price_date = price_current[0] if price_current else parse_date(front_values.cell(front_price_row, 6).value)
        pe_header = header_row_for(index_values, pe_row)
        pe_points = row_points(index_values, pe_row, pe_header)
        current = select_current(pe_points, price_date)
        if current is None:
            continue
        previous = select_previous(pe_points, current[0])
        current_date, current_value = current
        previous_date, previous_value = previous if previous else (None, None)
        change = current_value - previous_value if previous_value is not None else None
        data_rng = f"{index_name}!$F${pe_row}:${get_column_letter(max_col)}${pe_row}"
        date_rng = f"{index_name}!$F${pe_header}:${get_column_letter(max_col)}${pe_header}"
        updates[code] = {
            "front_row": front_row,
            "index_row": pe_row,
            "price_row": front_price_row,
            "price_date": price_date,
            "current": current_value,
            "current_date": current_date,
            "change": change,
            "previous": previous_value,
            "previous_date": previous_date,
            "formulas": {
                5: f'IFERROR(LOOKUP(2,1/({date_rng}<=$F${front_price_row})/({data_rng}<>""),{data_rng}),"")',
                6: f'IFERROR(LOOKUP(2,1/({date_rng}<=$F${front_price_row})/({data_rng}<>""),{date_rng}),"")',
                7: f'IFERROR(E{front_row}-H{front_row},"")',
                8: f'IFERROR(LOOKUP(2,1/({date_rng}<F{front_row})/({data_rng}<>""),{data_rng}),"")',
                9: f'IFERROR(LOOKUP(2,1/({date_rng}<F{front_row})/({data_rng}<>""),{date_rng}),"")',
            },
            "cache_values": {
                5: current_value,
                6: current_date,
                7: change,
                8: previous_value,
                9: previous_date,
            },
            "label": front_values.cell(front_row, 1).value,
        }
    wb_values.close()
    wb_formula.close()

    with ZipFile(input_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, front_name)
        sheet_root = read_xml(zin, sheet_path)
        for item in updates.values():
            row = item["front_row"]
            for col, formula in item["formulas"].items():
                cell = find_cell(sheet_root, row, col)
                set_formula_and_cache(cell, formula, item["cache_values"][col])
        sheet_xml = write_xml(sheet_root)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, sheet_xml)
                else:
                    zout.writestr(name, zin.read(name))

    return {
        "input": str(input_path),
        "output": str(output_path),
        "updated_count": len(updates),
        "updates": [
            {
                "code": code,
                "label": item["label"],
                "price_date": date_str(item["price_date"]),
                "pe_date": date_str(item["current_date"]),
                "pe": item["current"],
                "previous_date": date_str(item["previous_date"]),
                "previous": item["previous"],
                "change": item["change"],
            }
            for code, item in sorted(updates.items())
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Fix daily front PE formula caches from the index PE block.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    result = update_daily_pe_front(Path(args.input), Path(args.output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
