from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "（6月11日V12）TREE宏观分析_新增重点跟踪项.xlsx"
OUTPUT = ROOT / "（6月12日V13）TREE宏观分析_重点跟踪项补全.xlsx"
AUDIT = ROOT / "20260612_TREE重点跟踪项补全复核.xlsx"
SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def compact(value: Any) -> str:
    return text(value).replace("\n", "").replace(" ", "")


def is_macro_row(big: str, dim: str) -> bool:
    if big not in {"中国基本面", "美国经济基本面"}:
        return False
    if big == "中国基本面" and dim == "产业":
        return False
    return True


def fill_focus_for_gap(name: str, code: str, caliber: str) -> tuple[str, str]:
    t = compact(f"{name} {code} {caliber}")
    if "货币政策报告" in t:
        return "定性跟踪", "货币政策报告不是连续数值指标，核心是跟踪政策措辞、政策取向和边际表述变化。"
    if any(word in t for word in ["利率", "收益率", "利差", "SOFR", "IORB", "LPR"]):
        return "环比", "利率、收益率和利差类指标，市场通常关注相对上期的边际变化。"
    if any(word in t for word in ["PMI", "ISM", "景气", "新订单", "信心", "乐观指数"]):
        return "环比", "景气类扩散指数通常看本期较上期改善或走弱，而不是同比。"
    return "定性跟踪", "该行缺少可自动更新的数值代码，先作为定性跟踪项处理。"


def build_audit(rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重点跟踪项补全复核"
    headers = [
        "TREE行号",
        "一级分类",
        "维度",
        "子维度",
        "指标名称",
        "指标代码",
        "数据口径",
        "原重点跟踪项",
        "新重点跟踪项",
        "是否补充",
        "判断依据",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["row"],
                row["big"],
                row["dim"],
                row["sub"],
                row["name"],
                row["code"],
                row["caliber"],
                row["old_focus"],
                row["new_focus"],
                "是" if row["changed"] else "否",
                row["reason"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [10, 14, 14, 20, 34, 24, 16, 14, 14, 10, 62]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(AUDIT)


def main() -> None:
    wb = openpyxl.load_workbook(INPUT)
    ws = wb[SHEET]
    header_map = {text(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    focus_col = header_map["重点跟踪项"]

    rows: list[dict[str, Any]] = []
    big = dim = sub = ""
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        if text(ws.cell(row_idx, 1).value):
            big = text(ws.cell(row_idx, 1).value)
        if text(ws.cell(row_idx, 2).value):
            dim = text(ws.cell(row_idx, 2).value)
        if text(ws.cell(row_idx, 3).value):
            sub = text(ws.cell(row_idx, 3).value)

        name = text(ws.cell(row_idx, 4).value)
        if not name or not is_macro_row(big, dim):
            continue

        code = text(ws.cell(row_idx, 10).value)
        caliber = text(ws.cell(row_idx, 11).value)
        old_focus = text(ws.cell(row_idx, focus_col).value)
        new_focus = old_focus
        reason = "原判断已覆盖，保留。"
        changed = False
        if old_focus in {"", "-", "—"}:
            new_focus, reason = fill_focus_for_gap(name, code, caliber)
            ws.cell(row_idx, focus_col).value = new_focus
            changed = True

        rows.append(
            {
                "row": row_idx,
                "big": big,
                "dim": dim,
                "sub": sub,
                "name": name,
                "code": code,
                "caliber": caliber,
                "old_focus": old_focus,
                "new_focus": new_focus,
                "changed": changed,
                "reason": reason,
            }
        )

    wb.save(OUTPUT)
    build_audit(rows)
    changed_rows = [row for row in rows if row["changed"]]
    print(f"saved: {OUTPUT}")
    print(f"audit: {AUDIT}")
    print(f"changed: {len(changed_rows)}")
    for row in changed_rows:
        print(f'{row["row"]}: {row["name"]} -> {row["new_focus"]}')


if __name__ == "__main__":
    main()
