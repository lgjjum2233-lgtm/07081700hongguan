from __future__ import annotations

import argparse
import copy
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


DATE_FIRST_COL = 6
DATE_LAST_COL = 534  # TN, matching the existing front-sheet formulas.
MACRO_SHEET = "宏观数据"
FRONT_SHEETS = ["A股港股", "中债", "中国宏观", "海外数据"]
FISCAL_IMPULSE_CODE = "M0046168-M0046166/M0001395"
CHINA_US_10Y_SPREAD_CODE = "M0325687-G0000891"
PREFERRED_TEMPLATE_ROWS = {
    "A股港股": 15,
    "中债": 24,
    "中国宏观": 68,
    "海外数据": 58,
}

NS = {
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}

for prefix, uri in NS.items():
    ET.register_namespace(prefix, uri)


@dataclass(frozen=True)
class Indicator:
    name: str
    code: str
    front_sheet: str
    freq: str
    unit: str | None
    kind: str = "wind"
    note: str = ""


INDICATORS: list[Indicator] = [
    Indicator("SHIBOR:3个月", "M0017142", "中债", "日度", "%"),
    Indicator("中国:居民部门杠杆率", "M6404533", "中国宏观", "季度", "%"),
    Indicator("中国:政府部门杠杆率", "M6404535", "中国宏观", "季度", "%"),
    Indicator("中国:非金融企业部门杠杆率", "M6404534", "中国宏观", "季度", "%"),
    Indicator("中国:贷款市场报价利率(LPR):1年", "M0096870", "中债", "月度", "%"),
    Indicator("中国:国债收益率:10年", "M0325687", "中债", "日度", "%"),
    Indicator("美国:国债收益率:10年", "G0000891", "海外数据", "日度", "%"),
    Indicator("中美10年期国债利差 (中-美)", CHINA_US_10Y_SPREAD_CODE, "中债", "日度", "%", "formula"),
    Indicator(
        "中国:公共财政收入:累计值",
        "M0046168",
        "中国宏观",
        "月度",
        "亿元",
        note="用户清单写作累计同比，但该代码在旧表中已作为财政收入累计金额使用。",
    ),
    Indicator("中国:公共财政支出:累计同比", "M0046167", "中国宏观", "月度", "%"),
    Indicator("中国:地方政府性基金收入:国有土地使用权出让收入:累计同比", "M0096886", "中国宏观", "月度", "%"),
    Indicator("中国:中央政府债务余额", "M5567950", "中国宏观", "月度", "亿元"),
    Indicator("中国:财政赤字率 (赤字/GDP)", "M5405502", "中国宏观", "年度", "%"),
    Indicator("美国:财政赤字率 (赤字/GDP)", "G1147446", "海外数据", "年度", "%"),
    Indicator("GDP总量：中国", "M5567876", "中国宏观", "季度", "亿元"),
    Indicator("中国:GDP累计同比贡献率:最终消费支出:支出法", "M6001128", "中国宏观", "季度", "%"),
    Indicator("中国:GDP累计同比贡献率:资本形成总额:支出法", "M6001129", "中国宏观", "季度", "%"),
    Indicator("中国:GDP累计同比贡献率:货物和服务净出口:支出法", "M6001130", "中国宏观", "季度", "%"),
    Indicator("中国:房地产开发投资完成额:累计同比", "S0029657", "中国宏观", "月度", "%"),
    Indicator("中国:固定资产投资完成额:制造业:累计同比", "M0000357", "中国宏观", "月度", "%"),
    Indicator("中国:固定资产投资完成额:基础设施建设投资:累计同比", "M5440435", "中国宏观", "月度", "%"),
    Indicator("中国:出口金额:美国:累计同比", "M0008499", "中国宏观", "月度", "%"),
    Indicator("中国:出口金额:东南亚国家联盟:累计同比", "M0007911", "中国宏观", "月度", "%"),
    Indicator("进出口：中国:贸易差额", "M0000610", "中国宏观", "月度", "亿美元"),
    Indicator("中国:农产品批发价格指数:猪肉", "V6842305", "中国宏观", "日度", "点"),
    Indicator("期货结算价(活跃合约):布伦特原油", "S0031525", "海外数据", "日度", "美元/桶"),
    Indicator("中国:产成品存货:规模以上工业企业:同比", "M0000561", "中国宏观", "月度", "%"),
    Indicator("中国:制造业PMI", "M0017126", "中国宏观", "月度", "%"),
    Indicator("中国:PPI:全部工业品:当月同比", "M0001227", "中国宏观", "月度", "%"),
    Indicator("中国:利润总额:制造业:累计同比", "S0206721", "中国宏观", "月度", "%"),
]


def norm(value: Any) -> str:
    return str(value or "").strip().upper()


def formula_text(value: Any) -> str:
    return value.text if isinstance(value, ArrayFormula) else str(value or "")


def copy_cell(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.protection = copy.copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        copy_cell(ws.cell(src_row, col), ws.cell(dst_row, col))
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    ws.row_dimensions[dst_row].hidden = ws.row_dimensions[src_row].hidden
    ws.row_dimensions[dst_row].outlineLevel = ws.row_dimensions[src_row].outlineLevel


def copy_row_all(ws, src_row: int, dst_row: int, max_col: int) -> None:
    copy_row_style(ws, src_row, dst_row, max_col)
    for col in range(1, max_col + 1):
        ws.cell(dst_row, col).value = ws.cell(src_row, col).value


def clear_row_values(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def find_row_by_code(ws, code: str, code_col: int = 5) -> int | None:
    target = norm(code)
    for row in range(1, ws.max_row + 1):
        if norm(ws.cell(row, code_col).value) == target:
            return row
    return None


def all_front_rows_by_code(wb) -> dict[str, tuple[str, int]]:
    out: dict[str, tuple[str, int]] = {}
    for sheet in FRONT_SHEETS:
        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            code = norm(ws.cell(row, 3).value)
            if code and code not in out:
                out[code] = (sheet, row)
    return out


def replace_macro_row_refs(wb, old_row: int, new_row: int) -> int:
    old = str(old_row)
    new = str(new_row)
    replaced = 0
    patterns = [
        (f"{MACRO_SHEET}!$F${old}:$TN${old}", f"{MACRO_SHEET}!$F${new}:$TN${new}"),
        (f"{MACRO_SHEET}!$F${old},", f"{MACRO_SHEET}!$F${new},"),
        (f"COLUMN({MACRO_SHEET}!$F${old})", f"COLUMN({MACRO_SHEET}!$F${new})"),
        (f"COLUMN({MACRO_SHEET}!$F${old}:$TN${old})", f"COLUMN({MACRO_SHEET}!$F${new}:$TN${new})"),
        (f"{MACRO_SHEET}!$F${old}:$TR${old}", f"{MACRO_SHEET}!$F${new}:$TR${new}"),
        (f"COLUMN({MACRO_SHEET}!$F${old}:$TR${old})", f"COLUMN({MACRO_SHEET}!$F${new}:$TR${new})"),
    ]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                new_value = value
                for src, dst in patterns:
                    new_value = new_value.replace(src, dst)
                if new_value != value:
                    cell.value = new_value
                    replaced += 1
    return replaced


def update_wsd_formula(macro, raw_end_row: int) -> str:
    cell = macro["F2"]
    text = formula_text(cell.value)
    text = re.sub(r"E2:E\d+", f"E2:E{raw_end_row}", text)
    text = re.sub(r"rows=\d+", f"rows={raw_end_row - 1}", text)
    if isinstance(cell.value, ArrayFormula):
        cell.value.text = text
    else:
        cell.value = text
    return text


def front_formulas(front_row: int, src_row: int, divisor: float = 1.0) -> dict[int, str]:
    if divisor == 1:
        current_tail = ""
        previous_tail = ""
    else:
        current_tail = f"/{int(divisor) if divisor.is_integer() else divisor}"
        previous_tail = current_tail
    return {
        5: f'=IFERROR(LOOKUP(9.99E+307,{MACRO_SHEET}!$F${src_row}:$TN${src_row}){current_tail},"")',
        6: f'=IFERROR(LOOKUP(2,1/({MACRO_SHEET}!$F${src_row}:$TN${src_row}<>""),{MACRO_SHEET}!$F$1:$TN$1),"")',
        7: f'=IFERROR(E{front_row}-H{front_row},"")',
        8: (
            f'=IFERROR(LOOKUP(9.99E+307,OFFSET({MACRO_SHEET}!$F${src_row},0,0,1,'
            f'LOOKUP(2,1/({MACRO_SHEET}!$F${src_row}:$TN${src_row}<>""),'
            f'COLUMN({MACRO_SHEET}!$F${src_row}:$TN${src_row}))-COLUMN({MACRO_SHEET}!$F${src_row})))'
            f'{previous_tail},"")'
        ),
        9: (
            f'=IFERROR(LOOKUP(2,1/({MACRO_SHEET}!$F${src_row}:$TN${src_row}<>"")/'
            f'({MACRO_SHEET}!$F$1:$TN$1<F{front_row}),{MACRO_SHEET}!$F$1:$TN$1),"")'
        ),
    }


def macro_formula_row_spread(macro, target_row: int, china_row: int, us_row: int) -> None:
    for col in range(DATE_FIRST_COL, DATE_LAST_COL + 1):
        letter = openpyxl.utils.get_column_letter(col)
        macro.cell(target_row, col).value = (
            f'=IFERROR(IF(OR({letter}{china_row}="",{letter}{us_row}=""),"",'
            f'{letter}{china_row}-{letter}{us_row}),"")'
        )


def update_fiscal_impulse_formula_row(macro, row: int, revenue_row: int, spending_row: int, gdp_row: int) -> None:
    for col in range(DATE_FIRST_COL, DATE_LAST_COL + 1):
        letter = openpyxl.utils.get_column_letter(col)
        macro.cell(row, col).value = (
            f'=IFERROR(IF(OR({letter}{revenue_row}="",{letter}{spending_row}=""),"",'
            f'({letter}{revenue_row}-{letter}{spending_row})/LOOKUP(2,1/((YEAR($F$1:{letter}$1)=YEAR({letter}$1)-1)*'
            f'($F${gdp_row}:{letter}${gdp_row}<>"")),$F${gdp_row}:{letter}${gdp_row})*100),"")'
        )


def chart_by_row(ws, row: int):
    for chart in ws._charts:
        anchor = getattr(chart, "anchor", None)
        if hasattr(anchor, "_from") and anchor._from.row + 1 == row:
            return chart
    return None


def clone_template_chart(ws, macro, front_row: int, src_row: int) -> bool:
    template = chart_by_row(ws, PREFERRED_TEMPLATE_ROWS.get(ws.title, 0))
    for chart in ws._charts:
        anchor = getattr(chart, "anchor", None)
        if not hasattr(anchor, "_from"):
            continue
        if template is None and anchor._from.row + 1 < front_row:
            template = chart
    if template is None:
        return False
    chart = copy.deepcopy(template)
    if hasattr(chart.anchor, "_from"):
        old_from = template.anchor._from
        old_to = template.anchor.to
        row_delta = front_row - (old_from.row + 1)
        chart.anchor._from.row = old_from.row + row_delta
        chart.anchor.to.row = old_to.row + row_delta
    cats_ref = f"'{MACRO_SHEET}'!$F$1:$TN$1"
    vals_ref = f"'{MACRO_SHEET}'!$F${src_row}:$TN${src_row}"
    for series in chart.series:
        try:
            series.val.numRef.f = vals_ref
        except Exception:
            pass
        try:
            series.cat.numRef.f = cats_ref
        except Exception:
            try:
                series.cat.strRef.f = cats_ref
            except Exception:
                pass
        try:
            series.graphicalProperties.line.solidFill = "C00000"
            series.graphicalProperties.line.width = 15120
            series.marker.symbol = "none"
        except Exception:
            pass
    try:
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
    except Exception:
        pass
    ws.add_chart(chart)
    return True


def normalize_front_charts(wb) -> None:
    for sheet in FRONT_SHEETS:
        ws = wb[sheet]
        for chart in ws._charts:
            try:
                chart.x_axis.majorGridlines = None
                chart.y_axis.majorGridlines = None
            except Exception:
                pass
            try:
                chart.graphical_properties.noFill = True
                chart.graphical_properties.line.noFill = True
            except Exception:
                pass
            try:
                chart.plot_area.graphicalProperties.noFill = True
                chart.plot_area.graphicalProperties.line.noFill = True
            except Exception:
                pass
            for series in chart.series:
                try:
                    series.graphicalProperties.line.solidFill = "C00000"
                    series.graphicalProperties.line.width = 15120
                    series.marker.symbol = "none"
                except Exception:
                    pass


def remove_children(parent: ET.Element, names: set[str]) -> None:
    for child in list(parent):
        if child.tag in names:
            parent.remove(child)


def ensure_child(parent: ET.Element, tag: str) -> ET.Element:
    child = parent.find(tag, NS)
    if child is None:
        child = ET.SubElement(parent, f"{{{NS[tag.split(':', 1)[0]]}}}{tag.split(':', 1)[1]}")
    return child


def set_shape_no_fill(sppr: ET.Element) -> None:
    fills = {
        f"{{{NS['a']}}}solidFill",
        f"{{{NS['a']}}}gradFill",
        f"{{{NS['a']}}}blipFill",
        f"{{{NS['a']}}}pattFill",
        f"{{{NS['a']}}}grpFill",
    }
    remove_children(sppr, fills)
    if sppr.find("a:noFill", NS) is None:
        sppr.insert(0, ET.Element(f"{{{NS['a']}}}noFill"))
    line = sppr.find("a:ln", NS)
    if line is None:
        line = ET.SubElement(sppr, f"{{{NS['a']}}}ln")
    remove_children(line, fills)
    if line.find("a:noFill", NS) is None:
        line.insert(0, ET.Element(f"{{{NS['a']}}}noFill"))


def set_series_line(line: ET.Element) -> None:
    line.attrib["w"] = "15120"
    fills = {
        f"{{{NS['a']}}}noFill",
        f"{{{NS['a']}}}solidFill",
        f"{{{NS['a']}}}gradFill",
        f"{{{NS['a']}}}blipFill",
        f"{{{NS['a']}}}pattFill",
        f"{{{NS['a']}}}grpFill",
    }
    remove_children(line, fills)
    solid = ET.SubElement(line, f"{{{NS['a']}}}solidFill")
    ET.SubElement(solid, f"{{{NS['a']}}}srgbClr", {"val": "C00000"})


def standardize_chart_xmls(path: Path) -> None:
    tmp = path.with_suffix(".tmp.xlsx")
    with ZipFile(path, "r") as zin, ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.startswith("xl/charts/chart") and info.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for grid in root.findall(".//c:majorGridlines", NS):
                    parent = root.find(".//c:valAx", NS)
                    if parent is not None and grid in list(parent):
                        parent.remove(grid)
                chart_space_sppr = root.find("c:spPr", NS)
                if chart_space_sppr is None:
                    chart_space_sppr = ET.SubElement(root, f"{{{NS['c']}}}spPr")
                set_shape_no_fill(chart_space_sppr)
                plot_area = root.find(".//c:plotArea", NS)
                if plot_area is not None:
                    plot_sppr = plot_area.find("c:spPr", NS)
                    if plot_sppr is None:
                        plot_sppr = ET.SubElement(plot_area, f"{{{NS['c']}}}spPr")
                    set_shape_no_fill(plot_sppr)
                for line in root.findall(".//c:ser/c:spPr/a:ln", NS):
                    set_series_line(line)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            zout.writestr(info, data)
    tmp.replace(path)


def add_front_row(wb, item: Indicator, src_row: int) -> tuple[str, int, bool]:
    ws = wb[item.front_sheet]
    front_row = ws.max_row + 1
    copy_row_style(ws, front_row - 1, front_row, ws.max_column)
    for col in range(1, ws.max_column + 1):
        ws.cell(front_row, col).value = None
    ws.cell(front_row, 1).value = item.name
    ws.cell(front_row, 2).value = item.freq
    ws.cell(front_row, 3).value = item.code
    ws.cell(front_row, 4).value = item.unit
    for col, formula in front_formulas(front_row, src_row).items():
        ws.cell(front_row, col).value = formula
    chart_added = clone_template_chart(ws, wb[MACRO_SHEET], front_row, src_row)
    return item.front_sheet, front_row, chart_added


def build_indices(wb) -> tuple[dict[str, int], dict[str, tuple[str, int]]]:
    macro = wb[MACRO_SHEET]
    macro_rows: dict[str, int] = {}
    for row in range(1, macro.max_row + 1):
        code = norm(macro.cell(row, 5).value)
        if code and code not in macro_rows:
            macro_rows[code] = row
    return macro_rows, all_front_rows_by_code(wb)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    report_path = Path(args.report)

    wb = openpyxl.load_workbook(input_path, data_only=False)
    macro = wb[MACRO_SHEET]
    macro_max_col = max(macro.max_column, DATE_LAST_COL)

    macro_rows, front_rows = build_indices(wb)
    fiscal_old_row = find_row_by_code(macro, FISCAL_IMPULSE_CODE)
    if fiscal_old_row is None:
        raise RuntimeError(f"Cannot find {FISCAL_IMPULSE_CODE}")

    wind_to_add: list[Indicator] = []
    formula_to_add: list[Indicator] = []
    front_to_add: list[Indicator] = []
    skipped: list[dict[str, Any]] = []

    for item in INDICATORS:
        if item.kind == "formula":
            if norm(item.code) not in macro_rows:
                formula_to_add.append(item)
            if norm(item.code) not in front_rows:
                front_to_add.append(item)
            else:
                skipped.append({"code": item.code, "name": item.name, "reason": "front_exists", "where": front_rows[norm(item.code)]})
            continue
        if norm(item.code) not in macro_rows:
            wind_to_add.append(item)
        if norm(item.code) not in front_rows:
            front_to_add.append(item)
        else:
            skipped.append({"code": item.code, "name": item.name, "reason": "front_exists", "where": front_rows[norm(item.code)]})

    old_raw_end = fiscal_old_row - 1
    new_raw_end = old_raw_end + len(wind_to_add)
    formula_start = new_raw_end + 1
    fiscal_new_row = formula_start + len(formula_to_add)

    # Move the fiscal formula row to the new tail, then reuse the opened rows.
    copy_row_all(macro, fiscal_old_row, fiscal_new_row, macro_max_col)
    for row in range(fiscal_old_row, fiscal_new_row):
        copy_row_style(macro, old_raw_end, row, macro_max_col)
        clear_row_values(macro, row, macro_max_col)

    source_rows: dict[str, int] = dict(macro_rows)
    added_macro: list[dict[str, Any]] = []

    for offset, item in enumerate(wind_to_add):
        row = fiscal_old_row + offset
        macro.cell(row, 4).value = item.name
        macro.cell(row, 5).value = item.code
        source_rows[norm(item.code)] = row
        added_macro.append({"code": item.code, "name": item.name, "row": row, "kind": "wind"})

    for offset, item in enumerate(formula_to_add):
        row = formula_start + offset
        macro.cell(row, 4).value = item.name
        macro.cell(row, 5).value = item.code
        source_rows[norm(item.code)] = row
        if item.code == CHINA_US_10Y_SPREAD_CODE:
            china_row = source_rows.get("M0325687")
            us_row = source_rows.get("G0000891")
            if not china_row or not us_row:
                raise RuntimeError("Missing source rows for China-US 10Y spread")
            macro_formula_row_spread(macro, row, china_row, us_row)
        added_macro.append({"code": item.code, "name": item.name, "row": row, "kind": "formula"})

    revenue_row = source_rows.get("M0046168")
    spending_row = source_rows.get("M0046166")
    gdp_row = source_rows.get("M0001395")
    if revenue_row and spending_row and gdp_row:
        update_fiscal_impulse_formula_row(macro, fiscal_new_row, revenue_row, spending_row, gdp_row)
    source_rows[norm(FISCAL_IMPULSE_CODE)] = fiscal_new_row

    updated_f2 = update_wsd_formula(macro, new_raw_end)
    replaced_refs = replace_macro_row_refs(wb, fiscal_old_row, fiscal_new_row)

    added_front: list[dict[str, Any]] = []
    # Rebuild after macro additions, but keep front row state from the original workbook for de-duplication.
    for item in front_to_add:
        src_row = source_rows.get(norm(item.code))
        if src_row is None:
            raise RuntimeError(f"No macro source row for {item.code}")
        sheet, row, chart_added = add_front_row(wb, item, src_row)
        added_front.append(
            {
                "code": item.code,
                "name": item.name,
                "sheet": sheet,
                "row": row,
                "macro_row": src_row,
                "chart_added": chart_added,
                "note": item.note,
            }
        )

    normalize_front_charts(wb)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    standardize_chart_xmls(output_path)

    verify = openpyxl.load_workbook(output_path, data_only=False)
    verify_macro = verify[MACRO_SHEET]
    verify_front_rows = all_front_rows_by_code(verify)
    errors: list[str] = []

    f2_after = formula_text(verify_macro["F2"].value)
    if f"E2:E{new_raw_end}" not in f2_after or f"rows={new_raw_end - 1}" not in f2_after:
        errors.append("F2 WSD range was not updated as expected")
    for item in wind_to_add:
        row = find_row_by_code(verify_macro, item.code)
        if row is None or row > new_raw_end:
            errors.append(f"Wind code not inside WSD raw range: {item.code}")
    for item in formula_to_add:
        row = find_row_by_code(verify_macro, item.code)
        if row is None or row <= new_raw_end:
            errors.append(f"Formula code row placement invalid: {item.code}")
    fiscal_verify_row = find_row_by_code(verify_macro, FISCAL_IMPULSE_CODE)
    if fiscal_verify_row != fiscal_new_row:
        errors.append(f"Fiscal impulse row mismatch: expected {fiscal_new_row}, got {fiscal_verify_row}")
    for item in front_to_add:
        if norm(item.code) not in verify_front_rows:
            errors.append(f"Missing front row: {item.code}")
    chart_counts = {sheet: len(verify[sheet]._charts) for sheet in FRONT_SHEETS}
    sheet_rows = {sheet: verify[sheet].max_row for sheet in FRONT_SHEETS}

    report = {
        "input": str(input_path),
        "output": str(output_path),
        "added_macro": added_macro,
        "added_front": added_front,
        "skipped_existing_front": skipped,
        "wind_rows_added": len(wind_to_add),
        "formula_rows_added": len(formula_to_add),
        "front_rows_added": len(added_front),
        "fiscal_old_row": fiscal_old_row,
        "fiscal_new_row": fiscal_new_row,
        "new_raw_end": new_raw_end,
        "f2_formula": f2_after,
        "formula_refs_replaced": replaced_refs,
        "chart_counts": chart_counts,
        "sheet_rows": sheet_rows,
        "errors": errors,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if errors:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
