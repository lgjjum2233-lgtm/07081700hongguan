from __future__ import annotations

import argparse
import copy
import json
import math
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl


TREE_SHEET = "重点策略跟踪情况(V2.5)"
RED = "FFFF0000"
GREEN = "FF008000"
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {
    "main": SHEET_NS,
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

ET.register_namespace("", SHEET_NS)
ET.register_namespace("r", NS["officeRel"])


def q(tag: str) -> str:
    return f"{{{SHEET_NS}}}{tag}"


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def date_key(value: Any) -> tuple[int, Any]:
    if isinstance(value, datetime):
        return 2, datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return 2, datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        return 2, datetime(1899, 12, 30) + timedelta(days=float(value))
    if value is None:
        return 0, ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return 2, datetime.strptime(text, fmt)
        except ValueError:
            pass
    return (1 if text else 0), text


def date_text(value: Any) -> str:
    kind, key = date_key(value)
    if kind == 2:
        return key.strftime("%Y-%m-%d")
    return str(key) if key else ""


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib["name"] == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def cell_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    return openpyxl.utils.column_index_from_string(letters)


def find_row(sheet_root: ET.Element, row_num: int) -> ET.Element:
    sheet_data = sheet_root.find("main:sheetData", NS)
    if sheet_data is None:
        raise KeyError("sheetData")
    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib["r"]) == row_num:
            return row
    new_row = ET.Element(q("row"), {"r": str(row_num)})
    sheet_data.append(new_row)
    return new_row


def find_cell(sheet_root: ET.Element, row_num: int, col_num: int) -> ET.Element:
    row = find_row(sheet_root, row_num)
    ref = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell
    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        if cell_col_index(cell.attrib["r"]) > col_num:
            insert_at = idx
            break
    row.insert(insert_at, new_cell)
    return new_cell


def font_rgb(cell) -> str | None:
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return f"{color.type}:{color.value}"


def set_font_rgb(font: ET.Element, rgb: str) -> ET.Element:
    font = copy.deepcopy(font)
    for color in list(font.findall("main:color", NS)):
        font.remove(color)
    color_el = ET.Element(q("color"), {"rgb": rgb})
    insert_at = len(list(font))
    for idx, child in enumerate(list(font)):
        if child.tag.rsplit("}", 1)[-1] in {"sz", "u", "vertAlign", "scheme"}:
            insert_at = idx
            break
    font.insert(insert_at, color_el)
    return font


def ensure_colored_style(
    styles_root: ET.Element,
    base_style_id: int,
    rgb: str,
    cache: dict[tuple[int, str], int],
) -> int:
    key = (base_style_id, rgb)
    if key in cache:
        return cache[key]
    fonts = styles_root.find("main:fonts", NS)
    cell_xfs = styles_root.find("main:cellXfs", NS)
    if fonts is None or cell_xfs is None:
        raise KeyError("styles")
    xfs = cell_xfs.findall("main:xf", NS)
    base_xf = xfs[base_style_id] if 0 <= base_style_id < len(xfs) else xfs[0]
    font_id = int(base_xf.attrib.get("fontId", "0"))
    font_nodes = fonts.findall("main:font", NS)
    base_font = font_nodes[font_id] if 0 <= font_id < len(font_nodes) else font_nodes[0]
    fonts.append(set_font_rgb(base_font, rgb))
    new_font_id = len(font_nodes)
    fonts.attrib["count"] = str(len(font_nodes) + 1)
    new_xf = copy.deepcopy(base_xf)
    new_xf.attrib["fontId"] = str(new_font_id)
    new_xf.attrib["applyFont"] = "1"
    cell_xfs.append(new_xf)
    new_style_id = len(xfs)
    cell_xfs.attrib["count"] = str(len(xfs) + 1)
    cache[key] = new_style_id
    return new_style_id


def latest_nonzero_date_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    wf = openpyxl.load_workbook(path, data_only=False)
    ws = wb[TREE_SHEET]
    wsf = wf[TREE_SHEET]
    latest_key: tuple[int, Any] = (0, "")
    latest = ""
    for row in range(6, ws.max_row + 1):
        change = to_number(ws.cell(row, 13).value)
        if change is None or change == 0:
            continue
        key = date_key(ws.cell(row, 12).value)
        if key > latest_key:
            latest_key = key
            latest = date_text(ws.cell(row, 12).value)
    rows = []
    for row in range(6, ws.max_row + 1):
        change = to_number(ws.cell(row, 13).value)
        if change is None or change == 0:
            continue
        if date_text(ws.cell(row, 12).value) != latest:
            continue
        rgb = font_rgb(wsf.cell(row, 13))
        expected = RED if change > 0 else GREEN
        rows.append(
            {
                "row": row,
                "name": ws.cell(row, 4).value,
                "code": ws.cell(row, 10).value,
                "date": latest,
                "change": ws.cell(row, 13).value,
                "current_rgb": rgb,
                "expected_rgb": expected,
                "needs_fix": rgb != expected,
            }
        )
    return latest, rows


def apply_fix(input_path: Path, output_path: Path) -> dict[str, Any]:
    latest, rows = latest_nonzero_date_rows(input_path)
    target_rows = {item["row"]: item for item in rows if item["needs_fix"]}
    cache: dict[tuple[int, str], int] = {}
    with ZipFile(input_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)
        styles_root = read_xml(zin, "xl/styles.xml")
        for row, item in target_rows.items():
            cell = find_cell(sheet_root, row, 13)
            base_style_id = int(cell.attrib.get("s", "0"))
            cell.attrib["s"] = str(ensure_colored_style(styles_root, base_style_id, item["expected_rgb"], cache))
        sheet_xml = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
        styles_xml = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, sheet_xml)
                elif name == "xl/styles.xml":
                    zout.writestr(name, styles_xml)
                else:
                    zout.writestr(name, zin.read(name))
    return {
        "input": str(input_path),
        "output": str(output_path),
        "latest_nonzero_date": latest,
        "latest_date_rows": len(rows),
        "fixed_count": len(target_rows),
        "fixed_rows": list(target_rows.values()),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    result = apply_fix(Path(args.input), Path(args.output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
