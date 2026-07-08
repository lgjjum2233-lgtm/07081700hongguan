from __future__ import annotations

import argparse
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from sync_tree_new_indicators_from_daily import (
    TREE_SHEET,
    find_cell,
    find_existing_cell,
    read_xml,
    set_inline_string,
    set_number,
    workbook_sheet_path,
    write_xml,
)


METADATA = {
    19: {
        "weight": 0.012,
        "reason": "3个月SHIBOR反映银行中期同业负债成本，比隔夜和7天利率更能观察资金面压力向银行负债端的传导；但它仍是报价型利率，受同业结构和期限溢价影响，因此作为DR007、R007后的辅助期限指标。",
    },
    27: {
        "weight": 0.012,
        "reason": "居民部门杠杆率反映居民资产负债表扩张空间和消费、地产信用约束，对中期需求有指示意义；但季度更新且属于慢变量，权重低于社融、居民贷款和M1等更敏感的信用指标。",
    },
    28: {
        "weight": 0.014,
        "reason": "政府部门杠杆率衡量财政加杠杆空间和稳增长约束，能辅助判断基建、财政支出和政府债融资的持续性；但更新频率低，更多用于确认财政周期背景。",
    },
    29: {
        "weight": 0.018,
        "reason": "非金融企业部门杠杆率反映企业资产负债表压力和再融资空间，对投资扩张、信用风险和盈利修复都有中期影响，因此重要性高于居民杠杆率；但它仍是季度慢变量，不能替代月度信用数据。",
    },
    30: {
        "weight": 0.018,
        "reason": "1年LPR是企业和居民贷款定价的重要基准，直接影响实体融资成本和信用扩张意愿；但调整频率不高，更多用于判断政策利率向贷款端的传导效果。",
    },
    32: {
        "weight": 0.012,
        "reason": "中国10年期国债收益率是国内长期无风险利率锚，能反映增长预期、通胀预期和配置需求变化；在本模块中主要用于补充观察中长期利率环境。",
    },
    33: {
        "weight": 0.012,
        "reason": "美国10年期国债收益率代表全球美元长期利率锚，会影响人民币资产估值、汇率压力和跨境资金偏好；但它不是中国本土流动性变量，因此作为外部约束指标给予中低权重。",
    },
    34: {
        "weight": 0.016,
        "reason": "中美10年期国债利差直接衡量中美长期利率相对吸引力，对汇率、外资配置和跨境资金流向有较强解释力；相比单独看两国利率，利差更适合做外部流动性约束判断。",
    },
    36: {
        "weight": 0.014,
        "reason": "公共财政收入累计值反映财政收入端的资源基础，能观察经济活动和税收恢复情况；但收入本身不等于政策发力，需结合支出、赤字和财政存款一起判断，因此权重低于财政支出和财政脉冲。",
    },
    37: {
        "weight": 0.018,
        "reason": "公共财政支出累计同比直接反映财政资金投放力度，是判断财政托底和稳增长节奏的重要月度指标；相比收入端，它对当期需求的边际影响更直接。",
    },
    38: {
        "weight": 0.012,
        "reason": "土地出让收入累计同比反映地方政府性基金收入和地产链景气，对地方财政、基建配套和地产政策空间有辅助意义；但受土地成交节奏扰动较大，因此权重低于一般公共财政支出。",
    },
    39: {
        "weight": 0.010,
        "reason": "中央政府债务余额用于观察中央财政加杠杆程度和债务扩张空间，但它是余额型指标，边际政策信号弱于赤字率、政府债发行和财政支出，因此作为背景约束指标。",
    },
    40: {
        "weight": 0.014,
        "reason": "中国财政赤字率衡量财政扩张力度相对经济总量的强弱，是判断政策托底空间的重要年度指标；但发布频率低，短期交易信号弱于月度财政收支。",
    },
    41: {
        "weight": 0.008,
        "reason": "美国财政赤字率影响美债供给、美元利率和全球流动性环境，对国内资产属于外部背景变量；由于传导链条较长，在中国政策模块中只作为辅助观察。",
    },
    45: {
        "weight": 0.0222222222222222,
        "reason": "中国GDP总量是经济规模和产出水平的总口径，能校准增长判断的基础盘；但季度发布且偏滞后，方向判断仍需依赖社零、投资、出口、PMI等更高频指标。",
    },
    46: {
        "weight": 0.0333333333333333,
        "reason": "中国GDP同比是宏观增长最核心的总量指标之一，能直接验证经济周期方向和政策成效；但季度发布滞后，适合作为增长模块的锚，而不是高频边际信号。",
    },
    47: {
        "weight": 0.0166666666666667,
        "reason": "最终消费支出贡献率用于拆解GDP增长来源，能判断增长是否由内需和居民消费驱动；它是结构验证指标，重要性低于消费零售等更高频数据。",
    },
    48: {
        "weight": 0.0166666666666667,
        "reason": "资本形成总额贡献率用于观察投资对GDP增长的拉动，是判断基建、制造业和地产投资合力的结构指标；由于季度更新，主要用于验证投资链条结论。",
    },
    49: {
        "weight": 0.0111111111111111,
        "reason": "货物和服务净出口贡献率衡量外需对GDP增长的贡献，能补充出口链条判断；但贡献率受进口、价格和汇率共同影响，权重低于直接出口同比和贸易差额。",
    },
    51: {
        "weight": 0.0222222222222222,
        "reason": "房地产开发投资累计同比是地产链景气和固定资产投资的重要分项，对信用需求、地方财政和周期行业影响较大；在增长模块中属于地产方向的核心拆解指标。",
    },
    52: {
        "weight": 0.018,
        "reason": "制造业投资累计同比反映企业资本开支意愿和产业升级动能，是固定资产投资中更市场化的分项；用于补充判断实体投资扩张是否具备内生动力。",
    },
    53: {
        "weight": 0.018,
        "reason": "基础设施建设投资累计同比反映政策性投资和稳增长发力强度，是财政政策向实体需求传导的重要观察口径；但需结合政府债、财政支出和项目落地验证。",
    },
    57: {
        "weight": 0.0111111111111111,
        "reason": "对美国出口累计同比反映外需中美国方向的景气变化，对制造业订单和外需结构有参考价值；但单一地区口径窄于总出口，因此作为出口总指标的辅助拆解。",
    },
    58: {
        "weight": 0.0111111111111111,
        "reason": "对东盟出口累计同比反映区域贸易和产业链转移背景下的外需韧性，是出口结构的重要补充；但仍属于分地区指标，权重低于出口总额和贸易差额。",
    },
    59: {
        "weight": 0.014,
        "reason": "贸易差额直接反映净出口对外汇收入和总需求的影响，能辅助判断外需贡献和汇率基本面；但它同时受进口收缩、价格和汇率影响，需与出口、进口分项一起看。",
    },
    67: {
        "weight": 0.0111111111111111,
        "reason": "猪肉批发价格指数是中国CPI食品项的重要高频线索，能提前观察居民通胀压力；但它只覆盖单一品类，不能代表整体价格周期，因此作为CPI的辅助观察。",
    },
    68: {
        "weight": 0.0166666666666667,
        "reason": "布伦特原油价格是全球能源成本和输入性通胀的重要来源，会影响PPI、CPI、企业利润和风险偏好；但短期受地缘和供给扰动较大，需与国内价格指标一起判断。",
    },
    78: {
        "weight": 0.035,
        "reason": "A股全部上市公司ROE(TTM)衡量上市公司整体盈利质量和资本回报能力，是连接宏观增长、价格利润和权益估值的关键指标；但披露频率低、滞后于价格和景气指标，因此权重低于高频增长和利润数据。",
    },
    79: {
        "weight": 0.018,
        "reason": "产成品存货同比反映工业企业库存周期位置，能判断需求改善是主动补库还是被动累库；它对周期行业和利润弹性有解释力，但需要结合PMI、PPI和利润一起确认。",
    },
    80: {
        "weight": 0.018,
        "reason": "制造业PMI是行业景气的领先调查指标，发布时间早，能快速捕捉订单、生产和预期变化；这里作为行业景气拆解，权重低于增长模块中的PMI核心项。",
    },
    81: {
        "weight": 0.015,
        "reason": "PPI当月同比反映工业品价格和企业收入端变化，是制造业利润的重要前置变量；由于价格模块已有PPI核心指标，这里作为行业价格确认项，避免重复计数。",
    },
    82: {
        "weight": 0.025,
        "reason": "制造业利润总额累计同比直接反映制造业盈利修复质量，是判断企业扩产、投资和权益盈利的重要指标；相比价格和库存，它更接近最终利润结果，因此在产业观察中权重较高。",
    },
    137: {
        "weight": 0.012,
        "reason": "美国GDP总量用于观察美国经济规模和产出水平，是总需求背景指标；但它是季度总量口径，边际变化不如GDP同比、零售销售和就业数据敏感，因此只作辅助锚定。",
    },
    138: {
        "weight": 0.020,
        "reason": "美国GDP同比衡量美国经济增长的总方向，是验证软着陆、衰退和政策效果的核心季度指标；但表内已有GDP年化增速口径，为避免重复计数，这里给中低权重。",
    },
}


def cell_style(sheet_root: ET.Element, row: int, col: int, fallback: int | None = None) -> int | None:
    ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
    cell = find_existing_cell(sheet_root, ref)
    if cell is not None and "s" in cell.attrib:
        return int(cell.attrib["s"])
    return fallback


def apply_metadata(input_path: Path, output_path: Path) -> dict:
    with ZipFile(input_path, "r") as zin:
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)

        default_weight_style = cell_style(sheet_root, 65, 5)
        default_reason_style = cell_style(sheet_root, 65, 6)
        updated = []

        for row, item in METADATA.items():
            weight_style = cell_style(sheet_root, row, 5, default_weight_style)
            reason_style = cell_style(sheet_root, row, 6, default_reason_style)
            if row == 78:
                weight_style = default_weight_style
                reason_style = default_reason_style

            set_number(find_cell(sheet_root, row, 5), item["weight"], weight_style)
            set_inline_string(find_cell(sheet_root, row, 6), item["reason"], reason_style)
            updated.append(row)

        sheet_xml = write_xml(sheet_root)
        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_path:
                    zout.writestr(name, sheet_xml)
                else:
                    zout.writestr(name, zin.read(name))

    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[TREE_SHEET]
    rows = []
    missing = []
    for row in sorted(METADATA):
        record = {
            "row": row,
            "indicator": ws.cell(row, 4).value,
            "weight": ws.cell(row, 5).value,
            "reason": ws.cell(row, 6).value,
        }
        if record["weight"] is None or not record["reason"]:
            missing.append(record)
        rows.append(record)
    return {"output": str(output_path), "updated_rows": updated, "missing_after": missing, "rows": rows}


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill weights and weight reasons for newly added TREE indicators.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    result = apply_metadata(Path(args.input), Path(args.output))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
