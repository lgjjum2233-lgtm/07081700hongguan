from __future__ import annotations

import argparse
import json
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

TOOL_DIR = Path(__file__).resolve().parent
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))
WORKSPACE = TOOL_DIR.parents[1]
LEGACY_TOOLS = WORKSPACE / "codex_tmp" / "20260604_tree_update"
if str(LEGACY_TOOLS) not in sys.path:
    sys.path.insert(0, str(LEGACY_TOOLS))

import sync_tree_20260604 as base  # type: ignore  # noqa: E402
from sync_tree_new_indicators_from_daily import (  # noqa: E402
    NS,
    find_cell,
    find_existing_cell,
    read_xml,
    set_inline_string,
    workbook_sheet_path,
    write_xml,
)
from sync_tree_v30_full import build_sources  # noqa: E402


TREE_SHEET = "重点策略跟踪情况(V3.0)"
HEADER_ROW = 5
NAME_COL = 4
CODE_COL = 10
CALIBER_COL = 15  # O; keep existing J:N data/chart columns stable.


@dataclass
class CaliberResult:
    caliber: str
    basis: str
    confidence: str


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_code(value: Any) -> str:
    return as_text(value).upper()


def cell_style(sheet_root: ET.Element, row: int, col: int, fallback: int | None = None) -> int | None:
    cell = find_existing_cell(sheet_root, f"{openpyxl.utils.get_column_letter(col)}{row}")
    if cell is not None and "s" in cell.attrib:
        return int(cell.attrib["s"])
    return fallback


def build_macro_standard_names(daily_path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(daily_path, data_only=False, read_only=True)
    names: dict[str, str] = {}
    if "宏观数据" in wb.sheetnames:
        ws = wb["宏观数据"]
        for row in range(2, ws.max_row + 1):
            code = norm_code(ws.cell(row, 5).value)
            name = as_text(ws.cell(row, 4).value)
            if code and name and code not in names:
                names[code] = name
    wb.close()
    return names


def build_source_index(daily_path: Path) -> dict[str, Any]:
    by_code, _, _ = build_sources(daily_path, __import__("datetime").date(2026, 6, 11))
    return by_code


def has_formula_operator(code: str) -> bool:
    return bool(code and any(op in code for op in ["+", "/", "*"])) or bool(re.search(r"[A-Z]\d+[-][A-Z]\d+", code))


def result(caliber: str, basis: str, confidence: str = "高") -> CaliberResult:
    return CaliberResult(caliber, basis, confidence)


def classify(name: str, code: str, source_name: str, unit: str = "") -> CaliberResult:
    text = f"{source_name} {name}".upper()
    human_text = f"{source_name} {name}"
    code = code.upper()

    if not code or code in {"—", "-", "NONE"}:
        return result("定性指标", "TREE无有效Wind代码，按定性观察项处理", "中")

    if has_formula_operator(code):
        if "利差" in human_text:
            return result("利差", "指标代码为表达式且名称含“利差”", "高")
        if "赤字脉冲" in human_text or "赤字" in human_text:
            return result("衍生比值/脉冲", "指标代码为表达式，按财政收入、支出和GDP计算", "高")
        if "总投放" in human_text or "合计" in human_text:
            return result("合成流量值", "指标代码为多个Wind代码合成", "高")
        return result("衍生计算值", "指标代码为表达式，非单一Wind原始指标", "中")

    rules: list[tuple[str, str, str]] = [
        ("万得一致预测", "一致预期值", "日报宏观数据标准名称含“万得一致预测”"),
        ("同比多增", "同比多增", "日报宏观数据标准名称含“同比多增”"),
        ("同比少增", "同比多增", "日报宏观数据标准名称含“同比少增/同比多增”"),
        ("累计同比", "累计同比", "日报宏观数据标准名称含“累计同比”"),
        ("当月同比", "当月同比", "日报宏观数据标准名称含“当月同比”"),
        ("同比", "同比", "日报宏观数据标准名称含“同比”"),
        ("环比", "环比", "日报宏观数据标准名称含“环比”"),
        ("累计值", "累计值", "日报宏观数据标准名称含“累计值”"),
        ("当月值", "当月值", "日报宏观数据标准名称含“当月值”"),
        ("余额", "余额/存量", "日报宏观数据标准名称含“余额”"),
        ("净投放", "净投放", "日报宏观数据标准名称含“净投放”"),
        ("净流入", "净流量", "日报宏观数据标准名称含“净流入”"),
        ("差额", "差额/净额", "日报宏观数据标准名称含“差额”"),
        ("贸易差额", "差额/净额", "日报宏观数据标准名称含“贸易差额”"),
        ("利差", "利差", "日报宏观数据标准名称或TREE名称含“利差”"),
        ("利率", "利率水平", "日报宏观数据标准名称含“利率”"),
        ("收益率", "收益率水平", "日报宏观数据标准名称含“收益率”"),
        ("杠杆率", "比例水平", "日报宏观数据标准名称含“杠杆率”"),
        ("赤字率", "比例水平", "日报宏观数据标准名称含“赤字率”"),
        ("贡献率", "比例水平", "日报宏观数据标准名称含“贡献率”"),
        ("换手率", "比例水平", "日报宏观数据标准名称含“换手率”"),
        ("市盈率", "估值倍数", "日报宏观数据标准名称含“市盈率”"),
        ("ROE", "盈利比例", "指标名称含“ROE”"),
        ("PMI", "扩散指数", "指标名称含“PMI”"),
        ("ISM", "扩散指数", "指标名称含“ISM”"),
        ("指数", "指数水平", "日报宏观数据标准名称含“指数”"),
        ("价格指数", "指数水平", "日报宏观数据标准名称含“价格指数”"),
        ("现价", "名义现值", "日报宏观数据标准名称含“现价”"),
        ("不变价", "实际值/实际增速", "日报宏观数据标准名称含“不变价”，需结合是否含同比判断"),
        ("成交额", "成交金额", "指标名称含“成交额”"),
        ("交易总金额", "成交金额", "指标名称含“交易总金额”"),
        ("持仓", "持仓规模", "指标名称含“持仓”"),
        ("市值", "规模现值", "指标名称含“市值”"),
        ("销售额", "销售金额", "指标名称含“销售额”"),
        ("收入", "金额现值", "指标名称含“收入”且不含同比"),
        ("支出", "金额现值", "指标名称含“支出”且不含同比"),
        ("总额", "金额现值", "指标名称含“总额”且不含同比"),
        ("总量", "规模现值", "指标名称含“总量”"),
        ("资产", "资产规模/余额", "指标名称含“资产”"),
        ("GDP", "规模/增速", "指标名称含“GDP”，需结合标准名称细分"),
        ("债务", "余额/存量", "指标名称含“债务”"),
        ("融资", "流量值", "指标名称含“融资”"),
        ("新增", "流量值", "指标名称含“新增”"),
    ]
    for key, caliber, basis in rules:
        if key.upper() in text:
            if key in {"收入", "支出", "总额"} and "同比" in human_text:
                continue
            return result(caliber, basis, "高" if source_name else "中")

    if unit == "%":
        return result("比例/增速类", "单位为%，但标准名称未明确同比/环比，需人工复核", "中")
    return result("水平值/现值", "未命中同比、环比、累计、余额等关键词，按水平值/现值处理", "中")


def set_dimension(sheet_root: ET.Element, max_row: int) -> None:
    dim = sheet_root.find("main:dimension", NS)
    if dim is not None:
        dim.attrib["ref"] = f"A1:O{max_row}"


def ensure_column_width(sheet_root: ET.Element) -> None:
    cols = sheet_root.find("main:cols", NS)
    if cols is None:
        sheet_data = sheet_root.find("main:sheetData", NS)
        cols = ET.Element(f"{{{NS['main']}}}cols")
        if sheet_data is not None:
            sheet_root.insert(list(sheet_root).index(sheet_data), cols)
        else:
            sheet_root.insert(0, cols)
    for col in cols.findall("main:col", NS):
        if col.attrib.get("min") == "15" and col.attrib.get("max") == "15":
            col.attrib["width"] = "16"
            col.attrib["customWidth"] = "1"
            return
    cols.append(ET.Element(f"{{{NS['main']}}}col", {"min": "15", "max": "15", "width": "16", "customWidth": "1"}))


def build_records(tree_path: Path, daily_path: Path) -> list[dict[str, Any]]:
    macro_names = build_macro_standard_names(daily_path)
    by_code = build_source_index(daily_path)
    wb = openpyxl.load_workbook(tree_path, data_only=True, read_only=True)
    ws = wb[TREE_SHEET]
    records = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        tree_name = as_text(ws.cell(row, NAME_COL).value)
        code = norm_code(ws.cell(row, CODE_COL).value)
        if not tree_name and not code:
            continue
        source = by_code.get(code)
        source_name = macro_names.get(code) or as_text(getattr(source, "name", ""))
        unit = as_text(getattr(source, "unit", ""))
        cal = classify(tree_name, code, source_name, unit)
        records.append(
            {
                "row": row,
                "tree_name": tree_name,
                "code": code,
                "daily_standard_name": source_name,
                "unit": unit,
                "data_caliber": cal.caliber,
                "basis": cal.basis,
                "confidence": cal.confidence,
                "source_type": "日报宏观数据标准名称" if code in macro_names else ("日报前置/派生数据" if source else "未匹配到底层"),
            }
        )
    wb.close()
    return records


def write_tree_column(input_tree: Path, output_tree: Path, records: list[dict[str, Any]]) -> None:
    record_by_row = {int(item["row"]): item for item in records}
    with ZipFile(input_tree, "r") as zin:
        names = zin.namelist()
        sheet_path = workbook_sheet_path(zin, TREE_SHEET)
        sheet_root = read_xml(zin, sheet_path)
        header_style = cell_style(sheet_root, HEADER_ROW, 14, cell_style(sheet_root, HEADER_ROW, 9))
        body_style = cell_style(sheet_root, HEADER_ROW + 1, 9)

        set_inline_string(find_cell(sheet_root, HEADER_ROW, CALIBER_COL), "数据口径", header_style)
        max_row = max(record_by_row) if record_by_row else HEADER_ROW
        for row, item in record_by_row.items():
            set_inline_string(find_cell(sheet_root, row, CALIBER_COL), item["data_caliber"], cell_style(sheet_root, row, 9, body_style))

        set_dimension(sheet_root, max_row)
        ensure_column_width(sheet_root)
        replacements = {sheet_path: write_xml(sheet_root)}
        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, replacements[name] if name in replacements else zin.read(name))


def write_audit(audit_path: Path, records: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据口径判定复核"
    headers = ["TREE行号", "TREE名称", "指标代码", "日报标准名称", "单位", "数据口径", "判断依据", "可信度", "来源类型"]
    ws.append(headers)
    for item in records:
        ws.append(
            [
                item["row"],
                item["tree_name"],
                item["code"],
                item["daily_standard_name"],
                item["unit"],
                item["data_caliber"],
                item["basis"],
                item["confidence"],
                item["source_type"],
            ]
        )
    widths = [10, 34, 24, 42, 12, 16, 48, 10, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb.save(audit_path)


def verify(output_tree: Path, records: list[dict[str, Any]]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(output_tree, data_only=True, read_only=True)
    ws = wb[TREE_SHEET]
    missing = []
    for item in records:
        row = int(item["row"])
        if as_text(ws.cell(row, CALIBER_COL).value) != item["data_caliber"]:
            missing.append({"row": row, "expected": item["data_caliber"], "actual": ws.cell(row, CALIBER_COL).value})
    header = ws.cell(HEADER_ROW, CALIBER_COL).value
    wb.close()
    return {"header": header, "record_count": len(records), "mismatch": missing}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--audit-output", required=True)
    args = parser.parse_args()

    tree_path = Path(args.tree)
    daily_path = Path(args.daily)
    output_tree = Path(args.output_tree)
    audit_output = Path(args.audit_output)

    records = build_records(tree_path, daily_path)
    write_tree_column(tree_path, output_tree, records)
    write_audit(audit_output, records)
    check = verify(output_tree, records)
    result = {
        "tree": str(tree_path),
        "daily": str(daily_path),
        "output_tree": str(output_tree),
        "audit_output": str(audit_output),
        **check,
        "summary_by_caliber": {},
    }
    for item in records:
        result["summary_by_caliber"][item["data_caliber"]] = result["summary_by_caliber"].get(item["data_caliber"], 0) + 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if check["mismatch"] or check["header"] != "数据口径":
        raise SystemExit(2)


if __name__ == "__main__":
    main()
