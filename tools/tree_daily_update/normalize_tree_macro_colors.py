from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

import update_tree_v3_latest_layout as tree


def choose_series(
    code: str,
    daily_series: dict[str, tree.Series],
    monthly_series: dict[str, tree.Series],
) -> tree.Series | None:
    candidates = [
        series
        for series in (daily_series.get(code), monthly_series.get(code))
        if series is not None and series.points
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda series: series.points[-1][0])


def fill_rgb(cell: object) -> str | None:
    color = cell.fill.fgColor
    if color.type != "rgb" or not color.rgb or color.rgb in {"00000000", "FFFFFFFF"}:
        return None
    return color.rgb[-6:].upper()


def normalized_rgb(value: str | None) -> str | None:
    return value[-6:].upper() if value else None


def expected_styles(ws: object, row: int, series: tree.Series) -> list[tuple[str | None, str]]:
    name = ws.cell(row, tree.MASTER_COLS["name"]).value
    current_date = tree.parse_date(ws.cell(row, tree.MASTER_COLS["date"]).value)
    if current_date is None:
        return []
    current_display = ws.cell(row, tree.MASTER_COLS["current"]).value
    change_display = ws.cell(row, tree.MASTER_COLS["change"]).value
    current_point = tree.closest_point(series.points, current_date)
    display_scale = tree.raw_display_scale(current_display, current_point[1]) if current_point else 1.0
    yoy_value, yoy_hist = tree.yoy_value_and_history(series, current_date, current_display, name)
    yoy_inc_value, yoy_inc_hist = tree.yoy_increment_and_history(series, current_date, current_display)
    change_series = [(d, v * display_scale) for d, v in series.points if d <= current_date]
    return [
        tree.color_by_position(series.points, current_date)[:2],
        tree.color_by_abs_change(change_series, change_display, current_date)[:2],
        tree.color_by_position(yoy_hist, current_date)[:2],
        tree.color_by_abs_change(yoy_inc_hist, yoy_inc_value, current_date)[:2],
    ]


def run(tree_path: Path, daily_path: Path, monthly_path: Path, output_path: Path, as_of: date) -> dict[str, object]:
    daily_series = tree.read_full_series(daily_path, as_of)
    monthly_series = tree.read_full_series(monthly_path, as_of)
    wb = openpyxl.load_workbook(tree_path, data_only=True, read_only=True)
    sheet_name = tree.find_sheet_name(wb, "(V3.0)")
    ws = wb[sheet_name]

    replacements: dict[str, bytes] = {}
    normalized = 0
    source_counts = {"daily": 0, "monthly": 0}

    with ZipFile(tree_path, "r") as zin:
        styles_root = tree.read_xml(zin, "xl/styles.xml")
        sheet_path = tree.workbook_sheet_path(zin, sheet_name)
        sheet_root = tree.read_xml(zin, sheet_path)
        master_context = {col: "" for col in (1, 2, 3, 5)}

        for row in range(2, ws.max_row + 1):
            style_cache: dict[tuple[int, str, str], int] = {}
            name = ws.cell(row, tree.MASTER_COLS["name"]).value
            code = tree.norm_code(ws.cell(row, tree.MASTER_COLS["code"]).value)
            for col in master_context:
                value = tree.as_text(ws.cell(row, col).value)
                if value:
                    master_context[col] = value
            row_text = " ".join(
                [*(master_context[col] for col in (1, 2, 3, 5)), tree.as_text(name)]
            )
            if any(word in row_text for word in tree.NON_MACRO_SCOPE_WORDS):
                for col in (
                    tree.MASTER_COLS["current"],
                    tree.MASTER_COLS["change"],
                    tree.MASTER_COLS["yoy"],
                    tree.MASTER_COLS["yoy_inc"],
                ):
                    tree.set_style(sheet_root, styles_root, row, col, None, tree.BLACK, style_cache)
                continue
            if not tree.SIMPLE_WIND_CODE_RE.fullmatch(code):
                continue

            daily = daily_series.get(code)
            monthly = monthly_series.get(code)
            series = choose_series(code, daily_series, monthly_series)
            if series is None or series.kind != "macro":
                continue
            if daily is series:
                source_counts["daily"] += 1
            elif monthly is series:
                source_counts["monthly"] += 1

            current_date = tree.parse_date(ws.cell(row, tree.MASTER_COLS["date"]).value)
            if current_date is None or not tree.is_macro_row(code, row_text, {code: series}):
                continue

            styles = expected_styles(ws, row, series)
            for col, (fill, font) in zip(
                (
                    tree.MASTER_COLS["current"],
                    tree.MASTER_COLS["change"],
                    tree.MASTER_COLS["yoy"],
                    tree.MASTER_COLS["yoy_inc"],
                ),
                styles,
            ):
                tree.set_style(sheet_root, styles_root, row, col, fill, font, style_cache)
            normalized += 1

        replacements["xl/styles.xml"] = tree.write_xml(styles_root)
        replacements[sheet_path] = tree.write_xml(sheet_root)

        display_sheet = tree.find_sheet_name(wb, "(V3)")
        display_ws = wb[display_sheet]
        display_path = tree.workbook_sheet_path(zin, display_sheet)
        display_root = tree.read_xml(zin, display_path)
        display_context = {col: "" for col in (6, 7, 8)}
        for row in range(6, display_ws.max_row + 1):
            for col in display_context:
                value = tree.as_text(display_ws.cell(row, col).value)
                if value:
                    display_context[col] = value
            row_text = " ".join(
                [
                    *(display_context[col] for col in (6, 7, 8)),
                    tree.as_text(display_ws.cell(row, tree.DISPLAY_COLS["name"]).value),
                ]
            )
            if not any(word in row_text for word in tree.NON_MACRO_SCOPE_WORDS):
                continue
            row_cache: dict[tuple[int, str, str], int] = {}
            for col in (tree.DISPLAY_COLS["current"], tree.DISPLAY_COLS["change"]):
                tree.set_style(display_root, styles_root, row, col, None, tree.BLACK, row_cache)
        replacements["xl/styles.xml"] = tree.write_xml(styles_root)
        replacements[display_path] = tree.write_xml(display_root)

        with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
            for member in zin.namelist():
                if member in replacements:
                    zout.writestr(member, replacements[member])
                else:
                    zout.writestr(member, zin.read(member))

    wb.close()

    check_wb = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
    check_ws = check_wb[sheet_name]
    repair_rows: list[int] = []
    check_context = {col: "" for col in (1, 2, 3, 5)}
    for row in range(2, check_ws.max_row + 1):
        code = tree.norm_code(check_ws.cell(row, tree.MASTER_COLS["code"]).value)
        for col in check_context:
            value = tree.as_text(check_ws.cell(row, col).value)
            if value:
                check_context[col] = value
        row_text = " ".join(
            [
                *(check_context[col] for col in (1, 2, 3, 5)),
                tree.as_text(check_ws.cell(row, tree.MASTER_COLS["name"]).value),
            ]
        )
        if any(word in row_text for word in tree.NON_MACRO_SCOPE_WORDS):
            continue
        if not tree.SIMPLE_WIND_CODE_RE.fullmatch(code):
            continue
        series = choose_series(code, daily_series, monthly_series)
        if series is None or series.kind != "macro":
            continue
        styles = expected_styles(check_ws, row, series)
        expected = [normalized_rgb(fill) for fill, _font in styles]
        actual = [
            fill_rgb(check_ws.cell(row, col))
            for col in (
                tree.MASTER_COLS["current"],
                tree.MASTER_COLS["change"],
                tree.MASTER_COLS["yoy"],
                tree.MASTER_COLS["yoy_inc"],
            )
        ]
        if expected != actual:
            repair_rows.append(row)
    check_wb.close()

    if repair_rows:
        repair_wb = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
        repair_ws = repair_wb[sheet_name]
        repair_replacements: dict[str, bytes] = {}
        with ZipFile(output_path, "r") as zin:
            styles_root = tree.read_xml(zin, "xl/styles.xml")
            sheet_path = tree.workbook_sheet_path(zin, sheet_name)
            sheet_root = tree.read_xml(zin, sheet_path)
            for row in repair_rows:
                code = tree.norm_code(repair_ws.cell(row, tree.MASTER_COLS["code"]).value)
                series = choose_series(code, daily_series, monthly_series)
                if series is None:
                    continue
                styles = expected_styles(repair_ws, row, series)
                row_cache: dict[tuple[int, str, str], int] = {}
                for col, (fill, font) in zip(
                    (
                        tree.MASTER_COLS["current"],
                        tree.MASTER_COLS["change"],
                        tree.MASTER_COLS["yoy"],
                        tree.MASTER_COLS["yoy_inc"],
                    ),
                    styles,
                ):
                    tree.set_style(sheet_root, styles_root, row, col, fill, font, row_cache)
            repair_replacements["xl/styles.xml"] = tree.write_xml(styles_root)
            repair_replacements[sheet_path] = tree.write_xml(sheet_root)
            repair_path = output_path.with_name(output_path.stem + ".repaired.xlsx")
            with ZipFile(repair_path, "w", ZIP_DEFLATED) as zout:
                for member in zin.namelist():
                    if member in repair_replacements:
                        zout.writestr(member, repair_replacements[member])
                    else:
                        zout.writestr(member, zin.read(member))
        repair_wb.close()
        repair_path.replace(output_path)

    return {
        "input": str(tree_path),
        "output": str(output_path),
        "as_of": as_of.isoformat(),
        "normalized_rows": normalized,
        "repaired_rows": repair_rows,
        "source_counts": source_counts,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize TREE macro colors from refreshed source histories.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--monthly", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of", default=date.today().isoformat())
    args = parser.parse_args()
    result = run(
        Path(args.tree),
        Path(args.daily),
        Path(args.monthly),
        Path(args.output),
        datetime.strptime(args.as_of, "%Y-%m-%d").date(),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
