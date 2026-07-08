from __future__ import annotations

import json
from pathlib import Path

import openpyxl


TREE_PATH = Path("（6月9日）TREE宏观分析数据资讯新增指标.xlsx")
DAILY_PATH = Path("20260609日报数据版V2_新增指标Wind刷新_图表缓存修正.xlsx")
TREE_SHEET = "重点策略跟踪情况(V3.0)"
FRONT_SHEETS = ["A股港股", "中债", "中国宏观", "海外数据"]

ROW_CODE = {
    19: "M0017142",
    27: "M6404533",
    28: "M6404535",
    29: "M6404534",
    30: "M0096870",
    32: "M0325687",
    33: "G0000891",
    34: "M0325687-G0000891",
    36: "M0046168",
    37: "M0046167",
    38: "M0096886",
    39: "M5567950",
    40: "M5405502",
    41: "G1147446",
    45: "M5567876",
    47: "M6001128",
    48: "M6001129",
    49: "M6001130",
    51: "S0029657",
    52: "M0000357",
    53: "M5440435",
    57: "M0008499",
    58: "M0007911",
    59: "M0000610",
    67: "V6842305",
    68: "S0031525",
    79: "M0000561",
    80: "M0017126",
    81: "M0001227",
    82: "S0206721",
    138: "G1112986",
}


def txt(value):
    return "" if value is None else str(value).strip()


def build_daily_index():
    wb = openpyxl.load_workbook(DAILY_PATH, data_only=True)
    index = {}
    for sheet_name in FRONT_SHEETS:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            code = txt(ws.cell(row, 3).value).upper()
            if not code:
                continue
            index.setdefault(
                code,
                {
                    "sheet": sheet_name,
                    "row": row,
                    "name": ws.cell(row, 1).value,
                    "freq": ws.cell(row, 2).value,
                    "unit": ws.cell(row, 4).value,
                    "current": ws.cell(row, 5).value,
                    "date": ws.cell(row, 6).value,
                    "change": ws.cell(row, 7).value,
                    "previous": ws.cell(row, 8).value,
                    "previous_date": ws.cell(row, 9).value,
                },
            )
    return index


def main():
    daily_index = build_daily_index()
    tree_wb = openpyxl.load_workbook(TREE_PATH, data_only=False)
    ws = tree_wb[TREE_SHEET]
    out = []
    for row, code in ROW_CODE.items():
        source = daily_index.get(code.upper())
        out.append(
            {
                "row": row,
                "tree_name": ws.cell(row, 4).value,
                "expected_code": code,
                "tree_freq": ws.cell(row, 9).value,
                "tree_code": ws.cell(row, 10).value,
                "tree_current": ws.cell(row, 11).value,
                "tree_date": ws.cell(row, 12).value,
                "tree_change": ws.cell(row, 13).value,
                "fmt_current": ws.cell(row, 11).number_format,
                "fmt_change": ws.cell(row, 13).number_format,
                "source": source,
            }
        )
    missing_rows = [46, 78, 137]
    missing = [
        {
            "row": row,
            "tree_name": ws.cell(row, 4).value,
            "tree_freq": ws.cell(row, 9).value,
            "tree_code": ws.cell(row, 10).value,
        }
        for row in missing_rows
    ]
    print(json.dumps({"mapped": out, "no_daily_source_rows": missing}, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
