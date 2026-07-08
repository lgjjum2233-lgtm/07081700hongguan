from __future__ import annotations

import argparse
import ast
import json
import math
import re
import xml.etree.ElementTree as ET
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils.datetime import from_excel

import apply_macro_change_history_colors as style


PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

ET.register_namespace("", style.SHEET_NS)
ET.register_namespace("r", style.REL_NS)

LOOKBACK_DAYS = 365
MIN_PRIOR_DIFFS = 3
MIN_PERCENTILE_TO_FILL = 0.35
AUDIT_SHEET_NAME = "近一年边际变化复核"


@dataclass(frozen=True)
class Point:
    when: datetime
    value: float
    source_col: int


@dataclass(frozen=True)
class ChangePoint:
    when: datetime
    value: float
    previous_when: datetime
    previous_value: float
    change: float


@dataclass(frozen=True)
class MacroSource:
    name: str
    code: str
    row: int
    points: tuple[Point, ...]
    changes: tuple[ChangePoint, ...]
    match_note: str = ""

    @property
    def current(self) -> float | None:
        return self.points[-1].value if self.points else None

    @property
    def data_date(self) -> datetime | None:
        return self.points[-1].when if self.points else None

    @property
    def previous(self) -> float | None:
        return self.points[-2].value if len(self.points) >= 2 else None

    @property
    def previous_date(self) -> datetime | None:
        return self.points[-2].when if len(self.points) >= 2 else None

    @property
    def latest_change(self) -> float | None:
        return self.changes[-1].change if self.changes else None


def as_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        try:
            parsed = from_excel(value)
            return datetime(parsed.year, parsed.month, parsed.day)
        except Exception:
            return None
    text = style.as_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return datetime(parsed.year, parsed.month, parsed.day)
        except ValueError:
            continue
    return None


def fmt_date(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y-%m-%d")


def fmt_num(value: Any) -> Any:
    number = style.to_number(value)
    if number is None:
        return "" if value is None else value
    if abs(number) >= 100:
        return round(number, 4)
    return round(number, 8)


def percentile(sorted_values: list[float], p: float) -> float | None:
    if not sorted_values:
        return None
    if len(sorted_values) == 1:
        return sorted_values[0]
    pos = (len(sorted_values) - 1) * p
    lower = math.floor(pos)
    upper = math.ceil(pos)
    if lower == upper:
        return sorted_values[lower]
    return sorted_values[lower] + (sorted_values[upper] - sorted_values[lower]) * (pos - lower)


def build_change_points(points: list[Point]) -> tuple[ChangePoint, ...]:
    changes: list[ChangePoint] = []
    for idx in range(1, len(points)):
        prev = points[idx - 1]
        cur = points[idx]
        changes.append(
            ChangePoint(
                when=cur.when,
                value=cur.value,
                previous_when=prev.when,
                previous_value=prev.value,
                change=cur.value - prev.value,
            )
        )
    return tuple(changes)


def prefer_source(old: MacroSource | None, new: MacroSource) -> MacroSource:
    if old is None:
        return new
    old_date = old.data_date or datetime.min
    new_date = new.data_date or datetime.min
    if new_date != old_date:
        return new if new_date > old_date else old
    if len(new.points) != len(old.points):
        return new if len(new.points) > len(old.points) else old
    return old


def build_macro_sources(daily_path: Path, as_of: date) -> tuple[dict[str, MacroSource], dict[str, list[MacroSource]], list[MacroSource]]:
    wb = openpyxl.load_workbook(daily_path, data_only=True, read_only=True)
    if "宏观数据" not in wb.sheetnames:
        raise RuntimeError("日报 workbook 中没有找到“宏观数据”sheet。")
    ws = wb["宏观数据"]

    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    date_cols: list[tuple[int, datetime]] = []
    for col, value in enumerate(header_values, start=1):
        if col < 6:
            continue
        parsed = as_date(value)
        if parsed is not None and parsed.date() <= as_of:
            date_cols.append((col, parsed))
    date_col_indexes = [(col - 1, col, parsed) for col, parsed in date_cols]

    sources: list[MacroSource] = []
    for row, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = style.as_text(row_values[3] if len(row_values) >= 4 else None)
        code = style.as_text(row_values[4] if len(row_values) >= 5 else None)
        if not name and not code:
            continue
        points: list[Point] = []
        for value_idx, col, parsed in date_col_indexes:
            if value_idx >= len(row_values):
                continue
            value = style.to_number(row_values[value_idx])
            if value is not None:
                points.append(Point(parsed, value, col))
        points.sort(key=lambda point: (point.when, point.source_col))
        if not points:
            continue
        sources.append(
            MacroSource(
                name=name,
                code=code,
                row=row,
                points=tuple(points),
                changes=build_change_points(points),
            )
        )

    wb.close()

    by_code: dict[str, MacroSource] = {}
    by_name: dict[str, list[MacroSource]] = {}
    for source in sources:
        if source.code:
            key = style.base.code_key(source.code)
            by_code[key] = prefer_source(by_code.get(key), source)
        nkey = style.base.norm(source.name)
        if nkey:
            by_name.setdefault(nkey, []).append(source)
    return by_code, by_name, sources


CODE_TOKEN_RE = re.compile(r"[A-Z][A-Z0-9]*(?:\.[A-Z0-9]+)?")


def safe_eval(node: ast.AST, values: dict[str, float]) -> float:
    if isinstance(node, ast.Expression):
        return safe_eval(node.body, values)
    if isinstance(node, ast.BinOp):
        left = safe_eval(node.left, values)
        right = safe_eval(node.right, values)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            if abs(right) <= 1e-12:
                raise ZeroDivisionError
            return left / right
    if isinstance(node, ast.UnaryOp):
        value = safe_eval(node.operand, values)
        if isinstance(node.op, ast.USub):
            return -value
        if isinstance(node.op, ast.UAdd):
            return value
    if isinstance(node, ast.Name):
        return values[node.id]
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    raise ValueError(f"Unsupported expression node: {type(node).__name__}")


def build_expression_source(expr: str, by_code: dict[str, MacroSource]) -> tuple[MacroSource | None, str]:
    if not any(op in expr for op in "+-*/"):
        return None, ""
    codes = []
    for token in CODE_TOKEN_RE.findall(expr.replace(" ", "")):
        key = style.base.code_key(token)
        if key not in by_code:
            return None, f"表达式代码未匹配: {token}"
        if key not in codes:
            codes.append(key)
    if not codes:
        return None, "表达式中没有识别到代码"

    replaced = expr.replace(" ", "")
    token_names: dict[str, str] = {}
    for idx, code in enumerate(sorted(codes, key=len, reverse=True)):
        var = f"V{idx}"
        token_names[code] = var
        replaced = re.sub(rf"(?<![A-Z0-9.]){re.escape(code)}(?![A-Z0-9.])", var, replaced)
    try:
        parsed = ast.parse(replaced, mode="eval")
    except SyntaxError:
        return None, "表达式无法解析"

    value_maps: dict[str, dict[datetime, float]] = {}
    common_dates: set[datetime] | None = None
    for code, var in token_names.items():
        source = by_code[code]
        current_map: dict[datetime, float] = {point.when: point.value for point in source.points}
        value_maps[var] = current_map
        dates = set(current_map)
        common_dates = dates if common_dates is None else common_dates & dates
    if not common_dates:
        return None, "表达式分项没有共同日期"

    points: list[Point] = []
    for idx, when in enumerate(sorted(common_dates)):
        env = {var: values[when] for var, values in value_maps.items()}
        try:
            value = safe_eval(parsed, env)
        except (ZeroDivisionError, ValueError, KeyError):
            continue
        if math.isfinite(value):
            points.append(Point(when, value, idx))
    if not points:
        return None, "表达式没有可用历史值"
    source = MacroSource(
        name=expr,
        code=expr,
        row=0,
        points=tuple(points),
        changes=build_change_points(points),
        match_note="按宏观数据代码表达式计算",
    )
    return source, "代码表达式"


def resolve_source(
    row_info: dict[str, Any],
    by_code: dict[str, MacroSource],
    by_name: dict[str, list[MacroSource]],
) -> tuple[MacroSource | None, str, str]:
    code = style.as_text(row_info.get("code"))
    name = style.as_text(row_info.get("name"))
    if code and code not in {"—", "-", "None"}:
        expr_source, expr_note = build_expression_source(code, by_code)
        if expr_source is not None:
            return expr_source, "代码表达式", expr_note
        if expr_note:
            key = style.base.code_key(code)
            if key not in by_code:
                return None, "未匹配", expr_note
        source = by_code.get(style.base.code_key(code))
        if source is not None:
            return source, "代码", "按指标代码匹配宏观数据"

    nrow = style.base.norm(name)
    if nrow in by_name:
        return by_name[nrow][0], "名称", "按指标名称精确匹配宏观数据"

    candidates: list[MacroSource] = []
    for nkey, items in by_name.items():
        if len(nkey) < 4:
            continue
        if nkey in nrow or nrow in nkey:
            candidates.extend(items)
    if candidates:
        candidates.sort(key=lambda src: (src.data_date or datetime.min, len(src.points)), reverse=True)
        return candidates[0], "名称模糊", "按指标名称包含关系匹配宏观数据"
    return None, "未匹配", "宏观数据sheet中没有匹配代码或名称"


def analyze_source(source: MacroSource | None) -> dict[str, Any]:
    if source is None:
        return {
            "current_date": "",
            "window_start": "",
            "current": "",
            "previous_date": "",
            "previous": "",
            "latest_change": "",
            "direction": "",
            "prior_count": 0,
            "p50": "",
            "p75": "",
            "p90": "",
            "p95": "",
            "max_abs": "",
            "rank": "",
            "colored": "否",
            "fill_direction": "",
            "intensity": "",
            "fill": "",
            "reason": "未匹配到底层宏观数据",
            "recent_changes": "",
        }

    latest = source.changes[-1] if source.changes else None
    window_start: datetime | None = None
    prior_changes: list[ChangePoint] = []
    if latest is not None:
        window_start = latest.when - timedelta(days=LOOKBACK_DAYS)
        prior_changes = [
            point
            for point in source.changes[:-1]
            if point.when >= window_start and math.isfinite(point.change)
        ]
    prior_abs = sorted(abs(point.change) for point in prior_changes)
    if latest is None:
        rank = None
        intensity = None
        reason = "历史点不足，无法计算边际变化"
    elif abs(latest.change) <= 1e-12:
        rank = 0.0
        intensity = None
        reason = "最新边际变化为0，保持白色"
    elif len(prior_abs) < MIN_PRIOR_DIFFS:
        rank = None
        intensity = None
        reason = f"近一年历史边际变化样本不足{MIN_PRIOR_DIFFS}个，保持白色"
    else:
        rank = bisect_right(prior_abs, abs(latest.change)) / len(prior_abs)
        if rank < MIN_PERCENTILE_TO_FILL:
            intensity = None
            reason = f"近一年历史分位低于{MIN_PERCENTILE_TO_FILL:.0%}，保持白色"
        else:
            intensity = max(0.15, min(1.0, (rank - MIN_PERCENTILE_TO_FILL) / (1.0 - MIN_PERCENTILE_TO_FILL)))
            reason = "达到着色阈值"

    fill = ""
    fill_direction = ""
    colored = "否"
    if latest is not None and intensity is not None:
        if latest.change > 0:
            fill = style.interpolate_color(style.RED_LIGHT, style.RED_DARK, intensity)
            fill_direction = "正增长-红色"
        else:
            fill = style.interpolate_color(style.GREEN_LIGHT, style.GREEN_DARK, intensity)
            fill_direction = "负增长-绿色"
        colored = "是"

    window_changes = [point for point in source.changes if window_start is None or point.when >= window_start]
    recent_changes = " | ".join(f"{fmt_date(point.when)}:{fmt_num(point.change)}" for point in window_changes[-12:])
    return {
        "current_date": fmt_date(source.data_date),
        "window_start": fmt_date(window_start),
        "current": fmt_num(source.current),
        "previous_date": fmt_date(source.previous_date),
        "previous": fmt_num(source.previous),
        "latest_change": fmt_num(latest.change) if latest is not None else "",
        "direction": "正增长" if latest is not None and latest.change > 0 else ("负增长" if latest is not None and latest.change < 0 else "持平"),
        "prior_count": len(prior_abs),
        "p50": fmt_num(percentile(prior_abs, 0.50)),
        "p75": fmt_num(percentile(prior_abs, 0.75)),
        "p90": fmt_num(percentile(prior_abs, 0.90)),
        "p95": fmt_num(percentile(prior_abs, 0.95)),
        "max_abs": fmt_num(max(prior_abs) if prior_abs else None),
        "rank": round(rank, 6) if rank is not None else "",
        "colored": colored,
        "fill_direction": fill_direction,
        "intensity": round(intensity, 6) if intensity is not None else "",
        "fill": fill,
        "reason": reason,
        "recent_changes": recent_changes,
    }


AUDIT_HEADERS = [
    "TREE页",
    "TREE行号",
    "大类",
    "维度",
    "子维度",
    "TREE指标名称",
    "TREE指标代码",
    "TREE边际变化单元格值",
    "匹配方式",
    "宏观数据行号",
    "宏观数据指标名称",
    "宏观数据代码",
    "当前日期",
    "近一年窗口开始",
    "当前值",
    "上期日期",
    "上期值",
    "最新边际变化",
    "方向",
    "近一年样本数(不含本期)",
    "近一年绝对变化P50",
    "近一年绝对变化P75",
    "近一年绝对变化P90",
    "近一年绝对变化P95",
    "近一年最大绝对变化",
    "最新变化近一年分位(不含本期)",
    "是否着色",
    "填充方向",
    "填充强度(0-1)",
    "填充色",
    "留白/着色说明",
    "近一年边际变化(最近12期)",
]


def build_audit_rows(tree_path: Path, daily_path: Path, as_of: date, sheet_name: str) -> tuple[list[dict[str, Any]], dict[int, dict[str, Any]]]:
    by_code, by_name, _sources = build_macro_sources(daily_path, as_of)
    wb_tree = openpyxl.load_workbook(tree_path, data_only=True, read_only=False)
    config = style.SHEET_CONFIGS[sheet_name]
    ws = wb_tree[sheet_name]
    rows = style.iter_indicator_rows(ws, config)

    audit: list[dict[str, Any]] = []
    by_tree_row: dict[int, dict[str, Any]] = {}
    for row_info in rows:
        if not style.include_macro_scope(row_info["big"], row_info["dim"], row_info["sub"]):
            continue
        source, match_type, note = resolve_source(row_info, by_code, by_name)
        analysis = analyze_source(source)
        tree_change = style.to_number(ws.cell(row_info["row"], config.change_col).value)
        item = {
            "sheet": sheet_name,
            "tree_row": row_info["row"],
            "big": row_info["big"],
            "dim": row_info["dim"],
            "sub": row_info["sub"],
            "tree_name": row_info["name"],
            "tree_code": row_info["code"],
            "tree_change": fmt_num(tree_change),
            "match_type": match_type,
            "source_row": source.row if source is not None else "",
            "source_name": source.name if source is not None else "",
            "source_code": source.code if source is not None else "",
            "match_note": note,
            **analysis,
        }
        audit.append(item)
        by_tree_row[row_info["row"]] = item
    wb_tree.close()
    return audit, by_tree_row


def audit_to_sheet_rows(audit: list[dict[str, Any]]) -> list[list[Any]]:
    output = [AUDIT_HEADERS]
    for item in audit:
        output.append(
            [
                item["sheet"],
                item["tree_row"],
                item["big"],
                item["dim"],
                item["sub"],
                item["tree_name"],
                item["tree_code"],
                item["tree_change"],
                item["match_type"],
                item["source_row"],
                item["source_name"],
                item["source_code"],
                item["current_date"],
                item["window_start"],
                item["current"],
                item["previous_date"],
                item["previous"],
                item["latest_change"],
                item["direction"],
                item["prior_count"],
                item["p50"],
                item["p75"],
                item["p90"],
                item["p95"],
                item["max_abs"],
                item["rank"],
                item["colored"],
                item["fill_direction"],
                item["intensity"],
                item["fill"],
                item["reason"],
                item["recent_changes"],
            ]
        )
    return output


def col_letter(index: int) -> str:
    return openpyxl.utils.get_column_letter(index)


def xml_text(value: str) -> str:
    return value


def sheet_cell(row_idx: int, col_idx: int, value: Any) -> ET.Element | None:
    if value is None or value == "":
        return None
    cell = ET.Element(style.q("c"), {"r": f"{col_letter(col_idx)}{row_idx}"})
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        ET.SubElement(cell, style.q("v")).text = str(value)
        return cell
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, style.q("is"))
    text = ET.SubElement(inline, style.q("t"))
    text.text = xml_text(str(value))
    if str(value).strip() != str(value):
        text.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    return cell


def build_audit_sheet_xml(rows: list[list[Any]]) -> bytes:
    max_row = len(rows)
    max_col = max(len(row) for row in rows)
    root = ET.Element(style.q("worksheet"))
    dimension = ET.SubElement(root, style.q("dimension"))
    dimension.attrib["ref"] = f"A1:{col_letter(max_col)}{max_row}"
    sheet_views = ET.SubElement(root, style.q("sheetViews"))
    view = ET.SubElement(sheet_views, style.q("sheetView"), {"workbookViewId": "0"})
    ET.SubElement(view, style.q("pane"), {"ySplit": "1", "topLeftCell": "A2", "activePane": "bottomLeft", "state": "frozen"})
    ET.SubElement(view, style.q("selection"), {"pane": "bottomLeft", "activeCell": "A2", "sqref": "A2"})
    ET.SubElement(root, style.q("sheetFormatPr"), {"defaultRowHeight": "15"})
    cols = ET.SubElement(root, style.q("cols"))
    widths = {
        1: 20,
        2: 8,
        3: 16,
        4: 18,
        5: 18,
        6: 34,
        7: 18,
        8: 14,
        9: 12,
        10: 10,
        11: 34,
        12: 18,
        30: 24,
        31: 90,
    }
    for idx in range(1, max_col + 1):
        width = widths.get(idx, 14)
        ET.SubElement(cols, style.q("col"), {"min": str(idx), "max": str(idx), "width": str(width), "customWidth": "1"})
    sheet_data = ET.SubElement(root, style.q("sheetData"))
    for row_idx, row_values in enumerate(rows, start=1):
        row_el = ET.SubElement(sheet_data, style.q("row"), {"r": str(row_idx)})
        if row_idx == 1:
            row_el.attrib["ht"] = "28"
            row_el.attrib["customHeight"] = "1"
        for col_idx, value in enumerate(row_values, start=1):
            cell = sheet_cell(row_idx, col_idx, value)
            if cell is not None:
                row_el.append(cell)
    ET.SubElement(root, style.q("autoFilter"), {"ref": f"A1:{col_letter(max_col)}{max_row}"})
    ET.SubElement(root, style.q("pageMargins"), {"left": "0.7", "right": "0.7", "top": "0.75", "bottom": "0.75", "header": "0.3", "footer": "0.3"})
    return style.write_xml(root)


def unique_sheet_name(existing: set[str], desired: str) -> str:
    if desired not in existing:
        return desired
    base = desired[:28]
    idx = 2
    while f"{base}{idx}" in existing:
        idx += 1
    return f"{base}{idx}"


def add_audit_sheet(input_daily: Path, output_daily: Path, rows: list[list[Any]], desired_name: str = AUDIT_SHEET_NAME) -> str:
    sheet_xml = build_audit_sheet_xml(rows)
    with ZipFile(input_daily, "r") as zin:
        workbook = style.read_xml(zin, "xl/workbook.xml")
        rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        content_types = ET.fromstring(zin.read("[Content_Types].xml"))

        sheets_el = workbook.find("main:sheets", style.NS)
        if sheets_el is None:
            raise KeyError("workbook sheets")
        existing_names = {sheet.attrib["name"] for sheet in sheets_el.findall("main:sheet", style.NS)}
        sheet_name = unique_sheet_name(existing_names, desired_name)
        existing_sheet_ids = [int(sheet.attrib.get("sheetId", "0")) for sheet in sheets_el.findall("main:sheet", style.NS)]
        next_sheet_id = max(existing_sheet_ids or [0]) + 1

        worksheet_nums = []
        for name in zin.namelist():
            match = re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", name)
            if match:
                worksheet_nums.append(int(match.group(1)))
        next_sheet_num = max(worksheet_nums or [0]) + 1
        new_sheet_path = f"xl/worksheets/sheet{next_sheet_num}.xml"

        rid_nums = []
        for rel in rels:
            match = re.fullmatch(r"rId(\d+)", rel.attrib.get("Id", ""))
            if match:
                rid_nums.append(int(match.group(1)))
        new_rid = f"rId{max(rid_nums or [0]) + 1}"

        ET.SubElement(
            sheets_el,
            style.q("sheet"),
            {"name": sheet_name, "sheetId": str(next_sheet_id), f"{{{style.REL_NS}}}id": new_rid},
        )
        ET.SubElement(
            rels,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": new_rid,
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
                "Target": f"worksheets/sheet{next_sheet_num}.xml",
            },
        )
        ET.SubElement(
            content_types,
            f"{{{CONTENT_TYPES_NS}}}Override",
            {
                "PartName": f"/{new_sheet_path}",
                "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
            },
        )

        replacements = {
            "xl/workbook.xml": style.write_xml(workbook),
            "xl/_rels/workbook.xml.rels": ET.tostring(rels, encoding="utf-8", xml_declaration=True),
            "[Content_Types].xml": ET.tostring(content_types, encoding="utf-8", xml_declaration=True),
            new_sheet_path: sheet_xml,
        }
        with ZipFile(output_daily, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))
            zout.writestr(new_sheet_path, sheet_xml)
    return sheet_name


def apply_tree_colors_from_audit(input_tree: Path, output_tree: Path, audit_by_row: dict[int, dict[str, Any]], sheet_name: str) -> dict[str, Any]:
    wb_tree = openpyxl.load_workbook(input_tree, data_only=True, read_only=False)
    config = style.SHEET_CONFIGS[sheet_name]
    ws = wb_tree[sheet_name]
    indicator_rows = style.iter_indicator_rows(ws, config)

    replacements: dict[str, bytes] = {}
    style_cache: dict[tuple[int, str, str], int] = {}
    styled_rows = 0
    cleared_rows = 0
    with ZipFile(input_tree, "r") as zin:
        styles_root = style.read_xml(zin, "xl/styles.xml")
        sheet_path = style.workbook_sheet_path(zin, sheet_name)
        sheet_root = style.read_xml(zin, sheet_path)
        for row_info in indicator_rows:
            row = row_info["row"]
            cell = style.find_cell(sheet_root, row, config.change_col)
            base_style_id = int(cell.attrib.get("s", "0"))
            item = audit_by_row.get(row)
            fill = item.get("fill") if item else ""
            if fill:
                font_rgb = style.text_color_for_fill(fill)
                styled_rows += 1
                cell.attrib["s"] = str(style.ensure_cell_style(styles_root, base_style_id, fill, font_rgb, style_cache))
            else:
                cleared_rows += 1
                cell.attrib["s"] = str(style.ensure_cell_style(styles_root, base_style_id, None, style.BLACK, style_cache))
        replacements["xl/styles.xml"] = style.write_xml(styles_root)
        replacements[sheet_path] = style.write_xml(sheet_root)
        with ZipFile(output_tree, "w", ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in replacements:
                    zout.writestr(name, replacements[name])
                else:
                    zout.writestr(name, zin.read(name))
    wb_tree.close()
    return {"sheet": sheet_name, "styled_rows": styled_rows, "cleared_rows": cleared_rows}


def run(tree: Path, daily: Path, output_daily: Path, output_tree: Path, as_of: date, sheet_name: str) -> dict[str, Any]:
    audit, audit_by_row = build_audit_rows(tree, daily, as_of, sheet_name)
    sheet_rows = audit_to_sheet_rows(audit)
    audit_sheet_name = add_audit_sheet(daily, output_daily, sheet_rows)
    color_summary = apply_tree_colors_from_audit(tree, output_tree, audit_by_row, sheet_name)
    colored = sum(1 for item in audit if item["colored"] == "是")
    unmatched = sum(1 for item in audit if item["match_type"] == "未匹配")
    insufficient = sum(1 for item in audit if "样本不足" in item["reason"])
    low_rank = sum(1 for item in audit if "历史分位低于" in item["reason"])
    zero = sum(1 for item in audit if "为0" in item["reason"])
    return {
        "tree_input": str(tree),
        "daily_input": str(daily),
        "output_daily": str(output_daily),
        "output_tree": str(output_tree),
        "audit_sheet": audit_sheet_name,
        "tree_sheet": sheet_name,
        "as_of": as_of.isoformat(),
        "lookback_days": LOOKBACK_DAYS,
        "macro_rows": len(audit),
        "colored_rows": colored,
        "unmatched_rows": unmatched,
        "insufficient_history_rows": insufficient,
        "low_rank_rows": low_rank,
        "zero_change_rows": zero,
        "color_summary": color_summary,
        "sample": audit[:12],
    }


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a macro-data-only marginal change audit sheet and recolor TREE.")
    parser.add_argument("--tree", required=True)
    parser.add_argument("--daily", required=True)
    parser.add_argument("--output-daily", required=True)
    parser.add_argument("--output-tree", required=True)
    parser.add_argument("--as-of", default=datetime.today().strftime("%Y-%m-%d"))
    parser.add_argument("--sheet", default="重点策略跟踪情况(V3.0)")
    args = parser.parse_args()

    result = run(
        Path(args.tree),
        Path(args.daily),
        Path(args.output_daily),
        Path(args.output_tree),
        datetime.strptime(args.as_of, "%Y-%m-%d").date(),
        args.sheet,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=json_default))


if __name__ == "__main__":
    main()
