from __future__ import annotations

import argparse
import json
import math
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZipFile

import openpyxl

from postprocess_daily import load_config


ERROR_TERMS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A")
TREE_SHEET = "重点策略跟踪情况(V2.5)"
RED_LINE = "C00000"

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def close_enough(a: Any, b: Any, rel_tol: float = 1e-7, abs_tol: float = 1e-8) -> bool:
    na = to_number(a)
    nb = to_number(b)
    if na is None or nb is None:
        return na is None and nb is None
    return abs(na - nb) <= max(abs_tol, rel_tol * max(1.0, abs(na), abs(nb)))


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        dt = datetime(1899, 12, 30) + timedelta(days=float(value))
        return dt.strftime("%Y-%m-%d")
    text = as_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def is_percent_format(fmt: str) -> bool:
    return "%" in as_text(fmt)


def expected_for_cell(raw: Any, number_format: str, already_decimal: bool = False) -> Any:
    n = to_number(raw)
    if n is None:
        return raw
    if is_percent_format(number_format) and not already_decimal:
        return n / 100
    return n


def read_xml(z: ZipFile, name: str) -> ET.Element:
    return ET.fromstring(z.read(name))


def rel_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(PurePosixPath(base).parent)
    parts: list[str] = []
    for part in (base_dir + "/" + target).split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def workbook_sheet_path(z: ZipFile, sheet_name: str) -> str:
    wb = read_xml(z, "xl/workbook.xml")
    rels = read_xml(z, "xl/_rels/workbook.xml.rels")
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in wb.findall(".//main:sheets/main:sheet", NS):
        if sheet.attrib["name"] == sheet_name:
            rid = sheet.attrib[f"{{{NS['officeRel']}}}id"]
            return rel_target("xl/workbook.xml", rel_map[rid])
    raise KeyError(sheet_name)


def sheet_drawing_path(z: ZipFile, sheet_path: str) -> str | None:
    sheet = read_xml(z, sheet_path)
    drawing = sheet.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib[f"{{{NS['officeRel']}}}id"]
    rels_name = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
    rels = read_xml(z, rels_name)
    for rel in rels:
        if rel.attrib["Id"] == rid:
            return rel_target(sheet_path, rel.attrib["Target"])
    return None


def drawing_rels_path(drawing_path: str) -> str:
    p = PurePosixPath(drawing_path)
    return str(p.parent / "_rels" / (p.name + ".rels"))


def chart_row_paths(tree_path: Path, tree_sheet: str) -> dict[int, str]:
    with ZipFile(tree_path, "r") as z:
        sheet_path = workbook_sheet_path(z, tree_sheet)
        drawing_path = sheet_drawing_path(z, sheet_path)
        if drawing_path is None:
            return {}
        rels = read_xml(z, drawing_rels_path(drawing_path))
        rel_map = {rel.attrib["Id"]: rel_target(drawing_path, rel.attrib["Target"]) for rel in rels}
        drawing = read_xml(z, drawing_path)
        rows: dict[int, str] = {}
        for anchor in list(drawing):
            frm = anchor.find("xdr:from", NS)
            chart_ref = anchor.find(".//a:graphicData/c:chart", NS)
            if frm is None or chart_ref is None:
                continue
            row = int(frm.findtext("xdr:row", namespaces=NS)) + 1
            rid = chart_ref.attrib[f"{{{NS['officeRel']}}}id"]
            rows[row] = rel_map[rid]
        return rows


def chart_values_and_style(tree_path: Path, chart_path: str) -> dict[str, Any]:
    with ZipFile(tree_path, "r") as z:
        root = read_xml(z, chart_path)
    points = []
    for pt in root.findall(".//c:numLit/c:pt", NS) + root.findall(".//c:numCache/c:pt", NS):
        n = to_number(pt.findtext("c:v", namespaces=NS))
        if n is not None:
            points.append(n)
    return {
        "point_count": len(points),
        "last": points[-1] if points else None,
        "srgb": [node.attrib.get("val") for node in root.findall(".//a:srgbClr", NS)],
        "majorGridlines_count": len(root.findall(".//c:majorGridlines", NS)),
        "noFill_count": len(root.findall(".//a:noFill", NS)),
    }


def formula_errors(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(term in value for term in ERROR_TERMS):
                    hits.append({"sheet": ws.title, "cell": cell.coordinate, "value": value})
    return {
        "count": len(hits),
        "by_sheet": dict(Counter(hit["sheet"] for hit in hits)),
        "sample": hits[:20],
    }


def build_sources(daily_path: Path, tree_path: Path, workspace: Path):
    tools = workspace / "codex_tmp" / "20260604_tree_update"
    if str(tools) not in sys.path:
        sys.path.insert(0, str(tools))
    import sync_tree_20260604 as base  # type: ignore
    import sync_tree_20260604_legacy_charts as legacy  # type: ignore

    base.DAILY_PATH = daily_path
    legacy.DAILY_PATH = daily_path
    base.TREE_PATH = tree_path
    legacy.TREE_PATH = tree_path
    return legacy.build_sources(), base


def audit_tree_against_daily(daily_path: Path, tree_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    (row_sources, trends, source_meta), base = build_sources(daily_path, tree_path, Path(config["workspace"]))
    pct_change_codes = {str(code).upper() for code in config.get("stock_index_pct_change_codes", [])}
    tree_sheet = config["tree_sheet"]
    wb = openpyxl.load_workbook(tree_path, data_only=False)
    ws = wb[tree_sheet]
    chart_paths = chart_row_paths(tree_path, tree_sheet)
    value_mismatches = []
    date_mismatches = []
    change_mismatches = []
    missing_charts = []
    chart_last_mismatches = []
    chart_style_issues = []

    for row, source in sorted(row_sources.items()):
        row_name = ws.cell(row, 4).value
        row_code = as_text(ws.cell(row, 10).value)
        current_cell = ws.cell(row, 11)
        expected_current = expected_for_cell(source.current, current_cell.number_format)
        if not close_enough(current_cell.value, expected_current):
            value_mismatches.append(
                {
                    "row": row,
                    "name": row_name,
                    "code": row_code,
                    "tree": current_cell.value,
                    "expected": expected_current,
                    "source": f"{source.sheet}!{source.row}",
                }
            )

        tree_date = date_text(ws.cell(row, 12).value)
        source_date = date_text(source.data_date)
        if tree_date != source_date:
            date_mismatches.append(
                {
                    "row": row,
                    "name": row_name,
                    "code": row_code,
                    "tree": tree_date,
                    "expected": source_date,
                    "source": f"{source.sheet}!{source.row}",
                }
            )

        expected_change = source.change
        already_decimal = False
        if base.stock_index_change_row(as_text(row_name), row_code) or row_code.upper() in pct_change_codes:
            cur = to_number(source.current)
            prev = to_number(source.previous)
            if cur is not None and prev not in (None, 0):
                expected_change = (cur - prev) / prev
                already_decimal = True
        expected_change = expected_for_cell(expected_change, ws.cell(row, 13).number_format, already_decimal=already_decimal)
        if to_number(expected_change) is not None and not close_enough(ws.cell(row, 13).value, expected_change):
            change_mismatches.append(
                {
                    "row": row,
                    "name": row_name,
                    "code": row_code,
                    "tree": ws.cell(row, 13).value,
                    "expected": expected_change,
                    "source": f"{source.sheet}!{source.row}",
                }
            )

    for row, values in sorted(trends.items()):
        if not values:
            continue
        chart_path = chart_paths.get(row)
        if chart_path is None:
            source = row_sources.get(row)
            missing_charts.append(
                {
                    "row": row,
                    "name": ws.cell(row, 4).value,
                    "code": ws.cell(row, 10).value,
                    "source": f"{source.sheet}!{source.row}" if source else None,
                    "trend_points": len(values),
                }
            )
            continue
        chart = chart_values_and_style(tree_path, chart_path)
        if not close_enough(chart["last"], values[-1]):
            chart_last_mismatches.append(
                {
                    "row": row,
                    "name": ws.cell(row, 4).value,
                    "chart_last": chart["last"],
                    "expected": values[-1],
                    "chart_path": chart_path,
                }
            )
        if chart["majorGridlines_count"] != 0 or RED_LINE not in chart["srgb"] or chart["noFill_count"] < 2:
            chart_style_issues.append({"row": row, "name": ws.cell(row, 4).value, **chart})

    return {
        "resolved_rows": len(row_sources),
        "trend_rows": len(trends),
        "tree_chart_rows": len(chart_paths),
        "unmatched_count": source_meta.get("unmatched_count"),
        "unmatched_sample": source_meta.get("unmatched_sample"),
        "value_mismatch_count": len(value_mismatches),
        "value_mismatch_sample": value_mismatches[:20],
        "date_mismatch_count": len(date_mismatches),
        "date_mismatch_sample": date_mismatches[:20],
        "change_mismatch_count": len(change_mismatches),
        "change_mismatch_sample": change_mismatches[:20],
        "missing_chart_count": len(missing_charts),
        "missing_chart_sample": missing_charts[:20],
        "chart_last_mismatch_count": len(chart_last_mismatches),
        "chart_last_mismatch_sample": chart_last_mismatches[:20],
        "chart_style_issue_count": len(chart_style_issues),
        "chart_style_issue_sample": chart_style_issues[:20],
        "tga": {
            "row": 85,
            "current": ws.cell(85, 11).value,
            "date": date_text(ws.cell(85, 12).value),
            "has_chart": 85 in chart_paths,
            "chart": chart_values_and_style(tree_path, chart_paths[85]) if 85 in chart_paths else None,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--daily", required=True)
    parser.add_argument("--tree", required=True)
    parser.add_argument("--config", default=str(Path(__file__).with_name("config.json")))
    parser.add_argument("--output")
    args = parser.parse_args()
    config = load_config(Path(args.config))
    daily_path = Path(args.daily)
    tree_path = Path(args.tree)
    result = {
        "daily": {
            "path": str(daily_path),
            "zip_test": ZipFile(daily_path).testzip(),
            "formula_errors": formula_errors(daily_path),
        },
        "tree": {
            "path": str(tree_path),
            "zip_test": ZipFile(tree_path).testzip(),
            "formula_errors": formula_errors(tree_path),
            "against_daily": audit_tree_against_daily(daily_path, tree_path, config),
        },
    }
    checks = result["tree"]["against_daily"]
    result["summary"] = {
        "pass": (
            result["daily"]["zip_test"] is None
            and result["tree"]["zip_test"] is None
            and checks["value_mismatch_count"] == 0
            and checks["date_mismatch_count"] == 0
            and checks["change_mismatch_count"] == 0
            and checks["missing_chart_count"] == 0
            and checks["chart_last_mismatch_count"] == 0
            and checks["chart_style_issue_count"] == 0
        )
    }
    text = json.dumps(result, ensure_ascii=False, indent=2, default=json_default)
    if args.output:
        Path(args.output).write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
